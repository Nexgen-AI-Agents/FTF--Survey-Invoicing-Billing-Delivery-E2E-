#!/usr/bin/env python3
"""verify-a2-output: Scan data_collected/draft_posted orders for sentinel values."""
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed. Run: pip install openpyxl")

BASE_DIR   = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
EXCEL_PATH = os.path.join(BASE_DIR, "data", "invoice_pipeline_state.xlsx")

SENTINEL_VALUES  = frozenset({"unknown", "n/a", "none", "not available", "not found", ""})
CHECK_STATUSES   = frozenset({"data_collected", "invoice_draft_posted"})


def _is_sentinel(val) -> bool:
    return str(val or "").strip().lower() in SENTINEL_VALUES


def verify() -> tuple[int, list[tuple]]:
    """
    Returns (checked_count, failures).
    failures = [(order_id, status, [bad_field_descriptions]), ...]
    Importable by other skills.
    """
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Excel state not found: {EXCEL_PATH}")

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}

    checked, failures = 0, []
    for row in ws.iter_rows(min_row=2, values_only=True):
        status = str(row[col.get("status", 1)] or "")
        if status not in CHECK_STATUSES:
            continue
        checked += 1
        order_id = str(row[col.get("order_id", 0)] or "")
        client   = row[col.get("client_name", 4)]
        address  = row[col.get("property_address", 5)]

        bad = []
        if _is_sentinel(client):
            bad.append(f"client_name='{client}'")
        if _is_sentinel(address):
            bad.append(f"property_address='{address}'")
        if bad:
            failures.append((order_id, status, bad))

    wb.close()
    return checked, failures


def main():
    try:
        checked, failures = verify()
    except FileNotFoundError as exc:
        sys.exit(str(exc))

    width = 70
    print(f"\n{'='*width}")
    print(f"  A2 OUTPUT VERIFICATION")
    print(f"  Checked {checked} orders (statuses: {', '.join(sorted(CHECK_STATUSES))})")
    print(f"{'='*width}")

    if not failures:
        print(f"\n  PASS — All {checked} orders have valid client_name and property_address.\n")
    else:
        print(f"\n  FAIL — {len(failures)} order(s) with sentinel/empty values:\n")
        print(f"  {'Order ID':<16} {'Status':<26} Bad Fields")
        print(f"  {'-'*68}")
        for oid, status, bad in failures:
            print(f"  {oid:<16} {status:<26} {'; '.join(bad)}")
        print()

    print(f"{'='*width}\n")
    sys.exit(0 if not failures else 1)


if __name__ == "__main__":
    main()
