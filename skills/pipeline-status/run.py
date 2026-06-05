#!/usr/bin/env python3
"""pipeline-status: Snapshot of all order statuses in the pipeline Excel state."""
import os
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed. Run: pip install openpyxl")

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
EXCEL_PATH = os.path.join(BASE_DIR, "data", "invoice_pipeline_state.xlsx")


def snapshot() -> dict:
    """Return {status: [order_id, ...]} dict. Importable by other skills."""
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Excel state not found: {EXCEL_PATH}")

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        status_col = headers.index("status")
        order_col  = headers.index("order_id")
    except ValueError as exc:
        wb.close()
        raise ValueError(f"Column not found: {exc}") from exc

    groups: dict = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        status   = str(row[status_col] or "unknown")
        order_id = str(row[order_col]  or "")
        groups[status].append(order_id)

    wb.close()
    return dict(groups)


def print_snapshot(groups: dict) -> None:
    total = sum(len(v) for v in groups.values())
    print(f"\n{'='*62}")
    print(f"  PIPELINE STATUS SNAPSHOT  ({total} total orders)")
    print(f"{'='*62}")
    print(f"  {'Status':<30} {'Count':>6}  {'Sample Order IDs'}")
    print(f"  {'-'*60}")
    for status, orders in sorted(groups.items(), key=lambda x: -len(x[1])):
        samples = ", ".join(orders[:3])
        if len(orders) > 3:
            samples += f"  (+{len(orders) - 3} more)"
        print(f"  {status:<30} {len(orders):>6}  {samples}")
    print(f"{'='*62}\n")


def main():
    try:
        groups = snapshot()
    except FileNotFoundError as exc:
        sys.exit(str(exc))
    print_snapshot(groups)


if __name__ == "__main__":
    main()
