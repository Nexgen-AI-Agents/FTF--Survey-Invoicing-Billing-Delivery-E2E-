#!/usr/bin/env python3
"""requeue-orders: Reset specific orders to a target status for reprocessing."""
import argparse
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed. Run: pip install openpyxl")

BASE_DIR   = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
EXCEL_PATH = os.path.join(BASE_DIR, "data", "invoice_pipeline_state.xlsx")


def requeue(
    target_status: str,
    order_ids: set[str] | None = None,
    from_status: str | None = None,
    clear_cols: list[str] | None = None,
    dry_run: bool = False,
) -> list[tuple[str, str, str]]:
    """
    Reset matching rows. Returns list of (order_id, old_status, new_status).
    Importable by other skills.
    """
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Excel state not found: {EXCEL_PATH}")

    clear_cols = clear_cols or []

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers) if h}

    if "status" not in col or "order_id" not in col:
        wb.close()
        raise ValueError("Required columns 'status' and 'order_id' not found.")

    changed = []
    for row in ws.iter_rows(min_row=2):
        oid     = str(row[col["order_id"]].value or "").strip()
        current = str(row[col["status"]].value   or "").strip()

        matched = (order_ids and oid in order_ids) or \
                  (from_status and current == from_status)
        if not matched:
            continue

        changed.append((oid, current, target_status))

        if not dry_run:
            row[col["status"]].value = target_status
            for cname in clear_cols:
                if cname in col:
                    row[col[cname]].value = ""

    if changed and not dry_run:
        wb.save(EXCEL_PATH)
    wb.close()
    return changed


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--orders",      help="Comma-separated order IDs to reset")
    group.add_argument("--from-status", help="Reset ALL orders in this status")
    parser.add_argument("--target-status", required=True)
    parser.add_argument("--clear",  default="",
                        help="Comma-separated column names to blank out")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    order_ids  = {o.strip() for o in args.orders.split(",") if o.strip()} \
                 if args.orders else None
    clear_cols = [c.strip() for c in args.clear.split(",") if c.strip()]

    try:
        changed = requeue(
            target_status=args.target_status,
            order_ids=order_ids,
            from_status=getattr(args, "from_status", None),
            clear_cols=clear_cols,
            dry_run=args.dry_run,
        )
    except (FileNotFoundError, ValueError) as exc:
        sys.exit(str(exc))

    if not changed:
        print("No matching orders found.")
        return

    label = "DRY RUN — " if args.dry_run else ""
    print(f"\n{label}Requeued {len(changed)} order(s):")
    print(f"  {'Order ID':<16} {'From':<26} {'To'}")
    print(f"  {'-'*66}")
    for oid, old, new in changed:
        print(f"  {oid:<16} {old:<26} {new}")
    if clear_cols:
        print(f"\n  Cleared columns: {', '.join(clear_cols)}")
    if not args.dry_run:
        print(f"\nSaved: {EXCEL_PATH}")
    else:
        print("\n(Dry run — no changes written)")


if __name__ == "__main__":
    main()
