#!/usr/bin/env python3
"""check-dollar-sign-orders: Find orders with no invoice amount (the $ sign in FTF)."""
import argparse
import os
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed. Run: pip install openpyxl")

BASE_DIR  = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
EXCEL_PATH = os.path.join(BASE_DIR, "data", "invoice_pipeline_state.xlsx")

_ZERO = {None, "", "0", "0.0", 0, 0.0}


def _is_missing(val) -> bool:
    if val in _ZERO:
        return True
    try:
        return float(val) == 0.0
    except (ValueError, TypeError):
        return False


def find_missing(status_filter: str = "all") -> list[dict]:
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Excel state not found: {EXCEL_PATH}")

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}

    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        status = str(row[col.get("status", 1)] or "")
        if status_filter != "all" and status != status_filter:
            continue
        if _is_missing(row[col.get("estimate_amount")]):
            results.append({
                "order_id":        str(row[col.get("order_id", 0)] or ""),
                "status":          status,
                "client_name":     str(row[col.get("client_name", 4)] or ""),
                "property_address": str(row[col.get("property_address", 5)] or ""),
            })

    wb.close()
    return results


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--status", default="all",
                        help="Filter by pipeline status (default: all)")
    args = parser.parse_args()

    try:
        results = find_missing(args.status)
    except FileNotFoundError as exc:
        sys.exit(str(exc))

    print(f"\n{'='*80}")
    print(f"  ORDERS WITH NO INVOICE AMOUNT  ({len(results)} found)")
    if args.status != "all":
        print(f"  Filtered by status: {args.status}")
    print(f"{'='*80}")

    if not results:
        print("  None found — all orders have amounts.\n")
        print(f"{'='*80}\n")
        return

    by_status: dict = defaultdict(list)
    for r in results:
        by_status[r["status"]].append(r)

    for status, orders in sorted(by_status.items(), key=lambda x: -len(x[1])):
        print(f"\n  [{status}] — {len(orders)} orders")
        print(f"  {'Order ID':<16} {'Client Name':<34} {'Address'}")
        print(f"  {'-'*76}")
        for o in orders[:25]:
            name = str(o["client_name"])[:33]
            addr = str(o["property_address"])[:34]
            print(f"  {o['order_id']:<16} {name:<34} {addr}")
        if len(orders) > 25:
            print(f"  ... and {len(orders) - 25} more")

    print(f"\n  TOTAL: {len(results)} orders missing invoice amount")
    print(f"{'='*80}\n")


if __name__ == "__main__":
    main()
