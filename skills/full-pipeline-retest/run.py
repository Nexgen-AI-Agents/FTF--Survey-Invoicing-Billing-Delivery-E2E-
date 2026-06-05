#!/usr/bin/env python3
"""full-pipeline-retest: End-to-end retest for specific orders after a bug fix."""
import argparse
import os
import sys

# ── Project path setup ────────────────────────────────────────────────────────
BASE_DIR    = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
SHARED_DIR  = os.path.join(BASE_DIR, "code", "shared")
SPRINT_DIR  = os.path.join(BASE_DIR, "code", "sprint_11_invoice_pipeline")
SKILLS_DIR  = os.path.join(BASE_DIR, "skills")

for path in (SHARED_DIR, SPRINT_DIR, SKILLS_DIR):
    if path not in sys.path:
        sys.path.insert(0, path)

# Load .env before importing anything that reads env vars
_env_path = os.path.join(BASE_DIR, ".env")
if os.path.exists(_env_path):
    try:
        from dotenv import load_dotenv
        load_dotenv(_env_path, override=False)
    except ImportError:
        pass  # dotenv optional; rely on shell env if not installed


def _import_agents():
    """Lazy import so --dry-run doesn't require agent dependencies."""
    from agents.agent_a2_data_collector import collect_for_order
    from agents.agent_a3_invoice_compiler import run as a3_run
    return collect_for_order, a3_run


def _read_order_ids_by_status(status: str) -> list[str]:
    """Read all order IDs with the given status from Excel state."""
    import openpyxl
    from pipeline_status import run as ps_run  # noqa: unused import — just for path check
    excel_path = os.path.join(BASE_DIR, "data", "invoice_pipeline_state.xlsx")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}
    ids = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[col.get("status", 1)] or "") == status:
            ids.append(str(row[col.get("order_id", 0)] or ""))
    wb.close()
    return [i for i in ids if i]


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--orders",      help="Comma-separated order IDs")
    group.add_argument("--from-status", help="Retest all orders in this status")
    parser.add_argument("--dry-run", action="store_true",
                        help="Show plan without running pipeline steps")
    args = parser.parse_args()

    if args.orders:
        order_ids = [o.strip() for o in args.orders.split(",") if o.strip()]
    else:
        order_ids = _read_order_ids_by_status(args.from_status)

    if not order_ids:
        sys.exit("No matching orders found.")

    print(f"\n{'='*66}")
    print(f"  FULL PIPELINE RETEST  ({len(order_ids)} orders)")
    print(f"  {', '.join(order_ids[:6])}{'...' if len(order_ids) > 6 else ''}")
    print(f"{'='*66}\n")

    # ── Step 1: Before snapshot ──────────────────────────────────────────────
    print("STEP 1 — Before snapshot")
    from pipeline_status.run import snapshot, print_snapshot  # type: ignore
    before = snapshot()
    print_snapshot(before)

    if args.dry_run:
        print("(Dry run — stopping before pipeline steps)\n")
        return

    # ── Step 2: Run A2 for each order ────────────────────────────────────────
    print("STEP 2 — Running A2 data collection")
    collect_for_order, a3_run = _import_agents()
    a2_results = {}
    for oid in order_ids:
        print(f"  A2 → {oid} ...", end=" ", flush=True)
        try:
            packet = collect_for_order(oid)
            status = packet.get("source_of_truth", "done")
            print(f"OK  (source={status})")
            a2_results[oid] = "ok"
        except Exception as exc:
            print(f"FAIL  ({exc})")
            a2_results[oid] = f"error: {exc}"

    ok  = sum(1 for v in a2_results.values() if v == "ok")
    bad = len(a2_results) - ok
    print(f"\n  A2 summary: {ok} ok, {bad} failed\n")

    # ── Step 3: Run A3 ───────────────────────────────────────────────────────
    print("STEP 3 — Running A3 invoice compiler (all data_collected orders)")
    try:
        a3_run()
        print("  A3 complete\n")
    except Exception as exc:
        print(f"  A3 failed: {exc}\n")

    # ── Step 4: Verify A2 output quality ─────────────────────────────────────
    print("STEP 4 — Verifying A2 output quality")
    from verify_a2_output.run import verify  # type: ignore
    checked, failures = verify()
    if not failures:
        print(f"  PASS — {checked} orders, all have valid client_name + property_address\n")
    else:
        print(f"  FAIL — {len(failures)} order(s) still have sentinel values:")
        for oid, status, bad in failures:
            if oid in set(order_ids):
                print(f"    {oid}  {'; '.join(bad)}")
        print()

    # ── Step 5: After snapshot + diff ────────────────────────────────────────
    print("STEP 5 — After snapshot")
    after = snapshot()
    print_snapshot(after)

    print("STATUS DIFF (before → after):")
    all_statuses = set(before) | set(after)
    changed = False
    for s in sorted(all_statuses):
        b = len(before.get(s, []))
        a = len(after.get(s, []))
        if b != a:
            delta = a - b
            sign  = "+" if delta > 0 else ""
            print(f"  {s:<30} {b:>5} → {a:>5}  ({sign}{delta})")
            changed = True
    if not changed:
        print("  No status changes detected.")
    print()


if __name__ == "__main__":
    main()
