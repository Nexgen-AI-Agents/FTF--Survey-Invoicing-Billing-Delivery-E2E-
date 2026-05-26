"""
FTF AR Report — builds Unpaid_AR_Report_MM.DD.YYYY.xlsx from the Books module.

Usage:
    python3 skills/ftf_ar_report/scripts/build_ar_report.py --output <dir>

Env overrides:
    FTF_BOOKS_BASE_URL   (default: https://stage.fieldtofinish.jobs)
    FTF_BOOKS_USER       (default: pchandra)
    FTF_BOOKS_PASSWORD   (default: prateek@123)
"""

import argparse
import io
import os
import sys
from datetime import date, datetime
from pathlib import Path

import httpx
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ─── Config ──────────────────────────────────────────────────────────────────
_BASE     = os.getenv("FTF_BOOKS_BASE_URL", "https://stage.fieldtofinish.jobs")
_USER     = os.getenv("FTF_BOOKS_USER")
_PASSWORD = os.getenv("FTF_BOOKS_PASSWORD")

if not _USER or not _PASSWORD:
    print("ERROR: FTF_BOOKS_USER and FTF_BOOKS_PASSWORD must be set in .env or environment.")
    sys.exit(1)

LOGIN_URL    = f"{_BASE}/admin/login"
DOWNLOAD_URL = f"{_BASE}/books/get_data_excel"
ORDER_LINK   = "https://fieldtofinish.jobs/order/?order={order}"
XLSX_MAGIC   = b"PK\x03\x04"

BUCKET_ORDER = ["0-29", "30-59", "60-89", "90-364", "Over 365"]

# Header style
HEADER_FILL = PatternFill("solid", fgColor="1F3864")  # dark navy
HEADER_FONT = Font(bold=True, color="FFFFFF")


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _days_and_group(date_delivered_str: str | None) -> tuple[int, str]:
    if not date_delivered_str or str(date_delivered_str).strip() in ("", "None", "null"):
        return 0, "0-29"
    try:
        delivered = datetime.strptime(str(date_delivered_str).strip(), "%Y-%m-%d").date()
        days = (date.today() - delivered).days
    except ValueError:
        return 0, "0-29"
    if days < 0:
        days = 0
    if days <= 29:
        return days, "0-29"
    if days <= 59:
        return days, "30-59"
    if days <= 89:
        return days, "60-89"
    if days <= 364:
        return days, "90-364"
    return days, "Over 365"


def _currency(ws, row: int, col: int, value) -> None:
    try:
        num = float(str(value).replace(",", "").strip()) if value else 0.0
    except ValueError:
        num = 0.0
    cell = ws.cell(row=row, column=col, value=num)
    cell.number_format = '#,##0.00'


def _auto_fit(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)


# ─── Main ─────────────────────────────────────────────────────────────────────

def fetch_raw_xlsx(client: httpx.Client) -> bytes:
    # Login
    r = client.post(
        LOGIN_URL,
        data={"user": _USER, "password": _PASSWORD},
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    if r.status_code not in (200, 302):
        print(f"ERROR: Login failed — HTTP {r.status_code}")
        print(f"       Response: {r.text[:200]}")
        sys.exit(1)
    print(f"  Login OK ({_USER}@{_BASE})")

    # Download
    r2 = client.get(DOWNLOAD_URL, params={"show_all": "1"})
    if r2.status_code != 200:
        print(f"ERROR: Download failed — HTTP {r2.status_code}")
        sys.exit(1)
    if r2.content[:4] != XLSX_MAGIC:
        print(f"ERROR: Response is not an xlsx file.")
        print(f"       Content-Type: {r2.headers.get('content-type')}")
        print(f"       First 200 bytes: {r2.content[:200]}")
        sys.exit(1)

    print(f"  Downloaded {len(r2.content):,} bytes")
    return r2.content


def parse_raw(raw: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb.active
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rec = dict(zip(headers, row))
        rows.append(rec)
    print(f"  Parsed {len(rows):,} rows from raw download (columns: {headers})")
    return rows


def build_unpaid_sheet(ws, rows: list[dict], today_str: str) -> list[dict]:
    """Write the Unpaid detail sheet. Returns enriched rows for Summary."""
    COLS = [
        "Company Name", "Order", "File", "Name", "Address",
        "Date Ordered", "Date Delivered", "Days Since Delivery", "Group",
        "Amount", "Paid", "Owed",
    ]
    # Write header
    for ci, col in enumerate(COLS, 1):
        ws.cell(row=1, column=ci, value=col)

    enriched = []
    for ri, rec in enumerate(rows, 2):
        company = str(rec.get("Company Name") or "").strip().lstrip("\xa0").strip()
        order   = str(rec.get("Order") or "").strip()
        file_   = str(rec.get("File") or "").strip()
        name    = str(rec.get("Name") or "").strip().lstrip("\xa0").strip()
        address = str(rec.get("Address") or "").strip()
        date_ordered   = str(rec.get("Date Ordered") or "").strip()
        date_delivered = str(rec.get("Date Delivered") or "").strip()

        days, group = _days_and_group(date_delivered)

        # Company Name
        ws.cell(row=ri, column=1, value=company)
        # Order — hyperlink
        order_cell = ws.cell(row=ri, column=2, value=order)
        if order:
            order_cell.hyperlink = ORDER_LINK.format(order=order)
            order_cell.font = Font(color="0563C1", underline="single")
        # File, Name, Address
        ws.cell(row=ri, column=3, value=file_)
        ws.cell(row=ri, column=4, value=name)
        ws.cell(row=ri, column=5, value=address)
        # Dates
        ws.cell(row=ri, column=6, value=date_ordered)
        ws.cell(row=ri, column=7, value=date_delivered if date_delivered not in ("None", "") else "")
        # Calculated
        ws.cell(row=ri, column=8, value=days)
        ws.cell(row=ri, column=9, value=group)
        # Currency
        _currency(ws, ri, 10, rec.get("Amount"))
        _currency(ws, ri, 11, rec.get("Paid"))
        _currency(ws, ri, 12, rec.get("Owed"))

        enriched.append({
            "company": company or "(no company)",
            "group": group,
            "owed": float(str(rec.get("Owed") or "0").replace(",", "").strip() or "0"),
        })

    _style_header(ws)
    ws.freeze_panes = "A2"
    _auto_fit(ws)
    return enriched


def build_summary_sheet(ws, enriched: list[dict], today_str: str) -> None:
    """Write the pivot-by-company summary sheet."""
    # Build pivot: company -> bucket -> owed
    pivot: dict[str, dict[str, float]] = {}
    for rec in enriched:
        company = rec["company"]
        bucket  = rec["group"]
        owed    = rec["owed"]
        if company not in pivot:
            pivot[company] = {b: 0.0 for b in BUCKET_ORDER}
        pivot[company][bucket] = pivot[company].get(bucket, 0.0) + owed

    # Grand totals per company
    rows = []
    for company, buckets in pivot.items():
        grand = sum(buckets[b] for b in BUCKET_ORDER)
        rows.append((company, buckets, grand))

    # Sort by Grand Total descending
    rows.sort(key=lambda x: x[2], reverse=True)

    # Header
    header = ["Company"] + BUCKET_ORDER + ["Grand Total"]
    for ci, h in enumerate(header, 1):
        ws.cell(row=1, column=ci, value=h)

    # Data rows
    for ri, (company, buckets, grand) in enumerate(rows, 2):
        ws.cell(row=ri, column=1, value=company)
        for ci, bucket in enumerate(BUCKET_ORDER, 2):
            cell = ws.cell(row=ri, column=ci, value=buckets.get(bucket, 0.0))
            cell.number_format = '#,##0.00'
        grand_cell = ws.cell(row=ri, column=len(BUCKET_ORDER) + 2, value=grand)
        grand_cell.number_format = '#,##0.00'
        grand_cell.font = Font(bold=True)

    # Totals row
    totals_row = len(rows) + 2
    ws.cell(row=totals_row, column=1, value="TOTAL").font = Font(bold=True)
    for ci, bucket in enumerate(BUCKET_ORDER, 2):
        total = sum(r[1].get(bucket, 0.0) for r in rows)
        cell = ws.cell(row=totals_row, column=ci, value=total)
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)
    grand_total = sum(r[2] for r in rows)
    gt_cell = ws.cell(row=totals_row, column=len(BUCKET_ORDER) + 2, value=grand_total)
    gt_cell.number_format = '#,##0.00'
    gt_cell.font = Font(bold=True)

    _style_header(ws)
    ws.freeze_panes = "B2"
    _auto_fit(ws)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build FTF Unpaid AR Report Excel")
    parser.add_argument("--output", required=True, help="Output directory path")
    args = parser.parse_args()

    out_dir = Path(args.output)
    out_dir.mkdir(parents=True, exist_ok=True)

    today      = date.today()
    today_str  = today.strftime("%m.%d.%Y")
    sheet_date = today_str.replace(".", "")  # MMDDYYYY for sheet names
    filename   = f"Unpaid_AR_Report_{today_str}.xlsx"
    out_path   = out_dir / filename

    print(f"FTF AR Report — {today_str}")
    print(f"  Output: {out_path}")
    print()

    client = httpx.Client(follow_redirects=True, timeout=60.0)
    try:
        raw = fetch_raw_xlsx(client)
    finally:
        client.close()

    rows = parse_raw(raw)

    wb = openpyxl.Workbook()

    # Unpaid sheet
    unpaid_ws = wb.active
    unpaid_ws.title = f"Unpaid_{today_str}"
    print(f"  Building Unpaid_{today_str} sheet...")
    enriched = build_unpaid_sheet(unpaid_ws, rows, today_str)

    # Summary sheet
    summary_ws = wb.create_sheet(title=f"Summary_{today_str}")
    print(f"  Building Summary_{today_str} sheet...")
    build_summary_sheet(summary_ws, enriched, today_str)

    wb.save(out_path)
    print()
    print(f"Done: {out_path}")
    print(f"computer://{out_path}")


if __name__ == "__main__":
    main()
