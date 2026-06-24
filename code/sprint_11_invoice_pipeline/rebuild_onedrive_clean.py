"""rebuild_onedrive_clean.py — rebuild the OneDrive workbook from scratch (cruft-free).

WHY: the live 'FTF-Invoicing Agent.xlsx' accumulated XML cruft over months of mixed
Graph-API + Excel + openpyxl edits. Microsoft Graph's Excel (workbook) API rejected the
whole file with 501 UnsupportedWorkbook/FileCorruptTryRepair even after the inline->shared
string fix and even with tables/validation/CF removed — but a freshly *built* openpyxl
workbook carrying the exact same features opens fine. So the cure is a clean rebuild:
copy only cell values + styles into a brand-new Workbook (which regenerates every
workbook-level part) and re-add the Approvals dropdown / row-colors / table fresh.

Preserves: Pipeline Guide, Pricing Rules (data + table), How-to docs. Approvals is left
EMPTY (header row only) per the 'process from today onward' reset.

Run (file must be UNLOCKED — no Office session holding it):
  cd code/sprint_11_invoice_pipeline && python rebuild_onedrive_clean.py
"""
import io
import os
import sys
from copy import copy

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "shared"))

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from core.logger import get_logger
from core import onedrive_excel_client as od

log = get_logger("rebuild_onedrive_clean")


def _copy_sheet(src_ws, new_wb, name):
    """Create `name` in new_wb and copy values + per-cell styles + geometry from src_ws."""
    ws = new_wb.create_sheet(name)
    for row in src_ws.iter_rows():
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            nc = ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                nc.font = copy(cell.font)
                nc.fill = copy(cell.fill)
                nc.alignment = copy(cell.alignment)
                nc.border = copy(cell.border)
                nc.number_format = cell.number_format
    for k, dim in src_ws.column_dimensions.items():
        if dim.width:
            ws.column_dimensions[k].width = dim.width
    for k, dim in src_ws.row_dimensions.items():
        if dim.height:
            ws.row_dimensions[k].height = dim.height
    for mc in list(src_ws.merged_cells.ranges):
        ws.merge_cells(str(mc))
    return ws


def _add_approvals_features(ws):
    """Re-create the Approvals dropdown + row-color rules fresh.

    NO Excel table here — a header-only openpyxl table makes Graph reject the whole
    workbook (501). The ApprovalTable is created via the Graph API after upload.
    """
    end = od._END_COL
    act = chr(ord("A") + od._COL_ACTION)

    dv = DataValidation(type="list", formula1='"Approve,Reject,On-hold"',
                        allow_blank=True, showDropDown=False)
    dv.error = "Select Approve, Reject, or On-hold"
    dv.errorTitle = "Invalid action"
    dv.sqref = f"{act}2:{act}10000"
    ws.add_data_validation(dv)

    for action_val, hex_color in [("Approve", "C6EFCE"), ("Reject", "FFC7CE"), ("On-hold", "FFEB9C")]:
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        ws.conditional_formatting.add(
            f"A2:{end}10000",
            FormulaRule(formula=[f'${act}2="{action_val}"'], fill=fill),
        )


def build_clean_bytes(src_bytes: bytes) -> bytes:
    """Return a freshly-built, cruft-free copy of the workbook (still openpyxl/inline —
    the caller's upload path converts to shared strings)."""
    src = openpyxl.load_workbook(io.BytesIO(src_bytes))
    new = openpyxl.Workbook()
    new.remove(new.active)

    approvals = od.ONEDRIVE_SHEET_NAME
    pricing = od.PRICING_RULES_SHEET_NAME

    # rebuild every sheet in original order; Approvals is rebuilt empty
    for name in src.sheetnames:
        if name == approvals:
            ws = new.create_sheet(approvals)
            for c, h in enumerate(od.APPROVAL_HEADERS, 1):
                ws.cell(row=1, column=c, value=h).font = Font(bold=True)
            _add_approvals_features(ws)
        else:
            ws = _copy_sheet(src[name], new, name)
            if name == pricing:
                # re-add the Pricing Rules table fresh, matching the data extent
                last = max(ws.max_row, 1)
                t = Table(displayName=od.PRICING_RULES_TABLE_NAME, ref=f"A1:H{last}")
                t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
                ws.add_table(t)

    buf = io.BytesIO()
    new.save(buf)
    return buf.getvalue()


def main():
    raw = od._download_workbook_bytes()
    log.info("downloaded current workbook: %d bytes", len(raw))
    clean = build_clean_bytes(raw)
    log.info("built clean workbook: %d bytes", len(clean))
    od._upload_workbook_bytes(clean)   # converts to shared strings inside
    log.info("clean workbook uploaded — verifying workbook API…")
    import time
    import httpx
    for _ in range(6):
        time.sleep(8)
        r = httpx.post(od._wb_base() + "/createSession", headers=od._headers(),
                       json={"persistChanges": True}, timeout=30.0)
        if r.status_code in (200, 201):
            sid = r.json().get("id")
            try:
                httpx.post(od._wb_base() + "/closeSession",
                           headers={**od._headers(), "workbook-session-id": sid}, timeout=10.0)
            except Exception:
                pass
            # workbook API healthy → create the ApprovalTable via Graph API
            od._ensure_approval_table_via_api()
            print("OK — workbook API healthy (createSession 200) + ApprovalTable ensured. Sheet is live.")
            return 0
    print(f"STILL FAILING — createSession {r.status_code}: {r.text[:160]}")
    return 1


if __name__ == "__main__":
    sys.exit(main())
