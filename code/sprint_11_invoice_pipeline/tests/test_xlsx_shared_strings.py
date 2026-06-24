"""Tests for xlsx_shared_strings — the openpyxl inline-string -> shared-string fix.

openpyxl writes inline strings with no sharedStrings.xml, which Graph's Excel API
rejects (501 UnsupportedWorkbook). to_shared_strings() must produce a file that
(a) has a shared-string table, (b) has no inline strings left, (c) still opens in
openpyxl with identical cell values, and (d) is idempotent.
"""
import io
import os
import sys
import zipfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "shared"))

import openpyxl

from core.xlsx_shared_strings import to_shared_strings


def _make_inline_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Approvals"
    ws.append(["Order ID", "Client Name", "Amount ($) by AI", "Notes"])
    ws.append(["48217", "Old Republic Title", 475.0, "ESCALATE — verify scope"])
    ws.append(["48218", "Old Republic Title", 300.0, ""])  # repeat client → dedup
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _has_shared_strings(raw: bytes) -> bool:
    return "xl/sharedStrings.xml" in zipfile.ZipFile(io.BytesIO(raw)).namelist()


def _sheet_xml(raw: bytes) -> str:
    z = zipfile.ZipFile(io.BytesIO(raw))
    name = next(n for n in z.namelist() if n.startswith("xl/worksheets/sheet"))
    return z.read(name).decode("utf-8")


def test_openpyxl_baseline_is_inline_with_no_shared_strings():
    raw = _make_inline_xlsx()
    assert not _has_shared_strings(raw)
    assert "inlineStr" in _sheet_xml(raw)


def test_conversion_adds_shared_string_table_and_removes_inline():
    out = to_shared_strings(_make_inline_xlsx())
    assert _has_shared_strings(out)
    sheet = _sheet_xml(out)
    assert "inlineStr" not in sheet
    assert 't="s"' in sheet


def test_conversion_preserves_cell_values():
    out = to_shared_strings(_make_inline_xlsx())
    wb = openpyxl.load_workbook(io.BytesIO(out))
    ws = wb["Approvals"]
    assert [c.value for c in ws[1]] == ["Order ID", "Client Name", "Amount ($) by AI", "Notes"]
    assert ws["A2"].value == "48217"
    assert ws["B2"].value == "Old Republic Title"
    assert ws["C2"].value == 475.0
    assert ws["D2"].value == "ESCALATE — verify scope"


def test_dedup_shared_strings():
    out = to_shared_strings(_make_inline_xlsx())
    sst = zipfile.ZipFile(io.BytesIO(out)).read("xl/sharedStrings.xml").decode("utf-8")
    # "Old Republic Title" appears in two rows but must be stored once
    assert sst.count("Old Republic Title") == 1
    assert 'uniqueCount=' in sst and 'count=' in sst


def test_content_types_and_rels_registered():
    out = to_shared_strings(_make_inline_xlsx())
    z = zipfile.ZipFile(io.BytesIO(out))
    ct = z.read("[Content_Types].xml").decode("utf-8")
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    assert "sharedStrings.xml" in ct
    assert "sharedStrings.xml" in rels


def test_idempotent_on_already_shared():
    once = to_shared_strings(_make_inline_xlsx())
    twice = to_shared_strings(once)
    assert twice == once  # already shared-string → returned unchanged
