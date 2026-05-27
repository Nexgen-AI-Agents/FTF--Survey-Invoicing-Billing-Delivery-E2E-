"""Tests for agent_15_statement_generator."""

from datetime import date
from unittest.mock import patch

import openpyxl
import pytest

from agents.agent_15_statement_generator import (
    _build_excel,
    _build_pdf,
    _group_by_client,
    _order_row,
)


def _sample_orders(n=2, email="billing@acme.com", paid=False):
    status = "Paid" if paid else "Unpaid"
    return [
        {
            "order_id": f"ORD-{100 + i}",
            "service_type": "Boundary Survey",
            "order_date": "2026-04-15",
            "invoice_amount": 500.0,
            "payment_status": status,
            "billing_email": email,
            "customer_type": "b2b",
        }
        for i in range(n)
    ]


# ── _order_row ────────────────────────────────────────────────────────────────

def test_order_row_unpaid_includes_balance():
    order = {
        "order_id": "ORD-001", "service_type": "Survey",
        "order_date": "2026-04-10", "invoice_amount": 400.0,
        "payment_status": "Unpaid",
    }
    row = _order_row(order)
    assert row[0] == "ORD-001"
    assert "$400.00" in row[3]
    assert "$400.00" in row[5]


def test_order_row_paid_zero_balance():
    order = {
        "order_id": "ORD-002", "service_type": "Survey",
        "order_date": "2026-04-11", "invoice_amount": 300.0,
        "payment_status": "Paid",
    }
    row = _order_row(order)
    assert "$0.00" in row[5]


# ── _build_excel ──────────────────────────────────────────────────────────────

def test_build_excel_creates_file(tmp_path):
    orders = _sample_orders(3)
    month = date(2026, 4, 1)
    path = _build_excel("billing@acme.com", orders, month, str(tmp_path))
    assert path.endswith(".xlsx")
    assert (tmp_path / f"2026-04_billing_acme.com.xlsx").exists() or \
           any(f.suffix == ".xlsx" for f in tmp_path.iterdir())


def test_build_excel_has_two_sheets(tmp_path):
    orders = _sample_orders(2)
    month = date(2026, 4, 1)
    path = _build_excel("billing@acme.com", orders, month, str(tmp_path))
    wb = openpyxl.load_workbook(path)
    assert "Unpaid Detail" in wb.sheetnames
    assert "Summary" in wb.sheetnames


def test_build_excel_detail_row_count(tmp_path):
    orders = _sample_orders(4)
    month = date(2026, 4, 1)
    path = _build_excel("billing@acme.com", orders, month, str(tmp_path))
    wb = openpyxl.load_workbook(path)
    ws = wb["Unpaid Detail"]
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 4


def test_build_excel_summary_total_reflects_orders(tmp_path):
    orders = _sample_orders(3)  # 3 × $500 unpaid = $1500 balance
    month = date(2026, 4, 1)
    path = _build_excel("billing@acme.com", orders, month, str(tmp_path))
    wb = openpyxl.load_workbook(path)
    ws = wb["Summary"]
    row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert row[2] == 3  # order count
    assert "1,500.00" in str(row[3])  # total invoiced


# ── _build_pdf ────────────────────────────────────────────────────────────────

def test_build_pdf_creates_file(tmp_path):
    orders = _sample_orders(2)
    month = date(2026, 4, 1)
    path = _build_pdf("billing@acme.com", orders, month, str(tmp_path))
    assert path.endswith(".pdf")
    assert any(f.suffix == ".pdf" for f in tmp_path.iterdir())


# ── _group_by_client ──────────────────────────────────────────────────────────

def test_group_by_client_groups_correctly():
    orders = [
        {"billing_email": "a@co.com", "order_id": "1"},
        {"billing_email": "a@co.com", "order_id": "2"},
        {"billing_email": "b@co.com", "order_id": "3"},
    ]
    groups = _group_by_client(orders)
    assert len(groups) == 2
    assert len(groups["a@co.com"]) == 2


def test_group_by_client_skips_no_email():
    orders = [
        {"billing_email": "", "order_id": "1"},
        {"customer_email": "x@y.com", "order_id": "2"},
    ]
    groups = _group_by_client(orders)
    assert "x@y.com" in groups
    assert "" not in groups


# ── agent_15 run (mocked) ─────────────────────────────────────────────────────

def test_run_generates_statement_per_client(tmp_path):
    month = date(2026, 4, 1)
    mock_orders = _sample_orders(2, email="a@co.com") + _sample_orders(1, email="b@co.com")

    with patch("agents.agent_15_statement_generator.get_b2b_orders_for_month",
               return_value=mock_orders), \
         patch("agents.agent_15_statement_generator.upsert_monthly_statement") as mock_upsert:

        from agents.agent_15_statement_generator import run
        result = run(statement_month=month, output_dir=str(tmp_path))

    assert result["clients_processed"] == 2
    assert result["total_orders"] == 3
    assert mock_upsert.call_count == 2
