"""Agent 15 — Statement Generator

On the 1st of each month, fetches all B2B orders for the prior month,
groups by client billing email, builds an Excel (2 tabs) + PDF (summary)
per client, saves to STATEMENT_OUTPUT_DIR, and upserts monthly_statements.

Columns: Order#, Service Type, Date of Service, Invoice Amount,
         Payment Status, Balance Due
"""

import os
import re
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors as rl_colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from config.settings import STATEMENT_OUTPUT_DIR
from core.db import upsert_monthly_statement
from core.ftf_client import get_b2b_orders_for_month
from core.logger import get_logger

logger = get_logger(__name__)

_HEADERS = ["Order #", "Service Type", "Date of Service",
            "Invoice Amount", "Payment Status", "Balance Due"]

_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT  = Font(color="FFFFFF", bold=True)
_SUMMARY_FILL = PatternFill("solid", fgColor="D6E4F0")


def _safe_filename(email: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]", "_", email)


def _order_row(order: dict) -> list:
    order_id      = order.get("order_id") or order.get("id", "")
    service_type  = order.get("service_type", "")
    raw_date      = order.get("order_date") or order.get("service_date") or order.get("created_at", "")
    if isinstance(raw_date, (date, datetime)):
        svc_date = raw_date.strftime("%Y-%m-%d") if isinstance(raw_date, datetime) else str(raw_date)
    else:
        svc_date = str(raw_date)[:10] if raw_date else ""

    amount        = float(order.get("invoice_amount") or order.get("amount") or 0)
    pay_status    = order.get("payment_status") or order.get("status") or "Unpaid"
    balance_due   = 0.0 if str(pay_status).lower() in ("paid", "complete") else amount

    return [str(order_id), str(service_type), svc_date,
            f"${amount:,.2f}", str(pay_status), f"${balance_due:,.2f}"]


def _build_excel(client_email: str, orders: list[dict], statement_month: date,
                 output_dir: str) -> str:
    os.makedirs(output_dir, exist_ok=True)
    filename = f"{statement_month.strftime('%Y-%m')}_{_safe_filename(client_email)}.xlsx"
    path = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()

    # Tab 1 — Unpaid Detail (all orders for this client this month)
    ws1 = wb.active
    ws1.title = "Unpaid Detail"
    ws1.append(_HEADERS)
    for cell in ws1[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    total_invoiced = 0.0
    total_balance  = 0.0
    for order in orders:
        row = _order_row(order)
        ws1.append(row)
        total_invoiced += float(order.get("invoice_amount") or order.get("amount") or 0)
        pay_status = order.get("payment_status") or order.get("status") or "Unpaid"
        if str(pay_status).lower() not in ("paid", "complete"):
            total_balance += float(order.get("invoice_amount") or order.get("amount") or 0)

    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 20

    # Tab 2 — Summary by Account
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Client", "Period", "Order Count", "Total Invoiced", "Balance Due"])
    for cell in ws2[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    ws2.append([
        client_email,
        statement_month.strftime("%B %Y"),
        len(orders),
        f"${total_invoiced:,.2f}",
        f"${total_balance:,.2f}",
    ])
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.fill = _SUMMARY_FILL

    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 22

    wb.save(path)
    return path


def _build_pdf(client_email: str, orders: list[dict], statement_month: date,
               output_dir: str) -> str:
    os.makedirs(output_dir, exist_ok=True)
    filename = f"{statement_month.strftime('%Y-%m')}_{_safe_filename(client_email)}.pdf"
    path = os.path.join(output_dir, filename)

    styles = getSampleStyleSheet()
    doc    = SimpleDocTemplate(path, pagesize=letter,
                               leftMargin=40, rightMargin=40,
                               topMargin=40, bottomMargin=40)
    elements = []

    elements.append(Paragraph(f"Monthly Statement — {statement_month.strftime('%B %Y')}", styles["Title"]))
    elements.append(Paragraph(f"Client: {client_email}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    table_data = [_HEADERS] + [_order_row(o) for o in orders]
    tbl = Table(table_data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1F4E79")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("GRID",       (0, 0), (-1, -1), 0.5, rl_colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#EEF4FB")]),
    ]))
    elements.append(tbl)

    total_invoiced = sum(float(o.get("invoice_amount") or o.get("amount") or 0) for o in orders)
    total_balance  = sum(
        float(o.get("invoice_amount") or o.get("amount") or 0)
        for o in orders
        if str(o.get("payment_status") or o.get("status") or "").lower() not in ("paid", "complete")
    )
    elements.append(Spacer(1, 16))
    elements.append(Paragraph(f"Total Orders: {len(orders)}", styles["Normal"]))
    elements.append(Paragraph(f"Total Invoiced: ${total_invoiced:,.2f}", styles["Normal"]))
    elements.append(Paragraph(f"Balance Due: ${total_balance:,.2f}", styles["Normal"]))

    doc.build(elements)
    return path


def _group_by_client(orders: list[dict]) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = {}
    for order in orders:
        email = (order.get("billing_email") or order.get("customer_email") or "").strip().lower()
        if not email:
            continue
        grouped.setdefault(email, []).append(order)
    return grouped


def run(statement_month: date | None = None, output_dir: str | None = None) -> dict:
    if statement_month is None:
        today = date.today()
        # run on 1st — generate statement for the just-completed prior month
        if today.month == 1:
            statement_month = date(today.year - 1, 12, 1)
        else:
            statement_month = date(today.year, today.month - 1, 1)

    if output_dir is None:
        output_dir = os.path.abspath(STATEMENT_OUTPUT_DIR)

    orders = get_b2b_orders_for_month(statement_month)
    clients = _group_by_client(orders)

    generated = 0
    for email, client_orders in clients.items():
        excel_path = _build_excel(email, client_orders, statement_month, output_dir)
        pdf_path   = _build_pdf(email, client_orders, statement_month, output_dir)
        total_amount = sum(float(o.get("invoice_amount") or o.get("amount") or 0)
                           for o in client_orders)
        upsert_monthly_statement(
            client_email=email,
            statement_month=statement_month,
            order_count=len(client_orders),
            total_amount=total_amount,
            excel_path=excel_path,
            pdf_path=pdf_path,
        )
        generated += 1
        logger.info("statement generated client=%s orders=%d", email, len(client_orders))

    logger.info("agent_15_statement_generator: %d clients, %d orders", generated, len(orders))
    return {"clients_processed": generated, "total_orders": len(orders)}


if __name__ == "__main__":
    print(run())
