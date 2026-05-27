"""Agent 16 — Statement Reviewer

Validates generated monthly statements before delivery:
  - File exists on disk (Excel + PDF)
  - order_count matches actual rows in Excel Tab 1
  - total_amount matches sum of Invoice Amount column
  - No duplicate order IDs in the statement

Passes → status 'reviewed'  (ready to send)
Fails  → status 'failed', logs decision, Teams alert to Ryan
"""

import os
from datetime import date, datetime, timezone

import httpx
import openpyxl

from config.settings import TEAMS_WEBHOOK_URL
from core.db import get_generated_statements, log_decision, update_statement_status
from core.logger import get_logger

logger = get_logger(__name__)

_TOLERANCE = 0.01  # dollar rounding tolerance


def _parse_amount(val: str) -> float:
    try:
        return float(str(val).replace("$", "").replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _validate_excel(excel_path: str, expected_count: int, expected_total: float) -> list[str]:
    """Return list of validation errors; empty list = pass."""
    errors = []

    if not os.path.exists(excel_path):
        return [f"Excel file not found: {excel_path}"]

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if "Unpaid Detail" not in wb.sheetnames:
        errors.append("Missing sheet: Unpaid Detail")
        return errors

    ws = wb["Unpaid Detail"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data_rows = [r for r in rows if any(v is not None for v in r)]

    if len(data_rows) != expected_count:
        errors.append(
            f"Row count mismatch: expected {expected_count}, found {len(data_rows)}"
        )

    # column index 3 = Invoice Amount (0-indexed)
    amounts = [_parse_amount(r[3]) for r in data_rows if r and len(r) > 3]
    actual_total = round(sum(amounts), 2)
    if abs(actual_total - round(expected_total, 2)) > _TOLERANCE:
        errors.append(
            f"Amount mismatch: expected ${expected_total:,.2f}, found ${actual_total:,.2f}"
        )

    # duplicate order IDs — column index 0
    order_ids = [str(r[0]).strip() for r in data_rows if r and r[0]]
    if len(order_ids) != len(set(order_ids)):
        errors.append("Duplicate order IDs found in statement")

    return errors


def _alert_failure(client_email: str, errors: list[str]) -> None:
    if not TEAMS_WEBHOOK_URL:
        return
    body = "\n".join(f"- {e}" for e in errors)
    payload = {
        "@type":      "MessageCard",
        "@context":   "https://schema.org/extensions",
        "themeColor": "FF0000",
        "summary":    f"Statement validation FAILED — {client_email}",
        "title":      f"Statement Validation Failed: {client_email}",
        "text":       f"**Errors:**\n{body}\n\nRyan — please review before sending.",
    }
    try:
        httpx.post(TEAMS_WEBHOOK_URL, json=payload, timeout=15).raise_for_status()
    except Exception as exc:
        logger.error("teams alert failed client=%s: %s", client_email, exc)


def run(statement_month: date | None = None) -> dict:
    if statement_month is None:
        today = date.today()
        if today.month == 1:
            statement_month = date(today.year - 1, 12, 1)
        else:
            statement_month = date(today.year, today.month - 1, 1)

    statements = get_generated_statements(statement_month)
    passed = failed = 0

    for stmt in statements:
        excel_path = stmt.get("excel_path") or ""
        pdf_path   = stmt.get("pdf_path") or ""
        errors     = _validate_excel(
            excel_path,
            expected_count=stmt["order_count"],
            expected_total=float(stmt["total_amount"] or 0),
        )

        if not os.path.exists(pdf_path):
            errors.append(f"PDF file not found: {pdf_path}")

        if errors:
            update_statement_status(stmt["client_email"], statement_month, "failed")
            log_decision(
                agent_name="agent_16_statement_reviewer",
                decision="statement_validation_failed",
                reason="; ".join(errors),
                output_summary=f"client={stmt['client_email']}",
            )
            _alert_failure(stmt["client_email"], errors)
            failed += 1
            logger.warning("statement failed validation client=%s errors=%s",
                           stmt["client_email"], errors)
        else:
            update_statement_status(stmt["client_email"], statement_month, "reviewed")
            log_decision(
                agent_name="agent_16_statement_reviewer",
                decision="statement_validated",
                reason="all checks passed",
                output_summary=f"client={stmt['client_email']} orders={stmt['order_count']}",
            )
            passed += 1
            logger.info("statement validated client=%s", stmt["client_email"])

    logger.info("agent_16_statement_reviewer: passed=%d failed=%d", passed, failed)
    return {"passed": passed, "failed": failed}


if __name__ == "__main__":
    print(run())
