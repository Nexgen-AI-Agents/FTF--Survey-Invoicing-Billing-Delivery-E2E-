"""Excel-based state store for the Sprint 11 invoice pipeline.

Replaces PostgreSQL (core/db.py) with an openpyxl workbook committed to Git.
File: <repo_root>/data/invoice_pipeline_state.xlsx
Sheets: pipeline_state, learnings
"""
import json
import os
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from openpyxl import Workbook

from core.logger import get_logger

logger = get_logger(__name__)

_REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
EXCEL_PATH = os.path.join(_REPO_ROOT, "data", "invoice_pipeline_state.xlsx")

_STATE_COLS = [
    "order_id", "status", "service_type", "customer_email", "client_name",
    "property_address", "invoice_draft", "data_sources", "approval_message_id",
    "modification_count", "invoice_id", "data_collected_at", "draft_posted_at",
    "invoice_created_at", "processed_reply_ids", "approved_by", "sent_at",
    "pay_link", "estimate_amount", "deferred_until", "created_at", "updated_at",
]

_LEARNINGS_COLS = [
    "id", "order_id", "original_draft", "human_correction", "learned_rule",
    "service_type", "county", "entered_by", "created_at",
]

_VALID_STATE_FIELDS = set(_STATE_COLS) - {"order_id", "created_at"}


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _ensure_columns(ws, cols: list[str]) -> bool:
    """Add any missing column headers to the sheet. Returns True if any were added."""
    existing = _get_headers(ws)
    added = False
    for col in cols:
        if col not in existing:
            ws.cell(row=1, column=len(existing) + 1).value = col
            existing.append(col)
            added = True
    return added


def _load_workbook() -> Workbook:
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        _init_sheets(wb)
        _ensure_columns(wb["pipeline_state"], _STATE_COLS)
        return wb
    wb = Workbook()
    _init_sheets(wb)
    return wb


def _init_sheets(wb: Workbook) -> None:
    if "pipeline_state" not in wb.sheetnames:
        ws = wb.create_sheet("pipeline_state")
        ws.append(_STATE_COLS)
    if "learnings" not in wb.sheetnames:
        ws2 = wb.create_sheet("learnings")
        ws2.append(_LEARNINGS_COLS)
    for default in ["Sheet", "Sheet1"]:
        if default in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb[default]


def _save(wb: Workbook) -> None:
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    wb.save(EXCEL_PATH)


def _get_headers(ws) -> list[str]:
    row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not row or not row[0]:
        return []
    return [str(h) if h is not None else "" for h in row[0]]


def _sheet_to_dicts(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    return [
        {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        for row in rows[1:]
    ]


def _find_row(ws, col_name: str, value: str) -> Optional[int]:
    if ws.max_row < 2:
        return None
    headers = _get_headers(ws)
    if col_name not in headers:
        return None
    col_idx = headers.index(col_name) + 1
    for row_num in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row_num, column=col_idx).value
        if str(cell_val) == str(value):
            return row_num
    return None


_STR_COLS = {"order_id", "approval_message_id", "invoice_id"}


def _row_to_dict(ws, row_num: int) -> dict:
    headers = _get_headers(ws)
    out = {}
    for i in range(len(headers)):
        val = ws.cell(row=row_num, column=i + 1).value
        if headers[i] in _STR_COLS and val is not None:
            val = str(val).strip()
        out[headers[i]] = val
    return out


# ── Public API ────────────────────────────────────────────────────────────────

def order_exists(order_id: str) -> bool:
    wb = _load_workbook()
    _init_sheets(wb)
    return _find_row(wb["pipeline_state"], "order_id", order_id) is not None


def save_order_state(order_id: str, **fields) -> None:
    invalid = set(fields.keys()) - _VALID_STATE_FIELDS
    if invalid:
        logger.warning("save_order_state: ignoring unknown fields %s", invalid)
        fields = {k: v for k, v in fields.items() if k not in invalid}

    wb = _load_workbook()
    _init_sheets(wb)
    ws = wb["pipeline_state"]
    row_num = _find_row(ws, "order_id", order_id)

    if row_num:
        headers = _get_headers(ws)
        for col_name, value in fields.items():
            if col_name in headers:
                ws.cell(row=row_num, column=headers.index(col_name) + 1).value = value
        if "updated_at" in headers:
            ws.cell(row=row_num, column=headers.index("updated_at") + 1).value = _now()
    else:
        now = _now()
        row_data = {col: None for col in _STATE_COLS}
        row_data["order_id"] = order_id
        row_data["created_at"] = now
        row_data["updated_at"] = now
        row_data.update(fields)
        ws.append([row_data.get(col) for col in _STATE_COLS])

    _save(wb)


def get_order_by_id(order_id: str) -> Optional[dict]:
    wb = _load_workbook()
    _init_sheets(wb)
    ws = wb["pipeline_state"]
    row_num = _find_row(ws, "order_id", order_id)
    return _row_to_dict(ws, row_num) if row_num else None


def get_orders_by_status(status: str) -> list[dict]:
    wb = _load_workbook()
    _init_sheets(wb)
    rows = _sheet_to_dicts(wb["pipeline_state"])
    return [r for r in rows if str(r.get("status", "")) == status]


def get_orders_awaiting_invoice_approval() -> list[dict]:
    wb = _load_workbook()
    _init_sheets(wb)
    rows = _sheet_to_dicts(wb["pipeline_state"])
    waiting = {"invoice_draft_posted", "invoice_modification_requested", "on_hold"}
    return [r for r in rows if str(r.get("status", "")) in waiting]


def increment_modification_count(order_id: str) -> int:
    wb = _load_workbook()
    _init_sheets(wb)
    ws = wb["pipeline_state"]
    row_num = _find_row(ws, "order_id", order_id)
    if not row_num:
        return 1
    headers = _get_headers(ws)
    if "modification_count" not in headers:
        return 1
    col_idx = headers.index("modification_count") + 1
    current = ws.cell(row=row_num, column=col_idx).value
    try:
        new_val = int(float(str(current))) + 1 if current is not None else 1
    except (ValueError, TypeError):
        new_val = 1
    ws.cell(row=row_num, column=col_idx).value = new_val
    if "updated_at" in headers:
        ws.cell(row=row_num, column=headers.index("updated_at") + 1).value = _now()
    _save(wb)
    return new_val


def get_processed_reply_ids(order_id: str) -> set:
    row = get_order_by_id(order_id)
    if not row or not row.get("processed_reply_ids"):
        return set()
    try:
        return set(json.loads(str(row["processed_reply_ids"])))
    except Exception:
        return set()


def mark_reply_processed(order_id: str, reply_id: str) -> None:
    existing = get_processed_reply_ids(order_id)
    existing.add(reply_id)
    save_order_state(order_id, processed_reply_ids=json.dumps(list(existing)))


def save_invoice_learning(
    order_id: str,
    original_draft: str,
    human_correction: str,
    learned_rule: str,
    service_type: Optional[str] = None,
    county: Optional[str] = None,
    entered_by: str = "system",
) -> int:
    wb = _load_workbook()
    _init_sheets(wb)
    ws = wb["learnings"]
    new_id = ws.max_row  # header row = 1, so first data row id = 1
    ws.append([
        new_id, order_id, original_draft, human_correction, learned_rule,
        service_type, county, entered_by, _now(),
    ])
    _save(wb)
    return new_id


def get_invoice_learnings(
    service_type: Optional[str] = None,
    county: Optional[str] = None,
    limit: int = 10,
) -> list[dict]:
    wb = _load_workbook()
    _init_sheets(wb)
    rows = _sheet_to_dicts(wb["learnings"])
    rows.reverse()
    if service_type:
        rows = [r for r in rows if str(r.get("service_type") or "").lower() == service_type.lower()]
    if county:
        rows = [r for r in rows if str(r.get("county") or "").lower() == county.lower()]
    return rows[:limit]


def get_pricing_examples(
    service_type: Optional[str] = None,
    county: Optional[str] = None,
    limit: int = 20,
) -> list[dict]:
    """Pricing examples populated by Robert via admin tool. Returns empty until populated."""
    try:
        wb = _load_workbook()
        if "pricing_examples" not in wb.sheetnames:
            return []
        rows = _sheet_to_dicts(wb["pricing_examples"])
        if service_type:
            rows = [r for r in rows if str(r.get("service_type") or "").lower() == service_type.lower()]
        if county:
            rows = [r for r in rows if str(r.get("county") or "").lower() == county.lower()]
        return rows[:limit]
    except Exception:
        return []


def log_decision(
    agent_name: str,
    decision: str,
    order_id: Optional[str] = None,
    reason: Optional[str] = None,
    input_summary: Optional[str] = None,
    output_summary: Optional[str] = None,
    model_used: Optional[str] = None,
) -> None:
    logger.info(
        "DECISION agent=%s order=%s decision=%s reason=%s",
        agent_name, order_id, decision, reason,
    )


