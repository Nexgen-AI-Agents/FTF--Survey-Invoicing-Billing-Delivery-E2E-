"""
onedrive_excel_client.py — Microsoft Graph Workbook API client.

Reads/writes rows in FTF-Invoicing Agent.xlsx on nesa@nexgenlogix.com's OneDrive.
Uses Azure AD app credentials via client_credentials grant (AZURE_TENANT_ID / AZURE_APP_ID / AZURE_CLIENT_SECRET).

Required Graph permissions (application):
  Files.ReadWrite.All  — read/write files in any user's OneDrive
"""

import base64
import hashlib
import json
import os
import re
import time
import urllib.parse
from datetime import datetime, timezone
from typing import Optional
from zoneinfo import ZoneInfo

_EASTERN = ZoneInfo("America/New_York")

import httpx

from config.settings import (
    AZURE_APP_ID, AZURE_CLIENT_SECRET, AZURE_TENANT_ID,
    FTF_ORDER_URL,
    ONEDRIVE_FILE_USER, ONEDRIVE_FILE_PATH, ONEDRIVE_SHARE_URL,
    ONEDRIVE_SHEET_NAME, ONEDRIVE_TABLE_NAME,
)
from core.exceptions import AgentError
from core.logger import get_logger

log = get_logger("onedrive_excel_client")

_GRAPH     = "https://graph.microsoft.com/v1.0"
_TOKEN_URL = "https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"
_SCOPE     = "https://graph.microsoft.com/.default"

_cache: dict = {}

# Guide tab — bump version string whenever guide content changes to force a re-write
GUIDE_SHEET_NAME = "Pipeline Guide"
_GUIDE_VERSION   = "v14"  # increment when guide content changes

# How-To tab — plain-language step-by-step guide for end users
HOWTO_SHEET_NAME = "How to use Invoicing agent"
_HOWTO_VERSION   = "v7"   # increment when how-to content changes

# Pricing Rules tab — user-editable table of override prices
PRICING_RULES_SHEET_NAME  = "Pricing Rules"
PRICING_RULES_TABLE_NAME  = "PricingRulesTable"
PRICING_RULES_HEADERS = [
    "Rule ID", "Status", "Service Pattern", "County", "Client Pattern",
    "Price ($)", "Priority", "Notes",
]
_PR_COL_COUNT = len(PRICING_RULES_HEADERS)   # 8

# Column order must stay in sync with append_approval_row() values list.
# 2026-06-26: dropped "FTF Link" (the Order ID cell is now the clickable link, so the
# link can never drift from its row); added "Learning provided by user" (approver writes
# feedback here → the AI folds it into its learning on the next run).
# 2026-06-26: split the single "Service / Breakdown" into TWO columns — "Service /
# Breakdown by AI" (E, locked AI baseline) and "Service / Breakdown by User" (G, the
# editable copy that is the SOURCE OF TRUTH for invoicing + learning). "Amount ($) by
# User" (H) is now AUTO-COMPUTED from the user breakdown's total (no longer typed).
# 2026-07-09: added four property-context columns right after "Property Address" at Robert's
# request — "Property Size", "Map Link" (clickable Regrid parcel link), "FEMA Zone" and
# "Service Type". All four are AI-managed (gray) reference fields populated by A3 from the
# FTF order + county appraiser data. They sit BEFORE the Service/Amount columns so every
# _COL_* index below is DERIVED (never hardcoded) and shifts automatically. Order ID (0),
# Order Status (1), Client Name (2) and Property Address (3) keep their positions, so the
# helper scripts that read those fixed indices are unaffected.
APPROVAL_HEADERS = [
    "Order ID", "Order Status", "Client Name", "Property Address",
    "Property Size", "Map Link", "FEMA Zone", "Service Type",
    "Service / Breakdown by AI", "Amount ($) by AI",
    "Service / Breakdown by User", "Amount ($) by User",
    "Confidence", "Escalate",
    "Action", "Notes", "Posted At", "Processed At", "AI Learning", "Learning provided by user",
]
_COL_COUNT        = len(APPROVAL_HEADERS)   # 20
_END_COL          = chr(ord("A") + _COL_COUNT - 1)   # "T"

# All column indices are DERIVED from APPROVAL_HEADERS (never hardcoded) so the schema
# can change without silently breaking absolute-index arithmetic across this module.
_COL_ORDER_ID      = APPROVAL_HEADERS.index("Order ID")                    # 0  (A)
_COL_ORDER_STATUS  = APPROVAL_HEADERS.index("Order Status")                # 1  (B)
_COL_CLIENT        = APPROVAL_HEADERS.index("Client Name")                 # 2  (C)  title company / one-off client
_COL_ADDRESS       = APPROVAL_HEADERS.index("Property Address")            # 3  (D)
_COL_PROPERTY_SIZE = APPROVAL_HEADERS.index("Property Size")               # 4  (E)  lot size (acres / sq ft) — AI-managed
_COL_MAP_LINK      = APPROVAL_HEADERS.index("Map Link")                    # 5  (F)  clickable Regrid parcel/address link — AI-managed
_COL_FEMA_ZONE     = APPROVAL_HEADERS.index("FEMA Zone")                   # 6  (G)  FEMA flood zone — AI-managed
_COL_SERVICE_TYPE  = APPROVAL_HEADERS.index("Service Type")                # 7  (H)  requested service type — AI-managed
_COL_SERVICE_AI    = APPROVAL_HEADERS.index("Service / Breakdown by AI")   # 8  (I)  AI's proposed breakdown — LOCKED baseline (for AI-vs-user learning)
_COL_AMOUNT_AI     = APPROVAL_HEADERS.index("Amount ($) by AI")            # 9  (J)  AI's proposed total — LOCKED
_COL_SERVICE_USER  = APPROVAL_HEADERS.index("Service / Breakdown by User") # 10 (K)  approver's breakdown — EDITABLE, SOURCE OF TRUTH for invoice
_COL_AMOUNT_USER   = APPROVAL_HEADERS.index("Amount ($) by User")          # 11 (L)  total of user breakdown — AUTO-COMPUTED (not typed)
_COL_CONFIDENCE    = APPROVAL_HEADERS.index("Confidence")                  # 12 (M)
_COL_ESCALATE      = APPROVAL_HEADERS.index("Escalate")                    # 13 (N)
_COL_ACTION        = APPROVAL_HEADERS.index("Action")                      # 14 (O)  dropdown: Approve / Reject / On-hold
_COL_NOTES         = APPROVAL_HEADERS.index("Notes")                       # 15 (P)
_COL_POSTED_AT     = APPROVAL_HEADERS.index("Posted At")                   # 16 (Q)
_COL_PROCESSED_AT  = APPROVAL_HEADERS.index("Processed At")                # 17 (R)
_COL_AI_LEARNING   = APPROVAL_HEADERS.index("AI Learning")                 # 18 (S)  AI's per-order learning record (AI-filled)
_COL_USER_LEARNING = APPROVAL_HEADERS.index("Learning provided by user")   # 19 (T)  approver feedback → AI learns next run

# Cell-fill ownership: approver-editable columns are tinted blue ("edit here"); every
# other column is AI-managed and tinted gray ("locked — don't edit by hand"). Applied
# per-row so a recycled row position can never inherit a stale fill.
_USER_EDITABLE_HEADERS = {
    "Service / Breakdown by User", "Action", "Notes", "Learning provided by user",
}
_AI_LOCKED_FILL = "#D9D9D9"   # gray  — AI-managed / read-only
_USER_EDIT_FILL = "#DDEBF7"   # blue  — approver edits here

# Neutral table style so Excel's default blue banded style ("TableStyleMedium2") does NOT
# paint every column blue. With banding off, the per-column ownership fills (gray = AI,
# blue = approver) are what the eye sees — including on the header row.
_APPROVAL_TABLE_STYLE = "TableStyleLight1"

# Row fill colors for Action dropdown choices (light palette, Excel-compatible hex)
_ACTION_COLORS = {
    "Approve": "#C6EFCE",
    "Reject":  "#FFC7CE",
    "On-hold": "#FFEB9C",
}


# ── Auth ──────────────────────────────────────────────────────────────────────

def _get_token() -> str:
    now = time.monotonic()
    if _cache.get("od_token") and _cache.get("od_exp", 0) > now + 60:
        return _cache["od_token"]

    if not all([AZURE_TENANT_ID, AZURE_APP_ID, AZURE_CLIENT_SECRET]):
        raise AgentError("Graph API credentials not configured (AZURE_TENANT_ID / AZURE_APP_ID / AZURE_CLIENT_SECRET)")

    r = httpx.post(
        _TOKEN_URL.format(tenant=AZURE_TENANT_ID),
        data={
            "grant_type":    "client_credentials",
            "client_id":     AZURE_APP_ID,
            "client_secret": AZURE_CLIENT_SECRET,
            "scope":         _SCOPE,
        },
        timeout=15.0,
    )
    r.raise_for_status()
    data = r.json()
    _cache["od_token"] = data["access_token"]
    _cache["od_exp"]   = now + int(data.get("expires_in", 3600))
    log.debug("onedrive token refreshed")
    return _cache["od_token"]


def _headers() -> dict:
    return {"Authorization": f"Bearer {_get_token()}", "Content-Type": "application/json"}


# ── File + session resolution ─────────────────────────────────────────────────

def _share_id(share_url: str) -> str:
    """Encode a sharing URL into a Graph API share ID (u! prefix + base64url, no padding)."""
    encoded = base64.urlsafe_b64encode(share_url.encode("utf-8")).decode("utf-8").rstrip("=")
    return f"u!{encoded}"


def _get_item_id() -> str:
    """Return the OneDrive driveItem ID, caching drive_id alongside it.

    Tries ONEDRIVE_SHARE_URL first (resolves directly by sharing link — no path guessing).
    Falls back to ONEDRIVE_FILE_PATH lookup if the share URL is not set.
    """
    if _cache.get("od_item_id"):
        return _cache["od_item_id"]

    if ONEDRIVE_SHARE_URL:
        r = httpx.get(
            f"{_GRAPH}/shares/{_share_id(ONEDRIVE_SHARE_URL)}/driveItem",
            headers=_headers(),
            timeout=15.0,
        )
        if r.status_code == 403:
            raise AgentError(
                "Graph API cannot access this file via sharing link. "
                "Ensure the Azure app has Files.ReadWrite.All permission."
            )
        r.raise_for_status()
        item = r.json()
        _cache["od_item_id"] = item["id"]
        _cache["od_drive_id"] = item["parentReference"]["driveId"]
        log.info("onedrive item resolved via share URL: item=%s drive=%s",
                 item["id"], item["parentReference"]["driveId"])
        return _cache["od_item_id"]

    # Fallback: path-based lookup
    encoded = urllib.parse.quote(ONEDRIVE_FILE_PATH, safe="/")
    r = httpx.get(
        f"{_GRAPH}/users/{ONEDRIVE_FILE_USER}/drive/root:/{encoded}",
        headers=_headers(),
        timeout=15.0,
    )
    if r.status_code == 404:
        raise AgentError(
            f"OneDrive file not found: '{ONEDRIVE_FILE_PATH}' for user '{ONEDRIVE_FILE_USER}'. "
            "Set ONEDRIVE_SHARE_URL with the file's sharing link or fix ONEDRIVE_FILE_PATH."
        )
    r.raise_for_status()
    _cache["od_item_id"] = r.json()["id"]
    log.info("onedrive item resolved via path: %s", _cache["od_item_id"])
    return _cache["od_item_id"]


def _wb_base() -> str:
    item_id  = _get_item_id()
    drive_id = _cache.get("od_drive_id")
    if drive_id:
        return f"{_GRAPH}/drives/{drive_id}/items/{item_id}/workbook"
    return f"{_GRAPH}/users/{ONEDRIVE_FILE_USER}/drive/items/{item_id}/workbook"


_TRANSIENT_STATUS = {429, 500, 501, 502, 503, 504}


def _graph_get_retry(url: str, headers: dict, timeout: float = 15.0, tries: int = 4):
    """GET with retry on transient Graph errors (429/5xx, incl. the 501 the workbook API
    returns while RE-INDEXING a freshly-uploaded file). Returns the final httpx.Response
    (the caller decides how to handle a still-bad status). Brief linear backoff."""
    import time
    resp = None
    for attempt in range(1, tries + 1):
        try:
            resp = httpx.get(url, headers=headers, timeout=timeout)
        except Exception as exc:
            if attempt == tries:
                raise
            log.warning("graph GET error (attempt %d/%d) %s: %s", attempt, tries, url[-60:], exc)
            time.sleep(2 * attempt)
            continue
        if resp.status_code not in _TRANSIENT_STATUS:
            return resp
        if attempt < tries:
            log.warning("graph GET %d (attempt %d/%d) — transient, retrying %s",
                        resp.status_code, attempt, tries, url[-60:])
            time.sleep(2 * attempt)
    return resp


def _session_headers() -> dict:
    """Add workbook-session-id header for batched operations (faster)."""
    h = _headers()
    if not _cache.get("od_session"):
        try:
            r = httpx.post(
                f"{_wb_base()}/createSession",
                headers=h,
                json={"persistChanges": True},
                timeout=20.0,
            )
            r.raise_for_status()
            _cache["od_session"] = r.json()["id"]
            log.debug("workbook session created")
        except Exception as exc:
            log.warning("workbook session failed (sessionless fallback): %s", exc)
    if _cache.get("od_session"):
        h["workbook-session-id"] = _cache["od_session"]
    return h


# ── Sheet + table setup ───────────────────────────────────────────────────────

def _setup_full_sheet_via_openpyxl() -> None:
    """Create/recreate the Approvals sheet with all formatting via openpyxl download → modify → upload.

    This is the ONLY way to set data validation and conditional formatting on an Excel workbook
    in OneDrive — the Graph API workbook REST endpoints do not expose these operations.

    The file must NOT have an active workbook session when this runs (PUT upload returns 423 if
    the session lock is held). Call _close_session() before invoking.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.worksheet.datavalidation import DataValidation

    try:
        raw = _download_workbook_bytes()
    except Exception as exc:
        log.warning("_setup_full_sheet_via_openpyxl: download failed — skipping: %s", exc)
        return

    wb = openpyxl.load_workbook(io.BytesIO(raw))

    # Remove stale Approvals sheet (clean slate on every schema change)
    if ONEDRIVE_SHEET_NAME in wb.sheetnames:
        del wb[ONEDRIVE_SHEET_NAME]

    ws = wb.create_sheet(ONEDRIVE_SHEET_NAME)

    # ── Headers (tinted by ownership: gray = AI-managed, blue = approver-editable) ──
    _ai_hdr_fill   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _user_hdr_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    for col_idx, header in enumerate(APPROVAL_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = _user_hdr_fill if header in _USER_EDITABLE_HEADERS else _ai_hdr_fill

    # ── NO Excel Table here ───────────────────────────────────────────────────
    # A header-only openpyxl table (zero data rows) makes Microsoft Graph reject the
    # ENTIRE workbook with 501 UnsupportedWorkbook/FileCorruptTryRepair (verified:
    # two header-only tables in one file => 501; same tables with >=1 data row => 200).
    # The ApprovalTable is therefore created via the Graph API AFTER upload, by
    # _ensure_approval_table_via_api(), where Graph manages an empty table cleanly.

    # ── Dropdown on Action column ─────────────────────────────────────────────
    action_letter = chr(ord("A") + _COL_ACTION)   # "J"
    dv = DataValidation(
        type="list",
        formula1='"Approve,Reject,On-hold"',
        allow_blank=True,
        showDropDown=False,   # False = show the dropdown arrow in Excel
    )
    dv.error      = "Select Approve, Reject, or On-hold"
    dv.errorTitle = "Invalid action"
    dv.sqref      = f"{action_letter}2:{action_letter}10000"
    ws.add_data_validation(dv)

    # ── Row colors by Action value ────────────────────────────────────────────
    full_range = f"A2:{_END_COL}10000"
    for action_val, hex_color in [("Approve", "C6EFCE"), ("Reject", "FFC7CE"), ("On-hold", "FFEB9C")]:
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        ws.conditional_formatting.add(
            full_range,
            FormulaRule(formula=[f'${action_letter}2="{action_val}"'], fill=fill),
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    try:
        _upload_workbook_bytes(buf.read())
        log.info(
            "_setup_full_sheet_via_openpyxl: Approvals sheet created — %d cols, dropdown + row colors applied",
            _COL_COUNT,
        )
        _ensure_approval_table_via_api()
    except Exception as exc:
        log.warning("_setup_full_sheet_via_openpyxl: upload failed: %s", exc)


def _ensure_approval_table_via_api() -> None:
    """Create the ApprovalTable over the header row via the Graph API (idempotent).

    Must run AFTER the file is uploaded and the workbook API is healthy. We create
    the table here rather than in openpyxl because a header-only openpyxl table makes
    Graph reject the whole workbook (501). Graph manages an empty table fine.
    The runtime's append_approval_row()/get_pending_approvals() need this table.
    """
    base = _wb_base()
    h = _headers()
    try:
        existing = _graph_get_retry(
            f"{base}/worksheets/{ONEDRIVE_SHEET_NAME}/tables", h, timeout=15.0
        )
        if existing is not None and existing.status_code == 200:
            names = [t.get("name") for t in existing.json().get("value", [])]
            if ONEDRIVE_TABLE_NAME in names:
                log.info("ApprovalTable already present — skipping create")
                return
        add = httpx.post(
            f"{base}/worksheets/{ONEDRIVE_SHEET_NAME}/tables/add",
            headers=h, json={"address": f"A1:{_END_COL}1", "hasHeaders": True}, timeout=20.0,
        )
        add.raise_for_status()
        tid = add.json().get("id")
        if tid:
            httpx.patch(f"{base}/tables/{tid}", headers=h,
                        json={"name": ONEDRIVE_TABLE_NAME}, timeout=15.0)
        log.info("ApprovalTable created via Graph API")
    except Exception as exc:
        log.warning("_ensure_approval_table_via_api failed (non-fatal): %s", exc)

    _neutralize_table_and_color_header()


def _neutralize_table_and_color_header() -> None:
    """Make AI columns visibly GRAY and approver columns BLUE on the Approvals tab.

    Excel's default table style ("TableStyleMedium2") paints the whole table blue, which
    hides the per-column ownership tint. This switches the table to a neutral style with
    banding OFF, then tints the header row by ownership (gray = AI-managed, blue = editable).
    Data rows are tinted per-row by _apply_row_colors() as they are appended.

    Idempotent and self-healing: a GET-guard skips the work once the style is already
    neutral, so it is cheap to call on every pipeline run. Uses plain (non-session) headers
    and never raises — formatting must never break the pipeline.
    """
    base = _wb_base()
    h    = _headers()
    try:
        r = _graph_get_retry(
            f"{base}/worksheets/{ONEDRIVE_SHEET_NAME}/tables/{ONEDRIVE_TABLE_NAME}",
            headers=h, timeout=15.0,
        )
        if r is None or not r.is_success:
            return
        tbl = r.json()
        already_neutral = (
            tbl.get("style") == _APPROVAL_TABLE_STYLE
            and tbl.get("showBandedRows") is False
        )
        if already_neutral:
            return   # nothing to do — already gray/blue

        tid = tbl.get("id")
        if tid:
            httpx.patch(
                f"{base}/tables/{tid}", headers=h,
                json={
                    "style": _APPROVAL_TABLE_STYLE,
                    "showBandedRows": False,
                    "showBandedColumns": False,
                },
                timeout=15.0,
            )
        # Tint the header row by ownership (gray AI / blue approver). Manual cell fills
        # override the table style, so the ownership colors win regardless of the style.
        for c1, c2, color in _color_runs():
            try:
                httpx.patch(
                    f"{base}/worksheets/{ONEDRIVE_SHEET_NAME}/range(address='{c1}1:{c2}1')/format/fill",
                    headers=h, json={"color": color}, timeout=10.0,
                )
            except Exception as exc:
                log.debug("header color run %s:%s failed (non-fatal): %s", c1, c2, exc)
        log.info("_neutralize_table_and_color_header: table style neutralized + header tinted by ownership")
    except Exception as exc:
        log.debug("_neutralize_table_and_color_header failed (non-fatal): %s", exc)


def ensure_pricing_rules_sheet() -> None:
    """Create the 'Pricing Rules' tab using Graph API (no file upload — works even when Excel is open).

    This tab is user-editable. Add rows directly in Excel:
      Service Pattern | County | Client Pattern | Price ($) | Priority | Notes
      Boundary Survey | Hillsborough | Hillsborough Title | 550.00 | 1 | Standard rate
      * = wildcard (matches anything). Lower Priority = higher priority.
    """
    if _cache.get("od_pricing_rules_done"):
        return

    h    = _session_headers()
    base = _wb_base()

    r_sheets = httpx.get(f"{base}/worksheets", headers=h, timeout=15.0)
    r_sheets.raise_for_status()
    existing_sheets = [s["name"] for s in r_sheets.json().get("value", [])]

    if PRICING_RULES_SHEET_NAME in existing_sheets:
        _cache["od_pricing_rules_done"] = True
        return

    log.info("Pricing Rules tab missing — creating via Graph API")
    try:
        # 1. Create worksheet
        httpx.post(
            f"{base}/worksheets/add",
            headers=h, json={"name": PRICING_RULES_SHEET_NAME}, timeout=15.0,
        ).raise_for_status()

        # 2. Write header + seed row in one call
        end_col = chr(ord("A") + _PR_COL_COUNT - 1)
        seed_row = ["1", "Active", "*", "*", "*", 0.0, 999,
                    "EXAMPLE — delete this. Set Price=0 to keep AI pricing. * = match anything."]
        httpx.patch(
            f"{base}/worksheets/{PRICING_RULES_SHEET_NAME}/range(address='A1:{end_col}2')",
            headers=h, json={"values": [PRICING_RULES_HEADERS, seed_row]}, timeout=15.0,
        ).raise_for_status()

        # 3. Create structured table so /tables/rows reads work
        httpx.post(
            f"{base}/worksheets/{PRICING_RULES_SHEET_NAME}/tables/add",
            headers=h, json={"address": f"A1:{end_col}2", "hasHeaders": True}, timeout=15.0,
        ).raise_for_status()

        # 4. Rename the table to our known name
        r_tabs = httpx.get(
            f"{base}/worksheets/{PRICING_RULES_SHEET_NAME}/tables",
            headers=h, timeout=10.0,
        )
        if r_tabs.is_success:
            tables = r_tabs.json().get("value", [])
            if tables:
                httpx.patch(
                    f"{base}/tables/{tables[-1]['id']}",
                    headers=h, json={"name": PRICING_RULES_TABLE_NAME}, timeout=10.0,
                )

        # 5. Bold header row
        httpx.patch(
            f"{base}/worksheets/{PRICING_RULES_SHEET_NAME}/range(address='A1:{end_col}1')/format/font",
            headers=h, json={"bold": True}, timeout=10.0,
        )

        log.info("Pricing Rules tab created via Graph API")
        _cache["od_pricing_rules_done"] = True

    except Exception as exc:
        log.warning("ensure_pricing_rules_sheet failed (non-fatal): %s", exc)


# ── Pricing rules — in-memory cache + git-backed JSON file ───────────────────
_pricing_rules_cache: list | None = None

# Committed to git every pipeline run — zero OneDrive dependency for reads
_PRICING_RULES_FILE = os.path.normpath(
    os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "pricing_rules.json")
)


def _parse_rules_from_excel_rows(rows: list) -> list[dict]:
    """Parse raw Graph API table rows into rule dicts. Shared by fetch and sync."""
    rules = []
    for row in rows:
        vals = (row.get("values") or [[]])[0]
        if len(vals) < _PR_COL_COUNT:
            vals = list(vals) + [""] * (_PR_COL_COUNT - len(vals))

        if str(vals[1]).strip().lower() != "active":
            continue

        price = 0.0
        try:
            price = float(vals[5])
        except (ValueError, TypeError):
            pass

        priority = 999
        try:
            priority = int(float(vals[6]))
        except (ValueError, TypeError):
            pass

        rules.append({
            "rule_id":  str(vals[0]).strip(),
            "service":  str(vals[2]).strip() or "*",
            "county":   str(vals[3]).strip() or "*",
            "client":   str(vals[4]).strip() or "*",
            "price":    price,
            "priority": priority,
            "notes":    str(vals[7]).strip() if len(vals) > 7 else "",
        })
    rules.sort(key=lambda r: r["priority"])
    return rules


def get_pricing_rules() -> list[dict]:
    """Return active pricing rules, sorted by Priority (lower = higher priority).

    Read order:
      1. In-memory cache (fastest — populated by sync_pricing_rules_to_json at run start)
      2. data/pricing_rules.json (git-backed — survives OneDrive outages)
      3. OneDrive Excel API (fresh fetch — used when JSON file is absent)

    Each rule dict: {rule_id, service, county, client, price, priority, notes}
    Pattern values: '*' matches anything; otherwise case-insensitive substring match.
    """
    global _pricing_rules_cache
    if _pricing_rules_cache is not None:
        return _pricing_rules_cache

    # Try git-backed JSON file first (no API call needed)
    if os.path.exists(_PRICING_RULES_FILE):
        try:
            with open(_PRICING_RULES_FILE) as f:
                data = json.load(f)
            rules = data.get("rules", [])
            _pricing_rules_cache = rules
            log.info("get_pricing_rules: %d rules loaded from pricing_rules.json", len(rules))
            return rules
        except Exception as exc:
            log.warning("get_pricing_rules: JSON file unreadable (%s) — falling back to Excel", exc)

    # Fall back to Excel API
    try:
        ensure_pricing_rules_sheet()
        r = httpx.get(
            f"{_wb_base()}/worksheets/{PRICING_RULES_SHEET_NAME}/tables/{PRICING_RULES_TABLE_NAME}/rows",
            headers=_session_headers(),
            timeout=15.0,
        )
        if not r.is_success:
            log.warning("get_pricing_rules: table read failed (%d) — no rules applied", r.status_code)
            _pricing_rules_cache = []
            return []

        rules = _parse_rules_from_excel_rows(r.json().get("value", []))
        _pricing_rules_cache = rules
        log.info("get_pricing_rules: %d active rules loaded from Excel", len(rules))
        return rules

    except Exception as exc:
        log.warning("get_pricing_rules failed — no rules applied: %s", exc)
        _pricing_rules_cache = []
        return []


def sync_pricing_rules_to_json() -> int:
    """Fetch latest pricing rules from Excel and write to data/pricing_rules.json.

    Called once at pipeline start (A0). Subsequent calls to get_pricing_rules()
    within the same run hit the in-memory cache — no further API calls.

    If OneDrive is unreachable, the existing pricing_rules.json (committed to git
    from the previous run) is left untouched and used as fallback.

    Returns number of active rules synced.
    """
    global _pricing_rules_cache
    # Clear cache so we force a fresh fetch from Excel
    _pricing_rules_cache = None

    try:
        ensure_pricing_rules_sheet()
        r = httpx.get(
            f"{_wb_base()}/worksheets/{PRICING_RULES_SHEET_NAME}/tables/{PRICING_RULES_TABLE_NAME}/rows",
            headers=_session_headers(),
            timeout=15.0,
        )
        if not r.is_success:
            log.warning("sync_pricing_rules_to_json: Excel read failed (%d) — keeping existing JSON", r.status_code)
            return 0

        rules = _parse_rules_from_excel_rows(r.json().get("value", []))

        os.makedirs(os.path.dirname(_PRICING_RULES_FILE), exist_ok=True)
        with open(_PRICING_RULES_FILE, "w") as f:
            json.dump({"rules": rules, "synced_at": datetime.now(_EASTERN).isoformat()}, f, indent=2)

        _pricing_rules_cache = rules
        log.info("sync_pricing_rules_to_json: %d rules synced → pricing_rules.json", len(rules))
        return len(rules)

    except Exception as exc:
        log.warning("sync_pricing_rules_to_json failed — existing JSON unchanged: %s", exc)
        return 0


def match_pricing_rule(service: str, county: str, client: str) -> dict | None:
    """Return the highest-priority matching rule, or None if no match.

    Matching logic (all three must match):
      * = wildcard (matches any non-empty string)
      Otherwise: case-insensitive substring match
      Rule with price=0 means "keep AI pricing" (no override).
    """
    s_lower = service.lower()
    c_lower = county.lower()
    cl_lower = client.lower()

    for rule in get_pricing_rules():
        svc_pat = rule["service"]
        cty_pat = rule["county"]
        cli_pat = rule["client"]

        svc_match = (svc_pat == "*") or (svc_pat.lower() in s_lower)
        cty_match = (cty_pat == "*") or (cty_pat.lower() in c_lower)
        cli_match = (cli_pat == "*") or (cli_pat.lower() in cl_lower)

        if svc_match and cty_match and cli_match:
            return rule

    return None


def ensure_approval_sheet() -> None:
    """Ensure the Approvals sheet and ApprovalTable exist with the correct schema.

    Uses plain auth headers (no workbook session) for read-only checks so the session
    is never opened before a potential file upload (_setup_full_sheet_via_openpyxl uploads
    the raw file — which returns 423 Locked if a session is held).
    """
    if _cache.get("od_formatting_done"):
        return   # Already verified this session

    base = _wb_base()
    h    = _headers()   # intentionally NOT _session_headers() — avoid opening a session

    # ── Check sheet + table ───────────────────────────────────────────────────
    needs_setup = False

    r_sheets = _graph_get_retry(f"{base}/worksheets", headers=h, timeout=15.0)
    # Transient Graph workbook-API errors (e.g. 501 while re-indexing a just-uploaded file)
    # must NOT crash the caller. Skip the schema check this cycle and proceed — the file
    # itself is fine (it was written via /content); the API recovers shortly.
    if r_sheets is None or r_sheets.status_code in _TRANSIENT_STATUS:
        log.warning("ensure_approval_sheet: worksheets list unavailable (status %s) — "
                    "skipping schema check this cycle (Graph workbook API transient)",
                    getattr(r_sheets, "status_code", "n/a"))
        _cache["od_formatting_done"] = True
        return
    r_sheets.raise_for_status()
    existing_sheets = [s["name"] for s in r_sheets.json().get("value", [])]

    if ONEDRIVE_SHEET_NAME not in existing_sheets:
        log.info("ensure_approval_sheet: sheet missing — will create via openpyxl")
        needs_setup = True
    else:
        r_cols = _graph_get_retry(
            f"{base}/worksheets/{ONEDRIVE_SHEET_NAME}/tables/{ONEDRIVE_TABLE_NAME}/columns",
            headers=h, timeout=15.0,
        )
        if r_cols is not None and r_cols.status_code in _TRANSIENT_STATUS:
            log.warning("ensure_approval_sheet: columns read transient (status %d) — skipping check this cycle",
                        r_cols.status_code)
            _cache["od_formatting_done"] = True
            return
        if not r_cols.is_success:
            log.info("ensure_approval_sheet: table missing or unreadable (status %d) — will recreate", r_cols.status_code)
            needs_setup = True
        else:
            actual_cols = len(r_cols.json().get("value", []))
            if actual_cols != _COL_COUNT:
                log.info(
                    "ensure_approval_sheet: schema mismatch (%d cols vs expected %d) — will recreate",
                    actual_cols, _COL_COUNT,
                )
                needs_setup = True

    if needs_setup:
        _close_session()   # release any server-side lock before raw file upload
        _setup_full_sheet_via_openpyxl()

    _cache["od_formatting_done"] = True
    # Always ensure the guide tab is current (version-gated — skips if already up to date)
    ensure_guide_sheet()
    # Ensure the user-editable Pricing Rules tab exists
    ensure_pricing_rules_sheet()
    # Ensure the plain-language How-To tab is current (version-gated)
    ensure_howto_sheet()
    # Self-heal the Action-column dropdown (validation can be stripped by row deletes)
    ensure_action_dropdown()
    # Self-heal column ownership colors: gray = AI-managed, blue = approver-editable
    # (neutralizes Excel's default all-blue table style; idempotent GET-guard)
    _neutralize_table_and_color_header()


def ensure_action_dropdown() -> bool:
    """Maintain the two things that require an openpyxl download/upload (the Graph
    workbook REST API exposes neither): (1) the Approve/Reject/On-hold list validation on
    the Action column, and (2) NATIVE clickable hyperlinks on each Order ID cell (col A).
    Both are done in ONE download/save/upload so the dropdown, conditional formatting and
    links always stay consistent, and the upload happens only when something changed.

    Returns True if the dropdown/links are present/correct at the end of the call (either
    already correct, or successfully re-applied), False if it could not be applied
    (e.g. upload 423-locked because the file is open in Excel). Callers can surface
    this to the operator instead of falsely reporting success.

    The dropdown is normally written once by _setup_full_sheet_via_openpyxl, but Excel
    drops the validation when whole table rows are deleted (e.g. a full sheet reset).
    Since the sheet/table still exist after that, _setup_full_sheet_via_openpyxl won't
    re-run — so the dropdown silently disappears.

    The Graph workbook REST API does NOT expose data validation, so this re-applies it
    the only way possible: download the file, add the validation via openpyxl (only if
    it is actually missing), and re-upload. Idempotent — when the dropdown already
    covers column J it does nothing (no upload, no 423-lock risk). The PUT upload needs
    the file closed/unlocked; if it is open in Excel the upload may 423 — logged as
    non-fatal so the pipeline still runs.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.worksheet.datavalidation import DataValidation

    action_letter = chr(ord("A") + _COL_ACTION)   # "J"
    addr          = f"{action_letter}2:{action_letter}10000"

    try:
        _close_session()   # release lock before any potential upload
        raw = _download_workbook_bytes()
        wb  = openpyxl.load_workbook(io.BytesIO(raw))
        if ONEDRIVE_SHEET_NAME not in wb.sheetnames:
            log.info("ensure_action_dropdown: Approvals sheet absent — nothing to do")
            return False
        ws = wb[ONEDRIVE_SHEET_NAME]

        changed = False

        # ── 1. Action dropdown (list validation) ──────────────────────────────────
        # Idempotency: skip only if a list validation covers the EXACT canonical range.
        # Graph row-deletes shift/shrink the range (e.g. J2:J10000 → J3:J9991), which
        # leaves the first data row (J2) with no dropdown — so anything that isn't an
        # exact match must be removed and rebuilt.
        existing = [
            d for d in list(ws.data_validations.dataValidation)
            if d.type == "list" and action_letter in str(d.sqref)
        ]
        if not (len(existing) == 1 and str(existing[0].sqref).replace(" ", "") == addr):
            for d in existing:
                ws.data_validations.dataValidation.remove(d)   # drop drifted/partial ranges
            dv = DataValidation(
                type="list",
                formula1='"Approve,Reject,On-hold"',
                allow_blank=True,
                showDropDown=False,   # False = arrow IS shown in Excel
            )
            dv.error      = "Select Approve, Reject, or On-hold"
            dv.errorTitle = "Invalid action"
            dv.sqref      = addr
            ws.add_data_validation(dv)
            # Re-apply the Action-value row colors too (conditional formatting is range-based)
            for action_val, hex_color in [("Approve", "C6EFCE"), ("Reject", "FFC7CE"), ("On-hold", "FFEB9C")]:
                fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                ws.conditional_formatting.add(
                    f"A2:{_END_COL}10000",
                    FormulaRule(formula=[f'${action_letter}2="{action_val}"'], fill=fill),
                )
            changed = True
            log.info("ensure_action_dropdown: re-applied list validation on %s", addr)

        # ── 2. Order ID clickable links (col A) ────────────────────────────────────
        # NATIVE Excel cell hyperlinks — NOT =HYPERLINK() formulas. A formula in a table
        # column is treated by Excel as a "calculated column" and auto-propagated across
        # the WHOLE column, which collapses every Order ID to the last one written. A
        # native cell hyperlink carries no formula, moves with the row on sort, and never
        # propagates. Set per row only when missing/stale so the upload stays idempotent.
        link_font = Font(color="0563C1", underline="single")
        oid_col   = _COL_ORDER_ID + 1   # openpyxl is 1-based
        map_col   = _COL_MAP_LINK + 1   # Map Link column — made clickable the same way
        for rr in range(2, ws.max_row + 1):
            cell = ws.cell(row=rr, column=oid_col)
            oid  = str(cell.value or "").strip()
            if oid and not oid.startswith("="):
                want = f"{FTF_ORDER_URL}/?order={oid}"
                cur  = cell.hyperlink.target if cell.hyperlink else None
                if cur != want:
                    cell.hyperlink = want
                    cell.font      = link_font
                    changed = True

            # Map Link (col F): the cell value IS the target URL — set a native hyperlink so
            # it is clickable. Same native-link approach as Order ID (never a formula, so it
            # can't propagate across the table column). Blank/non-URL cells are left alone.
            mcell = ws.cell(row=rr, column=map_col)
            murl  = str(mcell.value or "").strip()
            if murl.lower().startswith("http"):
                mcur = mcell.hyperlink.target if mcell.hyperlink else None
                if mcur != murl:
                    mcell.hyperlink = murl
                    mcell.font      = link_font
                    changed = True

        if not changed:
            log.debug("ensure_action_dropdown: dropdown + Order ID links already current — skip upload")
            return True

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        _upload_workbook_bytes(buf.read())
        log.info("ensure_action_dropdown: uploaded via openpyxl (dropdown/Order-ID links refreshed)")
        return True
    except Exception as exc:
        log.warning("ensure_action_dropdown failed (non-fatal): %s", exc)
        return False


def ensure_guide_sheet() -> None:
    """Create or update the 'Pipeline Guide' tab using Graph API.

    No file download/upload — works even when Excel is open in browser or desktop.
    Version-gated: skips the write if the tab already has the current _GUIDE_VERSION.
    """
    # ── Version check — skip if already current ───────────────────────────────
    try:
        h    = _session_headers()
        base = _wb_base()
        r = httpx.get(
            f"{base}/worksheets/{GUIDE_SHEET_NAME}/range(address='B2')",
            headers=h, timeout=10.0,
        )
        if r.is_success:
            existing_version = (r.json().get("values") or [[""]])[0][0]
            if str(existing_version).strip() == _GUIDE_VERSION:
                log.debug("guide sheet already at %s — skipping update", _GUIDE_VERSION)
                return
    except Exception:
        pass

    log.info("guide sheet missing or outdated — writing %s via Graph API", _GUIDE_VERSION)

    try:
        h    = _session_headers()
        base = _wb_base()

        # Ensure worksheet exists
        r_sheets = httpx.get(f"{base}/worksheets", headers=h, timeout=15.0)
        r_sheets.raise_for_status()
        existing = [s["name"] for s in r_sheets.json().get("value", [])]

        if GUIDE_SHEET_NAME not in existing:
            httpx.post(
                f"{base}/worksheets/add",
                headers=h, json={"name": GUIDE_SHEET_NAME}, timeout=15.0,
            ).raise_for_status()
            log.info("guide sheet tab created")
        else:
            httpx.post(
                f"{base}/worksheets/{GUIDE_SHEET_NAME}/range(address='A1:B100')/clear",
                headers=h, json={"applyTo": "Contents"}, timeout=15.0,
            )

        stamp = f"{_GUIDE_VERSION} — {datetime.now(_EASTERN).strftime('%Y-%m-%d %H:%M %Z')}"
        rows = [
            ["FTF Invoice Pipeline — Field Guide", ""],
            ["Version", stamp],
            ["", ""],
            ["── COLUMN REFERENCE (Approvals tab) ──", ""],
            ["COLOR KEY",         "GRAY columns are AI-managed — please do NOT edit them. BLUE columns are yours: Service / Breakdown by User, Action, Notes, Learning provided by user."],
            ["Order ID",          "Unique FTF order number — CLICK IT to open this order in FieldToFinish. (gray / AI-managed)"],
            ["Order Status",      "Current FTF status (In Progress, Complete, Field, etc.). (gray / AI-managed)"],
            ["Client Name",       "Client or title company who placed the order (title company / one-off client). (gray / AI-managed)"],
            ["Property Address",  "Survey site address — the property to be surveyed. (gray / AI-managed)"],
            ["Property Size",     "Lot size of the property (acres or sq ft) from the FTF order / county appraiser. Reference only. (gray / AI-managed)"],
            ["Map Link",          "CLICKABLE link to the parcel on Regrid (app.regrid.com) — opens the property map/parcel. Uses the parcel number when known, otherwise an address search. Reference only. (gray / AI-managed)"],
            ["FEMA Zone",         "FEMA flood zone for the property (e.g. X, AE, VE). Drives elevation-certificate scope/pricing. Reference only. (gray / AI-managed)"],
            ["Service Type",      "The survey service requested on the order (e.g. Boundary Survey, Elevation Certificate). Reference only. (gray / AI-managed)"],
            ["Service / Breakdown by AI", "GRAY / locked. The AI's proposed services + prices, e.g. 'Boundary Survey: $475.00 | Elevation Cert: $150.00'. Reference only — do NOT edit. The AI keeps this to learn from your changes."],
            ["Amount ($) by AI",  "GRAY / locked. The AI's proposed total. Reference only — do NOT edit."],
            ["Service / Breakdown by User", "BLUE / editable — THIS is what gets invoiced (name AND price). Starts as a copy of the AI breakdown. To change, rename, add or remove a service, edit this cell using the STANDARD FORMAT (see the 'SERVICE ENTRY STANDARD' section below), then set Action = Approve. The invoice is regenerated from exactly what you type here — the service name you enter is the name that prints on the client invoice."],
            ["Amount ($) by User","GRAY / AUTO-CALCULATED. The total of 'Service / Breakdown by User' — refreshed automatically every run. Do NOT type here; change the breakdown instead."],
            ["Confidence",        "HIGH = very likely correct. MEDIUM = reasonable. LOW = limited data. N/A = condo or manual pricing. (gray)"],
            ["Escalate",          "Yes = AI flagged an unusual order; Robert or Ryan should review before approving. (gray)"],
            ["Action",            "BLUE / YOUR DECISION — Approve / Reject / On-hold. Leave blank to defer. Pipeline checks every few minutes."],
            ["Notes",             "BLUE / editable. The AI may pre-fill the reason for escalation/required action — read it. Anything YOU add here is also read by the AI as learning for this order/client."],
            ["Posted At",         "Date/time the pipeline posted this row (Eastern Time). (gray)"],
            ["Processed At",      "Auto-filled when the pipeline processes your decision. Once filled = complete. (gray)"],
            ["AI Learning",       "Gray / AI-managed (read-only). The AI's one-line learning note for this order — what it priced vs. what was actually charged, and its takeaway."],
            ["Learning provided by user", "BLUE / YOURS. Type anything you want the AI to learn for this order or client (e.g. 'this client is always $1,800 for a boundary'). The AI RE-CHECKS this column every run: whenever you add or change the text it learns it against THIS order and applies the same logic to SIMILAR future orders (same client / service / county)."],
            ["", ""],
            ["── SERVICE ENTRY STANDARD (Service / Breakdown by User) ──", ""],
            ["The rule",          "Type each service as  'Service Name: $Amount'  and separate multiple services with a pipe ' | '. Whatever you type here is EXACTLY what gets invoiced — both the service NAME and the amount print on the client invoice."],
            ["Format",            "Service Name: $Amount  |  Service Name: $Amount"],
            ["One service",       "Boundary Survey: $500.00"],
            ["Two services",      "Boundary Survey: $500.00 | Elevation Certificate: $150.00"],
            ["Rename a service",  "Just change the text before the colon, e.g. 'Boundary & Topographic Survey: $650.00'. That new name is what appears on the invoice."],
            ["Change a price",    "Edit the number after the $, e.g. 'Boundary Survey: $525.00'. 'Amount ($) by User' re-totals automatically."],
            ["Add a service",     "Add ' | New Service: $Amount' to the end of the line."],
            ["Remove a service",  "Delete that 'Name: $Amount' segment (and its surrounding ' | '). The removed line will not appear on the invoice."],
            ["Then",              "Set Action = Approve. The invoice is generated from this cell — your changes (names, prices, added/removed lines) all carry through."],
            ["Rules to keep",     "Always include the '$' and a colon between name and amount. Use a plain number (525 or 525.00), no letters. One space each side of the pipe ' | '."],
            ["", ""],
            ["── ACTION GUIDE ──", ""],
            ["Approve",           "Pipeline creates a real FTF invoice from 'Service / Breakdown by User' and emails the client. CANNOT be undone from pipeline."],
            ["Reject",            "No invoice created. No email sent. The AI LEARNS from it — add a reason in Notes if you can (optional). Manually clear ng_invoice_needed=1 in FTF if the order should not re-appear."],
            ["On-hold",           "Pipeline pauses this order (no invoice/email). The AI learns from it. Change to Approve or Reject when ready."],
            ["(leave blank)",     "Pipeline ignores this row every 30-min cycle until you select an action."],
            ["", ""],
            ["── NOTES FIELD GUIDE ──", ""],
            ["CONDO ORDER —",             "Cannot survey. Row is AUTO-REJECTED. Contact client — arrange refund or redirect to interior measurement."],
            ["⛔ CANCELED —",             "Order canceled in FTF. No invoice needed. Flagged red, not priced. Set Action = Reject to clear."],
            ["⚠️ DELIVERED —",            "Order already delivered. Flagged red, not auto-priced. Verify if an invoice is still needed; enter amount manually if so."],
            ["MANUAL PRICING REQUIRED —", "AI could not price. Type the service + amount in 'Service / Breakdown by User' as 'Survey: $X' (e.g. 'Boundary Survey: $500.00'); the total fills in automatically. Then set Action = Approve."],
            ["ESCALATE —",               "Unusual order (large lot, commercial, FEMA zone, duplicate). Get Robert or Ryan to review."],
            ["(empty notes)",             "Standard order. AI is confident. Review amount and service, then approve if correct."],
            ["", ""],
            ["── HOW TO TEACH THE AI — Pricing Rules tab ──", ""],
            ["What it is",     "The 'Pricing Rules' tab sets fixed prices for specific clients, counties, or service types. No coding required."],
            ["When to use it", "Use when: AI keeps getting a price wrong, you have a negotiated flat rate for a client, or you know the correct price."],
            ["How to add a rule", "Go to 'Pricing Rules' tab → add a row → Status=Active → fill in patterns → save. * = wildcard (matches anything)."],
            ["Example",        "Boundary Survey | Hillsborough | Hillsborough Title | $550 | priority 1"],
            ["Priority",       "Lower number = higher priority. Client-specific: 1-10. County/service: 50+. Global: 999."],
            ["Price = 0",      "If Price is 0, rule matches but AI still sets the price. Set non-zero to override AI."],
            ["", ""],
            ["── PIPELINE STATUS DEFINITIONS ──", ""],
            ["invoice_needed",        "Order picked up from FTF — queued for A2 data collection."],
            ["data_collected",        "A2 collected all data — ready for A3 pricing and Excel posting."],
            ["invoice_draft_posted",  "AI priced and posted to this sheet — awaiting your action."],
            ["pricing_needed",        "AI could not price. Row posted — enter amount manually."],
            ["condo_rejected",        "Condo detected — cannot survey. Posted as auto-rejected. Contact client."],
            ["canceled_flagged",      "Order canceled in FTF. Flagged red — no pricing performed. Set Action = Reject to clear."],
            ["delivered_flagged",     "Order already delivered. Flagged red — no automatic pricing. Verify if an invoice is still needed."],
            ["invoice_approved",      "You approved — pipeline creating FTF invoice (A5 running)."],
            ["invoice_sent",          "Invoice created in FTF and emailed to the client (A6 complete)."],
            ["already_invoiced",      "Order already had an invoice in FTF — skipped (not re-posted, not re-invoiced)."],
            ["invoice_rejected",      "Rejected (or auto-rejected: condo). No invoice created or sent."],
            ["on_hold",               "You selected On-hold. Pipeline paused for this order."],
            ["details_missing",       "FTF has insufficient data. Update in FTF or handle manually."],
            ["permanently_excluded",  "Order will never be processed — canceled in FTF (ng_status=0), internal email, etc."],
            ["", ""],
            ["── PIPELINE FLOW (prod server: intake every 30 min, approvals every ~5 min) ──", ""],
            ["A1 — Flag Hunter",      "Scans FTF DB for orders with ng_invoice_needed=1. Queues new orders."],
            ["A2 — Data Collector",   "Collects FTF API, emails, county appraiser, aerial image. AI builds order packet."],
            ["A3 — Invoice Compiler", "Detects condos and flags Canceled/Delivered orders (no pricing). Checks Pricing Rules tab first, then AI pricing. Posts to this sheet."],
            ["A4 — Human Gate",       "Reads your Action decision every ~5 min. Routes to A5 (approve) or rejected/on-hold."],
            ["A5 — Invoice Finalizer","Creates the real FTF invoice via API."],
            ["A6 — Email Sender",     "Delivers the invoice email to the client via the FTF portal (sent as 'nesa'). LIVE — emails go to real customers."],
            ["A7 — Feedback Learner", "Learns from your decisions to improve future pricing."],
        ]

        end_row = len(rows)
        httpx.patch(
            f"{base}/worksheets/{GUIDE_SHEET_NAME}/range(address='A1:B{end_row}')",
            headers=h, json={"values": rows}, timeout=30.0,
        ).raise_for_status()

        # Bold the section header rows (lines starting with "──")
        section_rows = [i + 1 for i, row in enumerate(rows) if str(row[0]).startswith("──")]
        for sr in section_rows:
            try:
                httpx.patch(
                    f"{base}/worksheets/{GUIDE_SHEET_NAME}/range(address='A{sr}:B{sr}')/format/font",
                    headers=h, json={"bold": True}, timeout=10.0,
                )
            except Exception:
                pass

        # Title row bold + larger font
        try:
            httpx.patch(
                f"{base}/worksheets/{GUIDE_SHEET_NAME}/range(address='A1:B1')/format/font",
                headers=h, json={"bold": True, "size": 14}, timeout=10.0,
            )
        except Exception:
            pass

        log.info("guide sheet '%s' written via Graph API (%s)", GUIDE_SHEET_NAME, _GUIDE_VERSION)

    except Exception as exc:
        log.warning("guide sheet write failed (non-fatal): %s", exc)


def ensure_howto_sheet() -> None:
    """Create or update the 'How to use Invoicing agent' tab — a plain-language,
    step-by-step guide for end users, with a worked dummy example.

    Version-gated (cell B2). Uses Graph API only — works even when the file is open.
    """
    try:
        h    = _session_headers()
        base = _wb_base()
        r = httpx.get(
            f"{base}/worksheets/{HOWTO_SHEET_NAME}/range(address='B2')",
            headers=h, timeout=10.0,
        )
        if r.is_success:
            existing_version = (r.json().get("values") or [[""]])[0][0]
            if str(existing_version).strip() == _HOWTO_VERSION:
                log.debug("how-to sheet already at %s — skipping update", _HOWTO_VERSION)
                return
    except Exception:
        pass

    log.info("how-to sheet missing or outdated — writing %s via Graph API", _HOWTO_VERSION)

    try:
        h    = _session_headers()
        base = _wb_base()

        r_sheets = httpx.get(f"{base}/worksheets", headers=h, timeout=15.0)
        r_sheets.raise_for_status()
        existing = [s["name"] for s in r_sheets.json().get("value", [])]

        if HOWTO_SHEET_NAME not in existing:
            httpx.post(
                f"{base}/worksheets/add",
                headers=h, json={"name": HOWTO_SHEET_NAME}, timeout=15.0,
            ).raise_for_status()
            log.info("how-to sheet tab created")
        else:
            httpx.post(
                f"{base}/worksheets/{HOWTO_SHEET_NAME}/range(address='A1:B200')/clear",
                headers=h, json={"applyTo": "Contents"}, timeout=15.0,
            )

        stamp = f"{_HOWTO_VERSION} — {datetime.now(_EASTERN).strftime('%Y-%m-%d %H:%M %Z')}"
        rows = [
            ["How to Use the Invoicing Agent", ""],
            ["Version", stamp],
            ["", ""],
            ["What is this?", "The Invoicing Agent reads new survey orders from FieldToFinish, works out the price, and lists each one on the 'Approvals' tab. You review each order and choose Approve, Reject, or On-hold. When you Approve, it creates the invoice in FTF and emails it to the customer automatically."],
            ["Where do I work?", "On the 'Approvals' tab — that is the only place you take action. This tab is just instructions."],
            ["", ""],
            ["── THE 30-SECOND VERSION ──", ""],
            ["1.", "Open the 'Approvals' tab."],
            ["2.", "Read each row: Client, Property, and Service / Breakdown by User (this is what gets billed)."],
            ["3.", "If the breakdown looks right, set the 'Action' column to Approve."],
            ["4.", "Within ~30 minutes the agent invoices the customer and emails them. Done."],
            ["", ""],
            ["── YOUR DAILY WORKFLOW (step by step) ──", ""],
            ["Step 1 — Open Approvals", "Go to the 'Approvals' tab. Each row is one order waiting for your decision."],
            ["Step 2 — Review the order", "Check Client Name, Property Address and Service / Breakdown by User (the AI's own breakdown sits in the 'by AI' column for reference). Four reference columns help you proof-read the quote: Property Size (lot size), Map Link (click to open the parcel on Regrid), FEMA Zone (flood zone), and Service Type. Click the Order ID to open the order in FieldToFinish if you need more detail."],
            ["Step 3 — Check the price", "If the breakdown looks right, leave it. To change it, see 'FIXING A PRICE' below. 'Amount ($) by User' totals your breakdown automatically — you never type it."],
            ["Step 4 — Read the Notes", "The Notes column tells you if anything needs attention (manual pricing, escalation, condo, canceled, delivered)."],
            ["Step 5 — Decide", "Set the 'Action' column to Approve, Reject, or On-hold. Leave blank to skip for now."],
            ["Step 6 — Wait ~30 min", "The agent runs every 30 minutes. It picks up your decision and acts on it. 'Processed At' fills in when it is done."],
            ["", ""],
            ["── THE 3 ACTIONS (Action column) ──", ""],
            ["Approve", "Creates a REAL invoice in FTF and emails it to the customer. This cannot be undone from here."],
            ["Reject", "No invoice, no email. Use for orders that should not be billed."],
            ["On-hold", "Pauses the order. Come back and change it to Approve or Reject later."],
            ["(leave blank)", "The agent ignores the row until you choose an action."],
            ["", ""],
            ["── HOW TO READ A ROW (key columns) ──", ""],
            ["Service / Breakdown by AI", "GRAY / locked. The AI's proposed services + prices, e.g. 'Boundary Survey: $475.00 | Elevation Cert: $275.00'. Reference only — don't edit. The AI keeps it to learn from your changes."],
            ["Amount ($) by AI", "GRAY / locked. The total the AI proposes. Reference only — don't edit."],
            ["Service / Breakdown by User", "BLUE / editable — THIS is what gets invoiced. Starts as a copy of the AI breakdown; change a number to reprice that service, e.g. 'Boundary Survey: $500.00 | Elevation Cert: $275.00'. Keep the 'Name: $amount' format (| between services)."],
            ["Amount ($) by User", "GRAY / AUTO-CALCULATED. The total of your breakdown — fills in automatically every run. Don't type here; edit the breakdown instead."],
            ["AI Learning", "Gray / read-only. The AI's one-line note on what it learned from this order (its price vs. the real one). You don't act on this."],
            ["Learning provided by user", "BLUE / yours to fill. Type anything you want the AI to learn for this order or client — it re-checks this column every run, learns any new/changed note against this order, and applies the same logic to similar future orders. No coding needed."],
            ["COLOR KEY", "Gray cells = AI-managed, please don't edit. Blue cells = yours: Service / Breakdown by User, Action, Notes, Learning provided by user."],
            ["Confidence", "HIGH / MEDIUM / LOW — how sure the AI is about the price. LOW means double-check it."],
            ["Escalate", "Yes = unusual order; have Robert or Ryan look before approving."],
            ["Notes", "Plain-language reason or instruction from the agent. Always read this."],
            ["", ""],
            ["── WORKED EXAMPLE (a dummy order) ──", ""],
            ["1. The row appears", "Order 1000299001 | Sunshine Title | 123 Palm Ave, Naples FL | Service / Breakdown by User: 'Boundary Survey: $475.00 | Elevation Cert: $275.00' | Amount by User: $750.00 | Confidence: HIGH | Action: (blank)"],
            ["2. You review it", "The services and the $750.00 total look correct for a boundary survey + elevation certificate in Naples."],
            ["3. (Optional) fix price", "Say you agreed $700 with this client. Edit 'Service / Breakdown by User' to 'Boundary Survey: $425.00 | Elevation Cert: $275.00' — 'Amount ($) by User' updates to $700 automatically; the AI keeps its own $750 baseline and learns the difference."],
            ["4. You approve", "Set the Action cell to 'Approve'."],
            ["5. What happens next", "Within ~30 min the agent creates the invoice in FTF, emails it to Sunshine Title's address on file, and fills in 'Processed At'."],
            ["6. Result", "Sunshine Title receives a $700 invoice by email. You did nothing except review and click Approve."],
            ["", ""],
            ["── WHAT HAPPENS AFTER YOU APPROVE ──", ""],
            ["1. Human Gate (A4)", "Reads your Approve and locks in the amount."],
            ["2. Finalizer (A5)", "Creates the real invoice inside FieldToFinish."],
            ["3. Sender (A6)", "Emails the invoice to the customer (sent as 'nesa')."],
            ["Processed At", "Fills in automatically — that is your confirmation it is done."],
            ["", ""],
            ["── SPECIAL ROWS YOU MIGHT SEE ──", ""],
            ["MANUAL PRICING REQUIRED", "The AI could not set a price. Type the service + amount in 'Service / Breakdown by User' as 'Survey: $X' (e.g. 'Boundary Survey: $500.00'); the total fills in automatically. Then Approve."],
            ["CONDO — Cannot Survey", "A condo / airspace unit. Auto-rejected. Contact the client; do not approve."],
            ["⛔ CANCELED (red row)", "Order was canceled in FTF. No invoice needed. Set Action = Reject to clear it."],
            ["⚠️ DELIVERED (red row)", "Order already delivered. Not auto-priced. Only invoice if one is genuinely still owed — enter the amount and Approve."],
            ["ESCALATE", "Unusual order. Get Robert or Ryan to review before approving."],
            ["", ""],
            ["── SERVICE ENTRY STANDARD (how to type services) ──", ""],
            ["Format", "Type each service in 'Service / Breakdown by User' as  'Service Name: $Amount'  and separate services with ' | '.  Example: 'Boundary Survey: $500.00 | Elevation Certificate: $150.00'."],
            ["Why it matters", "The invoice is built from EXACTLY what you type — the service NAME you enter is the name that prints on the client's invoice, and the amount is what they are billed. If you rename, add, remove or reprice a service here and Approve, the invoice comes out with your changed values."],
            ["Rename", "Change the words before the colon — e.g. 'Boundary & Topographic Survey: $650.00'. The new name prints on the invoice."],
            ["Add / remove", "Add ' | New Service: $Amount' to add a line; delete a 'Name: $Amount' segment to remove one."],
            ["Keep it valid", "Always keep the colon and the '$', use a plain number (e.g. 500 or 500.00), and one space each side of the pipe ' | '. Then set Action = Approve."],
            ["", ""],
            ["── FIXING A PRICE / TEACHING THE AI ──", ""],
            ["One-off change", "Edit the amount(s) in 'Service / Breakdown by User' on the row (e.g. 'Boundary Survey: $700.00'); 'Amount ($) by User' totals it automatically. Then Approve."],
            ["Permanent rule", "Use the 'Pricing Rules' tab to set a fixed price for a client, county, or service — so the AI gets it right next time. No coding needed."],
            ["Example rule", "Service=Boundary Survey | County=Collier | Client=Sunshine Title | Price=700 | Priority=1 | Status=Active"],
            ["", ""],
            ["── TIMING ──", ""],
            ["How often", "The agent runs every 30 minutes, around the clock."],
            ["How many", "Up to 10 new orders are added to the Approvals tab each run."],
            ["Your decisions", "Picked up on the next run after you set the Action — usually within 30 minutes."],
            ["", ""],
            ["── GOLDEN RULES ──", ""],
            ["Approve = real email", "Approving sends a real invoice to a real customer. Only approve when the price is right."],
            ["When unsure", "Use On-hold, or ask Robert / Ryan. Nothing happens until you choose an action."],
            ["Always read Notes", "The agent flags anything unusual there."],
        ]

        end_row = len(rows)
        httpx.patch(
            f"{base}/worksheets/{HOWTO_SHEET_NAME}/range(address='A1:B{end_row}')",
            headers=h, json={"values": rows}, timeout=30.0,
        ).raise_for_status()

        # Column widths + wrap long text in column B
        try:
            httpx.patch(
                f"{base}/worksheets/{HOWTO_SHEET_NAME}/range(address='A1:A{end_row}')/format",
                headers=h, json={"columnWidth": 190}, timeout=10.0,
            )
            httpx.patch(
                f"{base}/worksheets/{HOWTO_SHEET_NAME}/range(address='B1:B{end_row}')/format",
                headers=h, json={"columnWidth": 680, "wrapText": True}, timeout=10.0,
            )
        except Exception:
            pass

        # Bold section header rows ("──")
        section_rows = [i + 1 for i, row in enumerate(rows) if str(row[0]).startswith("──")]
        for sr in section_rows:
            try:
                httpx.patch(
                    f"{base}/worksheets/{HOWTO_SHEET_NAME}/range(address='A{sr}:B{sr}')/format/font",
                    headers=h, json={"bold": True}, timeout=10.0,
                )
            except Exception:
                pass

        # Title row bold + large
        try:
            httpx.patch(
                f"{base}/worksheets/{HOWTO_SHEET_NAME}/range(address='A1:B1')/format/font",
                headers=h, json={"bold": True, "size": 16}, timeout=10.0,
            )
        except Exception:
            pass

        log.info("how-to sheet '%s' written via Graph API (%s)", HOWTO_SHEET_NAME, _HOWTO_VERSION)

    except Exception as exc:
        log.warning("how-to sheet write failed (non-fatal): %s", exc)


def auto_reject_condo_row(order_id: str) -> None:
    """Set Action = 'Reject' and Processed At on the condo row in the approval table.

    Called immediately after append_approval_row for condo orders so the row is
    auto-rejected without waiting for a human to click the dropdown.
    A4 will see Processed At is already filled and skip re-processing.
    """
    try:
        ensure_approval_sheet()
        r = httpx.get(
            f"{_wb_base()}/worksheets/{ONEDRIVE_SHEET_NAME}/tables/{ONEDRIVE_TABLE_NAME}/rows",
            headers=_session_headers(),
            timeout=15.0,
        )
        r.raise_for_status()
        rows = r.json().get("value", [])

        target_index = None
        for row in rows:
            vals = row.get("values", [[]])[0]
            if vals and str(vals[0]).strip() == str(order_id):
                target_index = row.get("index")
                break

        if target_index is None:
            log.warning("auto_reject_condo_row: order %s not found in Excel table", order_id)
            return

        excel_row  = target_index + 2   # 0-based table index + header + 1
        action_col = chr(ord("A") + _COL_ACTION)        # "K"
        proc_col   = chr(ord("A") + _COL_PROCESSED_AT)  # "N"
        stamped_at = datetime.now(_EASTERN).strftime("%Y-%m-%d %H:%M %Z")
        wb_b       = _wb_base()
        h          = _session_headers()

        # Update Action and Processed At in separate calls — they are not adjacent
        # (Notes and Posted At columns sit between them and must NOT be overwritten).
        httpx.patch(
            f"{wb_b}/worksheets/{ONEDRIVE_SHEET_NAME}/range(address='{action_col}{excel_row}')",
            headers=h, json={"values": [["Reject"]]}, timeout=10.0,
        ).raise_for_status()
        httpx.patch(
            f"{wb_b}/worksheets/{ONEDRIVE_SHEET_NAME}/range(address='{proc_col}{excel_row}')",
            headers=h, json={"values": [[stamped_at]]}, timeout=10.0,
        ).raise_for_status()

        log.info("auto_reject_condo_row: order %s auto-rejected in Excel row %d", order_id, excel_row)
    except Exception as exc:
        log.warning("auto_reject_condo_row failed order=%s (non-fatal): %s", order_id, exc)


# ── Public write API ──────────────────────────────────────────────────────────

def get_pending_order_ids(strict: bool = False) -> set:
    """Return order IDs that already have a row in the approval table with blank Action (awaiting decision).

    strict=True re-raises on a fetch failure instead of returning an empty set. A3 uses
    this so it can ABORT the batch rather than treat a failed read as "sheet empty" — which
    would re-post orders that are already in the sheet (duplicate rows).
    """
    try:
        ensure_approval_sheet()
        r = httpx.get(
            f"{_wb_base()}/worksheets/{ONEDRIVE_SHEET_NAME}/tables/{ONEDRIVE_TABLE_NAME}/rows",
            headers=_session_headers(),
            timeout=15.0,
        )
        r.raise_for_status()
        pending = set()
        for row in r.json().get("value", []):
            vals = row.get("values", [[]])[0]
            # Action column (index _COL_ACTION) is blank → awaiting decision
            if len(vals) > _COL_ACTION and not str(vals[_COL_ACTION]).strip():
                pending.add(str(vals[0]).strip())
        log.info("get_pending_order_ids: %d awaiting-action rows in Excel", len(pending))
        return pending
    except Exception as exc:
        log.warning("get_pending_order_ids failed (dedup disabled): %s", exc)
        if strict:
            raise
        return set()


def get_all_approval_order_ids(strict: bool = False) -> set:
    """Return ALL order IDs already in the approval table regardless of action status.

    Used by backfill scripts and A3 dedup to avoid writing duplicate rows for orders that
    were already posted. strict=True re-raises on a fetch failure (see get_pending_order_ids).
    """
    try:
        ensure_approval_sheet()
        r = httpx.get(
            f"{_wb_base()}/worksheets/{ONEDRIVE_SHEET_NAME}/tables/{ONEDRIVE_TABLE_NAME}/rows",
            headers=_session_headers(),
            timeout=15.0,
        )
        r.raise_for_status()
        ids = set()
        for row in r.json().get("value", []):
            vals = row.get("values", [[]])[0]
            if vals:
                ids.add(str(vals[0]).strip())
        log.info("get_all_approval_order_ids: %d total rows in Excel", len(ids))
        return ids
    except Exception as exc:
        log.warning("get_all_approval_order_ids failed: %s", exc)
        if strict:
            raise
        return set()


def _close_session() -> None:
    """Close the current workbook session on the Graph API server, releasing the file lock."""
    session_id = _cache.pop("od_session", None)
    if not session_id:
        return
    try:
        base = _wb_base()
        httpx.post(
            f"{base}/closeSession",
            headers={
                "Authorization": f"Bearer {_get_token()}",
                "workbook-session-id": session_id,
            },
            timeout=10.0,
        )
        log.debug("workbook session closed")
    except Exception as exc:
        log.debug("session close error (ignored): %s", exc)


def _download_workbook_bytes() -> bytes:
    """Download the workbook file content as raw bytes."""
    drive_id = _cache.get("od_drive_id")
    item_id  = _get_item_id()
    url = (
        f"{_GRAPH}/drives/{drive_id}/items/{item_id}/content"
        if drive_id else
        f"{_GRAPH}/users/{ONEDRIVE_FILE_USER}/drive/items/{item_id}/content"
    )
    r = httpx.get(url, headers={"Authorization": f"Bearer {_get_token()}"}, follow_redirects=True, timeout=30.0)
    r.raise_for_status()
    return r.content


def _upload_workbook_bytes(data: bytes) -> None:
    """Replace the workbook file with new content (PUT upload). Retries once on 423.

    The bytes are first converted to a shared-string workbook. openpyxl writes inline
    strings with NO sharedStrings.xml, which Microsoft Graph's Excel (workbook) API
    REJECTS with HTTP 501 UnsupportedWorkbook/FileCorruptTryRepair — breaking every
    /workbook/* call until a human re-saves the file in Excel. to_shared_strings()
    makes the upload Graph-compatible up front so that never happens again. Idempotent
    on files that already have a shared-string table.
    """
    import time
    from core.xlsx_shared_strings import to_shared_strings
    try:
        data = to_shared_strings(data)
    except Exception as exc:
        log.warning("shared-string conversion failed (uploading as-is): %s", exc)
    drive_id = _cache.get("od_drive_id")
    item_id  = _get_item_id()
    url = (
        f"{_GRAPH}/drives/{drive_id}/items/{item_id}/content"
        if drive_id else
        f"{_GRAPH}/users/{ONEDRIVE_FILE_USER}/drive/items/{item_id}/content"
    )
    put_headers = {"Authorization": f"Bearer {_get_token()}", "Content-Type": "application/octet-stream"}
    for attempt in range(3):
        r = httpx.put(url, headers=put_headers, content=data, timeout=60.0)
        if r.status_code == 423:
            wait = (attempt + 1) * 4
            log.warning("upload 423 Locked — retrying in %ds (attempt %d/3)", wait, attempt + 1)
            time.sleep(wait)
            put_headers["Authorization"] = f"Bearer {_get_token()}"  # refresh token
            continue
        r.raise_for_status()
        # Invalidate session — file changed on disk
        _cache.pop("od_session", None)
        log.info("workbook re-uploaded (%d bytes)", len(data))
        return
    raise AgentError("workbook upload failed: file locked after 3 retries (423)")


def _col_letter(idx: int) -> str:
    return chr(ord("A") + idx)


def _color_runs() -> list:
    """Contiguous column runs grouped by owner fill-color: (start_letter, end_letter, color).
    Lets a whole row be tinted in a few PATCHes (one per run) instead of 15."""
    runs = []
    start = 0
    cur = None
    for i, hdr in enumerate(APPROVAL_HEADERS):
        color = _USER_EDIT_FILL if hdr in _USER_EDITABLE_HEADERS else _AI_LOCKED_FILL
        if cur is None:
            cur, start = color, i
        elif color != cur:
            runs.append((_col_letter(start), _col_letter(i - 1), cur))
            cur, start = color, i
    runs.append((_col_letter(start), _col_letter(len(APPROVAL_HEADERS) - 1), cur))
    return runs


def _apply_row_colors(excel_row: int) -> None:
    """Tint one row by column ownership (gray = AI-managed, blue = approver-editable).
    Overwrites any prior fill, so a recycled row position never keeps a stale color."""
    wb = _wb_base(); h = _session_headers()
    for c1, c2, color in _color_runs():
        try:
            httpx.patch(
                f"{wb}/worksheets/{ONEDRIVE_SHEET_NAME}/range(address='{c1}{excel_row}:{c2}{excel_row}')/format/fill",
                headers=h, json={"color": color}, timeout=10.0,
            ).raise_for_status()
        except Exception as exc:  # noqa: BLE001
            log.debug("row %d color run %s:%s failed (non-fatal): %s", excel_row, c1, c2, exc)


def _format_new_row(table_row_index: int, ftf_link: str = "", order_id: str = "") -> None:
    """Borders + currency + clickable Order-ID hyperlink + ownership colors on a new row."""
    excel_row = table_row_index + 2  # 0-based table index + header row + 1
    row_range = f"A{excel_row}:{_END_COL}{excel_row}"
    wb = _wb_base()
    h  = _session_headers()

    for side in ["EdgeTop", "EdgeBottom", "EdgeLeft", "EdgeRight", "InsideVertical"]:
        httpx.patch(
            f"{wb}/worksheets/{ONEDRIVE_SHEET_NAME}/range(address='{row_range}')/format/borders/{side}",
            headers=h,
            json={"style": "Continuous", "weight": "Thin", "color": "#000000"},
            timeout=10.0,
        ).raise_for_status()

    # Currency format on BOTH amount columns: "Amount ($) by AI" (F) and "Amount ($) by
    # User" (H). They are NOT adjacent anymore — "Service / Breakdown by User" (G, text)
    # sits between them — so format each cell separately (a single F:H range would also
    # number-format the text breakdown in G).
    for amt_idx in (_COL_AMOUNT_AI, _COL_AMOUNT_USER):
        amt_letter = _col_letter(amt_idx)             # "F" / "H"
        httpx.patch(
            f"{wb}/worksheets/{ONEDRIVE_SHEET_NAME}/range(address='{amt_letter}{excel_row}')/format",
            headers=h,
            json={"numberFormat": [["$#,##0.00"]]},
            timeout=10.0,
        ).raise_for_status()

    # Order ID stays a PLAIN VALUE — NEVER a =HYPERLINK() formula. Excel treats a formula
    # inside a table column as a "calculated column" and auto-propagates ONE row's formula
    # across the WHOLE column. Because the link is a per-row literal (order=<id>), that
    # propagation collapsed every Order ID to the last one written — the "same order id in
    # multiple rows" bug. Plain values never propagate. We (re)write the value here so the
    # cell also overwrites any formula a previously-propagated cell may have inherited.
    # (Clickability is handled out-of-band, not via an in-cell formula — see notes.)
    if order_id:
        httpx.patch(
            f"{wb}/worksheets/{ONEDRIVE_SHEET_NAME}/range(address='{_col_letter(_COL_ORDER_ID)}{excel_row}')",
            headers=h,
            json={"values": [[str(order_id)]]},
            timeout=10.0,
        ).raise_for_status()

    _apply_row_colors(excel_row)
    log.debug("row %d formatted (borders + currency + plain order-id value + colors)", excel_row)


def _record_ai_prefilled_note(order_id: str, notes: str) -> None:
    """Remember the Notes text the AI pre-fills on a row (escalation/condo/etc.), so
    consume_user_learnings() can tell AI-written Notes from operator-written ones and NEVER
    learn from the AI's own words. Stored in learned_rules.json['ai_prefilled_notes']. Non-fatal.
    """
    note = (notes or "").strip()
    if not order_id or not note:
        return
    path = os.path.normpath(
        os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "learned_rules.json")
    )
    try:
        data = {}
        if os.path.exists(path):
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
        prefills = data.setdefault("ai_prefilled_notes", {})
        if prefills.get(order_id) == note:
            return   # already recorded
        prefills[order_id] = note
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, default=str)
    except Exception as exc:  # noqa: BLE001
        log.debug("record ai_prefilled_note failed order=%s (non-fatal): %s", order_id, exc)


def append_approval_row(
    order_id:     str,
    client_name:  str,
    address:      str,
    service:      str,
    amount:       float,
    confidence:   str,
    escalate:     bool,
    ftf_link:     str,
    order_status: str = "",
    posted_at:    Optional[str] = None,
    notes:        str = "",
    highlight_red: bool = False,
    amount_user:  Optional[float] = None,
    ai_learning:  str = "",
    service_user: Optional[str] = None,
    property_size: str = "",
    map_link:     str = "",
    fema_zone:    str = "",
    service_type: str = "",
) -> None:
    """Append a new row to the approval table. Action column is blank — user picks from dropdown.

    service       → written to BOTH "Service / Breakdown by AI" (E, locked baseline) AND
                    "Service / Breakdown by User" (G, editable copy). G starts identical to
                    E; the approver edits G to change services/prices — G is the source of
                    truth for the invoice.
    amount        → "Amount ($) by AI" (F): the AI's proposed total — LOCKED (gray).
    amount_user   → "Amount ($) by User" (H): total of col G — AUTO-COMPUTED (gray), not
                    typed. Seeded to the AI total at creation (G == E), then recomputed from
                    G each run. Pass a value only for already-invoiced learning rows.
    ai_learning   → "AI Learning": the AI's per-order learning record (AI-filled).

    highlight_red=True fills the entire row with red (#FF4444) to flag critical issues
    (e.g. Delivered orders that returned a $0 invoice — cannot be processed automatically).
    The red fill is overridden once the user picks an Action value (conditional formatting wins).
    """
    ensure_approval_sheet()
    _record_ai_prefilled_note(order_id, notes)   # so the AI never learns from its own pre-filled Notes

    if not posted_at:
        posted_at = datetime.now(_EASTERN).strftime("%Y-%m-%d %H:%M %Z")

    # Build the row by column index so it can never drift out of sync with APPROVAL_HEADERS.
    row = [""] * _COL_COUNT
    row[_COL_ORDER_ID]      = str(order_id)       # turned into a clickable HYPERLINK in _format_new_row
    row[_COL_ORDER_STATUS]  = str(order_status)   # FTF stage status
    row[_COL_CLIENT]        = str(client_name)
    row[_COL_ADDRESS]       = str(address)[:120]
    # Property-context reference columns (AI-managed / gray). Map Link is written as the raw
    # URL here; ensure_action_dropdown() turns it into a clickable native hyperlink on every
    # cycle (a =HYPERLINK() formula can't be used — Excel auto-propagates formulas across a
    # whole table column, the same trap the Order ID link avoids).
    row[_COL_PROPERTY_SIZE] = str(property_size or "")
    row[_COL_MAP_LINK]      = str(map_link or "")
    row[_COL_FEMA_ZONE]     = str(fema_zone or "")
    row[_COL_SERVICE_TYPE]  = str(service_type or "")
    row[_COL_SERVICE_AI]    = str(service)        # AI's locked breakdown baseline
    row[_COL_AMOUNT_AI]     = float(amount)       # F — AI's locked total
    # G — editable copy of E by default (source of truth). Callers may pass service_user to
    # seed G differently (e.g. learning rows set it to the real human breakdown so col H,
    # which auto-sums G, stays equal to the human amount).
    row[_COL_SERVICE_USER]  = str(service_user) if service_user is not None else str(service)
    # H — total of G; auto-computed. Seed to AI total (G==E at creation), or the explicit
    # amount_user for already-invoiced learning rows.
    row[_COL_AMOUNT_USER]   = float(amount_user) if amount_user is not None else float(amount)
    row[_COL_CONFIDENCE]    = str(confidence)
    row[_COL_ESCALATE]      = "Yes" if escalate else "No"
    row[_COL_ACTION]        = ""                  # blank; user selects Approve/Reject/On-hold
    row[_COL_NOTES]         = str(notes)          # pre-filled for escalations
    row[_COL_POSTED_AT]     = posted_at
    row[_COL_PROCESSED_AT]  = ""                  # filled after pipeline processes decision
    row[_COL_AI_LEARNING]   = str(ai_learning)
    row[_COL_USER_LEARNING] = ""                  # approver fills this; AI reads it next run
    values = [row]

    r = httpx.post(
        f"{_wb_base()}/worksheets/{ONEDRIVE_SHEET_NAME}/tables/{ONEDRIVE_TABLE_NAME}/rows/add",
        headers=_session_headers(),
        json={"values": values},
        timeout=15.0,
    )
    r.raise_for_status()
    log.info("excel row appended order_id=%s amount=%.2f status=%s", order_id, amount, order_status)

    # Apply borders, currency, the Order-ID hyperlink, ownership colors, and optional red flag.
    try:
        new_index = r.json().get("index")   # 0-based table row index
        if new_index is not None:
            # _format_new_row applies the per-column ownership tint (gray/blue), which also
            # overwrites any stale fill a recycled row position might have inherited.
            _format_new_row(new_index, ftf_link=ftf_link, order_id=str(order_id))
            if highlight_red:
                # Red flag overrides the per-column tint for the whole row.
                excel_row = new_index + 2
                row_range = f"A{excel_row}:{_END_COL}{excel_row}"
                httpx.patch(
                    f"{_wb_base()}/worksheets/{ONEDRIVE_SHEET_NAME}/range(address='{row_range}')/format/fill",
                    headers=_session_headers(),
                    json={"color": "#FF4444"},
                    timeout=10.0,
                ).raise_for_status()
                log.info("row %d highlighted red (flagged order=%s)", excel_row, order_id)
    except Exception as exc:
        log.warning("row formatting failed (non-fatal) order_id=%s: %s", order_id, exc)


_ACTION_NORMALIZE = {
    "approve": "approve",
    "reject":  "reject",
    "on-hold": "hold",
    "on hold": "hold",
    "hold":    "hold",
}


def consume_user_learnings() -> dict:
    """Fold operator free-text into the AI's learning so it applies on the NEXT pricing run.

    Reads TWO user columns: "Learning provided by user" (P, always operator-written) and
    "Notes" (L, operator-written only when it differs from the AI's own pre-filled escalation
    text — tracked in learned_rules.json['ai_prefilled_notes'] so the AI never learns from
    its own words). Each absorbed note is stored in data/learned_rules.json as (a) an
    order-override for that order and (b) a general 'user_guidance' entry — BOTH of which
    _load_learned_rules() injects straight into the pricing prompt, so the model itself
    reasons with the operator's feedback. The AI Learning cell is stamped so the operator
    sees the feedback was absorbed. Idempotent per (order_id, source, note). Never raises.
    """
    summary = {"scanned": 0, "absorbed": 0}
    rules_path = os.path.normpath(
        os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "learned_rules.json")
    )
    try:
        ensure_approval_sheet()
        r = _graph_get_retry(
            f"{_wb_base()}/worksheets/{ONEDRIVE_SHEET_NAME}/tables/{ONEDRIVE_TABLE_NAME}/rows",
            headers=_headers(), timeout=15.0,
        )
        if r is None or r.status_code in _TRANSIENT_STATUS:
            return summary
        r.raise_for_status()
        rows = r.json().get("value", [])

        data = {}
        if os.path.exists(rules_path):
            with open(rules_path, encoding="utf-8") as f:
                data = json.load(f)
        overrides    = data.setdefault("order_overrides", {})
        guidance     = data.setdefault("user_guidance", [])
        consumed     = set(data.setdefault("user_notes_consumed", []))
        ai_prefills  = data.get("ai_prefilled_notes", {})   # {order_id: AI's pre-filled Notes text}

        now = datetime.now(_EASTERN).strftime("%Y-%m-%d %H:%M %Z")
        changed = False
        for row in rows:
            vals = row.get("values", [[]])[0]
            if len(vals) < _COL_COUNT:
                vals = list(vals) + [""] * (_COL_COUNT - len(vals))
            summary["scanned"] += 1
            order_id = str(vals[_COL_ORDER_ID]).strip()
            if not order_id:
                continue
            client  = str(vals[_COL_CLIENT]).strip()
            service = str(vals[_COL_SERVICE_AI]).strip()   # AI baseline service name

            # Candidate learning text from the two user columns. "Learning provided by
            # user" (P) is always operator-written. "Notes" (L) is operator-written ONLY
            # when it differs from the AI's pre-filled text (else it's the AI's own words).
            candidates = []
            p_note = str(vals[_COL_USER_LEARNING]).strip()
            if p_note:
                candidates.append(("learning", p_note))
            l_note     = str(vals[_COL_NOTES]).strip()
            ai_prefill = str(ai_prefills.get(order_id, "")).strip()
            if l_note and l_note != ai_prefill:
                candidates.append(("notes", l_note))

            absorbed_here = False
            for src, text in candidates:
                key = f"{order_id}:{src}:{hashlib.sha1(text.encode('utf-8')).hexdigest()[:12]}"
                if key in consumed:
                    continue   # already absorbed this exact note
                consumed.add(key)
                changed = True
                absorbed_here = True
                summary["absorbed"] += 1
                tag = "operator note" if src == "learning" else "operator note (Notes)"
                overrides.setdefault(order_id, []).append(f"[{tag} {now}] {text}")
                guidance.append({"order_id": order_id, "client": client, "service": service,
                                 "note": text, "source": src, "observed_at": now})

            # Stamp the AI Learning cell once if anything was absorbed for this row.
            if absorbed_here:
                row_index = row.get("index")
                if row_index is not None:
                    excel_row = row_index + 2
                    prev  = str(vals[_COL_AI_LEARNING]).strip()
                    stamp = f"AI_LEARN_USER | {now} | absorbed operator feedback — applies next pricing run"
                    new_val = (prev + " || " + stamp) if prev else stamp
                    try:
                        httpx.patch(
                            f"{_wb_base()}/worksheets/{ONEDRIVE_SHEET_NAME}/range(address='{_col_letter(_COL_AI_LEARNING)}{excel_row}')",
                            headers=_session_headers(), json={"values": [[new_val]]}, timeout=10.0,
                        ).raise_for_status()
                    except Exception as exc:  # noqa: BLE001
                        log.debug("consume_user_learnings: stamp row %s failed (non-fatal): %s", excel_row, exc)

        if changed:
            data["user_notes_consumed"] = sorted(consumed)
            with open(rules_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, default=str)
            log.info("consume_user_learnings: absorbed %d operator note(s)", summary["absorbed"])
        return summary
    except Exception as exc:  # noqa: BLE001
        log.warning("consume_user_learnings failed (non-fatal): %s", exc)
        return summary


def get_pending_approvals() -> list[dict]:
    """Return rows where Action is set (Approve/Reject/On-hold) and Processed At is empty.

    Used by the Excel Approval Watcher to pick up decisions made in the spreadsheet
    without needing Power Automate or any external trigger.

    Returns list of dicts: {order_id, action (normalized), notes}
    """
    ensure_approval_sheet()
    # Plain headers (no session) + retry: the workbook session createSession itself returns
    # 501 while Graph re-indexes a freshly-uploaded file. On a persistent transient error we
    # return [] (no pending this cycle) rather than crash the watcher — it recovers next cycle.
    r = _graph_get_retry(
        f"{_wb_base()}/worksheets/{ONEDRIVE_SHEET_NAME}/tables/{ONEDRIVE_TABLE_NAME}/rows",
        headers=_headers(),
        timeout=15.0,
    )
    if r is None or r.status_code in _TRANSIENT_STATUS:
        log.warning("get_pending_approvals: rows read unavailable (status %s) — treating as 0 pending "
                    "this cycle (Graph workbook API transient)", getattr(r, "status_code", "n/a"))
        return []
    r.raise_for_status()

    results = []
    for row in r.json().get("value", []):
        vals = row.get("values", [[]])[0]
        if len(vals) < _COL_COUNT:
            vals = list(vals) + [""] * (_COL_COUNT - len(vals))

        order_id        = str(vals[_COL_ORDER_ID]).strip()
        action_raw      = str(vals[_COL_ACTION]).strip()
        processed_at    = str(vals[_COL_PROCESSED_AT]).strip()
        notes           = str(vals[_COL_NOTES]).strip()
        # SOURCE OF TRUTH = the user breakdown (G) + its auto-computed total (H). The AI
        # baseline (E breakdown / F total) is read only so A4 can learn the AI-vs-user gap.
        amount_cell     = str(vals[_COL_AMOUNT_USER]).strip()   # H — Amount ($) by User (auto-computed total)
        breakdown_cell  = str(vals[_COL_SERVICE_USER]).strip()  # G — Service / Breakdown by User (editable)
        breakdown_ai_cell = str(vals[_COL_SERVICE_AI]).strip()  # E — Service / Breakdown by AI (baseline, for learning)
        learning_user   = str(vals[_COL_USER_LEARNING]).strip() # P — Learning provided by user (plain English)

        if not order_id or not action_raw or processed_at:
            continue   # blank action or already processed

        action_norm = _ACTION_NORMALIZE.get(action_raw.lower())
        if not action_norm:
            continue   # unrecognised value in Action column

        results.append({
            "order_id":       order_id,
            "action":         action_norm,
            "notes":          notes,
            "amount_cell":    amount_cell,
            "breakdown_cell": breakdown_cell,
            "breakdown_ai_cell": breakdown_ai_cell,
            "learning_user":  learning_user,
        })

    log.info("get_pending_approvals: %d pending decisions found", len(results))
    return results


def mark_row_processed(order_id: str) -> None:
    """Set Processed At timestamp on the most recent blank-Action row matching order_id."""
    base = _wb_base()
    h    = _session_headers()

    r = httpx.get(
        f"{base}/worksheets/{ONEDRIVE_SHEET_NAME}/tables/{ONEDRIVE_TABLE_NAME}/rows",
        headers=h, timeout=15.0,
    )
    r.raise_for_status()
    processed_at = datetime.now(_EASTERN).strftime("%Y-%m-%d %H:%M %Z")

    for row in reversed(r.json().get("value", [])):
        vals = row.get("values", [[]])[0]
        if len(vals) >= 1 and str(vals[0]) == str(order_id):
            # Single-cell range PATCH (Processed At only). A full-row table PATCH would
            # rewrite col A's =HYPERLINK() formula as plain text and kill the clickable
            # Order ID link — so we touch ONLY the one cell we're changing.
            excel_row   = row["index"] + 2   # 0-based table index + header + 1
            proc_letter = _col_letter(_COL_PROCESSED_AT)   # "M"
            httpx.patch(
                f"{base}/worksheets/{ONEDRIVE_SHEET_NAME}/range(address='{proc_letter}{excel_row}')",
                headers=h,
                json={"values": [[processed_at]]},
                timeout=15.0,
            ).raise_for_status()
            log.info("marked processed order_id=%s at=%s (cell %s%d)", order_id, processed_at, proc_letter, excel_row)
            return

    log.warning("mark_row_processed: order_id=%s not found in Excel", order_id)


def update_approval_notes(order_id: str, note: str, mark_processed: bool = True) -> bool:
    """Write `note` into the Notes column (K) for an order's row, and optionally stamp
    Processed At (M). Used to record outcomes the human should see — e.g. "invoice already
    exists in FTF, none generated". Returns True if a row was updated.
    """
    base = _wb_base()
    h    = _session_headers()
    try:
        r = httpx.get(
            f"{base}/worksheets/{ONEDRIVE_SHEET_NAME}/tables/{ONEDRIVE_TABLE_NAME}/rows",
            headers=h, timeout=15.0,
        )
        r.raise_for_status()
        processed_at = datetime.now(_EASTERN).strftime("%Y-%m-%d %H:%M %Z")
        for row in reversed(r.json().get("value", [])):
            vals = row.get("values", [[]])[0]
            if len(vals) >= 1 and str(vals[0]).strip() == str(order_id):
                # Single-cell range PATCHes (Notes, optionally Processed At). A full-row
                # table PATCH would rewrite col A's =HYPERLINK() formula as plain text and
                # kill the clickable Order ID link — so we touch ONLY the cells we change.
                excel_row    = row["index"] + 2   # 0-based table index + header + 1
                notes_letter = _col_letter(_COL_NOTES)   # "K"
                httpx.patch(
                    f"{base}/worksheets/{ONEDRIVE_SHEET_NAME}/range(address='{notes_letter}{excel_row}')",
                    headers=h, json={"values": [[str(note)]]}, timeout=15.0,
                ).raise_for_status()
                if mark_processed:
                    proc_letter = _col_letter(_COL_PROCESSED_AT)   # "M"
                    httpx.patch(
                        f"{base}/worksheets/{ONEDRIVE_SHEET_NAME}/range(address='{proc_letter}{excel_row}')",
                        headers=h, json={"values": [[processed_at]]}, timeout=15.0,
                    ).raise_for_status()
                log.info("update_approval_notes: order_id=%s note set (processed=%s)", order_id, mark_processed)
                return True
        log.warning("update_approval_notes: order_id=%s not found in Excel", order_id)
        return False
    except Exception as exc:
        log.warning("update_approval_notes failed order_id=%s (non-fatal): %s", order_id, exc)
        return False


def sync_approval_amounts(order_id: str, new_total: float, new_breakdown_str: str) -> None:
    """Write the reconciled USER breakdown (col G, "Service / Breakdown by User") and its
    FINAL invoiced total (col H, "Amount ($) by User") back to Excel after A4 reconciliation.

    The AI baseline — col E ("Service / Breakdown by AI") and col F ("Amount ($) by AI") —
    is left UNTOUCHED so the AI-vs-approver gap stays visible for learning. Cols G/H record
    what was actually invoiced.
    """
    base = _wb_base()
    h    = _session_headers()

    r = httpx.get(
        f"{base}/worksheets/{ONEDRIVE_SHEET_NAME}/tables/{ONEDRIVE_TABLE_NAME}/rows",
        headers=h, timeout=15.0,
    )
    r.raise_for_status()

    for row in reversed(r.json().get("value", [])):
        vals = row.get("values", [[]])[0]
        if len(vals) >= 1 and str(vals[0]) == str(order_id):
            idx       = row["index"]
            excel_row = idx + 2   # 0-based table index + header + 1
            svc_letter  = _col_letter(_COL_SERVICE_USER)  # "G"  Service / Breakdown by User
            user_letter = _col_letter(_COL_AMOUNT_USER)   # "H"  Amount ($) by User
            # G and H ARE adjacent, but patch separately (G is text, H is a number) to keep
            # each cell's type/format clean. E/F (AI baseline) are never touched.
            httpx.patch(
                f"{base}/worksheets/{ONEDRIVE_SHEET_NAME}/range(address='{svc_letter}{excel_row}')",
                headers=h, json={"values": [[new_breakdown_str]]}, timeout=15.0,
            ).raise_for_status()
            httpx.patch(
                f"{base}/worksheets/{ONEDRIVE_SHEET_NAME}/range(address='{user_letter}{excel_row}')",
                headers=h, json={"values": [[new_total]]}, timeout=15.0,
            ).raise_for_status()
            log.info(
                "sync_approval_amounts: order=%s total=%.2f → col H (Amount by User); breakdown → col G; E/F (AI) preserved",
                order_id, new_total,
            )
            return

    log.warning("sync_approval_amounts: order_id=%s not found in Excel", order_id)


# Parse a "Name: $amount | Name: $amount" user breakdown (same pattern A4 uses).
_USER_BREAKDOWN_RE = re.compile(r'([^|$:]+?):\s*\$\s*([\d,]+(?:\.\d{1,2})?)\s*(?:\||$)')


def _sum_breakdown(s: str):
    """Sum the $amounts in a 'Name: $amt | Name: $amt' breakdown string.

    Returns the float total, or None when the string has NO priced items (blank, a bare
    service name, or unparseable) — callers MUST treat None as 'leave the total alone'.
    """
    items = _USER_BREAKDOWN_RE.findall(s or "")
    if not items:
        return None
    try:
        return round(sum(float(amt.replace(",", "")) for _, amt in items), 2)
    except (TypeError, ValueError):
        return None


def recompute_user_amounts() -> dict:
    """Refresh "Amount ($) by User" (H) = sum of "Service / Breakdown by User" (G) for
    every data row, on every pipeline run — so the approver always sees the live total of
    their edited breakdown (BEFORE and AFTER approval).

    Idempotent: writes a cell only when the total actually changed (> $0.005). Rows whose G
    has no priced items (blank / bare service name / garbled) are SKIPPED so H is never
    clobbered. Single-cell Graph range PATCH only — no openpyxl/session/upload, so no sheet
    rebuild and no 423-lock risk. Never raises; returns {scanned, updated}.
    """
    summary = {"scanned": 0, "updated": 0}
    try:
        ensure_approval_sheet()
        base = _wb_base()
        h    = _headers()
        r = _graph_get_retry(
            f"{base}/worksheets/{ONEDRIVE_SHEET_NAME}/tables/{ONEDRIVE_TABLE_NAME}/rows",
            headers=h, timeout=15.0,
        )
        if r is None or r.status_code in _TRANSIENT_STATUS:
            log.warning("recompute_user_amounts: rows read unavailable (status %s) — skip this cycle",
                        getattr(r, "status_code", "n/a"))
            return summary
        r.raise_for_status()

        user_letter = _col_letter(_COL_AMOUNT_USER)   # "H"
        for row in r.json().get("value", []):
            vals = row.get("values", [[]])[0]
            if len(vals) < _COL_COUNT:
                vals = list(vals) + [""] * (_COL_COUNT - len(vals))
            order_id = str(vals[_COL_ORDER_ID]).strip()
            if not order_id:
                continue
            summary["scanned"] += 1
            new_h = _sum_breakdown(str(vals[_COL_SERVICE_USER]))
            if new_h is None:
                continue   # no priced items in G → never overwrite H
            try:
                cur_h = float(str(vals[_COL_AMOUNT_USER]).replace(",", "").replace("$", "") or 0)
            except (TypeError, ValueError):
                cur_h = 0.0
            if abs(new_h - cur_h) <= 0.005:
                continue   # already in sync
            excel_row = row["index"] + 2
            httpx.patch(
                f"{base}/worksheets/{ONEDRIVE_SHEET_NAME}/range(address='{user_letter}{excel_row}')",
                headers=h, json={"values": [[new_h]]}, timeout=15.0,
            ).raise_for_status()
            summary["updated"] += 1
        log.info("recompute_user_amounts: scanned=%d updated=%d", summary["scanned"], summary["updated"])
    except Exception as exc:
        log.warning("recompute_user_amounts failed (non-fatal): %s", exc)
    return summary
