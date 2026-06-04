"""
Generate FTF_Order_Scenarios.xlsx — a clean, non-tech-friendly reference sheet
covering all estimate / order-validation scenarios for Nexgen Land Solutions.

Run: python scripts/generate_scenarios_sheet.py
Output: Resources/FTF_Order_Scenarios.xlsx
"""

import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT  = os.path.join(ROOT, "Resources", "FTF_Order_Scenarios.xlsx")

# ── Brand colours ────────────────────────────────────────────────────────────
HEADER_BG   = "1A3A5C"
HEADER_FG   = "FFFFFF"

CAT_FG = "1A1A1A"
CAT_BG = {
    "Can Proceed":      "D4EDDA",   # green
    "Needs Review":     "FFF3CD",   # yellow
    "Special Case":     "D1ECF1",   # cyan
    "Cannot Proceed":   "F8D7DA",   # red
    "Payment / Invoice":"E2D9F3",   # purple
}

CONF_BG = {
    "HIGH":   "28A745",
    "MEDIUM": "FFC107",
    "LOW":    "DC3545",
    "N/A":    "AAAAAA",
}

# ── Thin border ──────────────────────────────────────────────────────────────
_side  = Side(style="thin", color="CCCCCC")
BORDER = Border(left=_side, right=_side, top=_side, bottom=_side)

# ── Scenarios data ───────────────────────────────────────────────────────────
# Columns: Category | Scenario Name | What Triggers This | What AI Checks
#          | AI Decision | Confidence | Est. Amount | What Happens Next
#          | Client Message / Outcome

SCENARIOS = [
    # ── CAN PROCEED ──────────────────────────────────────────────────────────
    ("Can Proceed", "Standard Residential Boundary Survey — Small Lot",
     "Order comes in for a residential property ≤ 0.30 acres",
     "Lot size, county, job type, client email present",
     "APPROVE — price at standard tier",
     "HIGH", "$475",
     "Invoice drafted and sent to Prateek for approval, then emailed to client",
     "Invoice for $475 sent to client."),

    ("Can Proceed", "Standard Residential Boundary Survey — Medium Lot",
     "Residential property between 0.31 and 0.50 acres",
     "Lot size, county, job type, client email present",
     "APPROVE — price at medium tier",
     "HIGH", "$575",
     "Invoice drafted and posted for Prateek approval",
     "Invoice for $575 sent to client."),

    ("Can Proceed", "Standard Residential Boundary Survey — Large Lot",
     "Residential property between 0.51 and 1.00 acres",
     "Lot size, county, job type, client email present",
     "APPROVE — price at large lot tier",
     "HIGH", "$675",
     "Invoice drafted and posted for Prateek approval",
     "Invoice for $675 sent to client."),

    ("Can Proceed", "Residential Boundary Survey — Over 1 Acre",
     "Residential property exceeding 1.00 acre",
     "Lot size > 1 acre detected; checks for complexity flags",
     "APPROVE with manual review note",
     "MEDIUM", "$750 – $1,200+",
     "Invoice drafted with ESCALATE flag; Prateek manually sets final price",
     "Custom quote provided to client after Prateek review."),

    ("Can Proceed", "Commercial Boundary Survey",
     "Order is for a commercial property (office, warehouse, retail, etc.)",
     "Property type = commercial; lot size and county present",
     "APPROVE — price at commercial tier",
     "MEDIUM", "$900 – $2,000+",
     "Escalated to Prateek for custom pricing; invoice drafted after confirmation",
     "Custom invoice sent to client after internal review."),

    ("Can Proceed", "Vacant Land / Lot Survey",
     "Order is for an undeveloped or vacant parcel",
     "Job type, acreage, county, client email",
     "APPROVE — standard pricing by acreage",
     "HIGH", "$475 – $750",
     "Invoice drafted automatically using standard tiers",
     "Invoice sent to client."),

    ("Can Proceed", "Property Has Prior Survey on Record",
     "Client mentions or system shows a previous survey exists for the parcel",
     "Prior survey flag in notes; reduces field complexity",
     "APPROVE — may apply discount or note in invoice",
     "HIGH", "Standard tier",
     "Invoice drafted; prior survey noted in record",
     "Invoice sent. Faster turnaround communicated to client."),

    ("Can Proceed", "Multiple Parcels — Same Client, Same Area",
     "Client orders surveys for 2 or more adjacent or nearby parcels",
     "Number of parcels, whether they share boundaries, client email",
     "APPROVE — bundle pricing may apply",
     "MEDIUM", "Custom (per-parcel rate)",
     "Escalated to Prateek for bundle pricing review",
     "Bundle quote sent to client after Prateek confirmation."),

    ("Can Proceed", "Rush / Expedited Order",
     "Client needs survey completed faster than standard lead time",
     "Requested date vs. current capacity; rush flag in notes",
     "APPROVE with rush surcharge",
     "MEDIUM", "Standard + 20–30% rush fee",
     "Invoice drafted with rush surcharge; Prateek confirms before sending",
     "Expedited invoice sent. Rush timeline confirmed with client."),

    # ── NEEDS REVIEW ─────────────────────────────────────────────────────────
    ("Needs Review", "Missing Lot Size / Acreage",
     "Order has no lot size or acreage information",
     "Checks ng_orders fields — lot size, parcel data, notes",
     "HOLD — cannot price without acreage",
     "LOW", "Unknown",
     "Order moved to 'details_missing'; Prateek notified in Teams to gather info",
     "Client may be contacted for property details before invoice is issued."),

    ("Needs Review", "Ambiguous Job Type",
     "Client description doesn't clearly match a known survey type",
     "Job type field, order notes, client description keywords",
     "ESCALATE — flag for human classification",
     "LOW", "TBD",
     "Posted to Teams with escalation flag; Robert/Ryan assigned to clarify",
     "NexGen team will reach out to client to confirm the right survey type."),

    ("Needs Review", "Missing Client Email Address",
     "Order has no valid email on record for the client",
     "customer_email field empty or invalid format",
     "HOLD — cannot send invoice without email",
     "N/A", "N/A",
     "Order flagged as 'details_missing'; team notified to find contact",
     "Invoice will be sent once valid email is confirmed."),

    ("Needs Review", "Estimate Amount Is $0 or Blank",
     "No pricing data exists for the order",
     "estimate_amount = 0 or NULL in order record",
     "HOLD — needs manual pricing by Prateek",
     "LOW", "$0 (needs input)",
     "Order listed as 'pricing_needed'; Prateek sets price via Teams command",
     "Invoice issued after Prateek sets price."),

    ("Needs Review", "Unusually High Estimated Value (> $2,000)",
     "Pricing algorithm returns a very high amount",
     "Calculated price exceeds $2,000 threshold",
     "ESCALATE — flag for Prateek confirmation",
     "MEDIUM", "> $2,000",
     "Prateek must manually confirm before invoice is created",
     "High-value quote sent to client only after Prateek approves."),

    ("Needs Review", "Complex Terrain / Access Issues Noted",
     "Client or field notes mention wetlands, steep slopes, or difficult access",
     "Keywords in notes: 'wetlands', 'swamp', 'steep', 'no road access', etc.",
     "ESCALATE — complexity upcharge may apply",
     "MEDIUM", "Standard + complexity upcharge",
     "Prateek reviews; adds surcharge if needed before approving invoice",
     "Adjusted quote sent to client explaining complexity factor."),

    ("Needs Review", "Client Suggests Their Own Price",
     "Client includes a budget or price expectation in their order notes",
     "Price keywords in notes; compared against NexGen standard rates",
     "ESCALATE — flag mismatch if client price ≠ our rate",
     "MEDIUM", "Varies",
     "Prateek reviews; accepts or overrides client suggested price",
     "Invoice reflects NexGen pricing; any differences explained to client."),

    ("Needs Review", "Incomplete Property Address",
     "Address is missing street number, city, state, or zip",
     "Validates address fields for completeness",
     "HOLD — cannot locate or price without valid address",
     "LOW", "N/A",
     "Order held; team contacts client or searches public records for address",
     "Survey cannot proceed without a confirmed property address."),

    ("Needs Review", "Duplicate Order Detected",
     "Same property address appears in another active order",
     "Checks for duplicate address or parcel ID across active orders",
     "FLAG — possible duplicate, needs human review",
     "MEDIUM", "N/A",
     "Both orders flagged; Prateek or Robert/Ryan decides which to keep",
     "Client contacted to confirm if one or both orders should proceed."),

    # ── SPECIAL CASES ─────────────────────────────────────────────────────────
    ("Special Case", "Condo Property",
     "Property is a condominium unit or condo complex",
     "Address or notes contain 'condo', 'unit #', 'apt', or condo parcel ID pattern",
     "REJECT — NexGen does not survey condo units",
     "N/A", "$0",
     "Order moved to 'condo_rejected'; client notified we do not handle condo surveys",
     "We're sorry — we do not provide surveys for condominium units. Please contact a condo-specialist surveyor."),

    ("Special Case", "Waterfront / Riparian Property",
     "Property borders a lake, river, canal, or ocean",
     "Keywords: 'waterfront', 'canal', 'lakefront', 'riparian', 'seawall'",
     "ESCALATE — additional mean high water line work may be needed",
     "MEDIUM", "Standard + $150–$400 riparian add-on",
     "Prateek reviews; confirms if mean high water line survey is included",
     "Custom quote including waterfront work sent to client."),

    ("Special Case", "New Construction / Builder Order",
     "Order is from a builder for a new development or subdivision",
     "Client type = builder; or keywords 'new construction', 'subdivision', 'developer'",
     "ESCALATE — builder pricing and scope apply",
     "MEDIUM", "Custom (project-based)",
     "Robert/Ryan handles builder relationship; custom scope and pricing agreed",
     "Project proposal sent to builder after NexGen team review."),

    ("Special Case", "Large Subdivision (10+ Lots)",
     "Order involves a subdivision with many individual lots",
     "Number of lots > 10 or 'subdivision' keyword in notes",
     "ESCALATE — project scope too large for standard pipeline",
     "LOW", "Custom project quote",
     "Handed off to Robert/Ryan for dedicated project management",
     "Project team will contact you directly to scope and quote this subdivision."),

    ("Special Case", "Government / Municipal Property",
     "Property is owned by a city, county, state, or federal entity",
     "Owner name contains 'county', 'city of', 'state of', 'USA', 'FDOT', etc.",
     "ESCALATE — government contract requirements may apply",
     "LOW", "Custom / contract-based",
     "Prateek or Robert handles; government PO or contract may be required",
     "NexGen will follow up regarding government contract requirements."),

    ("Special Case", "Client Requests Specific Surveyor",
     "Client mentions a specific NexGen surveyor by name",
     "Surveyor name appears in order notes",
     "NOTE — log preference, do not guarantee assignment",
     "N/A", "Standard pricing",
     "Preference noted in order file; field coordinator assigns if possible",
     "We'll do our best to accommodate your surveyor preference, subject to availability."),

    ("Special Case", "Litigation / Disputed Property",
     "Survey is needed for a legal dispute or boundary conflict",
     "Keywords: 'attorney', 'dispute', 'litigation', 'court', 'boundary conflict'",
     "ESCALATE — legal survey has stricter requirements and higher liability",
     "LOW", "Custom / attorney-rate",
     "Prateek reviews; expert witness and documentation requirements assessed",
     "Litigation surveys require additional scope. NexGen team will contact you."),

    # ── CANNOT PROCEED ────────────────────────────────────────────────────────
    ("Cannot Proceed", "Property Outside Florida",
     "Property address is in a state other than Florida",
     "State field in address; NexGen is licensed in Florida only",
     "REJECT — out of service area",
     "N/A", "$0",
     "Order closed; client notified we only operate in Florida",
     "We're sorry — Nexgen Land Solutions is licensed in Florida only. We cannot survey out-of-state properties."),

    ("Cannot Proceed", "Survey Type Not Offered",
     "Client requests a survey type NexGen does not provide (e.g., hydrographic, marine, ALTA without authorization)",
     "Job type matched against NexGen's service catalog",
     "REJECT — service not available",
     "N/A", "$0",
     "Order closed; client notified of available services",
     "We don't currently offer this survey type. Our available services are: Boundary, Topographic, Elevation Certificate, and more."),

    ("Cannot Proceed", "No Property Address Provided",
     "Order submitted with no property address at all",
     "Address fields completely empty",
     "REJECT / HOLD — cannot locate property",
     "N/A", "$0",
     "Order held pending address; if not received within 5 business days, closed",
     "We need a valid property address to begin your survey. Please provide the address to continue."),

    ("Cannot Proceed", "Client Explicitly Canceled",
     "Client contacts NexGen to cancel the survey order",
     "Cancellation keyword in Teams reply or email; Prateek or team confirms",
     "CANCEL — order closed",
     "N/A", "$0",
     "Order moved to 'invoice_rejected'; no invoice or email sent",
     "Your order has been canceled. Please contact us if you'd like to restart in the future."),

    # ── PAYMENT / INVOICE ─────────────────────────────────────────────────────
    ("Payment / Invoice", "Invoice Approved by Prateek",
     "Prateek replies 'APPROVE [order ID]' in Teams after reviewing the draft",
     "Approval keyword + order ID from authorized sender",
     "APPROVE — create invoice in FTF system and send to client",
     "HIGH", "As drafted",
     "FTF invoice created; Pay Now link generated; email sent to client",
     "Invoice emailed to client with Pay Now link for secure online payment."),

    ("Payment / Invoice", "Invoice Rejected by Prateek",
     "Prateek replies 'REJECT [order ID] [reason]' in Teams",
     "Reject keyword + order ID from authorized sender",
     "REJECT — no invoice created, no email sent",
     "N/A", "$0",
     "Order moved to 'invoice_rejected'; reason logged; no client contact",
     "No invoice sent. Order held pending further instructions."),

    ("Payment / Invoice", "Invoice Held for More Info",
     "Prateek or team replies 'HOLD [order ID]' in Teams",
     "Hold keyword + order ID",
     "HOLD — invoice paused, order stays in queue",
     "N/A", "TBD",
     "Order stays in 'invoice_draft_posted'; no action until explicitly approved or rejected",
     "Invoice on hold. NexGen team will follow up once all information is confirmed."),

    ("Payment / Invoice", "Client Disputes Invoice Amount",
     "Client replies to invoice email questioning the price",
     "Dispute keywords in email reply or Teams message",
     "ESCALATE — human review required",
     "N/A", "Varies",
     "Prateek or Robert/Ryan reviews dispute and responds to client",
     "Thank you for reaching out. Our team will review your invoice and respond within 1 business day."),

    ("Payment / Invoice", "Payment Received — Invoice Closed",
     "Client pays via Pay Now link or check; payment recorded in FTF",
     "Payment status updated to 'paid' in ng_payments",
     "CLOSE — order complete",
     "HIGH", "Full amount paid",
     "Order marked complete; confirmation sent to client",
     "Thank you — your payment has been received. Your survey is confirmed and scheduled."),

    ("Payment / Invoice", "Payment Overdue (> 30 Days)",
     "Invoice sent more than 30 days ago with no payment",
     "Invoice send date vs. today; payment status still 'unpaid'",
     "FLAG — overdue, send reminder",
     "N/A", "Outstanding balance",
     "Automated reminder email sent to client; Prateek notified in Teams",
     "Friendly reminder: your invoice is past due. Please make payment at your earliest convenience."),
]

HEADERS = [
    "#",
    "Category",
    "Scenario Name",
    "What Triggers This?",
    "What the AI Checks",
    "AI Decision",
    "Confidence",
    "Est. Invoice Amount",
    "What Happens Next",
    "Client Message / Outcome",
]

COL_WIDTHS = [4, 18, 32, 38, 38, 30, 12, 20, 45, 55]


def hex_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def make_border() -> Border:
    return BORDER


def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Order Scenarios"

    # ── Title row ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "Nexgen Land Solutions — FTF Order Scenarios & Estimate Logic"
    title_cell.font  = Font(name="Calibri", bold=True, size=14, color=HEADER_FG)
    title_cell.fill  = hex_fill(HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Sub-title / legend row ───────────────────────────────────────────────
    ws.merge_cells("A2:J2")
    sub = ws["A2"]
    sub.value = (
        "Color guide:   🟢 Can Proceed   🟡 Needs Review   🔵 Special Case   "
        "🔴 Cannot Proceed   🟣 Payment / Invoice"
    )
    sub.font      = Font(name="Calibri", italic=True, size=10, color="555555")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # ── Header row ───────────────────────────────────────────────────────────
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = Font(name="Calibri", bold=True, size=11, color=HEADER_FG)
        cell.fill      = hex_fill(HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = make_border()
    ws.row_dimensions[3].height = 28

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_num, scenario in enumerate(SCENARIOS, start=1):
        excel_row = row_num + 3
        cat = scenario[0]
        bg  = CAT_BG.get(cat, "FFFFFF")

        row_data = [row_num] + list(scenario)  # prepend row number

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font      = Font(name="Calibri", size=10, color=CAT_FG)
            cell.fill      = hex_fill(bg)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
            cell.border = make_border()

            # Centre the # column
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="top")

            # Confidence column — colour the text
            if col_idx == 7:  # Confidence column
                conf_color = CONF_BG.get(value, HEADER_BG)
                cell.font = Font(
                    name="Calibri", size=10, bold=True, color="FFFFFF"
                )
                cell.fill = hex_fill(conf_color)
                cell.alignment = Alignment(horizontal="center", vertical="top")

        ws.row_dimensions[excel_row].height = 70

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze panes (keep header visible when scrolling) ────────────────────
    ws.freeze_panes = "A4"

    # ── Auto-filter on header row ─────────────────────────────────────────────
    ws.auto_filter.ref = f"A3:J{len(SCENARIOS) + 3}"

    # ── Sheet tab colour ──────────────────────────────────────────────────────
    ws.sheet_properties.tabColor = HEADER_BG

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"OK Saved: {OUT}")
    print(f"   Rows: {len(SCENARIOS)} scenarios across {len(set(s[0] for s in SCENARIOS))} categories")


if __name__ == "__main__":
    build()
