#!/usr/bin/env python3
"""
Pipeline Excel Export
Generates docs/Pipeline_Report.xlsx — a plain-English snapshot of the entire
AI pipeline for the NexGen team to review without touching the database.

Usage:
  python scripts/export_pipeline_excel.py

Tabs produced:
  1. Summary          -- high-level counts + pipeline health
  2. Active Pipeline  -- orders currently in flight
  3. Sent Estimates   -- all estimates delivered to customers
  4. Needs Review     -- flagged / awaiting human approval
  5. AR Overdue       -- unpaid invoices being escalated
  6. Monthly Stmts    -- B2B statement delivery log
  7. AI Decision Log  -- last 30 days of every AI decision (audit trail)
"""

import os
import sys
from datetime import datetime, timezone

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "code", "shared"))

from dotenv import load_dotenv
load_dotenv()

import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from config.settings import DB_HOST, DB_PORT, DB_NAME, DB_USER, DB_PASSWORD

# ── Output path ───────────────────────────────────────────────────────────────
_HERE   = os.path.dirname(__file__)
_ROOT   = os.path.abspath(os.path.join(_HERE, ".."))
OUT_DIR = os.path.join(_ROOT, "docs")
os.makedirs(OUT_DIR, exist_ok=True)
OUT_PATH = os.path.join(OUT_DIR, "Pipeline_Report.xlsx")

# ── Colours ───────────────────────────────────────────────────────────────────
C_HEADER      = "1F3864"   # dark navy  — header row bg
C_HEADER_TEXT = "FFFFFF"   # white      — header text
C_SENT        = "C6EFCE"   # green
C_REVIEWED    = "BDD7EE"   # blue
C_WRITTEN     = "FFEB9C"   # yellow
C_PRICED      = "FCE4D6"   # peach
C_CLASSIFIED  = "EDEDED"   # light grey
C_FLAGGED     = "FFCC99"   # orange
C_AWAITING    = "FFF2CC"   # pale yellow
C_ERROR       = "FFC7CE"   # red
C_PENDING     = "FFFFFF"   # white
C_SECTION     = "D9E1F2"   # soft blue  — section label rows

STATUS_COLOUR = {
    "sent":              C_SENT,
    "reviewed":          C_REVIEWED,
    "written":           C_WRITTEN,
    "priced":            C_PRICED,
    "classified":        C_CLASSIFIED,
    "flagged":           C_FLAGGED,
    "awaiting_approval": C_AWAITING,
    "error":             C_ERROR,
    "pending":           C_PENDING,
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def _connect():
    return psycopg2.connect(
        host=DB_HOST, port=DB_PORT, dbname=DB_NAME,
        user=DB_USER, password=DB_PASSWORD,
    )


def _query(sql: str, params=None) -> list[dict]:
    conn = _connect()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(sql, params or ())
                return [dict(r) for r in cur.fetchall()]
    finally:
        conn.close()


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _header_font(bold: bool = True) -> Font:
    return Font(name="Calibri", bold=bold, color=C_HEADER_TEXT, size=11)


def _body_font(bold: bool = False) -> Font:
    return Font(name="Calibri", bold=bold, size=10)


def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_header(ws, row: int, cols: list[str]) -> None:
    for c, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.fill    = _fill(C_HEADER)
        cell.font    = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border  = _thin_border()
    ws.row_dimensions[row].height = 28


def _write_row(ws, row: int, values: list, status: str = "") -> None:
    bg = STATUS_COLOUR.get(status, C_PENDING)
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill   = _fill(bg)
        cell.font   = _body_font()
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.border = _thin_border()


def _auto_width(ws, min_w: int = 10, max_w: int = 45) -> None:
    for col in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = (
            min(max(length + 2, min_w), max_w)
        )


def _fmt_dt(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M")
    return str(val)[:16]


def _fmt_money(val) -> str:
    if val is None:
        return ""
    try:
        return f"${float(val):,.2f}"
    except (TypeError, ValueError):
        return str(val)


# ── Tab builders ──────────────────────────────────────────────────────────────

def build_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value     = "NexGen Surveying -- AI Pipeline Report"
    title.font      = Font(name="Calibri", bold=True, size=16, color=C_HEADER)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value     = f"Generated: {datetime.now().strftime('%A, %B %d %Y at %I:%M %p')}"
    sub.font      = Font(name="Calibri", italic=True, size=10, color="666666")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Pipeline status counts
    status_rows = _query(
        """
        SELECT status, COUNT(*) AS cnt
        FROM processed_orders
        GROUP BY status ORDER BY cnt DESC
        """
    )

    r = 4
    ws.merge_cells(f"A{r}:F{r}")
    ws[f"A{r}"].value = "PIPELINE STATUS BREAKDOWN"
    ws[f"A{r}"].font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{r}"].fill  = _fill(C_HEADER)
    ws[f"A{r}"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 22
    r += 1

    _write_header(ws, r, ["Status", "Order Count", "What This Means"])
    r += 1
    STATUS_MEANING = {
        "pending":           "Detected in FTF, waiting for classification",
        "classified":        "Service type + flood zone identified",
        "priced":            "Estimate amount calculated",
        "written":           "Estimate email drafted by AI",
        "reviewed":          "AI validated draft, ready to send",
        "sent":              "Estimate delivered to customer",
        "flagged":           "Needs human review before proceeding",
        "awaiting_approval": "Waiting for Bobby/Ryan sign-off",
        "error":             "Processing error -- needs attention",
    }
    for row_data in status_rows:
        st  = row_data["status"]
        cnt = row_data["cnt"]
        meaning = STATUS_MEANING.get(st, "")
        _write_row(ws, r, [st.replace("_", " ").title(), cnt, meaning], status=st)
        r += 1

    # Today's activity
    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    ws[f"A{r}"].value = "TODAY'S ACTIVITY"
    ws[f"A{r}"].font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{r}"].fill  = _fill(C_HEADER)
    ws[f"A{r}"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 22
    r += 1

    today_data = _query(
        """
        SELECT
            COUNT(*) FILTER (WHERE sent_at::date = CURRENT_DATE)     AS sent_today,
            COUNT(*) FILTER (WHERE flagged_at::date = CURRENT_DATE)  AS flagged_today,
            COUNT(*) FILTER (WHERE created_at::date = CURRENT_DATE)  AS new_today,
            COUNT(*) FILTER (WHERE status NOT IN ('sent','error'))   AS in_pipeline
        FROM processed_orders
        """
    )
    if today_data:
        td = today_data[0]
        _write_header(ws, r, ["Metric", "Count"])
        r += 1
        for label, val in [
            ("Estimates Sent Today",        td.get("sent_today", 0)),
            ("New Orders Detected Today",   td.get("new_today", 0)),
            ("Flagged for Human Review",    td.get("flagged_today", 0)),
            ("Orders Still in Pipeline",    td.get("in_pipeline", 0)),
        ]:
            ws.cell(row=r, column=1, value=label).font = _body_font()
            ws.cell(row=r, column=2, value=val).font   = _body_font(bold=True)
            r += 1

    # AR summary
    r += 1
    ar_data = _query(
        """
        SELECT
            COUNT(*) FILTER (WHERE days_overdue >= 90) AS overdue_90,
            COUNT(*) FILTER (WHERE days_overdue >= 60 AND days_overdue < 90) AS overdue_60,
            COALESCE(SUM(invoice_amount) FILTER (WHERE status != 'paid'), 0) AS total_outstanding
        FROM ar_reminders
        """
    )
    if ar_data:
        ws.merge_cells(f"A{r}:F{r}")
        ws[f"A{r}"].value = "AR (OVERDUE INVOICES)"
        ws[f"A{r}"].font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        ws[f"A{r}"].fill  = _fill(C_HEADER)
        ws[f"A{r}"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 22
        r += 1
        _write_header(ws, r, ["Metric", "Value"])
        r += 1
        ad = ar_data[0]
        for label, val in [
            ("90+ Days Overdue (critical)",  ad.get("overdue_90", 0)),
            ("60-89 Days Overdue (warning)", ad.get("overdue_60", 0)),
            ("Total Outstanding Amount",     _fmt_money(ad.get("total_outstanding", 0))),
        ]:
            ws.cell(row=r, column=1, value=label).font = _body_font()
            ws.cell(row=r, column=2, value=val).font   = _body_font(bold=True)
            r += 1

    _auto_width(ws)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 50


def build_active_pipeline(wb: Workbook) -> None:
    ws = wb.create_sheet("Active Pipeline")
    ws.freeze_panes = "A2"

    rows = _query(
        """
        SELECT order_id, status, service_type, customer_email,
               estimate_amount, is_flood_zone, flag_reason,
               created_at, classified_at, priced_at, written_at, reviewed_at
        FROM processed_orders
        WHERE status NOT IN ('sent', 'error')
        ORDER BY created_at DESC
        """
    )

    headers = [
        "Order ID", "Status", "Service Type", "Customer Email",
        "Estimate ($)", "Flood Zone", "Flag Reason",
        "Detected", "Classified", "Priced", "Written", "Reviewed",
    ]
    _write_header(ws, 1, headers)

    for i, r in enumerate(rows, 2):
        _write_row(ws, i, [
            r["order_id"],
            r["status"].replace("_", " ").title() if r["status"] else "",
            r.get("service_type") or "",
            r.get("customer_email") or "",
            float(r["estimate_amount"]) if r.get("estimate_amount") else "",
            "Yes" if r.get("is_flood_zone") else "No",
            r.get("flag_reason") or "",
            _fmt_dt(r.get("created_at")),
            _fmt_dt(r.get("classified_at")),
            _fmt_dt(r.get("priced_at")),
            _fmt_dt(r.get("written_at")),
            _fmt_dt(r.get("reviewed_at")),
        ], status=r.get("status", ""))

    _auto_width(ws)
    ws.column_dimensions["G"].width = 40   # Flag Reason
    ws.column_dimensions["D"].width = 30   # Email


def build_sent_estimates(wb: Workbook) -> None:
    ws = wb.create_sheet("Sent Estimates")
    ws.freeze_panes = "A2"

    rows = _query(
        """
        SELECT order_id, service_type, customer_email, estimate_amount,
               is_flood_zone, sent_at, created_at
        FROM processed_orders
        WHERE status = 'sent'
        ORDER BY sent_at DESC
        """
    )

    _write_header(ws, 1, [
        "Order ID", "Service Type", "Customer Email",
        "Estimate Amount", "Flood Zone", "Sent At", "Order Created",
    ])

    for i, r in enumerate(rows, 2):
        _write_row(ws, i, [
            r["order_id"],
            r.get("service_type") or "",
            r.get("customer_email") or "",
            float(r["estimate_amount"]) if r.get("estimate_amount") else "",
            "Yes" if r.get("is_flood_zone") else "No",
            _fmt_dt(r.get("sent_at")),
            _fmt_dt(r.get("created_at")),
        ], status="sent")

    _auto_width(ws)


def build_needs_review(wb: Workbook) -> None:
    ws = wb.create_sheet("Needs Review")
    ws.freeze_panes = "A2"

    rows = _query(
        """
        SELECT order_id, status, service_type, customer_email,
               estimate_amount, flag_reason, flagged_at, created_at
        FROM processed_orders
        WHERE status IN ('flagged', 'awaiting_approval')
        ORDER BY flagged_at ASC
        """
    )

    _write_header(ws, 1, [
        "Order ID", "Status", "Service Type", "Customer Email",
        "Estimate ($)", "Reason Flagged", "Flagged At", "Order Created",
    ])

    for i, r in enumerate(rows, 2):
        _write_row(ws, i, [
            r["order_id"],
            r.get("status", "").replace("_", " ").title(),
            r.get("service_type") or "",
            r.get("customer_email") or "",
            float(r["estimate_amount"]) if r.get("estimate_amount") else "",
            r.get("flag_reason") or "",
            _fmt_dt(r.get("flagged_at")),
            _fmt_dt(r.get("created_at")),
        ], status=r.get("status", ""))

    _auto_width(ws)
    ws.column_dimensions["F"].width = 50   # Reason


def build_ar_overdue(wb: Workbook) -> None:
    ws = wb.create_sheet("AR Overdue")
    ws.freeze_panes = "A2"

    rows = _query(
        """
        SELECT order_id, customer_email, invoice_amount, invoice_date,
               days_overdue, reminder_level, status,
               last_reminder_sent_at, updated_at
        FROM ar_reminders
        WHERE status NOT IN ('paid', 'excluded')
        ORDER BY days_overdue DESC
        """
    )

    _write_header(ws, 1, [
        "Order ID", "Customer Email", "Invoice Amount", "Invoice Date",
        "Days Overdue", "Alert Level", "Status",
        "Last Alert Sent", "Updated",
    ])

    for i, r in enumerate(rows, 2):
        days = r.get("days_overdue", 0)
        if days >= 90:
            row_status = "flagged"    # red
        elif days >= 60:
            row_status = "awaiting_approval"  # yellow
        else:
            row_status = "classified"

        _write_row(ws, i, [
            r["order_id"],
            r.get("customer_email") or "",
            float(r["invoice_amount"]) if r.get("invoice_amount") else "",
            _fmt_dt(r.get("invoice_date")),
            days,
            r.get("reminder_level", 1),
            r.get("status", "").title(),
            _fmt_dt(r.get("last_reminder_sent_at")),
            _fmt_dt(r.get("updated_at")),
        ], status=row_status)

    _auto_width(ws)


def build_monthly_statements(wb: Workbook) -> None:
    ws = wb.create_sheet("Monthly Statements")
    ws.freeze_panes = "A2"

    rows = _query(
        """
        SELECT client_email, statement_month, order_count,
               total_amount, status, sent_at, updated_at
        FROM monthly_statements
        ORDER BY statement_month DESC, client_email ASC
        """
    )

    _write_header(ws, 1, [
        "Client Email", "Statement Month", "Order Count",
        "Total Amount", "Status", "Sent At", "Last Updated",
    ])

    STATUS_MAP = {
        "generated": "classified",
        "reviewed":  "reviewed",
        "sent":      "sent",
        "failed":    "error",
    }

    for i, r in enumerate(rows, 2):
        st = r.get("status", "")
        _write_row(ws, i, [
            r.get("client_email") or "",
            str(r.get("statement_month") or ""),
            r.get("order_count", 0),
            float(r["total_amount"]) if r.get("total_amount") else "",
            st.title(),
            _fmt_dt(r.get("sent_at")),
            _fmt_dt(r.get("updated_at")),
        ], status=STATUS_MAP.get(st, "pending"))

    _auto_width(ws)


def build_ai_decision_log(wb: Workbook) -> None:
    ws = wb.create_sheet("AI Decision Log")
    ws.freeze_panes = "A2"

    rows = _query(
        """
        SELECT agent_name, order_id, decision, reason,
               input_summary, output_summary, model_used, created_at
        FROM agent_decision_log
        WHERE created_at >= NOW() - INTERVAL '30 days'
        ORDER BY created_at DESC
        """
    )

    _write_header(ws, 1, [
        "Agent", "Order ID", "Decision", "Reason",
        "Input Summary", "Output Summary", "Model", "Timestamp",
    ])

    DECISION_STATUS = {
        "sent":      "sent",
        "reviewed":  "reviewed",
        "written":   "written",
        "priced":    "priced",
        "classified": "classified",
        "flagged":   "flagged",
        "error":     "error",
    }

    for i, r in enumerate(rows, 2):
        decision = r.get("decision") or ""
        _write_row(ws, i, [
            (r.get("agent_name") or "").replace("agent_", "").replace("_", " ").title(),
            r.get("order_id") or "",
            decision.title(),
            r.get("reason") or "",
            r.get("input_summary") or "",
            r.get("output_summary") or "",
            r.get("model_used") or "",
            _fmt_dt(r.get("created_at")),
        ], status=DECISION_STATUS.get(decision, "pending"))

    _auto_width(ws)
    ws.column_dimensions["D"].width = 45   # Reason
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 30


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    print()
    print("=" * 56)
    print("  NexGen Pipeline Excel Export")
    print("=" * 56)
    print()
    print(f"  Connecting to DB: {DB_HOST}/{DB_NAME}")

    try:
        _connect().close()
    except Exception as exc:
        print(f"  ERROR: Cannot connect to database: {exc}")
        sys.exit(1)

    print("  DB connection OK")
    print()

    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    steps = [
        ("Summary",            build_summary),
        ("Active Pipeline",    build_active_pipeline),
        ("Sent Estimates",     build_sent_estimates),
        ("Needs Review",       build_needs_review),
        ("AR Overdue",         build_ar_overdue),
        ("Monthly Statements", build_monthly_statements),
        ("AI Decision Log",    build_ai_decision_log),
    ]

    for name, fn in steps:
        print(f"  Building: {name}...")
        try:
            fn(wb)
        except Exception as exc:
            print(f"    WARNING: {name} tab failed ({exc}) -- skipping")

    wb.save(OUT_PATH)
    print()
    print(f"  Saved: {OUT_PATH}")
    print()
    print("  Share this file with the NexGen team.")
    print("  Re-run this script any time for a fresh snapshot.")
    print()


if __name__ == "__main__":
    main()
