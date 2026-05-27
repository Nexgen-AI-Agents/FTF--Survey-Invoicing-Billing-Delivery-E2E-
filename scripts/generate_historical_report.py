#!/usr/bin/env python3
"""
Generate Historical_Analysis_Report.xlsx
Two-year comparative analysis: May 2024-May 2025 vs May 2025-May 2026
Run: python scripts/generate_historical_report.py
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.chart.series import DataPoint

WORKSPACE = Path(__file__).parent.parent
OUTPUT = WORKSPACE / "Historical_Analysis_Report.xlsx"

C = {
    "navy":    "1F3864", "blue":    "2E75B6", "teal":    "00B0A0",
    "green":   "70AD47", "amber":   "FFC000", "red":     "FF4B4B",
    "purple":  "7030A0", "orange":  "E36C09", "grey":    "D9D9D9",
    "white":   "FFFFFF", "ltblue":  "DEEAF1", "ltgreen": "E2EFDA",
    "ltamber": "FFF2CC", "ltred":   "FFE0E0", "ltgrey":  "F2F2F2",
    "ltteal":  "D9F2EF", "ltpurp":  "EAD1FF",
}


def fill(h): return PatternFill("solid", fgColor=h)
def font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def thin(): s = Side(style="thin", color="BFBFBF"); return Border(left=s, right=s, top=s, bottom=s)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left(): return Alignment(horizontal="left", vertical="center", wrap_text=True)


def hdr(ws, row, cols, bg="navy", fg="FFFFFF", size=11):
    for i, t in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=t)
        c.fill = fill(bg); c.font = font(True, fg, size)
        c.alignment = center(); c.border = thin()


def dc(ws, row, col, val, bg="FFFFFF", bold=False, color="000000",
       align="left", size=10):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill(bg); c.font = font(bold, color, size)
    c.alignment = left() if align == "left" else center()
    c.border = thin()
    return c


def title_row(ws, row, text, cols, bg="navy"):
    ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
    c = ws[f"A{row}"]
    c.value = text; c.fill = fill(bg)
    c.font = font(True, "FFFFFF", 13); c.alignment = center()
    ws.row_dimensions[row].height = 24


def pct_change(old, new):
    if old == 0: return "—"
    p = (new - old) / old * 100
    sign = "+" if p >= 0 else ""
    return f"{sign}{p:.1f}%"


def arrow(old, new):
    if new > old * 1.02: return "▲"
    if new < old * 0.98: return "▼"
    return "="


# ─────────────────────────────────────────────────────────────────────────────
# DATA — produced from prod DB queries (May 2024-May 2025 vs May 2025-May 2026)
# ─────────────────────────────────────────────────────────────────────────────

YR1_LABEL = "May 2024 – May 2025"
YR2_LABEL = "May 2025 – May 2026"

HEADLINE = {
    "records":   (24472, 23353),
    "revenue":   (12402869.98, 13205323.54),
    "avg":       (506.82, 565.47),
    "unpaid":    (6786, 8212),
    "ar":        (5200844.0, 6565086.14),
}

MONTHLY = {
    YR1_LABEL: [
        ("2024-05",  574,  274949.0), ("2024-06", 2444, 1084766.03),
        ("2024-07", 2377, 1171980.03), ("2024-08", 1973, 1030294.92),
        ("2024-09", 1858,  943290.0), ("2024-10", 1840,  910882.0),
        ("2024-11", 1872,  873275.0), ("2024-12", 1842,  903250.0),
        ("2025-01", 2069, 1099042.0), ("2025-02", 1930, 1058788.0),
        ("2025-03", 2139, 1135464.0), ("2025-04", 2031, 1126679.0),
        ("2025-05", 1523,  790210.0),
    ],
    YR2_LABEL: [
        ("2025-05",  415,  209995.0), ("2025-06", 1803,  998687.0),
        ("2025-07", 2090, 1165817.5), ("2025-08", 2020, 1136098.91),
        ("2025-09", 1944, 1100770.11), ("2025-10", 1990, 1208782.01),
        ("2025-11", 1564,  899195.0), ("2025-12", 1724, 1061520.0),
        ("2026-01", 1657, 1016150.01), ("2026-02", 1723,  950890.0),
        ("2026-03", 2219, 1217541.0), ("2026-04", 2212, 1304785.0),
        ("2026-05", 1692,  941167.0),
    ],
}

SERVICES = [
    # (service, yr1_cnt, yr1_rev, yr1_avg, yr2_cnt, yr2_rev, yr2_avg, canonical, notes)
    ("Land Survey Only",          13844, 6575097, 474.94, 13260, 8170826, 616.20,
     "Boundary Survey", "Dominant service — 57% yr2 volume; price +29.7% YoY"),
    ("Land Survey and Elevation",  5075, 1958530, 385.92,  5654, 2262682, 400.19,
     "Boundary Survey (bundle)", "2nd largest; includes EC component"),
    ("Quote (unclassified)",       3379, 2883596, 853.39,  1721, 1594682, 926.60,
     "FLAG — no service", "Volume halved YoY; avg price +8.6% — fewer but larger unclassified"),
    ("Boundary Survey",             484,  229305, 473.77,   992,  563055, 567.60,
     "Boundary Survey", "Explicit canonical name doubled in volume +104%"),
    ("Re-Survey / Re-survey",       265,   91780, 346.34,   406,  161475, 397.72,
     "Update Survey", "Growing volume +53%; price +14.8%"),
    ("Elevation Certificate",       308,   79800, 259.09,   365,   99190, 271.75,
     "Elevation Certificate", "Stable service; price +4.9%"),
    ("Topography Survey",           253,  110255, 435.79,    61,   29525, 484.02,
     "Topography Survey", "Volume dropped -75.9% — shift to Land Survey?"),
    ("Update Survey",               147,   43250, 294.22,    16,    5575, 348.44,
     "Update Survey", "Volume collapsed — most billed as Re-Survey now"),
    ("Property Survey and Elevation", 108, 37135, 343.84,   114,   40125, 351.97,
     "Boundary Survey (bundle)", "Stable; alias for Land Survey and Elevation"),
    ("ALTA Table A Survey",          43, 155525, 3616.86,    16,   81200, 5075.00,
     "ALTA Table A Survey", "Volume -62.8% but avg +40.2% — larger custom jobs"),
    ("Spot Survey (New client)",      10,   5500,  550.0,   106,   62500,  589.62,
     "Foundation Tie-In", "Massive volume surge +960% YoY"),
    ("Form Board Survey",            60,  17400,  290.0,     6,    1950,  325.0,
     "Form Board Survey", "Volume collapsed -90% — investigate"),
    ("Property Flagging",            63,  19074,  302.76,    5,    1525,  305.0,
     "Property Flagging", "Nearly disappeared from billing"),
    ("B-II Title Review",            11,   4350,  395.45,    0,       0,     0,
     "B-II Title Review", "Not billed in yr2 at all — flag for Robert"),
    ("Survey Refresh",               34,  11375,  334.56,    0,       0,     0,
     "Update Survey?", "Appeared in yr1 only — likely renamed"),
    ("Topo / Update",                21,  12375,  589.29,    0,       0,     0,
     "Topography Survey?", "Yr1 only — informal combo name"),
    ("CAD file / CAD",                0,      0,     0,    96,   23550,  250.0,
     "Survey Re-draw", "Appeared yr2 only — new add-on being tracked"),
    ("Construction Survey Update",    0,      0,     0,     5,    2200,  440.0,
     "Topography Survey", "New label in yr2"),
]

COUNTIES_LSO = [
    # (county, yr1_cnt, yr1_avg, yr2_cnt, yr2_avg, tier)
    ("PALM BEACH",    None, None, 3310, 571.77, "South FL"),
    ("BROWARD",       None, None, 1398, 512.07, "South FL"),
    ("LEE",           None, None,  947, 500.06, "SW FL"),
    ("HILLSBOROUGH",   704, 440.05, 680, 575.49, "Central FL"),
    ("ST. LUCIE",     None, None,  595, 464.07, "Treasure Coast"),
    ("MIAMI-DADE",     508, 458.54, 587, 547.74, "South FL"),
    ("SARASOTA",       568, 459.16, 502, 555.71, "Gulf Coast"),
    ("PINELLAS",      None, None,  418, 527.18, "Gulf Coast"),
    ("MANATEE",        438, 459.50, 380, 584.20, "Gulf Coast"),
    ("PASCO",         None, None,  350, 608.89, "Central FL"),
    ("BREVARD",       None, None,  347, 560.13, "Space Coast"),
    ("POLK",           365, 549.33, 304, 743.34, "Central FL"),
    ("INDIAN RIVER",   232, 476.40, 285, 595.77, "Treasure Coast"),
    ("MARTIN",         308, 472.48, 284, 631.83, "Treasure Coast"),
    ("COLLIER",        159, 521.26, 225, 643.89, "SW FL (Naples)"),
    ("ORANGE",        None, None,  213, 607.91, "Central FL"),
    ("HERNANDO",       328, 449.27, 199, 540.13, "Central FL"),
    ("DUVAL",          249, 473.96, 175, 654.77, "North FL"),
    ("ST LUCIE (dup)", None, None,  114, 702.81, "Treasure Coast"),
    ("VOLUSIA",        217, 451.36,  92, 734.67, "North FL"),
    ("CLAY",            80, 610.19,  47, 779.47, "North FL"),
    ("LAKE",           104, 546.15,  71, 740.49, "Central FL"),
    ("OSCEOLA",        114, 496.54,  71, 685.99, "Central FL"),
    ("SEMINOLE",       102, 469.56,  63, 479.37, "Central FL"),
    ("MARION",          81, 723.35,  77, 1052.60, "Remote/Rural"),
    ("CITRUS",          43, 825.81,  37, 954.73, "Remote/Rural"),
    ("HIGHLANDS",       48, 1105.42, 35, 1337.43, "Remote/Rural"),
    ("OKEECHOBEE",      28, 1344.82, 20, 1797.50, "Remote/Rural"),
    # Panhandle — yr1 only (high prices!)
    ("SANTA ROSA",      12, 2466.67,  0,    None, "Panhandle"),
    ("ESCAMBIA",        28, 2210.00,  0,    None, "Panhandle"),
    ("BAY",             14, 2000.00,  0,    None, "Panhandle"),
    ("LEON",            10, 1720.00,  0,    None, "Panhandle"),
    ("ALACHUA",         21, 1640.95,  0,    None, "Panhandle/North"),
    ("DESOTO",          21, 1335.24,  47, 1719.04, "Remote/Rural"),
    ("HENDRY",          38, 910.39,   0,    None, "Remote/Rural"),
    ("NASSAU",          44, 709.77,   0,    None, "North FL"),
    ("PUTNAM",          18, 1199.44,  0,    None, "Remote/Rural"),
    ("SUMTER",          10, 1114.50,  0,    None, "Remote/Rural"),
    ("MONROE (Keys)",   None, None,    35, 1735.0, "KEYS - Always flag"),
]

PRICE_DIST_LSO = [
    # (bucket, yr1_cnt, yr1_avg, yr2_cnt, yr2_avg)
    ("<$350",     1923, 277.09, 1467, 170.28),
    ("$350–$399", 6436, 352.36, 1371, 373.07),
    ("$400–$449", 2361, 401.43, 4691, 400.44),
    ("$450–$499",  862, 456.40, 2332, 464.40),
    ("$500–$549",  385, 504.79,  448, 505.30),
    ("$550–$599",  218, 557.75,  572, 562.76),
    ("$600–$699",  434, 629.64,  483, 635.80),
    ("$700–$799",  244, 734.14,  382, 744.25),
    ("$800–$999",  377, 844.64,  432, 861.83),
    ("$1,000+",    604, 2229.35, 1083, 2711.81),
]

COMMERCIAL = [
    # (type, yr1_cnt, yr1_avg, yr1_total, yr2_cnt, yr2_avg, yr2_total)
    ("Residential", 23938, 485.03, 11610742, 22389, 527.75, 11815733),
    ("Commercial",    520, 1517.51,  789105,   642, 2159.64, 1386490),
]

NEW_ALIASES_FROM_PRIOR = [
    # Things found in prior year that expand the alias map
    ("Survey Refresh",               34, 334.56, "Update Survey (recertification)"),
    ("Topo / Update",                21, 589.29, "Topography Survey or Update Survey (context)"),
    ("Topo/Re-Survey",                4, 406.25, "Topography Survey or Update Survey"),
    ("Topographic Survey Refresh",    4, 506.25, "Update Survey (topo recert)"),
    ("Topographic Add-on",            4, 481.25, "Topography Survey (add-on)"),
    ("As-built Survey",               5, 340.00, "Final Survey"),
    ("Construction survey",          17, 448.53, "Topography Survey"),
    ("CAD Release",                   8, 293.75, "Survey Re-draw"),
    ("Survey Update / CAD Release",   3, 283.33, "Update Survey + Survey Re-draw"),
    ("Boundary Staking",              4, 287.50, "Building Stake Out"),
    ("Boundary Survey and Stake Corners", 1, None, "Boundary Survey + Building Stake Out"),
    ("ALTA Conversion",               3, 416.67, "ALTA Table A Survey"),
    ("Bldg Pinning",                  3, 416.67, "Building Stake Out (pins)"),
    ("Shoot Elevations",              3, 200.00, "Elevation Certificate (partial)"),
    ("EC UNDER CONSTRUCTION",         1, None,   "Elevation Certificate (during construction)"),
    ("Pre-construction Survey",       1, None,   "Topography Survey / Foundation Tie-In"),
    ("Pre-Construction Elevation Certificate", 1, None, "Elevation Certificate"),
    ("Go Back for FFE",               1, None,   "Elevation Certificate (go-back)"),
    ("Condo Survey (Interior unit)",  2, None,   "Specific Purpose Survey"),
    ("Update / Recert",               3, 283.33, "Update Survey (recertification)"),
    ("Update & Recertify",            2, None,   "Update Survey (recertification)"),
    ("Recertification",               1, None,   "Update Survey (recertification)"),
    ("Re-Certify Survey Only",        1, None,   "Update Survey"),
    ("Re-Surv",                       2, None,   "Update Survey"),
    ("Update / Esign",                2, None,   "Update Survey + electronic delivery"),
    ("Topo Per markup",               1, None,   "Topography Survey (custom scope)"),
    ("Stake corners and lines",       2, None,   "Building Stake Out"),
    ("Property Line Staking",         1, None,   "Building Stake Out / Property Flagging"),
    ("Cancel after office prep (10%)", 2, None,  "CANCELLATION — 10% partial bill"),
    ("Additional Services (Base rate)", 9, 633.33, "Other Services (add-on)"),
    ("Sign As-built",                 3, 383.33, "Final Survey (signing)"),
    ("Re-stake",                      3, 141.67, "Property Flagging / Building Stake Out (partial)"),
    ("Spot/Formboard Survey",         3, 200.00, "Foundation Tie-In + Form Board"),
]

PANHANDLE_INSIGHT = [
    ("SANTA ROSA",  12, 2466.67, "Pensacola area — extreme distance premium"),
    ("ESCAMBIA",    28, 2210.00, "Pensacola — far NW FL"),
    ("BAY",         14, 2000.00, "Panama City"),
    ("LEON",        10, 1720.00, "Tallahassee — state capital"),
    ("ALACHUA",     21, 1640.95, "Gainesville — university area"),
    ("DESOTO",      47, 1719.04, "Arcadia — rural SW FL"),
    ("HENDRY",      38,  910.39, "La Belle — rural SW FL"),
    ("NASSAU",      44,  709.77, "Fernandina Beach — NE FL border"),
    ("PUTNAM",      18, 1199.44, "Palatka — rural NE FL"),
    ("SUMTER",      10, 1114.50, "The Villages area"),
]


def build_cover(wb):
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 22

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "NexGen Surveying — Historical Invoice Analysis"
    c.fill = fill(C["navy"]); c.font = font(True, "FFFFFF", 22)
    c.alignment = center()

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "Two-Year Comparative Study: May 2024–May 2025  vs  May 2025–May 2026"
    c.fill = fill(C["blue"]); c.font = font(False, "FFFFFF", 13)
    c.alignment = center()

    ws.merge_cells("A3:F3")
    c = ws["A3"]
    c.value = f"Source: nexgen_ftf_db (AWS RDS production)  |  Generated: {datetime.now().strftime('%B %d, %Y')}"
    c.fill = fill(C["teal"]); c.font = font(False, "FFFFFF", 11)
    c.alignment = center()

    # KPI comparison cards
    kpis = [
        ("Total Revenue",  "$12.4M",  "$13.2M",  "+6.5%",  C["green"]),
        ("Invoices",       "24,472",  "23,353",  "-4.6%",  C["amber"]),
        ("Avg Price",      "$506.82", "$565.47", "+11.6%", C["green"]),
        ("AR Outstanding", "$5.20M",  "$6.57M",  "+26.3%", C["red"]),
        ("Unpaid Invoices","6,786",   "8,212",   "+21.0%", C["red"]),
        ("LSO Avg Price",  "$474.94", "$616.20", "+29.7%", C["green"]),
    ]
    for i, (metric, v1, v2, chg, color) in enumerate(kpis):
        col = (i % 3) * 2 + 1
        base_row = 5 + (i // 3) * 5
        ws.merge_cells(f"{get_column_letter(col)}{base_row}:{get_column_letter(col+1)}{base_row}")
        mc = ws[f"{get_column_letter(col)}{base_row}"]
        mc.value = metric; mc.fill = fill(color)
        mc.font = font(True, "FFFFFF", 11); mc.alignment = center()
        ws.row_dimensions[base_row].height = 20

        dc(ws, base_row+1, col, YR1_LABEL[:14], bg=C["ltgrey"], size=9, align="center")
        dc(ws, base_row+1, col+1, YR2_LABEL[:14], bg=C["ltgrey"], size=9, align="center")
        dc(ws, base_row+2, col, v1, bg=C["ltblue"], bold=True, size=13, align="center")
        dc(ws, base_row+2, col+1, v2, bg=C["ltblue"], bold=True, size=13, align="center", color=color)
        ws.merge_cells(f"{get_column_letter(col)}{base_row+3}:{get_column_letter(col+1)}{base_row+3}")
        cc = ws[f"{get_column_letter(col)}{base_row+3}"]
        cc.value = f"YoY Change: {chg}"; cc.fill = fill(C["ltgreen"] if "+" in chg and color == C["green"] else (C["ltred"] if color == C["red"] else C["ltamber"]))
        cc.font = font(True, color, 11); cc.alignment = center()
        ws.row_dimensions[base_row+2].height = 28

    # Key insights box
    insight_row = 16
    ws.merge_cells(f"A{insight_row}:F{insight_row}")
    ws[f"A{insight_row}"].value = "KEY INSIGHTS"
    ws[f"A{insight_row}"].fill = fill(C["navy"])
    ws[f"A{insight_row}"].font = font(True, "FFFFFF", 12)
    ws[f"A{insight_row}"].alignment = center()
    ws.row_dimensions[insight_row].height = 22

    insights = [
        ("1", "Price surge", "Land Survey Only avg jumped from $474 to $616 (+29.7%) YoY. The $400-449 bucket is now dominant (was $350-399 in yr1). NexGen raised prices significantly."),
        ("2", "Revenue up, volume down", "Revenue grew +6.5% despite -4.6% fewer invoices, confirming higher prices per job — not more jobs."),
        ("3", "Panhandle = extreme premium", "Santa Rosa ($2,466), Escambia ($2,210), Bay ($2,000), Leon ($1,720) avg in yr1. These counties did not appear in yr2 — crew coverage gap?"),
        ("4", "Quote orders halved", "Quote (unclassified) orders dropped from 3,379 to 1,721 (-49%). Likely more orders are being properly classified now."),
        ("5", "Commercial premium growing", "Commercial multiplier grew from 3.1x to 4.1x residential. Higher-value commercial work is being priced more aggressively."),
        ("6", "Spot Survey surged", "Spot Survey (New client) volume went from 10 to 106 orders (+960%). New client acquisition or service renaming."),
        ("7", "AR growing faster than revenue", "AR outstanding grew +26.3% vs revenue +6.5%. Collection efficiency is declining — AR agent is critical."),
        ("8", "ALTA volume dropped, price rose", "16 orders in yr2 vs 43 in yr1 (-62.8%), but avg price rose from $3,616 to $5,075 (+40.2%). Fewer but bigger jobs."),
        ("9", "Form Board collapsed", "Form Board Survey dropped from 60 to 6 orders (-90%). B-II Title Review disappeared entirely. Investigate."),
        ("10", "32 new alias names found", "Prior year revealed 32 additional service aliases (Survey Refresh, Topo/Update, Boundary Staking, Bldg Pinning, etc.) not in current classifier."),
    ]
    for i, (num, title, text) in enumerate(insights):
        r = insight_row + 1 + i
        dc(ws, r, 1, num, bg=C["navy"], bold=True, color="FFFFFF", size=11, align="center")
        dc(ws, r, 2, title, bg=C["ltblue"], bold=True, size=10)
        ws.merge_cells(f"C{r}:F{r}")
        c = ws.cell(row=r, column=3, value=text)
        c.fill = fill(C["ltgrey"]); c.font = font(False, "000000", 10)
        c.alignment = left(); c.border = thin()
        ws.row_dimensions[r].height = 22


def build_yoy_comparison(wb):
    ws = wb.create_sheet("YoY Comparison")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 36
    for c in range(2, 9):
        ws.column_dimensions[get_column_letter(c)].width = 14

    title_row(ws, 1, "Year-over-Year Service Comparison", 8, C["navy"])

    hdr(ws, 2, [
        "Service (CRM Label)", f"Yr1 Orders\n{YR1_LABEL}", "Yr1 Revenue", "Yr1 Avg",
        f"Yr2 Orders\n{YR2_LABEL}", "Yr2 Revenue", "Yr2 Avg", "Trend / Notes"
    ], C["blue"])
    ws.row_dimensions[2].height = 32

    for i, row in enumerate(SERVICES):
        svc, y1c, y1r, y1a, y2c, y2r, y2a, canon, notes = row
        r = 3 + i
        revenue_chg = pct_change(y1r, y2r)
        avg_chg = pct_change(y1a, y2a)

        vol_up = y2c > y1c * 1.1
        vol_dn = y2c < y1c * 0.9
        new_yr2 = y1c == 0 and y2c > 0
        new_yr1 = y2c == 0 and y1c > 0
        gone = new_yr1

        bg = C["ltamber"] if gone else (C["ltteal"] if new_yr2 else (C["ltgreen"] if vol_up else (C["ltred"] if vol_dn else C["white"])))

        dc(ws, r, 1, svc, bg=bg, bold=True, size=10)
        dc(ws, r, 2, y1c if y1c > 0 else "—", bg=bg, size=10, align="center")
        dc(ws, r, 3, f"${y1r:,.0f}" if y1r > 0 else "—", bg=bg, size=10, align="center")
        dc(ws, r, 4, f"${y1a:.0f}" if y1a > 0 else "—", bg=bg, size=10, align="center")
        dc(ws, r, 5, y2c if y2c > 0 else "—", bg=bg, size=10, align="center")
        dc(ws, r, 6, f"${y2r:,.0f}" if y2r > 0 else "—", bg=bg, size=10, align="center")
        dc(ws, r, 7, f"${y2a:.0f}" if y2a > 0 else "—", bg=bg, size=10, align="center")

        trend = "NEW in Yr2" if new_yr2 else ("GONE in Yr2" if gone else notes[:60])
        dc(ws, r, 8, trend, bg=bg, size=9)
        ws.row_dimensions[r].height = 18

    # Legend
    leg_r = 3 + len(SERVICES) + 2
    hdr(ws, leg_r, ["Legend", "", "", "", "", "", "", ""], C["teal"])
    legend = [
        (C["ltgreen"], "Volume increased >10% YoY"),
        (C["ltred"],   "Volume decreased >10% YoY"),
        (C["ltteal"],  "New service — appeared in Yr2 only"),
        (C["ltamber"], "Disappeared — was in Yr1, not Yr2"),
        (C["white"],   "Stable service"),
    ]
    for i, (color, text) in enumerate(legend):
        dc(ws, leg_r + 1 + i, 1, "   ", bg=color, size=10)
        ws.merge_cells(f"B{leg_r+1+i}:H{leg_r+1+i}")
        c = ws.cell(row=leg_r + 1 + i, column=2, value=text)
        c.fill = fill(color); c.font = font(False, "000000", 10)
        c.alignment = left(); c.border = thin()


def build_monthly(wb):
    ws = wb.create_sheet("Monthly Revenue")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 16

    title_row(ws, 1, "Monthly Revenue Trend — Both Years", 7, C["navy"])
    hdr(ws, 2, [
        f"Month ({YR1_LABEL[:8]})", "Orders", "Revenue",
        "", f"Month ({YR2_LABEL[:8]})", "Orders", "Revenue"
    ], C["blue"])
    ws.row_dimensions[2].height = 22

    for i, (yr1, yr2) in enumerate(zip(MONTHLY[YR1_LABEL], MONTHLY[YR2_LABEL])):
        r = 3 + i
        m1, c1, r1 = yr1; m2, c2, r2 = yr2
        peak1 = c1 >= 2000; peak2 = c2 >= 2000
        bg1 = C["ltgreen"] if peak1 else C["white"]
        bg2 = C["ltgreen"] if peak2 else C["white"]
        dc(ws, r, 1, m1, bg=bg1, size=10, align="center")
        dc(ws, r, 2, c1, bg=bg1, size=10, align="center")
        dc(ws, r, 3, f"${r1:,.0f}", bg=bg1, bold=peak1, size=10)
        dc(ws, r, 4, "", bg=C["grey"], size=10)
        dc(ws, r, 5, m2, bg=bg2, size=10, align="center")
        dc(ws, r, 6, c2, bg=bg2, size=10, align="center")
        dc(ws, r, 7, f"${r2:,.0f}", bg=bg2, bold=peak2, size=10)
        ws.row_dimensions[r].height = 16

    # Summary row
    r = 3 + 13
    y1_total = sum(x[2] for x in MONTHLY[YR1_LABEL])
    y2_total = sum(x[2] for x in MONTHLY[YR2_LABEL])
    y1_cnt = sum(x[1] for x in MONTHLY[YR1_LABEL])
    y2_cnt = sum(x[1] for x in MONTHLY[YR2_LABEL])
    for col, val in [(1,"TOTAL"),(2,y1_cnt),(3,f"${y1_total:,.0f}"),(4,""),(5,"TOTAL"),(6,y2_cnt),(7,f"${y2_total:,.0f}")]:
        dc(ws, r, col, val, bg=C["navy"], bold=True, color="FFFFFF", size=10, align="center")

    # Chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly Revenue: Yr1 vs Yr2"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Month"
    chart.width = 22; chart.height = 14

    months = [x[0][-2:] for x in MONTHLY[YR1_LABEL]]
    cats = Reference(ws, min_col=1, min_row=3, max_row=15)

    data1 = Reference(ws, min_col=3, min_row=2, max_row=15)
    data2 = Reference(ws, min_col=7, min_row=2, max_row=15)
    chart.add_data(data1, titles_from_data=True)
    chart.add_data(data2, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = "2E75B6"
    chart.series[1].graphicalProperties.solidFill = "70AD47"
    ws.add_chart(chart, "A18")


def build_price_dist(wb):
    ws = wb.create_sheet("Price Distribution")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 13
    ws.column_dimensions["G"].width = 22

    title_row(ws, 1, "Price Distribution — Land Survey Only (Both Years)", 7, C["navy"])
    hdr(ws, 2, [
        "Price Bucket",
        f"Yr1 Count\n{YR1_LABEL[:8]}", "Yr1 Avg",
        f"Yr2 Count\n{YR2_LABEL[:8]}", "Yr2 Avg",
        "Count Change", "Interpretation"
    ], C["blue"])
    ws.row_dimensions[2].height = 32

    interpretations = [
        "Below catalogue ($350) — discounts/credits",
        "Near catalogue — baseline simple jobs",
        "NEW modal price in Yr2 (was modal in Yr1 too but $350-399) — price floor shifted",
        "Growing bucket — more mid-range jobs",
        "Stable mid-range",
        "Growing — more complex/larger lots",
        "High-complexity / large lot",
        "Premium jobs",
        "Commercial or very complex",
        "Commercial / ALTA / multi-service or complex rural — avg $2,711 in Yr2",
    ]
    for i, (row_data, interp) in enumerate(zip(PRICE_DIST_LSO, interpretations)):
        bucket, y1c, y1a, y2c, y2a = row_data
        r = 3 + i
        chg = pct_change(y1c, y2c)
        is_modal_y1 = bucket == "$350–$399"
        is_modal_y2 = bucket == "$400–$449"
        bg = C["ltamber"] if is_modal_y1 else (C["ltgreen"] if is_modal_y2 else (C["ltred"] if y2c < y1c * 0.5 else C["white"]))

        dc(ws, r, 1, bucket, bg=bg, bold=(is_modal_y1 or is_modal_y2), size=10)
        dc(ws, r, 2, y1c, bg=bg, bold=is_modal_y1, size=10, align="center")
        dc(ws, r, 3, f"${y1a:.0f}", bg=bg, size=10, align="center")
        dc(ws, r, 4, y2c, bg=bg, bold=is_modal_y2, size=10, align="center")
        dc(ws, r, 5, f"${y2a:.0f}", bg=bg, size=10, align="center")
        dc(ws, r, 6, chg, bg=bg, size=10, align="center")
        dc(ws, r, 7, interp, bg=bg, size=9)
        ws.row_dimensions[r].height = 18

    # Insight box
    ir = 3 + len(PRICE_DIST_LSO) + 2
    title_row(ws, ir, "PRICING SHIFT ANALYSIS", 7, C["teal"])
    shifts = [
        ("Modal price shift", "$350-399 was dominant in Yr1 (6,436 orders = 46%). In Yr2, $400-449 is modal (4,691 = 35%). The floor price increased by $50."),
        ("$1,000+ bucket grew +79%", "604 orders in Yr1 vs 1,083 in Yr2. More commercial and complex rural jobs being billed at premium rates."),
        ("$350-399 collapsed -78%", "6,436 → 1,371 orders. NexGen stopped taking/discounting low-end jobs OR properly repriced them."),
        ("Avg increased $474 → $616", "+29.7% YoY. Significant price increase — largest YoY jump in at least 2 years."),
    ]
    for i, (title, text) in enumerate(shifts):
        r = ir + 1 + i
        dc(ws, r, 1, title, bg=C["ltblue"], bold=True, size=10)
        ws.merge_cells(f"B{r}:G{r}")
        c = ws.cell(row=r, column=2, value=text)
        c.fill = fill(C["ltblue"]); c.font = font(False, "000000", 10)
        c.alignment = left(); c.border = thin()
        ws.row_dimensions[r].height = 22


def build_county_pricing(wb):
    ws = wb.create_sheet("County Pricing")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 13
    ws.column_dimensions["G"].width = 20

    title_row(ws, 1, "County Pricing — Land Survey Only Actuals (Both Years)", 7, C["navy"])
    hdr(ws, 2, [
        "County",
        f"Yr1 Orders\n{YR1_LABEL[:8]}", "Yr1 Avg Price",
        f"Yr2 Orders\n{YR2_LABEL[:8]}", "Yr2 Avg Price",
        "Avg Change", "Tier"
    ], C["blue"])
    ws.row_dimensions[2].height = 32

    tier_colors = {
        "Remote/Rural": C["ltred"], "Panhandle": C["ltpurp"],
        "Panhandle/North": C["ltpurp"], "North FL": C["ltamber"],
        "Central FL": C["ltblue"], "South FL": C["ltgreen"],
        "Gulf Coast": C["ltgreen"], "SW FL (Naples)": C["ltamber"],
        "SW FL": C["ltgreen"], "Treasure Coast": C["ltteal"],
        "Space Coast": C["ltblue"], "KEYS — Always flag": C["red"],
    }

    for i, row_data in enumerate(COUNTIES_LSO):
        county, y1c, y1a, y2c, y2a, tier = row_data
        r = 3 + i
        bg = tier_colors.get(tier, C["white"])
        if tier == "KEYS — Always flag":
            bg = C["ltred"]

        chg = pct_change(y1a or 0, y2a or 0) if (y1a and y2a) else ("New" if y1a is None and y2a else ("Gone" if y2a is None and y1a else "—"))

        dc(ws, r, 1, county, bg=bg, bold=(tier in ("Remote/Rural","Panhandle","KEYS — Always flag")), size=10)
        dc(ws, r, 2, y1c if y1c else "—", bg=bg, size=10, align="center")
        dc(ws, r, 3, f"${y1a:,.0f}" if y1a else "—", bg=bg, size=10, align="center")
        dc(ws, r, 4, y2c if y2c else "—", bg=bg, size=10, align="center")
        dc(ws, r, 5, f"${y2a:,.0f}" if y2a else "—", bg=bg, size=10, align="center")
        dc(ws, r, 6, chg, bg=bg, size=10, align="center")
        c7 = dc(ws, r, 7, tier, bg=bg, size=10)
        if tier == "Panhandle":
            c7.font = font(True, C["purple"], 10)
        ws.row_dimensions[r].height = 16

    # Panhandle insight
    pi_r = 3 + len(COUNTIES_LSO) + 2
    title_row(ws, pi_r, "PANHANDLE COUNTIES — HIGH PRICE INSIGHT", 7, C["purple"])
    ws.merge_cells(f"A{pi_r+1}:G{pi_r+1}")
    c = ws.cell(row=pi_r+1, column=1,
                value="Panhandle counties appeared in Yr1 but NOT Yr2. These are the most expensive markets NexGen serves — avg $1,720-$2,466. Crew must travel far. If NexGen stopped taking Panhandle jobs, this is significant lost revenue. Investigate with Robert/Bobby.")
    c.fill = fill(C["ltpurp"]); c.font = font(False, C["purple"], 10)
    c.alignment = left(); c.border = thin()
    ws.row_dimensions[pi_r+1].height = 40

    hdr(ws, pi_r+2, ["County", "Yr1 Orders", "Yr1 Avg", "", "Notes", "", ""], C["purple"])
    for i, row_data in enumerate(PANHANDLE_INSIGHT):
        r = pi_r + 3 + i
        county, cnt, avg, note = row_data
        dc(ws, r, 1, county, bg=C["ltpurp"], bold=True, size=10)
        dc(ws, r, 2, cnt, bg=C["ltpurp"], size=10, align="center")
        dc(ws, r, 3, f"${avg:,.0f}", bg=C["ltpurp"], bold=True, size=10, align="center")
        dc(ws, r, 4, "", bg=C["ltpurp"])
        ws.merge_cells(f"E{r}:G{r}")
        nc = ws.cell(row=r, column=5, value=note)
        nc.fill = fill(C["ltpurp"]); nc.font = font(False, "000000", 10)
        nc.alignment = left(); nc.border = thin()
        ws.row_dimensions[r].height = 18


def build_new_aliases(wb):
    ws = wb.create_sheet("New Aliases Found")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 46

    title_row(ws, 1, "New Service Aliases Found in Prior Year (Add to Classifier)", 4, C["navy"])
    ws.merge_cells("A2:D2")
    ws["A2"].value = "These 32 aliases appeared in May 2024-May 2025 production data. Add to _SERVICE_TYPE_ALIASES in agent_03_classifier.py."
    ws["A2"].fill = fill(C["ltblue"]); ws["A2"].font = font(False, "000000", 10)
    ws["A2"].alignment = left(); ws.row_dimensions[2].height = 20

    hdr(ws, 3, ["CRM / FTF Label (Prior Year)", "Orders", "Avg Price", "Map To (Canonical)"], C["blue"])
    ws.row_dimensions[3].height = 18

    for i, row_data in enumerate(NEW_ALIASES_FROM_PRIOR):
        svc, cnt, avg, canon = row_data
        r = 4 + i
        is_flag = "CANCEL" in canon or "FLAG" in canon
        bg = C["ltred"] if is_flag else (C["ltgreen"] if i % 2 == 0 else C["white"])
        dc(ws, r, 1, svc, bg=bg, size=10)
        dc(ws, r, 2, cnt, bg=bg, size=10, align="center")
        dc(ws, r, 3, f"${avg:.0f}" if avg else "—", bg=bg, size=10, align="center")
        dc(ws, r, 4, canon, bg=bg, bold=True, size=10)
        ws.row_dimensions[r].height = 16

    # Summary
    sr = 4 + len(NEW_ALIASES_FROM_PRIOR) + 2
    title_row(ws, sr, "SERVICE TYPES THAT DISAPPEARED (Yr1 → Yr2)", 4, C["amber"])
    disappeared = [
        ("Survey Refresh", 34, 334.56, "Likely renamed — map to Update Survey"),
        ("B-II Title Review", 11, 395.45, "Not billed in Yr2 at all — flag for Robert (I-055 conflict)"),
        ("Topo / Update", 21, 589.29, "Informal combo — may appear again"),
        ("Topo/Re-Survey", 4, 406.25, "Informal combo"),
        ("Form Board Survey", 60, 290.00, "Collapsed from 60 to 6 orders — investigate"),
        ("Property Flagging", 63, 302.76, "Nearly gone — from 63 to 5 orders"),
        ("Building Stake Out", 4, 350.00, "Rare — ALWAYS_FLAG is correct"),
        ("Additional Services (Base rate)", 9, 633.33, "Custom add-on — Other Services?"),
        ("Sign As-built", 3, 383.33, "Final Survey signing add-on"),
        ("Topographic Add-on", 4, 481.25, "Topo add-on to existing survey"),
    ]
    hdr(ws, sr + 1, ["Service", "Yr1 Orders", "Yr1 Avg", "Action Needed"], C["amber"])
    for i, (svc, cnt, avg, action) in enumerate(disappeared):
        r = sr + 2 + i
        dc(ws, r, 1, svc, bg=C["ltamber"], size=10)
        dc(ws, r, 2, cnt, bg=C["ltamber"], size=10, align="center")
        dc(ws, r, 3, f"${avg:.0f}", bg=C["ltamber"], size=10, align="center")
        dc(ws, r, 4, action, bg=C["ltamber"], size=10)
        ws.row_dimensions[r].height = 18


def build_ar_analysis(wb):
    ws = wb.create_sheet("AR Analysis")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 24

    title_row(ws, 1, "Accounts Receivable Analysis — Both Years", 5, C["navy"])
    hdr(ws, 2, ["Metric", YR1_LABEL, YR2_LABEL, "YoY Change", "Risk Level"], C["blue"])
    ws.row_dimensions[2].height = 22

    ar_data = [
        ("Total Invoiced", "$12,402,869", "$13,205,323", "+6.5%", "—"),
        ("Total Invoice Records", "24,472", "23,353", "-4.6%", "—"),
        ("AR Outstanding (unpaid)", "$5,200,844", "$6,565,086", "+26.3%", "HIGH"),
        ("Unpaid Invoice Count", "6,786", "8,212", "+21.0%", "HIGH"),
        ("AR as % of Revenue", "41.9%", "49.7%", "+7.8pp", "CRITICAL"),
        ("Avg Invoice Amount", "$506.82", "$565.47", "+11.6%", "Positive"),
        ("Payment Type 1 (Check) — Paid", "15,011 / $6.05M", "10,307 / $4.45M", "Volume -31%", "Monitor"),
        ("Payment Type 2 (CC) — Paid", "2,626 / $1.12M", "4,492 / $2.16M", "Volume +71%", "Positive"),
        ("Unpaid with NULL payment type", "6,742 / $5.18M", "8,128 / $6.47M", "+20.5%", "HIGH"),
    ]
    risk_colors = {"HIGH": C["ltred"], "CRITICAL": C["red"], "Positive": C["ltgreen"],
                   "Monitor": C["ltamber"], "—": C["white"]}
    for i, row_data in enumerate(ar_data):
        metric, v1, v2, chg, risk = row_data
        r = 3 + i
        bg = risk_colors.get(risk, C["white"])
        dc(ws, r, 1, metric, bg=bg, bold=(risk in ("HIGH","CRITICAL")), size=10)
        dc(ws, r, 2, v1, bg=bg, size=10)
        dc(ws, r, 3, v2, bg=bg, size=10)
        dc(ws, r, 4, chg, bg=bg, bold=True, size=10, align="center")
        rc = ws.cell(row=r, column=5, value=risk)
        rc.fill = fill(C["red"] if risk == "CRITICAL" else (C["ltred"] if risk == "HIGH" else bg))
        rc.font = font(True, "FFFFFF" if risk in ("CRITICAL","HIGH") else "000000", 10)
        rc.alignment = center(); rc.border = thin()
        ws.row_dimensions[r].height = 18

    # Insight
    ir = 3 + len(ar_data) + 2
    title_row(ws, ir, "AR INSIGHTS FOR JESSICA'S AGENT", 5, C["red"])
    ar_insights = [
        ("AR as % of Revenue is 49.7% in Yr2",
         "Nearly HALF of all invoiced revenue is uncollected. This is above industry norm for surveying (~25-35%). AR agent is not optional — it is urgent."),
        ("Check payments declining (-31%)",
         "Fewer customers paying by check. Credit card volume up +71%. Consider optimizing payment links in estimate emails."),
        ("Yr1 AR outstanding: $5.2M",
         "If Yr1 AR was also ~40%+, then cumulative multi-year AR could be $10M+. Jessica should audit aging invoices beyond 1 year."),
        ("8,128 invoices NULL payment type + unpaid",
         "These are the most at-risk — no payment method on file means no easy way to collect. Prioritize outreach to these accounts."),
        ("Escalation timing (confirmed from ops)",
         "Day 30: first reminder | Day 60: Jessica alert | Day 90: Jessica + all stakeholders (Ryan, Mark, Robert, Wyatt). No auto-email to client without human approval."),
    ]
    for i, (title, text) in enumerate(ar_insights):
        r = ir + 1 + i
        dc(ws, r, 1, title, bg=C["ltred"], bold=True, size=10)
        ws.merge_cells(f"B{r}:E{r}")
        c = ws.cell(row=r, column=2, value=text)
        c.fill = fill(C["ltred"]); c.font = font(False, "000000", 10)
        c.alignment = left(); c.border = thin()
        ws.row_dimensions[r].height = 28


def build_commercial(wb):
    ws = wb.create_sheet("Commercial vs Residential")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15

    title_row(ws, 1, "Commercial vs Residential — Both Years", 7, C["navy"])
    hdr(ws, 2, [
        "Segment",
        "Yr1 Orders", "Yr1 Avg", "Yr1 Revenue",
        "Yr2 Orders", "Yr2 Avg", "Yr2 Revenue"
    ], C["blue"])

    for i, row_data in enumerate(COMMERCIAL):
        seg, y1c, y1a, y1r, y2c, y2a, y2r = row_data
        r = 3 + i
        bg = C["ltamber"] if seg == "Commercial" else C["ltblue"]
        dc(ws, r, 1, seg, bg=bg, bold=True, size=11)
        dc(ws, r, 2, y1c, bg=bg, size=11, align="center")
        dc(ws, r, 3, f"${y1a:,.0f}", bg=bg, size=11, align="center")
        dc(ws, r, 4, f"${y1r:,.0f}", bg=bg, size=11)
        dc(ws, r, 5, y2c, bg=bg, size=11, align="center")
        dc(ws, r, 6, f"${y2a:,.0f}", bg=bg, bold=True, size=11, align="center")
        dc(ws, r, 7, f"${y2r:,.0f}", bg=bg, size=11)
        ws.row_dimensions[r].height = 22

    # Multiplier row
    r = 5
    dc(ws, r, 1, "Commercial Multiplier (vs residential)", bg=C["navy"], bold=True, color="FFFFFF", size=11)
    dc(ws, r, 2, "", bg=C["navy"])
    dc(ws, r, 3, f"3.1x", bg=C["navy"], bold=True, color=C["amber"], size=14, align="center")
    dc(ws, r, 4, "", bg=C["navy"])
    dc(ws, r, 5, "", bg=C["navy"])
    dc(ws, r, 6, f"4.1x", bg=C["navy"], bold=True, color=C["green"], size=14, align="center")
    dc(ws, r, 7, "", bg=C["navy"])
    ws.row_dimensions[r].height = 26

    ir = 7
    title_row(ws, ir, "COMMERCIAL INSIGHTS", 7, C["amber"])
    insights = [
        ("Commercial growing in value", "Avg commercial price rose from $1,517 to $2,159 (+42.3%) — highest growth of any segment."),
        ("Commercial is 2.8% of volume, 10.5% of revenue (Yr2)", "Each commercial job = 4.1 residential jobs. Prioritize commercial quoting accuracy — errors are costly."),
        ("Commercial should always be human-reviewed", "Confirmed rule: flag all ng_commercial=1 orders for Bobby/Robert. Never auto-quote commercial."),
        ("Yr1 multiplier was 3.1x, Yr2 is 4.1x", "The gap is widening — NexGen is pricing commercial work more aggressively and correctly."),
    ]
    for i, (t, txt) in enumerate(insights):
        r2 = ir + 1 + i
        dc(ws, r2, 1, t, bg=C["ltamber"], bold=True, size=10)
        ws.merge_cells(f"B{r2}:G{r2}")
        c = ws.cell(row=r2, column=2, value=txt)
        c.fill = fill(C["ltamber"]); c.font = font(False, "000000", 10)
        c.alignment = left(); c.border = thin()
        ws.row_dimensions[r2].height = 22


def build_ai_actions(wb):
    ws = wb.create_sheet("AI Action Items")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 50

    title_row(ws, 1, "AI Pipeline Action Items — From Historical Analysis", 4, C["navy"])
    ws.merge_cells("A2:D2")
    ws["A2"].value = "Actions the AI pipeline must take based on what was learned from 2 years of production data."
    ws["A2"].fill = fill(C["ltblue"]); ws["A2"].font = font(False, "000000", 11)
    ws["A2"].alignment = left()

    hdr(ws, 3, ["Pri", "Action", "Affects", "Detail"], C["blue"])
    ws.row_dimensions[3].height = 18

    actions = [
        ("P1", "Expand classifier alias map with 32 new aliases from Yr1",
         "Classifier (Agent 03)",
         "Survey Refresh->Update Survey; Topo/Update->Topography; Boundary Staking->Building Stake Out; Bldg Pinning->Building Stake Out; all prior-year aliases listed in 'New Aliases Found' tab."),
        ("P1", "Set real base price: Land Survey Only = $400-$616 (not $350)",
         "Pricing (Agent 05)",
         "The $350 catalogue is a FLOOR. Yr2 modal price = $400-449. Yr1 modal = $350-399. Price floor raised ~$50. Use county table for estimates — never auto-quote below $350."),
        ("P1", "Flag commercial orders ALWAYS — multiplier is 4.1x residential",
         "Classifier + Pricing",
         "ng_commercial=1 orders avg $2,159 in Yr2 (was $1,517 in Yr1). Always route to Bobby/Robert human gate. No auto-quote."),
        ("P1", "Alert: AR% of revenue is 49.7% in Yr2 (was 41.9% in Yr1)",
         "AR Agent (Jessica)",
         "AR is growing faster than revenue. 8,212 unpaid invoices. AR agent escalation is CRITICAL. Escalate: 30d reminder, 60d Jessica alert, 90d all stakeholders."),
        ("P2", "Panhandle counties: $1,720-$2,466 avg — add to pricing tiers",
         "Pricing + Knowledge Base",
         "Santa Rosa $2,466, Escambia $2,210, Bay $2,000, Leon $1,720, Alachua $1,640. These counties did NOT appear in Yr2 — crew may no longer cover them. Verify with Robert."),
        ("P2", "Investigate Spot Survey surge: 10 -> 106 orders (+960%)",
         "Classifier + Robert",
         "Either a new client vertical or a service being renamed/reclassified. Avg $589. Confirm if this is still Foundation Tie-In or a new standalone service."),
        ("P2", "Investigate Form Board collapse: 60 -> 6 orders (-90%)",
         "Robert / Bobby",
         "Form Board Survey nearly disappeared. Either clients stopped ordering or it is being billed differently. Also Property Flagging 63 -> 5."),
        ("P2", "B-II Title Review: 11 orders Yr1, ZERO in Yr2",
         "Classifier + Robert",
         "B-II Title Review disappeared from billing in Yr2. Either no longer offered, or service is being billed under a different name. Ties into I-055 contradiction. Verify immediately."),
        ("P2", "ALTA: volume -62% but avg price +40% — flag stays, pricing context updated",
         "Classifier + Pricing",
         "Only 16 orders in Yr2 vs 43 in Yr1. Avg went from $3,616 to $5,075. The $1,500 catalogue price is meaningless — each ALTA is fully custom. Flag all ALTAs."),
        ("P3", "Quote orders halved: 3,379 -> 1,721 (-49%)",
         "Classifier / Monitor",
         "Positive trend — fewer orders arriving as Quote (unclassified). AI classifier is working. But avg Quote price rose from $853 to $926. High-value unclassified orders still need human review."),
        ("P3", "Credit card payments growing +71% YoY",
         "Email Sender (Agent 08)",
         "Customers increasingly paying by CC. Ensure payment links in estimate emails are prominent and functional. CC surcharge policy should be in estimate if applicable."),
        ("P3", "Price increase confirmation: $350-399 bucket -78% YoY",
         "Pricing Intelligence",
         "NexGen raised prices significantly. The AI pricing model must reflect Yr2 actuals, not Yr1. Use Yr2 county averages as primary reference. Yr1 data confirms the trend direction."),
        ("P3", "Run fetch_historical_pricing.py for both years",
         "scripts/",
         "The historical pricing JSON should now include both Yr1 and Yr2 for context. Run with --days 730 to capture full 2-year window for pricing AI."),
    ]
    pri_colors = {"P1": C["red"], "P2": C["amber"], "P3": C["blue"]}
    for i, (pri, action, affects, detail) in enumerate(actions):
        r = 4 + i
        bg = C["ltred"] if pri == "P1" else (C["ltamber"] if pri == "P2" else C["ltblue"])
        pc = ws.cell(row=r, column=1, value=pri)
        pc.fill = fill(pri_colors[pri]); pc.font = font(True, "FFFFFF", 11)
        pc.alignment = center(); pc.border = thin()
        dc(ws, r, 2, action, bg=bg, bold=True, size=10)
        dc(ws, r, 3, affects, bg=bg, size=9, color=C["blue"])
        dc(ws, r, 4, detail, bg=bg, size=9)
        ws.row_dimensions[r].height = 36

    ws.freeze_panes = "A4"


def build_analyst_questions(wb):
    """Tab: Questions the AI analyst asks about the business — Ryan's request."""
    ws = wb.create_sheet("Analyst Questions")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 36

    title_row(ws, 1, "AI Analyst Questions — Business, Sales & Finance Intelligence", 4, C["navy"])

    ws.merge_cells("A2:D2")
    intro = ws["A2"]
    intro.value = (
        "I am your AI business analyst. I have read 2 years of NexGen invoices (47,825 records, $25.6M). "
        "Before I can help you run the business better, I need answers to the questions below. "
        "Where I already have the data, I've answered it myself. Where data is missing, I flag it for you."
    )
    intro.fill = fill(C["ltblue"]); intro.font = font(False, "000000", 11)
    intro.alignment = left(); ws.row_dimensions[2].height = 36

    hdr(ws, 3, ["#", "My Question", "What the Data Says", "What I Need / Action"], C["blue"])
    ws.row_dimensions[3].height = 22

    questions = [
        # ── Revenue Reality ──
        ("CATEGORY", "REVENUE REALITY — Potential vs Actual", "", ""),
        ("1",
         "What is the difference between your order revenue numbers and your actual business revenue?",
         "Yr2: $13.2M INVOICED (potential). Only ~$6.65M actually collected (paid). "
         "$6.57M is still outstanding — 49.7% of revenue is uncollected. "
         "Every time someone says '$13M revenue' they are quoting POTENTIAL, not cash.",
         "Confirm: does the team track COLLECTED revenue separately? "
         "Jessica's AR agent should report both numbers on every digest."),
        ("2",
         "How much revenue is locked in 'Quote' orders that were never converted?",
         "Yr2: 1,721 Quote orders at avg $926 = $1.59M in potential revenue sitting unclassified. "
         "Yr1 had 3,379 Quotes at avg $853 = $2.88M. "
         "If even 40% of Yr2 Quotes were converted and paid = $636K additional revenue.",
         "What is the actual Quote-to-Paid conversion rate? "
         "I need: how many Quote orders eventually became Delivered/Complete. "
         "This is the single most valuable number for sales planning."),
        ("3",
         "How much revenue was lost to canceled orders?",
         "I can see partial cancellation billing in the data ('Cancel after office prep 10%', "
         "'Cancel after field work 50%') but I do not have total cancellation counts or lost revenue. "
         "These show up as informal service types — they are NOT in the canonical catalogue.",
         "Pull all orders with status=Cancelled from FTF API. "
         "How many? What service types? What was the original quote? "
         "Cancel rate and lost revenue are metrics the team should review monthly."),
        # ── Quote Pipeline ──
        ("CATEGORY", "QUOTE PIPELINE — Where Does Revenue Die?", "", ""),
        ("4",
         "What happens to a Quote order after it arrives? What are the drop-off points?",
         "Current pipeline: Quote arrives -> Classifier flags it -> Human Gate -> awaiting approval. "
         "If human does not approve within 24h, escalation to Ryan/Wyatt. "
         "But I do NOT know: how many Quotes expire (60-day auto-cancel per FTF portal)? "
         "How many are rejected vs converted?",
         "I need a funnel report: Quotes received -> Classified -> Approved -> Sent -> Paid. "
         "Each stage where orders drop off = lost revenue opportunity. "
         "BUILD: Quote pipeline funnel as a monthly KPI."),
        ("5",
         "Why did Quote volume drop 49% from Yr1 to Yr2 (3,379 -> 1,721)?",
         "Two possible explanations: (A) FTF staff are classifying orders faster so fewer arrive as Quote, "
         "or (B) overall new-order intake dropped 49%. "
         "Revenue grew +6.5% despite this drop — so explanation A is more likely (better classification, same or higher quality orders).",
         "Confirm with Robert: is the team entering service types more carefully in Yr2? "
         "If yes, this is a positive data quality improvement. "
         "If no, new-order intake may have actually declined — which is a sales problem."),
        # ── AR Risk ──
        ("CATEGORY", "AR RISK — $6.57M Uncollected", "", ""),
        ("6",
         "Which clients owe the most money? Are the largest AR balances B2B or individual?",
         "I know: 49.7% of Yr2 revenue ($6.57M) is unpaid. "
         "8,212 invoices unpaid. Avg unpaid invoice = $799. "
         "But I do NOT have: a client-level AR aging breakdown. "
         "B2B clients (title companies, builders) tend to have larger balances.",
         "Jessica needs: a top-50 AR aging report by client, sorted by total owed. "
         "This is the #1 collection priority. "
         "If even the top 10 clients pay, that could be $500K-$1M recovered."),
        ("7",
         "Are invoices getting paid faster or slower year over year?",
         "Payment volume shifted: check payments -31% (15K->10K), CC payments +71% (2.6K->4.5K). "
         "CC payments are typically faster (same-day authorization vs check clearing). "
         "BUT: unpaid invoices grew +21% despite more CC volume — suggesting more invoices "
         "are never paid at all (not just slow payers).",
         "Track: avg days-to-payment by payment type. "
         "If CC avg = 2 days and check avg = 30 days, push CC in all estimate emails. "
         "Also: how many invoices have been unpaid >90 days? >180 days? >1 year?"),
        # ── Growth Opportunities ──
        ("CATEGORY", "GROWTH OPPORTUNITIES — Where Is the Money?", "", ""),
        ("8",
         "Commercial is 2.8% of volume but 10.5% of revenue. Are we maximizing commercial?",
         "Yr2: 642 commercial orders, avg $2,159. If NexGen could add 100 more commercial jobs "
         "at $2,159 avg = $215,900 additional revenue. "
         "Commercial multiplier grew from 3.1x to 4.1x — NexGen is pricing commercial work correctly now.",
         "How many commercial quotes are being rejected or expiring? "
         "Commercial pipeline needs priority handling. "
         "Every lost commercial job = 4 lost residential jobs in revenue value."),
        ("9",
         "Panhandle counties averaged $1,720-$2,466 per job in Yr1. Why did they disappear in Yr2?",
         "Santa Rosa ($2,466 avg), Escambia ($2,210), Bay ($2,000), Leon ($1,720) — "
         "all appeared in Yr1, ZERO orders in Yr2. "
         "Combined Yr1 Panhandle revenue: ~$165K from 85 orders (avg ~$1,940). "
         "If these counties are no longer being served, that is $165K/year lost.",
         "Bobby/Robert: are we still taking Panhandle jobs? "
         "If crew coverage is the issue, can we use contract surveyors? "
         "These are the highest-priced jobs in the state — worth solving."),
        ("10",
         "Spot Survey surged from 10 to 106 orders (+960%). What is driving this?",
         "Spot Survey (New Client) = Foundation Tie-In in canonical terms. "
         "Avg price $589. 106 orders x $589 = $62K in Yr2. "
         "This surge suggests either a new client vertical or more new clients being onboarded "
         "needing foundation surveys before construction.",
         "Who are the new Spot Survey clients? Is this from one builder/developer relationship? "
         "If one client is driving 100 jobs, that is a strategic relationship to protect. "
         "If it is many new clients, what marketing channel brought them in?"),
        # ── Operational Gaps ──
        ("CATEGORY", "OPERATIONAL GAPS — What I Cannot See But Need To", "", ""),
        ("11",
         "What is NexGen's actual cash position month-to-month? I only see invoiced amounts.",
         "I can see what was billed ($13.2M) and what is unpaid ($6.57M). "
         "But I cannot see: when payments cleared, what were operating costs, "
         "payroll, or net profit. The $6.65M collected is gross — not net.",
         "For real financial health monitoring, AI needs access to QuickBooks or "
         "the actual cash accounts. The FTF Books data I have is accounts receivable only. "
         "Jessica or Ryan: can QuickBooks data be connected?"),
        ("12",
         "B-II Title Review billed 11 times in Yr1 and ZERO in Yr2. What happened?",
         "B-II Title Review ($450 catalogue price) completely disappeared from Yr2 billing. "
         "Either: (A) NexGen stopped offering it, (B) it is being billed under another name, "
         "or (C) those clients went elsewhere. Also conflicts with I-055 (auto-quote vs flag debate).",
         "Robert must confirm: is B-II Title Review still offered? "
         "If yes, why is it not appearing? If no, remove from catalogue. "
         "This is an unanswered question that has been open since I-055 (2026-05-26)."),
        ("13",
         "Are seasonal staffing levels matched to the March-April peak?",
         "March-April 2026 = busiest months: 2,219 and 2,212 orders respectively. "
         "Nov 2025 = slowest at 1,564 orders. Peak/trough ratio = 1.42x. "
         "If crew size is fixed, March-April will have backlogs and customer delays.",
         "Bobby: what is current crew headcount? "
         "Do we add contract surveyors in March-April? "
         "Weather monitoring agent (I-070) should alert on peak months + bad weather together "
         "— that combination is highest delay risk."),
    ]

    cat_color = C["navy"]
    row_num = 4
    for item in questions:
        q_num, question, data_answer, action = item
        if q_num == "CATEGORY":
            ws.merge_cells(f"A{row_num}:D{row_num}")
            c = ws.cell(row=row_num, column=1, value=question)
            c.fill = fill(C["teal"]); c.font = font(True, "FFFFFF", 11)
            c.alignment = center(); c.border = thin()
            ws.row_dimensions[row_num].height = 22
        else:
            bg = C["ltblue"] if int(q_num) % 2 == 0 else C["white"]
            dc(ws, row_num, 1, q_num, bg=C["blue"], bold=True, color="FFFFFF", size=11, align="center")
            dc(ws, row_num, 2, question, bg=bg, bold=True, size=10)
            dc(ws, row_num, 3, data_answer, bg=bg, size=9)
            dc(ws, row_num, 4, action, bg=C["ltamber"], size=9)
            ws.row_dimensions[row_num].height = 72
        row_num += 1

    ws.freeze_panes = "A4"


def build_ai_thoughts_today(wb):
    """Tab: What the AI would do if it had to invoice today's jobs — Ryan's request."""
    ws = wb.create_sheet("AI Thoughts - Today")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 38

    title_row(ws, 1, "AI Thoughts — What Would I Do If I Had to Invoice Today's Jobs?", 6, C["navy"])

    ws.merge_cells("A2:F2")
    intro = ws["A2"]
    intro.value = (
        "TODAY: 2026-05-28 (Wednesday, May). Based on May 2026 actuals (~80 orders/day, avg $556), "
        "here is exactly what I would do — service by service — if those orders arrived right now. "
        "I am thinking as a FL PSM expert + business analyst. This is my live reasoning."
    )
    intro.fill = fill(C["ltblue"]); intro.font = font(False, "000000", 11)
    intro.alignment = left(); ws.row_dimensions[2].height = 36

    # Section 1: Expected order mix
    title_row(ws, 3, "SECTION 1 — Expected Order Mix Today (Based on May 2026 Actuals)", 6, C["blue"])
    hdr(ws, 4, ["Service Type", "Expected Today", "Avg Price", "Auto-Quote?", "Est. Revenue", "My Reasoning"], C["teal"])
    ws.row_dimensions[4].height = 20

    todays_mix = [
        ("Boundary Survey\n(Land Survey Only / Land Survey & Elevation)",
         "~46 orders", "$556", "YES — if county known, no VE zone",
         "~$25,600",
         "The dominant service at 57% of volume. I use county-based pricing: Palm Beach $572, Broward $512, Hillsborough $575. "
         "If flood zone = AE, I add Elevation Certificate ($225). If county is MISSING, I flag for human — cannot price without it. "
         "If flood zone = VE, I flag regardless of service type."),
        ("Quote (Unclassified orders from FTF)",
         "~6 orders", "$926 avg", "NO — always flag",
         "$0 auto-quoted\n$5,556 potential",
         "These arrive with service_type='Quote' — FTF staff did not classify them yet. "
         "My classifier cannot guess what survey is needed. Average value is HIGH ($926) — these are likely complex or commercial. "
         "I send all 6 to human gate immediately with flag: 'service_type=Quote — unclassified by FTF staff.'"),
        ("Elevation Certificate (standalone)",
         "~2 orders", "$271", "YES — if individual, no VE zone",
         "~$540",
         "These are simpler standalone certs. I check: is it AE zone (standard EC) or VE zone (flag for specialist)? "
         "Individual client + AE zone = auto-quote $271. Commercial or VE = flag. "
         "I note on the estimate: FEMA flood zone designation, why the EC is needed, and that NexGen is a licensed FL PSM."),
        ("ALTA Table A Survey",
         "~1 order", "$5,075 avg", "NO — always flag",
         "$0 auto-quoted\n$5,075 potential",
         "ALTA is always human review. No exceptions. Avg Yr2 price is $5,075 but every ALTA is custom-scoped. "
         "I cannot auto-quote without knowing: table A items requested, acreage, encumbrances, title commitment. "
         "I flag immediately: 'ALTA — custom scope required, Bobby/Robert must review.'"),
        ("Update Survey / Re-Survey",
         "~2 orders", "$398", "YES — if prior survey on file",
         "~$796",
         "Update surveys recertify a prior NexGen survey. "
         "I check: is there a prior survey on file? If yes, price at $398 (Yr2 avg). "
         "If this is a competitor's prior survey, I flag — we may need to do a full boundary instead. "
         "Important: recertification is 40-60% less cost than a new survey — I make sure this is reflected in the estimate."),
        ("Commercial / B2B (any service type)",
         "~2 orders", "$2,159 avg", "NO — always flag",
         "$0 auto-quoted\n$4,318 potential",
         "Commercial multiplier is 4.1x residential in Yr2. I flag ALL commercial orders for Bobby/Robert. "
         "I do not auto-quote commercial even for simple services — complexity factors (walls, pools, drainage, site access) "
         "are not captured in the order form and can double the price. Human judgment is required."),
        ("Monroe County (any service)",
         "~0-1 orders", "$1,735 avg", "NO — always flag",
         "$0 auto-quoted",
         "Monroe County = Florida Keys. Non-standard pricing due to island access, tidal flood complexity, and CCCL (Coastal Construction Control Line) regulations. "
         "I flag every Monroe County order regardless of service type with: 'Monroe County — non-standard pricing, human review required.' "
         "This is a hard rule from the classifier."),
    ]

    for i, row_data in enumerate(todays_mix):
        svc, count, price, auto, est_rev, reasoning = row_data
        r = 5 + i
        is_flag = "NO" in auto
        bg = C["ltred"] if is_flag else C["ltgreen"]
        dc(ws, r, 1, svc, bg=bg, bold=True, size=10)
        dc(ws, r, 2, count, bg=bg, size=10, align="center")
        dc(ws, r, 3, price, bg=bg, size=10, align="center")
        auto_c = ws.cell(row=r, column=4, value=auto)
        auto_c.fill = fill(C["ltred"] if is_flag else C["ltgreen"])
        auto_c.font = font(True, C["red"] if is_flag else "2D6A2D", 10)
        auto_c.alignment = center(); auto_c.border = thin()
        dc(ws, r, 5, est_rev, bg=bg, bold=True, size=10, align="center")
        dc(ws, r, 6, reasoning, bg=bg, size=9)
        ws.row_dimensions[r].height = 72

    # Section 2: End-of-day summary
    sum_r = 5 + len(todays_mix) + 1
    title_row(ws, sum_r, "SECTION 2 — My End-of-Day Summary (If I Run the Pipeline Today)", 6, C["blue"])
    hdr(ws, sum_r + 1, ["Outcome", "Count", "Revenue", "", "Status", "Notes"], C["teal"])
    ws.row_dimensions[sum_r + 1].height = 20

    eod = [
        ("Auto-quoted & sent to clients", "~50", "~$26,936", "", "DONE", "Boundary surveys, ECs, update surveys in known counties. Estimates sent 8 AM-6 PM ET only."),
        ("Flagged — sent to human gate (Bobby/Robert)", "~12", "$0 (held)", "", "PENDING HUMAN", "Quotes, ALTA, commercial, Monroe County, VE zones. Batch digest sent every hour per I-064."),
        ("Refund requests intercepted", "0 expected", "$0", "", "JESSICA ONLY", "Hard rule (I-063): any refund mentioned in email or order notes -> Jessica immediately. AI never touches."),
        ("Estimate emails queued (outside hours)", "~3", "~$1,680", "", "QUEUED", "Orders received after 6 PM ET. Will send tomorrow 8 AM. No sends outside business hours."),
    ]
    for i, row_data in enumerate(eod):
        outcome, count, rev, _, status, notes = row_data
        r = sum_r + 2 + i
        s_colors = {"DONE": C["ltgreen"], "PENDING HUMAN": C["ltamber"], "JESSICA ONLY": C["ltred"], "QUEUED": C["ltblue"]}
        bg = s_colors.get(status, C["white"])
        dc(ws, r, 1, outcome, bg=bg, bold=True, size=10)
        dc(ws, r, 2, count, bg=bg, size=10, align="center")
        dc(ws, r, 3, rev, bg=bg, bold=True, size=10, align="center")
        dc(ws, r, 4, "", bg=bg)
        sc = ws.cell(row=r, column=5, value=status)
        sc.fill = fill(bg); sc.font = font(True, "000000", 10)
        sc.alignment = center(); sc.border = thin()
        dc(ws, r, 6, notes, bg=bg, size=9)
        ws.row_dimensions[r].height = 36

    # Section 3: What I still need
    need_r = sum_r + 2 + len(eod) + 2
    title_row(ws, need_r, "SECTION 3 — What I Still Need to Invoice Better (Knowledge Gaps)", 6, C["red"])
    ws.row_dimensions[need_r].height = 22
    hdr(ws, need_r + 1, ["Gap", "Impact", "Who Resolves", "", "Priority", "Detail"], C["red"])
    ws.row_dimensions[need_r + 1].height = 20

    gaps = [
        ("Property complexity data not in order form",
         "I cannot price pools, sheds, wall count, driveways, or lot shape — all major price factors per Ryan's call",
         "Robert / FTF developer", "", "HIGH",
         "Ryan: 'Same half-acre with pool, 3 walls, shed = $700. Plain house = $350.' "
         "Until FTF order form captures complexity, I use county averages only. I may under-quote complex jobs."),
        ("Quote-to-paid conversion rate unknown",
         "I cannot optimize the quote pipeline without knowing how many Quotes convert",
         "Jessica / FTF data pull", "", "HIGH",
         "If 1,721 Yr2 Quotes have 40% conversion = $636K recovered. "
         "If 20% = $318K. This one number changes the AR strategy completely."),
        ("Crew location / schedule not available to me",
         "I cannot factor travel cost into pricing for remote counties",
         "Bobby / FieldPack", "", "MEDIUM",
         "Ryan: 'Far from crew = charge more.' I have no access to crew schedule. "
         "Near-term: add county-based distance tier to pricing (Panhandle = always high, Central FL = baseline)."),
        ("Florida PSM standards wired in but not tested on real disputes",
         "My FL PSM persona was built from regulations — not from NexGen's actual field decisions",
         "Robert / Mark", "", "MEDIUM",
         "Feed me 5-10 real NexGen estimate examples that were disputed or revised. "
         "I will learn the gap between the FL standard and what NexGen actually does in the field."),
    ]
    for i, row_data in enumerate(gaps):
        gap, impact, who, _, pri, detail = row_data
        r = need_r + 2 + i
        bg = C["ltred"] if i % 2 == 0 else C["ltamber"]
        dc(ws, r, 1, gap, bg=bg, bold=True, size=10)
        dc(ws, r, 2, impact, bg=bg, size=9)
        dc(ws, r, 3, who, bg=bg, size=10, align="center")
        dc(ws, r, 4, "", bg=bg)
        pc = ws.cell(row=r, column=5, value=pri)
        pc.fill = fill(C["red"] if pri == "HIGH" else C["amber"])
        pc.font = font(True, "FFFFFF", 10); pc.alignment = center(); pc.border = thin()
        dc(ws, r, 6, detail, bg=bg, size=9)
        ws.row_dimensions[r].height = 54

    ws.freeze_panes = "A4"


def main():
    wb = Workbook()
    wb.remove(wb.active)

    build_cover(wb)
    build_yoy_comparison(wb)
    build_monthly(wb)
    build_price_dist(wb)
    build_county_pricing(wb)
    build_new_aliases(wb)
    build_ar_analysis(wb)
    build_commercial(wb)
    build_ai_actions(wb)
    build_analyst_questions(wb)
    build_ai_thoughts_today(wb)

    wb.save(OUTPUT)
    print(f"Saved -> {OUTPUT}")
    print(f"Tabs: {[ws.title for ws in wb.worksheets]}")


if __name__ == "__main__":
    main()
