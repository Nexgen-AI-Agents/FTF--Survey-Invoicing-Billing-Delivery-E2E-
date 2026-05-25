"""
Generate NexGen Surveying — Florida Competitive Analysis spreadsheet.
Output: TEAM/research/NexGen_Competitive_Analysis_2026.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT

# ── Palette ──────────────────────────────────────────────────────────────────
NAVY       = "1B3A5C"
TEAL       = "2196A6"
TEAL_LIGHT = "D6F0F4"
GREEN      = "217346"
GREEN_LIGHT = "E2EFDA"
ORANGE     = "E07B00"
ORANGE_LIGHT = "FCE4D6"
RED        = "C00000"
RED_LIGHT  = "FFE0E0"
YELLOW     = "FFCC00"
YELLOW_LIGHT = "FFF9D0"
GRAY       = "F5F5F5"
DGRAY      = "D9D9D9"
WHITE      = "FFFFFF"
BLACK      = "000000"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=WHITE, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def border_thin():
    s = Side(border_style="thin", color=DGRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(border_style="medium", color=NAVY)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def header_row(ws, row, values, bg=NAVY, fg=WHITE, height=22, bold=True, size=11):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(bold=bold, color=fg, size=size)
        c.alignment = align("center", "center", wrap=True)
        c.border = border_thin()
    ws.row_dimensions[row].height = height

def data_cell(ws, row, col, value, bg=WHITE, fg=BLACK, bold=False,
              h_align="left", wrap=True, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = font(bold=bold, color=fg, size=size)
    c.alignment = align(h_align, "center", wrap=wrap)
    c.border = border_thin()
    return c

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def merge_title(ws, row, max_col, text, bg=NAVY, fg=WHITE, height=30, size=14):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=max_col)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(bg)
    c.font = font(bold=True, color=fg, size=size)
    c.alignment = align("center", "center")
    ws.row_dimensions[row].height = height

def subtitle_row(ws, row, max_col, text, bg=TEAL, fg=WHITE, height=18, size=10):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=max_col)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(bg)
    c.font = font(bold=False, color=fg, size=size, italic=True)
    c.alignment = align("center", "center")
    ws.row_dimensions[row].height = height


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 1 — OVERVIEW
# ═════════════════════════════════════════════════════════════════════════════
def sheet_overview(wb):
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False

    merge_title(ws, 1, 4, "NexGen Surveying — Florida Competitive Analysis 2026")
    subtitle_row(ws, 2, 4,
        "Researched by: Competitor Analyst AI  |  Reviewed by: Ryan AI, Robert AI, Mark AI  |  Date: 2026-05-25")

    # NexGen profile
    ws.row_dimensions[3].height = 8
    header_row(ws, 4, ["Field", "Detail", "", ""], bg=NAVY, height=20)
    profile = [
        ("Company", "NexGen Surveying (nexgensurveying.com)"),
        ("Address", "1547 Prosperity Farms Road, Lake Park, FL 33403"),
        ("Phone", "(561) 508-6272"),
        ("Email", "info@nexgensurveying.com"),
        ("Hours", "Monday–Friday 9 AM–5 PM (closed weekends)"),
        ("Service Area", "Entire state of Florida (statewide claim — no county map published)"),
        ("Experience", "40+ years"),
        ("Published Services", "3 of 24 actual FTF services shown on website"),
        ("Mobile App", "Yes — Google Play + Apple App Store"),
        ("Online Ordering", "Yes — nexgensurveying.com/order-survey/"),
        ("Tagline", '"We not only meet, but beat, all of the expectations set by our industry."'),
    ]
    for i, (field, detail) in enumerate(profile):
        r = 5 + i
        bg = GRAY if i % 2 == 0 else WHITE
        data_cell(ws, r, 1, field, bg=bg, bold=True, fg=NAVY, size=10)
        ws.merge_cells(start_row=r, start_column=2,
                       end_row=r, end_column=4)
        data_cell(ws, r, 2, detail, bg=bg, size=10)
        ws.row_dimensions[r].height = 18

    # Key findings summary
    ws.row_dimensions[16].height = 8
    header_row(ws, 17, ["#", "Key Finding", "Category", "Priority"],
               bg=TEAL, height=20)
    findings = [
        (1, "Boundary Survey is the #1 searched FL survey type — NOT on NexGen website, despite being in 24-service list",
         "Service Gap", "P1"),
        (2, "GT Surveys (gtsurvey.com) operates 5 miles from NexGen HQ with same services + Saturday hours",
         "Competitor Threat", "P1"),
        (3, "Homepage lists 6 services; only 3 have pages — Geological, Hydrographic, Land Subdivisions are orphaned",
         "Website Fix", "P1"),
        (4, "LOMA/LOMR processing is the highest-ROI upsell to elevation certificate clients — not offered",
         "Revenue Gap", "P1"),
        (5, "No PSM license numbers or credentials published — Florida surveyor norm violation",
         "Trust Signal", "P1"),
        (6, "Apex Surveying (apexsurvey.us) covers all 67 FL counties with published pricing and LiDAR drones",
         "Statewide Threat", "P1"),
        (7, "NEVER_AUTO_QUOTE expanded from 3 to 7 services after stakeholder AI review",
         "System Config", "DONE"),
        (8, "Monroe County (Florida Keys) has no hard flag — auto-quoting Keys surveys will lose money",
         "Flag Trigger Gap", "I-034"),
        (9, "National firms (Terracon, Kimley-Horn, AECOM) added to competitor list",
         "Flag Trigger", "DONE"),
        (10, "8 Florida competitors identified with full service profiles, pricing benchmarks, differentiators",
         "Research", "DONE"),
    ]
    for i, (num, finding, cat, pri) in enumerate(findings):
        r = 18 + i
        bg = GRAY if i % 2 == 0 else WHITE
        pri_bg = (GREEN_LIGHT if pri == "DONE" else
                  RED_LIGHT if pri == "P1" else
                  ORANGE_LIGHT if pri == "P2" else
                  YELLOW_LIGHT if pri == "I-034" else WHITE)
        data_cell(ws, r, 1, num, bg=bg, h_align="center", size=10)
        data_cell(ws, r, 2, finding, bg=bg, size=10, wrap=True)
        data_cell(ws, r, 3, cat, bg=bg, size=10, h_align="center")
        data_cell(ws, r, 4, pri, bg=pri_bg,
                  bold=True, fg=(GREEN if pri == "DONE" else
                                 RED if pri == "P1" else BLACK),
                  h_align="center", size=10)
        ws.row_dimensions[r].height = 30

    set_col_widths(ws, [5, 70, 22, 10])
    ws.freeze_panes = "A5"


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 2 — COMPETITORS
# ═════════════════════════════════════════════════════════════════════════════
def sheet_competitors(wb):
    ws = wb.create_sheet("Competitors")
    ws.sheet_view.showGridLines = False

    merge_title(ws, 1, 8, "Florida Competitors — Full Profiles")
    subtitle_row(ws, 2, 8,
        "8 competitors researched via website crawl | Threat levels assigned by Ryan AI, Robert AI, Mark AI")

    headers = ["Company", "Domain", "HQ / Coverage", "Service Count",
               "Key Services", "Differentiator", "Threat Level", "Pricing"]
    header_row(ws, 3, headers, bg=NAVY)

    competitors = [
        ("GT Surveys", "gtsurvey.com",
         "West Palm Beach (~5 mi from NexGen HQ)\nPalm Beach + Broward counties", "7",
         "Mortgage, Boundary, Elevation Certs, Residential Construction, Topo, Sketch & Desc, ALTA",
         "Family-owned 25 yrs; Saturday hours; military/first responder discounts",
         "HIGH", "Not published"),
        ("Apex Surveying & Mapping", "apexsurvey.us",
         "All 67 Florida counties\nMiami, Tampa, Orlando, Jacksonville, Naples, Key West", "10",
         "Boundary, ALTA, Topo, Construction Staking, Elevation Certs, Drone/UAV, 3D Laser Scan, As-Built, Subdivision Platting, ROW",
         "Sub-1-hr quote response; published pricing; LiDAR drones; 500+ surveys listed",
         "HIGH", "Published: Boundary $400–$1,000; ALTA $2,500–$10,000+"),
        ("Land Surveying Palm Beach", "landsurveyingpalmbeach.com",
         "Palm Beach, Martin, St. Lucie, Indian River counties\n40+ municipalities", "9",
         "ALTA, Boundary, Flood Elev Certs, Topo, Construction Layout, Subdivision Plats, As-Built, Tree/Utility, Lot Split",
         "Named PSM (Rick Morales); FEMA fluent; 100% satisfaction guarantee; local SEO landing pages",
         "MEDIUM", "Not published; free phone quote"),
        ("Accurate Land Surveyors, Inc.", "accuratelandsurveyors.com",
         "Miami-Dade, Broward, Palm Beach,\nMartin, St. Lucie counties", "11",
         "Purchases/Refinances, Flood Insurance, New Construction, Docks/Seawall/MHW, Liquor License, Topo, Boundary Disputes, Platting",
         "Since 1983; 100,000+ surveys; 72-hr turnaround; free RE professional seminars (referral pipeline)",
         "MEDIUM", "Not published"),
        ("Suarez Surveying & Mapping", "suarezsurveying.com",
         "Miami, Fort Lauderdale, South Florida,\nWest Palm Beach + South FL cities", "18+",
         "Boundary, Topo, ALTA, Elevation Certs, Construction, Stormwater/Drainage Design, Roadway Design, ADA Compliance",
         "Hybrid surveying + civil engineering; 20+ yrs; 48-hr turnaround; FSMS + NSPS member",
         "MEDIUM", "Not published; 48-hr turnaround claim"),
        ("Stoner & Associates, Inc.", "stonersurveyors.com",
         "South Florida — HQ: Davie, Broward County", "15",
         "ALTA, As-Built, Aviation, Boundary, Construction, Drone, Engineering, Environmental, Litigation, Topo, Utility Mapping, Hydrographic",
         "3rd-generation firm since 1988; aviation surveys, litigation surveys, hydrographic specializations",
         "LOW-MEDIUM", "Not published"),
        ("SurvTech Solutions, Inc.", "survtechsolutions.com",
         "Tampa HQ + Sanford, Ft. Pierce, Jacksonville,\nFort Myers, Melbourne; 15 states", "21",
         "ALTA, Boundary, UAV Mapping, 3D Laser Scan, Mobile LiDAR, Hydrographic, SUE, GIS, LOMA/LOMR, Expert Witness, Oil & Gas",
         "FAA-approved UAV; TWIC/MICCS badged; government/commercial/infrastructure focus",
         "LOW", "Not published"),
        ("No Flood Florida", "nofloodflorida.com",
         "Florida statewide\n(Tampa Bay area focus)", "5",
         "Elevation Certificates, Flood Zone Correction, LOMA Processing, Flood Analysis Memos, Flood Insurance Placement",
         "Processes FEMA LOMA/LOMR on client's behalf; removes properties from flood zones; complimentary flood review; bilingual",
         "MEDIUM\n(elevation cert)", "Not published; free initial review"),
    ]

    threat_colors = {
        "HIGH": RED_LIGHT,
        "MEDIUM": ORANGE_LIGHT,
        "MEDIUM\n(elevation cert)": ORANGE_LIGHT,
        "LOW-MEDIUM": YELLOW_LIGHT,
        "LOW": GREEN_LIGHT,
    }

    for i, row_data in enumerate(competitors):
        r = 4 + i
        bg = GRAY if i % 2 == 0 else WHITE
        for col, val in enumerate(row_data, 1):
            if col == 7:  # threat level
                cell_bg = threat_colors.get(val, WHITE)
                data_cell(ws, r, col, val, bg=cell_bg, bold=True,
                          fg=(RED if "HIGH" in val else
                              ORANGE if "MEDIUM" in val else BLACK),
                          h_align="center", size=10)
            else:
                data_cell(ws, r, col, val, bg=bg, size=10, wrap=True)
        ws.row_dimensions[r].height = 55

    set_col_widths(ws, [28, 30, 28, 10, 50, 45, 14, 32])
    ws.freeze_panes = "A4"


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 3 — SERVICE COMPARISON
# ═════════════════════════════════════════════════════════════════════════════
def sheet_service_comparison(wb):
    ws = wb.create_sheet("Service Comparison")
    ws.sheet_view.showGridLines = False

    merge_title(ws, 1, 9, "Service Comparison Matrix — NexGen vs. Florida Competitors")
    subtitle_row(ws, 2, 9,
        "✅ = offered  ❌ = not offered / not published  ⚠️ = listed on homepage but no service page")

    headers = ["Service", "NexGen", "GT Surveys", "Apex", "Land Survey PB",
               "Accurate", "Suarez", "Stoner", "SurvTech"]
    header_row(ws, 3, headers, bg=NAVY)

    Y = "✅"
    N = "❌"
    W = "⚠️"
    services = [
        # (Service, NexGen, GT, Apex, LSPB, Accurate, Suarez, Stoner, SurvTech)
        ("Mortgage / Purchase Surveys",  Y, Y, N, N, Y, N, N, N),
        ("ALTA Surveys",                 Y, Y, Y, Y, N, Y, Y, Y),
        ("Elevation Certificates",       Y, Y, Y, Y, Y, Y, N, Y),
        ("Boundary Surveys",             N, Y, Y, Y, Y, Y, Y, Y),
        ("Topographic Surveys",          N, Y, Y, Y, Y, Y, Y, N),
        ("Construction Staking / Layout",N, Y, Y, Y, Y, Y, Y, Y),
        ("As-Built Surveys",             N, N, Y, Y, N, Y, Y, Y),
        ("Subdivision Platting",         N, N, Y, Y, Y, N, N, N),
        ("Lot Split Surveys",            N, N, N, Y, N, N, N, N),
        ("Drone / UAV Mapping",          N, N, Y, N, N, N, Y, Y),
        ("3D Laser Scanning",            N, N, Y, N, N, N, Y, Y),
        ("LOMA / LOMR Processing",       N, N, N, N, N, N, Y, Y),
        ("Dock / Seawall / MHW",         N, N, N, N, Y, N, N, N),
        ("Litigation / Expert Witness",  N, N, N, N, N, N, Y, Y),
        ("Right-of-Way Surveys",         N, N, Y, N, N, N, Y, N),
        ("Hydrographic Surveys",         W, N, N, N, N, N, Y, Y),
        ("Geological Surveying",         W, N, N, N, N, N, N, N),
        ("Land Subdivisions",            W, N, N, Y, Y, N, N, N),
        ("Tree & Utility Surveys",       N, N, N, Y, N, N, N, N),
        ("Liquor License Surveys",       N, N, N, N, Y, N, N, N),
        ("Wetland Delineation",          Y, N, N, N, N, N, Y, Y),
        ("Published Pricing",            "ALTA only", Y, Y, N, N, N, N, N),
        ("Saturday Hours",               N, Y, N, N, N, N, N, N),
        ("County/City Landing Pages",    N, N, Y, Y, N, N, N, N),
        ("Mobile App",                   Y, N, N, N, N, N, N, N),
    ]

    for i, row_data in enumerate(services):
        r = 4 + i
        bg = GRAY if i % 2 == 0 else WHITE
        svc = row_data[0]
        gap_row = (row_data[1] == N or row_data[1] == W) and any(
            v == Y for v in row_data[2:])
        for col, val in enumerate(row_data, 1):
            if col == 1:
                cell_fg = RED if gap_row else BLACK
                data_cell(ws, r, col, val, bg=bg, bold=gap_row,
                          fg=cell_fg, size=10)
            else:
                cell_bg = (GREEN_LIGHT if val == Y else
                           RED_LIGHT if val == N and gap_row else
                           ORANGE_LIGHT if val == W else bg)
                cell_fg = (GREEN if val == Y else
                           RED if val == N and gap_row else
                           ORANGE if val == W else BLACK)
                data_cell(ws, r, col, val, bg=cell_bg,
                          bold=(val in (Y, N, W)),
                          fg=cell_fg, h_align="center", size=11)
        ws.row_dimensions[r].height = 20

    # Legend
    r = 4 + len(services) + 1
    ws.row_dimensions[r].height = 8
    r += 1
    header_row(ws, r, ["Legend", "", "", "", "", "", "", "", ""], bg=DGRAY,
               fg=BLACK, height=16)
    legends = [
        ("✅ Green", "Service offered by this company"),
        ("❌ Red (NexGen row)", "NexGen gap — offered by competitors but not by NexGen"),
        ("⚠️ Orange", "Listed on NexGen homepage but no service page exists"),
        ("Bold red service name", "Gap row — NexGen does not offer but competitors do"),
    ]
    for j, (sym, desc) in enumerate(legends):
        ws.merge_cells(start_row=r+1+j, start_column=2,
                       end_row=r+1+j, end_column=9)
        data_cell(ws, r+1+j, 1, sym, bg=GRAY, bold=True, size=9)
        data_cell(ws, r+1+j, 2, desc, bg=GRAY, size=9)
        ws.row_dimensions[r+1+j].height = 16

    set_col_widths(ws, [35, 14, 12, 10, 14, 12, 12, 12, 12])
    ws.freeze_panes = "B4"


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 4 — IMPROVEMENT PRIORITIES
# ═════════════════════════════════════════════════════════════════════════════
def sheet_priorities(wb):
    ws = wb.create_sheet("Improvement Priorities")
    ws.sheet_view.showGridLines = False

    merge_title(ws, 1, 5, "Improvement Priorities — NexGen Surveying")
    subtitle_row(ws, 2, 5,
        "P1 = Do now (competitive necessity)  |  P2 = Next quarter  |  P3 = Future consideration")

    headers = ["#", "Priority", "Area", "Issue", "Recommendation"]
    header_row(ws, 3, headers, bg=NAVY)

    priorities = [
        # P1
        ("1", "P1", "Service Portfolio",
         "Boundary Survey — #1 searched FL survey type — not on website despite being in 24-service list",
         "Add Boundary Surveys as named service page immediately. Already in FTF system at $350. Pure marketing fix."),
        ("2", "P1", "Revenue — LOMA/LOMR",
         "Elevation cert clients need LOMA processing next. No Flood Florida built a business on this step NexGen skips.",
         "Add LOMA/LOMR submission as upsell to every elevation certificate order. High-value, minimal overhead."),
        ("3", "P1", "Pricing Transparency",
         "Only ALTA price published. Competitors showing numbers get called first by price-shoppers.",
         "Publish 'starting from $X' for Mortgage Surveys + Elevation Certificates. Keep others quote-based (per Ryan AI)."),
        ("4", "P1", "Service Area Clarity",
         "Claims statewide coverage but zero counties/cities published. Cannot rank in local searches.",
         "Publish explicit county list or coverage map. Add county + city names to service pages for local SEO."),
        ("5", "P1", "Homepage Content Fix",
         "Homepage lists Geological Surveying, Hydrographic Survey, Land Subdivisions — none have service pages.",
         "Either publish proper service pages OR remove from homepage. Every listed service must have a page."),
        ("6", "P1", "SEO — Local Landing Pages",
         "No county-level or service+location landing pages. Competitors rank for 'West Palm Beach boundary survey'.",
         "Create landing pages per county served x per service type to capture long-tail local searches."),
        ("7", "P1", "License Transparency",
         "No PSM license numbers, DBPR IDs, or professional association memberships published.",
         "Publish PSM license numbers, DBPR registration, FSMS/NSPS membership on About page and footer."),
        ("8", "P1", "GT Surveys Threat",
         "GT Surveys — 5 miles from HQ, same services, opens Saturday. Direct local competition.",
         "Differentiate on speed, technology, or service breadth. Adding Saturday hours directly neutralizes GT's main advantage."),
        # P2
        ("9", "P2", "Saturday Availability",
         "GT Surveys opens Saturday. Real estate closings frequently require urgent weekend quotes.",
         "Add Saturday morning hours OR a weekend quote request with guaranteed Monday AM response."),
        ("10", "P2", "Quote Response Speed",
         "Apex promotes sub-1-hour quotes. NexGen makes no speed claim on the quote process.",
         "Publish quote-response SLA (e.g., 'Quotes within 2 hours during business hours') and honor it."),
        ("11", "P2", "Technology Specifics",
         "Claims 'cutting-edge technology' without naming equipment. Every competitor says the same.",
         "Name actual equipment used. If NexGen uses GPS rovers, robotic total stations — say so explicitly."),
        ("12", "P2", "Google Reviews",
         "No Google review count or rating featured. First Choice Surveying leads with '4.9/5 (50+ reviews)'.",
         "Embed live Google review widget or display star rating + count on homepage. Top trust signal for local services."),
        ("13", "P2", "Realtor Referral Network",
         "Accurate Land Surveyors runs free RE professional seminars — a referral pipeline. NexGen has none.",
         "Create Realtor Resource page with free 'How to Read a Mortgage Survey' PDF. Email realtors in Palm Beach + Broward."),
        ("14", "P2", "Mobile App Promotion",
         "Mobile app exists (Play + App Store) but barely visible on site. Genuine differentiator vs. most competitors.",
         "Feature prominently: 'Track your survey order in real-time.' Build a landing page. Most competitors have no app."),
        # P3
        ("15", "P3", "As-Built + Construction Staking",
         "6/8 competitors offer these repeat-business services tied to builders and contractors.",
         "Add As-Built Surveys + Construction Staking. Required for CO in many municipalities. High repeat-business volume."),
        ("16", "P3", "Web Consolidation",
         "NexGen operates 3 web properties (nexgensurveying.com, nexgensurveyingllc.com, Square site) — fragmented SEO.",
         "Consolidate to one domain with 301 redirects from the secondary properties."),
    ]

    p_colors = {"P1": (RED_LIGHT, RED), "P2": (ORANGE_LIGHT, ORANGE), "P3": (YELLOW_LIGHT, BLACK)}

    for i, (num, pri, area, issue, rec) in enumerate(priorities):
        r = 4 + i
        bg = GRAY if i % 2 == 0 else WHITE
        p_bg, p_fg = p_colors.get(pri, (WHITE, BLACK))
        data_cell(ws, r, 1, num, bg=p_bg, bold=True, fg=p_fg,
                  h_align="center", size=10)
        data_cell(ws, r, 2, pri, bg=p_bg, bold=True, fg=p_fg,
                  h_align="center", size=11)
        data_cell(ws, r, 3, area, bg=bg, bold=True, fg=NAVY, size=10)
        data_cell(ws, r, 4, issue, bg=bg, size=10, wrap=True)
        data_cell(ws, r, 5, rec, bg=bg, size=10, wrap=True)
        ws.row_dimensions[r].height = 48

    set_col_widths(ws, [4, 9, 24, 52, 55])
    ws.freeze_panes = "A4"


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 5 — FLAG TRIGGERS
# ═════════════════════════════════════════════════════════════════════════════
def sheet_flag_triggers(wb):
    ws = wb.create_sheet("Flag Triggers Config")
    ws.sheet_view.showGridLines = False

    merge_title(ws, 1, 4, "Agent 4 Flag Triggers — Competitor Detection Configuration")
    subtitle_row(ws, 2, 4,
        "Bootstrapped 2026-05-25 | Reviewed by Ryan AI, Robert AI, Mark AI | Robert/Mark validation required before Sprint 3")

    # Section: NEVER_AUTO_QUOTE
    header_row(ws, 3, ["Service Type", "Reason", "Added By", "Validation"],
               bg=TEAL, height=20)
    naq = [
        ("Specific Purpose Survey",
         "Scope undefined until client describes purpose — could be $600 to $6,000+",
         "Competitor Analyst AI", "Confirmed by all 3 agents"),
        ("Lot Split",
         "County/municipality review required; public hearing coordination; unpredictable timeline",
         "Competitor Analyst AI", "Confirmed by all 3 agents"),
        ("Wetland Delineation",
         "FDEP/Army Corps/SFWMD jurisdiction; outcome not guaranteed",
         "Competitor Analyst AI", "Confirmed by all 3 agents"),
        ("B-II Title Review",
         "Scope varies with title commitment exceptions; not a standard field survey",
         "Ryan AI + Robert AI", "Confirmed by Ryan AI + Robert AI"),
        ("Acreage",
         "$250 base rate does not hold above ~2 acres; large acreage scope swings hard",
         "Mark AI", "Confirmed by Mark AI — real Mark to validate"),
        ("Legal Description",
         "Rural/metes-and-bounds parcels are not flat-rate safe; liability exposure on errors",
         "Robert AI + Mark AI", "Confirmed by Robert + Mark AI — real Robert to validate"),
        ("Topography Survey",
         "$225 listed price is below FL market rate ($500–$2,000); commercial scope not flat-rate safe",
         "Mark AI", "Confirmed by Mark AI — real Mark to validate"),
    ]
    for i, (svc, reason, addedby, validation) in enumerate(naq):
        r = 4 + i
        bg = GRAY if i % 2 == 0 else WHITE
        data_cell(ws, r, 1, svc, bg=bg, bold=True, fg=NAVY, size=10)
        data_cell(ws, r, 2, reason, bg=bg, size=10, wrap=True)
        data_cell(ws, r, 3, addedby, bg=bg, size=9, h_align="center")
        v_bg = (GREEN_LIGHT if "all 3" in validation else
                YELLOW_LIGHT if "real" in validation else WHITE)
        data_cell(ws, r, 4, validation, bg=v_bg, size=9, h_align="center")
        ws.row_dimensions[r].height = 32

    ws.row_dimensions[4 + len(naq)].height = 8

    # Section: Competitor Names
    r_start = 4 + len(naq) + 2
    header_row(ws, r_start - 1,
               ["Competitor Name", "Domain", "Category", "Validation Status"],
               bg=NAVY, height=20)
    competitors_list = [
        ("Apex Surveying & Mapping", "apexsurvey.us", "FL Statewide", "Web-confirmed"),
        ("Apex Surveying", "apexsurvey.us", "FL Statewide", "Alias — confirmed"),
        ("GT Surveys", "gtsurvey.com", "Palm Beach + Broward", "Web-confirmed"),
        ("GT Surveyors", "gtsurvey.com", "Palm Beach + Broward", "Alias — confirmed"),
        ("Land Surveying Palm Beach", "landsurveyingpalmbeach.com", "4-county South FL", "Web-confirmed"),
        ("Accurate Land Surveyors", "accuratelandsurveyors.com", "5-county South FL", "Web-confirmed"),
        ("Accurate Land Surveyors Inc", "accuratelandsurveyors.com", "5-county South FL", "Alias — confirmed"),
        ("Suarez Surveying & Mapping", "suarezsurveying.com", "South FL", "Web-confirmed"),
        ("Stoner & Associates", "stonersurveyors.com", "Broward County", "Web-confirmed"),
        ("SurvTech Solutions", "survtechsolutions.com", "Tampa + 5 FL offices", "Web-confirmed"),
        ("No Flood Florida", "nofloodflorida.com", "FL Statewide", "Web-confirmed"),
        ("National Flood Experts", "nofloodflorida.com", "FL Statewide", "Alias — confirmed"),
        ("First Choice Surveying", "firstchoicesurveying.com", "FL", "Web-confirmed"),
        ("John Ibarra & Associates", "ibarralandsurveyors.com", "FL", "Web-confirmed"),
        ("GeoPoint Surveying", "geopointsurvey.com", "FL", "Web-confirmed"),
        ("Target Surveying", "targetsurveying.com", "FL", "Web-confirmed"),
        ("Sliger & Associates", "—", "FL", "Web-confirmed"),
        ("Fordco Surveying", "—", "FL", "Web-confirmed — domain unknown"),
        ("Atlantic Coast Surveying", "acsiweb.net", "FL", "⚠️ NEEDS REAL ROBERT — generic name risk"),
        ("Florida Land Surveying", "floridalandsurveying.com", "FL", "⚠️ NEEDS REAL ROBERT — generic name risk"),
        ("Florida Builders Engineers & Inspectors", "landsurveyingpalmbeach.com", "4-county South FL", "DBA of Land Surveying PB"),
        ("Terracon", "terracon.com", "National (FL active)", "Mark AI — national firm FL survey division"),
        ("Terracon Consultants", "terracon.com", "National (FL active)", "Alias — confirmed"),
        ("Kimley-Horn", "kimley-horn.com", "National (FL active)", "Mark AI — significant FL presence"),
        ("Kimley-Horn and Associates", "kimley-horn.com", "National (FL active)", "Alias — confirmed"),
        ("AECOM", "aecom.com", "National (FL active)", "Mark AI — FL infrastructure projects"),
    ]

    valid_colors = {
        "Web-confirmed": GREEN_LIGHT,
        "Alias — confirmed": GREEN_LIGHT,
        "Alias — confirmed": GREEN_LIGHT,
        "Web-confirmed — domain unknown": YELLOW_LIGHT,
        "DBA of Land Surveying PB": TEAL_LIGHT,
    }

    for i, (name, domain, coverage, status) in enumerate(competitors_list):
        r = r_start + i
        bg = GRAY if i % 2 == 0 else WHITE
        warning = "NEEDS REAL ROBERT" in status
        v_bg = RED_LIGHT if warning else valid_colors.get(status, YELLOW_LIGHT)
        data_cell(ws, r, 1, name, bg=bg, bold=warning, fg=(RED if warning else BLACK), size=10)
        data_cell(ws, r, 2, domain, bg=bg, size=10, h_align="center")
        data_cell(ws, r, 3, coverage, bg=bg, size=10, h_align="center")
        data_cell(ws, r, 4, status, bg=v_bg, size=9,
                  bold=warning, fg=(RED if warning else BLACK))
        ws.row_dimensions[r].height = 18

    set_col_widths(ws, [36, 32, 26, 40])
    ws.freeze_panes = "A4"


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 6 — PRICING BENCHMARKS
# ═════════════════════════════════════════════════════════════════════════════
def sheet_pricing(wb):
    ws = wb.create_sheet("Pricing Benchmarks")
    ws.sheet_view.showGridLines = False

    merge_title(ws, 1, 5, "Florida Survey Pricing Benchmarks vs. NexGen FTF Rates")
    subtitle_row(ws, 2, 5,
        "NexGen prices from memory.md FTF service list | Market rates from FL competitor web research 2026-05-25")

    headers = ["Service", "NexGen FTF Rate", "FL Market Rate", "Delta", "Notes"]
    header_row(ws, 3, headers, bg=NAVY)

    pricing = [
        ("Mortgage / Purchase Survey", "$350–600 (varies)", "$200–$500 residential",
         "At or above market", "NexGen's core volume service. Price range competitive."),
        ("Elevation Certificate", "$225", "$175–$900 residential; up to $2,000 complex",
         "⚠️ LOW end of market", "NexGen price is low for a commodity service. Could increase without losing volume."),
        ("ALTA Survey", "$2,000–$3,000 (published)", "$2,500–$10,000+",
         "⚠️ Priced below market floor", "NexGen publishes $2,000–$3,000 but Apex quotes $2,500+ as floor. May be underpricing."),
        ("Boundary Survey", "$350 (FTF rate)", "$400–$1,000 residential",
         "Below market", "NexGen FTF rate is $350; FL market starts at $400. Not on NexGen website."),
        ("Topographic Survey", "$225 (FTF rate)", "$500–$2,000 residential/small lot",
         "⚠️ SIGNIFICANTLY BELOW market", "FTF rate is likely scoped for small residential lots only. Commercial/large site topo cannot be flat-rated at $225. Added to NEVER_AUTO_QUOTE."),
        ("Construction Staking", "$225 (Building Stake Out)", "$300–$1,500 per phase",
         "Below market — varies by scope", "Multiple FTF services cover construction phases. Each phase priced separately."),
        ("Lot Split", "$450 (FTF rate)", "Varies widely — county dependent",
         "Cannot compare — scope-dependent", "Added to NEVER_AUTO_QUOTE. Not flat-rate safe."),
        ("Specific Purpose Survey", "$600 (FTF rate)", "$600–$5,000+ depending on purpose",
         "Floor rate only", "Added to NEVER_AUTO_QUOTE. Scope unknown until client describes."),
        ("Wetland Delineation", "$300 (FTF rate)", "Regulatory — not market-priced",
         "Not comparable", "FDEP/Army Corps scope makes this non-quotable at flat rate. In NEVER_AUTO_QUOTE."),
        ("LOMA / LOMR Processing", "Not offered", "$500–$2,000 professional fee",
         "Revenue gap", "No Flood Florida has built a standalone business on LOMA processing. Natural upsell after elevation cert."),
        ("Legal Description", "$300 (FTF rate)", "Varies — simple to complex",
         "Simple OK; complex underpriced", "Metes-and-bounds descriptions in rural FL counties are not $300 jobs. Added to NEVER_AUTO_QUOTE."),
        ("Acreage Survey", "$250 (FTF rate)", "$400–$2,000+ depending on acreage",
         "⚠️ Underpriced for large lots", "FL market rate scales with acreage. $250 flat rate only holds for small residential. Added to NEVER_AUTO_QUOTE."),
    ]

    for i, (svc, nexgen, market, delta, notes) in enumerate(pricing):
        r = 4 + i
        bg = GRAY if i % 2 == 0 else WHITE
        risk = "BELOW" in delta or "gap" in delta.lower() or "under" in delta.lower()
        warn = "⚠️" in delta
        delta_bg = RED_LIGHT if risk or warn else GREEN_LIGHT if "competitive" in delta.lower() else YELLOW_LIGHT
        data_cell(ws, r, 1, svc, bg=bg, bold=True, fg=NAVY, size=10)
        data_cell(ws, r, 2, nexgen, bg=bg, size=10, h_align="center")
        data_cell(ws, r, 3, market, bg=bg, size=10, h_align="center")
        data_cell(ws, r, 4, delta, bg=delta_bg, bold=warn or risk,
                  fg=(RED if risk else BLACK), size=10, h_align="center")
        data_cell(ws, r, 5, notes, bg=bg, size=10, wrap=True)
        ws.row_dimensions[r].height = 36

    set_col_widths(ws, [30, 22, 30, 28, 52])
    ws.freeze_panes = "A4"


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 7 — STAKEHOLDER REVIEW
# ═════════════════════════════════════════════════════════════════════════════
def sheet_stakeholder_review(wb):
    ws = wb.create_sheet("Stakeholder AI Review")
    ws.sheet_view.showGridLines = False

    merge_title(ws, 1, 4, "Stakeholder AI Review — Ryan, Robert, Mark")
    subtitle_row(ws, 2, 4,
        "All three Stakeholder AI agents reviewed the competitive analysis on 2026-05-25 | Items marked [Real Human] require confirmation before Sprint 3")

    # Action items table
    header_row(ws, 3, ["Item", "Source", "Status", "Action Required"],
               bg=NAVY, height=20)
    actions = [
        ("Removed studioaeng.com + cwi-assoc.com from competitor domains",
         "Ryan AI + Robert AI", "DONE", "—"),
        ("Added B-II Title Review to NEVER_AUTO_QUOTE",
         "Ryan AI + Robert AI", "DONE", "—"),
        ("Added Acreage to NEVER_AUTO_QUOTE",
         "Mark AI", "DONE", "Real Mark to validate threshold"),
        ("Added Legal Description to NEVER_AUTO_QUOTE",
         "Robert AI + Mark AI", "DONE", "Real Robert to validate"),
        ("Added Topography Survey to NEVER_AUTO_QUOTE",
         "Mark AI", "DONE", "Real Mark to validate — $225 vs market $500+"),
        ("Added Terracon, Kimley-Horn, AECOM to COMPETITOR_NAMES",
         "Mark AI", "DONE", "Real Robert/Mark to confirm"),
        ("Monroe County (Keys) hard flag missing — I-034 logged",
         "Mark AI", "ISSUE LOGGED", "Real Mark: does NexGen service the Keys?"),
        ("VE flood zone not flagged as coastal trigger — I-035 logged",
         "Mark AI", "ISSUE LOGGED", "Sprint 2 dev: parse flood_zone for 'V' prefix"),
        ("Missing property_county guard — I-036 logged",
         "Mark AI", "ISSUE LOGGED", "Sprint 2 dev: flag null/empty county"),
        ("False FL address coordinate cross-check — I-037 logged",
         "Mark AI", "ISSUE LOGGED", "Sprint 2 dev: check lat > 31.0 if state = FL"),
        ('"Florida Land Surveying" + "Atlantic Coast Surveying" false positive risk — I-038',
         "Ryan AI + Robert AI", "NEEDS REAL ROBERT/MARK", "Confirm or remove — too generic, may block legitimate clients"),
        ("acsiweb.net — identity unconfirmed",
         "Robert AI", "NEEDS REAL ROBERT", "Confirm: is this Atlantic Coast Surveying Inc?"),
        ("Surveyor's Affidavit as NEVER_AUTO_QUOTE candidate",
         "Ryan AI", "NEEDS REAL RYAN", "Liability call — only real Ryan can decide"),
        ("LOMA/LOMR as Phase 1 scope item",
         "Ryan AI", "NEEDS REAL RYAN", "Is LOMA/LOMR upsell in scope for Phase 1 build?"),
        ("Small local Palm Beach/Broward competitors with no web presence",
         "Mark AI", "NEEDS REAL MARK", "Web research cannot find these — only institutional memory"),
        ("Competitor alert format: must include matched field + value + competitor name",
         "Ryan AI + Robert AI", "SPRINT 3 REQUIREMENT", "Dev Manager: add to Sprint 3 human gate notification payload"),
        ("B2B title company orders — soft flag for Robert review (not hard competitor block)",
         "Mark AI", "SPRINT 3 REQUIREMENT", "New flag logic: B2B + known title company name = soft flag"),
    ]

    status_colors = {
        "DONE": GREEN_LIGHT,
        "ISSUE LOGGED": TEAL_LIGHT,
        "NEEDS REAL ROBERT/MARK": RED_LIGHT,
        "NEEDS REAL ROBERT": RED_LIGHT,
        "NEEDS REAL RYAN": RED_LIGHT,
        "NEEDS REAL MARK": RED_LIGHT,
        "SPRINT 3 REQUIREMENT": YELLOW_LIGHT,
    }
    for i, (item, source, status, action) in enumerate(actions):
        r = 4 + i
        bg = GRAY if i % 2 == 0 else WHITE
        s_bg = status_colors.get(status, WHITE)
        needs_human = "NEEDS REAL" in status
        data_cell(ws, r, 1, item, bg=bg, size=10, wrap=True,
                  bold=needs_human, fg=(RED if needs_human else BLACK))
        data_cell(ws, r, 2, source, bg=bg, size=9, h_align="center")
        data_cell(ws, r, 3, status, bg=s_bg, bold=True,
                  fg=(GREEN if status == "DONE" else
                      RED if needs_human else BLACK),
                  h_align="center", size=9)
        data_cell(ws, r, 4, action, bg=bg, size=10, wrap=True)
        ws.row_dimensions[r].height = 32

    set_col_widths(ws, [62, 24, 26, 48])
    ws.freeze_panes = "A4"


# ═════════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()

    sheet_overview(wb)
    sheet_competitors(wb)
    sheet_service_comparison(wb)
    sheet_priorities(wb)
    sheet_flag_triggers(wb)
    sheet_pricing(wb)
    sheet_stakeholder_review(wb)

    out_path = (r"c:\Users\Prateek Chandra\OneDrive - NexGen Enterprises"
                r"\Claude\Agentic AI\FTF- Survey Invoicing & Billing Delivery (E2E)"
                r"\TEAM\research\NexGen_Competitive_Analysis_2026.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
