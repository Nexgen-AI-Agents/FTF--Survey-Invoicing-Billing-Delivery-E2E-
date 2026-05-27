#!/usr/bin/env python3
"""
Generate Pipeline_Report.xlsx — NexGen FTF Agentic AI project tracker.
Tabs: Dashboard | Issues | Sprints | Data Insights | QA Orders
Run: python scripts/generate_pipeline_report.py
"""

import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

WORKSPACE = Path(__file__).parent.parent
OUTPUT = WORKSPACE / "Pipeline_Report.xlsx"
ISSUES_MD = WORKSPACE / "issues" / "issue.md"

# ── Colour palette ─────────────────────────────────────────────────────────────
C = {
    "navy":    "1F3864",
    "blue":    "2E75B6",
    "teal":    "00B0A0",
    "green":   "70AD47",
    "amber":   "FFC000",
    "red":     "FF4B4B",
    "grey":    "D9D9D9",
    "white":   "FFFFFF",
    "ltblue":  "DEEAF1",
    "ltgreen": "E2EFDA",
    "ltamber": "FFF2CC",
    "ltred":   "FFE0E0",
    "ltgrey":  "F2F2F2",
}


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")


def border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def header_row(ws, row, cols, bg="navy", fg="FFFFFF", bold=True, size=11):
    for col_idx, text in enumerate(cols, 1):
        c = ws.cell(row=row, column=col_idx, value=text)
        c.fill = fill(bg)
        c.font = font(bold=bold, color=fg, size=size)
        c.alignment = center()
        c.border = border_thin()


def data_cell(ws, row, col, value, bg="FFFFFF", bold=False, color="000000",
              align="left", size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = font(bold=bold, color=color, size=size)
    c.alignment = left() if align == "left" else center()
    c.border = border_thin()
    return c


# ── Parse issues/issue.md ──────────────────────────────────────────────────────

def parse_issues():
    text = ISSUES_MD.read_text(encoding="utf-8")
    open_issues, closed_issues = [], []

    # Open issues table rows
    open_section = re.search(r"## Open Issues.*?## Closed Issues", text, re.DOTALL)
    if open_section:
        for m in re.finditer(
            r"\|\s*(I-\d+)\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|\s*(\w+)\s*\|\s*(\S+)\s*\|",
            open_section.group()
        ):
            open_issues.append({
                "id": m.group(1), "sprint": m.group(2), "severity": m.group(3),
                "title": m.group(4), "assigned": m.group(5),
                "status": m.group(6), "opened": m.group(7),
            })

    # Closed issues table rows
    closed_section = re.search(r"## Closed Issues(.*?)## Issue ID Format", text, re.DOTALL)
    if closed_section:
        for m in re.finditer(
            r"\|\s*(I-\d+)\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|\s*(\S+)\s*\|",
            closed_section.group()
        ):
            closed_issues.append({
                "id": m.group(1), "sprint": m.group(2), "severity": m.group(3),
                "title": m.group(4), "resolved_by": m.group(5), "closed": m.group(6),
            })

    return open_issues, closed_issues


# ── Sheet builders ─────────────────────────────────────────────────────────────

def build_dashboard(wb):
    ws = wb.create_sheet("📊 Dashboard", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22

    # Title
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "NexGen FTF — Agentic AI OS  |  Pipeline Report"
    t.fill = fill(C["navy"])
    t.font = Font(bold=True, color=C["white"], size=16, name="Calibri")
    t.alignment = center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:E2")
    d = ws["A2"]
    d.value = f"Generated: {datetime.now().strftime('%B %d, %Y')}  |  Prod DB: nexgen_ftf_db  |  1-yr analysis: May 2025 – May 2026"
    d.fill = fill(C["blue"])
    d.font = Font(bold=False, color=C["white"], size=11, name="Calibri")
    d.alignment = center()
    ws.row_dimensions[2].height = 22

    # ── KPI row ───────────────────────────────────────────────────────────────
    row = 4
    ws.row_dimensions[row].height = 14
    kpi_label_row = 5
    kpi_val_row = 6
    ws.row_dimensions[kpi_label_row].height = 22
    ws.row_dimensions[kpi_val_row].height = 30

    kpis = [
        ("Total Invoiced (1yr)", "$13.2M", C["teal"]),
        ("Orders Billed", "18,497", C["blue"]),
        ("AR Outstanding", "$6.56M", C["red"]),
        ("Avg Invoice", "$565", C["green"]),
        ("Closed Issues", "52", C["navy"]),
    ]
    for col_idx, (label, val, color) in enumerate(kpis, 1):
        lc = ws.cell(row=kpi_label_row, column=col_idx, value=label)
        lc.fill = fill(color)
        lc.font = font(bold=True, color=C["white"], size=10)
        lc.alignment = center()
        lc.border = border_thin()

        vc = ws.cell(row=kpi_val_row, column=col_idx, value=val)
        vc.fill = fill(C["ltblue"])
        vc.font = font(bold=True, size=16, color=color)
        vc.alignment = center()
        vc.border = border_thin()

    # ── Sprint status ─────────────────────────────────────────────────────────
    r = 8
    ws.merge_cells(f"A{r}:E{r}")
    ws[f"A{r}"].value = "SPRINT STATUS"
    ws[f"A{r}"].fill = fill(C["navy"])
    ws[f"A{r}"].font = font(bold=True, color=C["white"], size=12)
    ws[f"A{r}"].alignment = center()
    ws.row_dimensions[r].height = 22

    header_row(ws, r + 1, ["Sprint", "Focus", "Status", "Key Deliverable", "Notes"], C["blue"])
    ws.row_dimensions[r + 1].height = 18

    sprints = [
        ("Sprint 0", "Foundation", "✅ RELEASED", "DB schema, config, logger", ""),
        ("Sprint 1", "Monitor", "✅ RELEASED", "Agent 02 — FTF Quote monitor", ""),
        ("Sprint 2", "Classifier", "✅ RELEASED", "Agent 03 — 50+ alias map, LLM fallback", "Prod data expanded aliases 2026-05-27"),
        ("Sprint 3", "Human Gate", "✅ RELEASED", "Agent 04 — batch approval digest", ""),
        ("Sprint 4", "Pricing", "✅ RELEASED", "Agent 05 — county pricing matrix", "Updated with prod actuals"),
        ("Sprint 5", "Writer + Email", "✅ RELEASED", "Agent 06 + 12 — estimate writer, email monitor", ""),
        ("Sprint 6", "Reviewer + AR", "✅ RELEASED", "Agent 07 + 10 — reviewer, AR scanner", ""),
        ("Sprint 7", "AR Escalation", "✅ RELEASED", "Agent 11 — 30/60/90 day AR escalation", ""),
        ("Sprint 8", "Billing Report", "✅ RELEASED", "Agent 09 — monthly billing statement", ""),
        ("Sprint 9", "Orchestrator + Memory", "✅ RELEASED", "Agent 01 + 13 — orchestrator, pricing trainer", ""),
        ("Sprint 10", "Staging QA", "🔄 IN PROGRESS", "E2E test runner, QA scenarios (10 orders)", "Prateek Test 1 & 2 seeded"),
        ("Sprint 11", "Prod Hardening", "⏳ PLANNED", "Rate limiting, prompt caching, prod go-live", "I-019 (caching), I-017 (rate limit)"),
    ]
    for i, s in enumerate(sprints):
        bg = C["ltgreen"] if "RELEASED" in s[2] else (C["ltblue"] if "PROGRESS" in s[2] else C["ltgrey"])
        for col_idx, val in enumerate(s, 1):
            data_cell(ws, r + 2 + i, col_idx, val, bg=bg, size=10)
        ws.row_dimensions[r + 2 + i].height = 18

    # ── Open issues summary ───────────────────────────────────────────────────
    r2 = r + 2 + len(sprints) + 2
    ws.merge_cells(f"A{r2}:E{r2}")
    ws[f"A{r2}"].value = "OPEN ISSUES SUMMARY"
    ws[f"A{r2}"].fill = fill(C["navy"])
    ws[f"A{r2}"].font = font(bold=True, color=C["white"], size=12)
    ws[f"A{r2}"].alignment = center()
    ws.row_dimensions[r2].height = 22

    header_row(ws, r2 + 1, ["ID", "Sprint", "Severity", "Title", "Status"], C["blue"])

    open_issues, _ = parse_issues()
    sev_color = {"BLOCKER": C["red"], "CRITICAL": C["red"], "MAJOR": C["amber"], "MINOR": C["ltgrey"]}
    for i, issue in enumerate(open_issues):
        bg = C["ltamber"] if issue["severity"] in ("BLOCKER", "CRITICAL", "MAJOR") else C["ltgrey"]
        data_cell(ws, r2 + 2 + i, 1, issue["id"], bg=bg, bold=True, size=10)
        data_cell(ws, r2 + 2 + i, 2, issue["sprint"], bg=bg, size=10)
        sc = sev_color.get(issue["severity"], C["ltgrey"])
        c = ws.cell(row=r2 + 2 + i, column=3, value=issue["severity"])
        c.fill = fill(sc)
        c.font = font(bold=True, size=10, color=C["white"] if issue["severity"] in ("BLOCKER","CRITICAL") else "000000")
        c.alignment = center()
        c.border = border_thin()
        data_cell(ws, r2 + 2 + i, 4, issue["title"], bg=bg, size=10)
        data_cell(ws, r2 + 2 + i, 5, issue["status"], bg=bg, size=10)
        ws.row_dimensions[r2 + 2 + i].height = 18

    ws.freeze_panes = "A3"


def build_issues(wb):
    ws = wb.create_sheet("🐛 Issues")
    ws.sheet_view.showGridLines = False
    widths = [8, 14, 10, 55, 18, 14, 12, 65]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:H1")
    ws["A1"].value = "Issue Tracker — FTF Agentic AI OS"
    ws["A1"].fill = fill(C["navy"])
    ws["A1"].font = font(bold=True, color=C["white"], size=14)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 30

    open_issues, closed_issues = parse_issues()

    # Open
    ws.merge_cells("A2:H2")
    ws["A2"].value = f"OPEN ISSUES ({len(open_issues)})"
    ws["A2"].fill = fill(C["red"])
    ws["A2"].font = font(bold=True, color=C["white"], size=11)
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 20

    header_row(ws, 3, ["ID", "Sprint", "Severity", "Title", "Assigned", "Status", "Opened", "Notes"], C["blue"])
    ws.row_dimensions[3].height = 18

    sev_color = {"BLOCKER": C["red"], "CRITICAL": C["red"], "MAJOR": C["amber"], "MINOR": C["grey"]}
    for i, issue in enumerate(open_issues):
        r = 4 + i
        bg = C["ltred"] if issue["severity"] in ("BLOCKER","CRITICAL") else (C["ltamber"] if issue["severity"]=="MAJOR" else C["ltgrey"])
        data_cell(ws, r, 1, issue["id"], bg=bg, bold=True, size=10, align="center")
        data_cell(ws, r, 2, issue["sprint"], bg=bg, size=10, align="center")
        sc = sev_color.get(issue["severity"], C["grey"])
        c = ws.cell(row=r, column=3, value=issue["severity"])
        c.fill = fill(sc)
        c.font = font(bold=True, size=9, color=C["white"] if issue["severity"] in ("BLOCKER","CRITICAL") else "000000")
        c.alignment = center()
        c.border = border_thin()
        data_cell(ws, r, 4, issue["title"], bg=bg, size=10)
        data_cell(ws, r, 5, issue["assigned"], bg=bg, size=10)
        data_cell(ws, r, 6, issue["status"], bg=bg, size=10, align="center")
        data_cell(ws, r, 7, issue["opened"], bg=bg, size=10, align="center")
        ws.row_dimensions[r].height = 18

    # Closed
    gap = 4 + len(open_issues) + 1
    ws.merge_cells(f"A{gap}:H{gap}")
    ws[f"A{gap}"].value = f"CLOSED ISSUES ({len(closed_issues)})"
    ws[f"A{gap}"].fill = fill(C["green"])
    ws[f"A{gap}"].font = font(bold=True, color=C["white"], size=11)
    ws[f"A{gap}"].alignment = center()
    ws.row_dimensions[gap].height = 20

    header_row(ws, gap + 1, ["ID", "Sprint", "Severity", "Title", "Resolved By", "Closed", "", ""], C["teal"])
    ws.row_dimensions[gap + 1].height = 18

    for i, issue in enumerate(closed_issues):
        r = gap + 2 + i
        data_cell(ws, r, 1, issue["id"], bg=C["ltgreen"], bold=True, size=10, align="center")
        data_cell(ws, r, 2, issue["sprint"], bg=C["ltgreen"], size=10, align="center")
        data_cell(ws, r, 3, issue["severity"], bg=C["ltgreen"], size=10, align="center")
        data_cell(ws, r, 4, issue["title"], bg=C["ltgreen"], size=10)
        data_cell(ws, r, 5, issue["resolved_by"], bg=C["ltgreen"], size=10)
        data_cell(ws, r, 6, issue["closed"], bg=C["ltgreen"], size=10, align="center")
        ws.row_dimensions[r].height = 16

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:H{3 + len(open_issues)}"


def build_data_insights(wb):
    ws = wb.create_sheet("📈 Data Insights")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 20

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Production Invoice Analysis — May 2025 to May 2026  |  nexgen_ftf_db"
    ws["A1"].fill = fill(C["navy"])
    ws["A1"].font = font(bold=True, color=C["white"], size=14)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 30

    # KPIs
    r = 2
    kpis = [
        ("Total Invoiced", "$13,205,323", C["teal"]),
        ("Total Records", "23,353", C["blue"]),
        ("Unique Orders", "~18,497", C["blue"]),
        ("Avg Invoice", "$565", C["green"]),
        ("AR Outstanding", "$6,565,086", C["red"]),
        ("Unpaid Rate", "35.9%", C["red"]),
    ]
    header_row(ws, r, ["Metric", "Value", "Metric", "Value", "Metric", "Value"], C["blue"])
    ws.row_dimensions[r].height = 18
    row3 = r + 1
    for i in range(0, len(kpis), 2):
        for pair_i, (label, val, color) in enumerate(kpis[i:i+2]):
            col_offset = pair_i * 2
            lc = ws.cell(row=row3 + i // 2, column=col_offset + 1, value=label)
            lc.fill = fill(C["ltblue"])
            lc.font = font(bold=True, size=10)
            lc.alignment = left()
            lc.border = border_thin()
            vc = ws.cell(row=row3 + i // 2, column=col_offset + 2, value=val)
            vc.fill = fill(C["ltblue"])
            vc.font = font(bold=True, size=12, color=color)
            vc.alignment = center()
            vc.border = border_thin()
        ws.row_dimensions[row3 + i // 2].height = 22

    # Service type table
    r = row3 + len(kpis) // 2 + 2
    ws.merge_cells(f"A{r}:F{r}")
    ws[f"A{r}"].value = "TOP SERVICES — Production Actuals"
    ws[f"A{r}"].fill = fill(C["navy"])
    ws[f"A{r}"].font = font(bold=True, color=C["white"], size=12)
    ws[f"A{r}"].alignment = center()
    ws.row_dimensions[r].height = 22

    header_row(ws, r + 1, ["Service (CRM Label)", "Orders", "Total Revenue", "Avg Price", "Min", "Max"], C["blue"])
    services = [
        ("Land Survey Only", 13260, "$8,170,826", "$616", "$0.01", "$35,000"),
        ("Land Survey and Elevation", 5654, "$2,262,682", "$400", "$0.01", "$12,500"),
        ("Quote (unclassified)", 1721, "$1,594,682", "$926", "$2", "$13,800"),
        ("Boundary Survey", 992, "$563,055", "$567", "$50", "$10,000"),
        ("Re-Survey / Update Survey", 406, "$161,475", "$397", "$0.01", "$2,800"),
        ("Elevation Certificate", 365, "$99,190", "$271", "$0.01", "$1,500"),
        ("ALTA Table A Survey", 16, "$81,200", "$5,075", "$1,500", "$12,500"),
        ("Spot Survey (New client)", 106, "$62,500", "$589", "$350", "$2,500"),
        ("Topography Survey", 61, "$29,525", "$484", "$75", "$1,800"),
    ]
    for i, row_data in enumerate(services):
        bg = C["ltamber"] if row_data[0].startswith("Land Survey") else C["ltblue"] if i % 2 == 0 else C["white"]
        for col_idx, val in enumerate(row_data, 1):
            data_cell(ws, r + 2 + i, col_idx, val, bg=bg, size=10)
        ws.row_dimensions[r + 2 + i].height = 16

    # County pricing table
    r2 = r + 2 + len(services) + 2
    ws.merge_cells(f"A{r2}:F{r2}")
    ws[f"A{r2}"].value = "COUNTY PRICING — Land Survey Only Actuals (n≥10)"
    ws[f"A{r2}"].fill = fill(C["navy"])
    ws[f"A{r2}"].font = font(bold=True, color=C["white"], size=12)
    ws[f"A{r2}"].alignment = center()
    ws.row_dimensions[r2].height = 22

    header_row(ws, r2 + 1, ["County", "Orders", "Avg Price", "Min", "Max", "Tier"], C["blue"])
    counties = [
        ("OKEECHOBEE", 20, "$1,797", "$700", "$3,800", "Remote/Rural"),
        ("HIGHLANDS", 35, "$1,337", "$575", "$5,000", "Remote/Rural"),
        ("MARION", 77, "$1,052", "$300", "$5,000", "Remote/Rural"),
        ("CITRUS", 37, "$954", "$475", "$4,800", "Remote/Rural"),
        ("CLAY", 47, "$779", "$225", "$5,000", "North FL"),
        ("POLK", 304, "$743", "$200", "$4,800", "Central FL"),
        ("VOLUSIA", 92, "$734", "$250", "$4,500", "North FL"),
        ("LAKE", 71, "$740", "$250", "$3,500", "Central FL"),
        ("DUVAL", 175, "$654", "$200", "$4,800", "North FL"),
        ("COLLIER", 225, "$643", "$200", "$3,800", "SW FL (Naples)"),
        ("MARTIN", 284, "$631", "$200", "$4,800", "Treasure Coast"),
        ("ORANGE", 213, "$607", "$200", "$4,500", "Central FL"),
        ("PASCO", 350, "$608", "$200", "$4,800", "Central FL"),
        ("INDIAN RIVER", 285, "$595", "$200", "$4,800", "Treasure Coast"),
        ("CHARLOTTE", 234, "$591", "$250", "$4,800", "SW FL"),
        ("MANATEE", 380, "$584", "$250", "$4,800", "Gulf Coast"),
        ("HILLSBOROUGH", 680, "$575", "$200", "$4,800", "Central FL"),
        ("PALM BEACH", 3310, "$571", "$200", "$5,000", "South FL"),
        ("BREVARD", 347, "$560", "$200", "$4,800", "Space Coast"),
        ("SARASOTA", 502, "$555", "$200", "$4,800", "Gulf Coast"),
        ("MIAMI-DADE", 587, "$547", "$200", "$4,800", "South FL"),
        ("HERNANDO", 199, "$540", "$275", "$4,400", "Central FL"),
        ("PINELLAS", 418, "$527", "$200", "$3,500", "Gulf Coast"),
        ("LEE", 947, "$500", "$225", "$4,800", "SW FL"),
        ("BROWARD", 1398, "$512", "$200", "$5,000", "South FL"),
        ("ST. LUCIE", 595, "$464", "$200", "$3,800", "Treasure Coast"),
        ("SEMINOLE", 63, "$479", "$250", "$1,800", "Central FL"),
        ("MONROE (Keys)", 35, "$1,735*", "—", "—", "KEYS — Always flag"),
    ]
    tier_colors = {
        "Remote/Rural": C["ltred"], "North FL": C["ltamber"],
        "Central FL": C["ltblue"], "South FL": C["ltgreen"],
        "Gulf Coast": C["ltgreen"], "SW FL (Naples)": C["ltamber"],
        "Treasure Coast": C["ltblue"], "SW FL": C["ltgreen"],
        "Space Coast": C["ltblue"], "KEYS — Always flag": C["red"],
    }
    for i, row_data in enumerate(counties):
        bg = tier_colors.get(row_data[5], C["white"])
        for col_idx, val in enumerate(row_data, 1):
            c = data_cell(ws, r2 + 2 + i, col_idx, val, bg=bg, size=10)
            if row_data[5] == "KEYS — Always flag":
                c.font = font(bold=True, size=10, color=C["white"] if col_idx == 6 else "000000")
        if row_data[5] == "KEYS — Always flag":
            ws.cell(row=r2 + 2 + i, column=6).fill = fill(C["red"])
            ws.cell(row=r2 + 2 + i, column=6).font = font(bold=True, color=C["white"], size=10)
        ws.row_dimensions[r2 + 2 + i].height = 16

    # Monthly trend
    r3 = r2 + 2 + len(counties) + 2
    ws.merge_cells(f"A{r3}:F{r3}")
    ws[f"A{r3}"].value = "MONTHLY REVENUE TREND"
    ws[f"A{r3}"].fill = fill(C["navy"])
    ws[f"A{r3}"].font = font(bold=True, color=C["white"], size=12)
    ws[f"A{r3}"].alignment = center()
    ws.row_dimensions[r3].height = 22

    header_row(ws, r3 + 1, ["Month", "Orders Billed", "Revenue", "", "", ""], C["blue"])
    months = [
        ("2025-05", 415, "$209,995"), ("2025-06", 1803, "$998,687"),
        ("2025-07", 2090, "$1,165,817"), ("2025-08", 2020, "$1,136,098"),
        ("2025-09", 1944, "$1,100,770"), ("2025-10", 1990, "$1,208,782"),
        ("2025-11", 1564, "$899,195"), ("2025-12", 1724, "$1,061,520"),
        ("2026-01", 1657, "$1,016,150"), ("2026-02", 1723, "$950,890"),
        ("2026-03", 2219, "$1,217,541"), ("2026-04", 2212, "$1,304,785"),
        ("2026-05", 1692, "$941,167"),
    ]
    for i, (month, orders, rev) in enumerate(months):
        is_peak = orders >= 2000
        bg = C["ltgreen"] if is_peak else C["white"]
        data_cell(ws, r3 + 2 + i, 1, month, bg=bg, size=10, align="center")
        data_cell(ws, r3 + 2 + i, 2, orders, bg=bg, size=10, align="center")
        data_cell(ws, r3 + 2 + i, 3, rev, bg=bg, bold=is_peak, size=10)
        ws.row_dimensions[r3 + 2 + i].height = 16

    ws.freeze_panes = "A2"


def build_qa_orders(wb):
    ws = wb.create_sheet("🧪 QA Orders")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 48

    ws.merge_cells("A1:F1")
    ws["A1"].value = "QA Orders — NexGen Pipeline Test Scenarios"
    ws["A1"].fill = fill(C["navy"])
    ws["A1"].font = font(bold=True, color=C["white"], size=14)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 30

    header_row(ws, 2, ["QA Order ID", "Service Type", "Customer Type", "Flood Zone", "Status", "Notes"], C["blue"])
    ws.row_dimensions[2].height = 18

    scenarios = [
        ("QA-TODAY-boundary-individual", "Boundary Survey", "individual", "No", "classified", "Standard residential — clean path through pricing→writer→reviewer"),
        ("QA-TODAY-boundary-flood", "Boundary Survey", "individual", "Yes (AE)", "classified", "AE flood zone — elevation cert upcharge expected"),
        ("QA-TODAY-topo-b2b", "Topography Survey", "b2b", "No", "classified", "B2B topography — b2b pricing tier"),
        ("QA-TODAY-final-survey", "Final Survey", "individual", "No", "classified", "Final survey — Tampa-area residential"),
        ("QA-TODAY-form-board", "Form Board Survey", "individual", "No", "classified", "Form board — Jacksonville area"),
        ("QA-TODAY-update-survey", "Update Survey", "individual", "No", "classified", "Update survey — Gainesville area"),
        ("QA-TODAY-elevation-cert-flagged", "Elevation Certificate", "individual", "Yes (AE)", "flagged", "ALWAYS_FLAG — routes to human gate"),
        ("QA-TODAY-alta-b2b-flagged", "ALTA Table A Survey", "b2b", "No", "flagged", "ALWAYS_FLAG — human gate required"),
        ("QA-TODAY-prateek-test-1", "Boundary Survey", "individual", "No", "classified", "Prateek Test 1 — FTF order 1000276115 (100 Test Lane, Orlando)"),
        ("QA-TODAY-prateek-test-2", "Elevation Certificate", "individual", "Yes (AE)", "flagged", "Prateek Test 2 — FTF order 1000276116 (200 Test Blvd, Miami-Dade)"),
    ]
    status_colors = {"classified": C["ltgreen"], "flagged": C["ltamber"]}
    for i, row_data in enumerate(scenarios):
        bg = status_colors.get(row_data[4], C["white"])
        for col_idx, val in enumerate(row_data, 1):
            data_cell(ws, 3 + i, col_idx, val, bg=bg, size=10)
        ws.row_dimensions[3 + i].height = 18

    # FTF orders created note
    note_row = 3 + len(scenarios) + 2
    ws.merge_cells(f"A{note_row}:F{note_row}")
    ws[f"A{note_row}"].value = "FTF Staging Orders Created: 1000276115 (Prateek Test 1) | 1000276116 (Prateek Test 2) | customer_id: 201911011657090565 | status: Quote"
    ws[f"A{note_row}"].fill = fill(C["ltblue"])
    ws[f"A{note_row}"].font = font(italic=True, size=10)
    ws[f"A{note_row}"].alignment = left()
    ws.row_dimensions[note_row].height = 18


def build_aliases(wb):
    ws = wb.create_sheet("🗂 Service Aliases")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 40

    ws.merge_cells("A1:D1")
    ws["A1"].value = "Service Type Alias Map — Production-Confirmed (2026-05-27)"
    ws["A1"].fill = fill(C["navy"])
    ws["A1"].font = font(bold=True, color=C["white"], size=14)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 30

    header_row(ws, 2, ["CRM / FTF Label", "Canonical Name", "Prod Orders", "Note"], C["blue"])
    aliases = [
        ("Land Survey Only", "Boundary Survey", "12,119", "Most common — 57% of all orders"),
        ("Land Survey and Elevation", "Boundary Survey (bundle)", "2,878", "Includes EC; price varies"),
        ("Property Survey and Elevation", "Boundary Survey (bundle)", "65", "Same as above"),
        ("Re-Survey", "Update Survey", "391", "Confirmed from prod data"),
        ("Resurvey / Resurvey (trailing space)", "Update Survey", "2", ""),
        ("Spot Survey (New client)", "Foundation Tie-In", "106", "First-time client"),
        ("Spot Survey (Prior NexGen Survey)", "Foundation Tie-In", "3", "Return client"),
        ("CAD file / CAD", "Survey Re-draw", "72+29", "Fixed $250 add-on"),
        ("Construction Survey", "Topography Survey", "1", ""),
        ("Construction Survey Update", "Topography Survey", "5", "Topo update"),
        ("Topo Survey / Topographic Survey", "Topography Survey", "confirmed", ""),
        ("Update/Topo", "Update Survey or Topography", "2", "Context-dependent"),
        ("As-Built Survey", "Final Survey", "1", ""),
        ("Lot split Survey", "Lot Split", "4", ""),
        ("Re-Flag Corners", "Property Flagging", "2", ""),
        ("Line Stake Out / Offset Staking", "Building Stake Out", "2+1", ""),
        ("Elevation Code Only", "Elevation Only", "FTF UI label", "I-072 fixed"),
        ("Elevation Cert", "Elevation Certificate", "alias", ""),
        ("ALTA Survey / ALTA Conversion / Optional ALTA Table A", "ALTA Table A Survey", "confirmed", ""),
        ("Abstract survey (includes B-II review)", "B-II Title Review", "1", ""),
        ("Finished Floor Elevation on Survey", "Elevation Certificate", "1", "Add-on"),
        ("Add topo to existing survey", "Topography Survey", "1", ""),
        ("Topo upgrade", "Topography Survey", "1", ""),
        ("Quote", "⚠️ UNCLASSIFIED — flag", "2,135", "Human must identify service"),
        ("Table Survey", "❓ UNKNOWN — I-071 open", "—", "Not confirmed by Robert"),
        ("Expedited Rush Survey", "⚠️ FLAG — surcharge needed", "1", ""),
        ("Cancel after office prep (10%)", "⚠️ CANCELLATION billing", "2", "Partial billing"),
        ("Cancel after field work (50%)", "⚠️ CANCELLATION billing", "1", "Partial billing"),
    ]
    for i, row_data in enumerate(aliases):
        bg = C["ltred"] if "⚠️" in str(row_data[1]) or "❓" in str(row_data[1]) else (C["ltgreen"] if i % 2 == 0 else C["white"])
        for col_idx, val in enumerate(row_data, 1):
            data_cell(ws, 3 + i, col_idx, val, bg=bg, size=10)
        ws.row_dimensions[3 + i].height = 16

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A2:D2"


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    wb.remove(wb.active)

    build_dashboard(wb)
    build_issues(wb)
    build_data_insights(wb)
    build_qa_orders(wb)
    build_aliases(wb)

    wb.save(OUTPUT)
    print(f"Saved -> {OUTPUT}")


if __name__ == "__main__":
    main()
