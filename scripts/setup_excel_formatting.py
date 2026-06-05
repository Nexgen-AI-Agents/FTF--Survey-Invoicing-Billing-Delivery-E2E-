"""
One-time script: apply dropdown + row coloring to the OneDrive Approvals sheet.

Run locally (never in CI) whenever the sheet schema changes or formatting is reset.
Preserves all existing data rows.

Usage:
    cd code/sprint_11_invoice_pipeline
    python ../../scripts/setup_excel_formatting.py
"""

import io
import os
import sys
import time

# ── allow imports from shared/ ────────────────────────────────────────────────
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "code", "shared"))

import httpx
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from config.settings import (
    AZURE_TENANT_ID, AZURE_APP_ID, AZURE_CLIENT_SECRET,
    ONEDRIVE_FILE_USER, ONEDRIVE_SHARE_URL,
)

# ── mirror constants from onedrive_excel_client ───────────────────────────────
APPROVAL_HEADERS = [
    "Order ID", "Order Status", "Client Name", "Property Address", "Service",
    "Amount", "Confidence", "Escalate", "FTF Link",
    "Action", "Notes", "Posted At", "Processed At",
]
SHEET_NAME  = "Approvals"
TABLE_NAME  = "ApprovalTable"
COL_COUNT   = len(APPROVAL_HEADERS)
END_COL     = chr(ord("A") + COL_COUNT - 1)   # "M"
COL_ACTION  = 9                                # "J"

_GRAPH     = "https://graph.microsoft.com/v1.0"
_TOKEN_URL = "https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"


def get_token() -> str:
    r = httpx.post(
        _TOKEN_URL.format(tenant=AZURE_TENANT_ID),
        data={
            "grant_type":    "client_credentials",
            "client_id":     AZURE_APP_ID,
            "client_secret": AZURE_CLIENT_SECRET,
            "scope":         "https://graph.microsoft.com/.default",
        },
        timeout=15.0,
    )
    r.raise_for_status()
    return r.json()["access_token"]


def resolve_file(token: str) -> tuple[str, str]:
    """Return (item_id, drive_id)."""
    import base64
    encoded = base64.urlsafe_b64encode(ONEDRIVE_SHARE_URL.encode()).decode().rstrip("=")
    share_id = f"u!{encoded}"
    r = httpx.get(
        f"{_GRAPH}/shares/{share_id}/driveItem",
        headers={"Authorization": f"Bearer {token}"},
        timeout=15.0,
    )
    r.raise_for_status()
    item = r.json()
    return item["id"], item["parentReference"]["driveId"]


def download(token: str, drive_id: str, item_id: str) -> bytes:
    url = f"{_GRAPH}/drives/{drive_id}/items/{item_id}/content"
    r = httpx.get(url, headers={"Authorization": f"Bearer {token}"}, follow_redirects=True, timeout=30.0)
    r.raise_for_status()
    return r.content


def upload(token: str, drive_id: str, item_id: str, data: bytes) -> None:
    url = f"{_GRAPH}/drives/{drive_id}/items/{item_id}/content"
    for attempt in range(4):
        r = httpx.put(
            url,
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/octet-stream"},
            content=data,
            timeout=60.0,
        )
        if r.status_code == 423:
            wait = (attempt + 1) * 15
            print(f"  File locked (423) - retrying in {wait}s (attempt {attempt+1}/4) ...")
            print("  NOTE: Close FTF-Invoicing Agent.xlsx in Excel Online if it is open.")
            time.sleep(wait)
            token = get_token()
            continue
        r.raise_for_status()
        print(f"  Uploaded {len(data):,} bytes OK")
        return
    raise RuntimeError("Upload failed after 4 retries - file is locked. Close it in Excel Online first.")


def apply_formatting(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    if SHEET_NAME not in wb.sheetnames:
        print(f"  Creating new sheet: {SHEET_NAME}")
        ws = wb.create_sheet(SHEET_NAME)
        for col_idx, header in enumerate(APPROVAL_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
        tab = Table(displayName=TABLE_NAME, ref=f"A1:{END_COL}1")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(tab)
    else:
        ws = wb[SHEET_NAME]
        print(f"  Existing sheet found — preserving {ws.max_row - 1} data rows")

    action_letter = chr(ord("A") + COL_ACTION)  # "J"

    # ── Data validation dropdown ──────────────────────────────────────────────
    ws.data_validations.dataValidation.clear()
    dv = DataValidation(
        type="list",
        formula1='"Approve,Reject,On-hold"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error      = "Select Approve, Reject, or On-hold"
    dv.errorTitle = "Invalid action"
    dv.sqref      = f"{action_letter}2:{action_letter}10000"
    ws.add_data_validation(dv)
    print(f"  Dropdown validation applied: {action_letter}2:{action_letter}10000")

    # ── Conditional formatting ────────────────────────────────────────────────
    full_range = f"A2:{END_COL}10000"
    # Clear existing cf rules for this range
    ws.conditional_formatting._cf_rules.pop(full_range, None)
    for action_val, hex_color in [("Approve", "C6EFCE"), ("Reject", "FFC7CE"), ("On-hold", "FFEB9C")]:
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        ws.conditional_formatting.add(
            full_range,
            FormulaRule(formula=[f'${action_letter}2="{action_val}"'], fill=fill),
        )
    print(f"  Row colors: Approve=green, Reject=red, On-hold=yellow on {full_range}")

    return wb


def main():
    print("=== FTF-Invoicing Agent.xlsx — Formatting Setup ===\n")

    print("1. Getting token ...")
    token = get_token()

    print("2. Resolving file ...")
    item_id, drive_id = resolve_file(token)
    print(f"   item_id={item_id[:20]}... drive_id={drive_id[:20]}...")

    print("3. Downloading workbook ...")
    raw = download(token, drive_id, item_id)
    print(f"   {len(raw):,} bytes")

    print("4. Applying formatting ...")
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    wb = apply_formatting(wb)

    print("5. Saving modified workbook ...")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    data = buf.read()
    print(f"   {len(data):,} bytes")

    print("6. Uploading to OneDrive ...")
    upload(token, drive_id, item_id, data)

    print("\nDone - open FTF-Invoicing Agent.xlsx to verify dropdown and row colors.")


if __name__ == "__main__":
    main()
