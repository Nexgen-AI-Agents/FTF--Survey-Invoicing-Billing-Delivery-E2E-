"""Full pipeline reset — empty the approval sheet AND all pipeline state.

Use when you want the pipeline to start completely fresh from live FTF data,
posting ONLY orders that still carry the invoice-needed ($) flag.

What it clears:
  1. OneDrive "Approvals" table — every data row deleted (header/schema kept)
  2. data/invoice_pipeline_state.xlsx "pipeline_state" sheet — data rows wiped
     (learnings / pricing_examples / any other sheets are PRESERVED)
  3. data/pipeline_state.json — orders[] emptied (dashboard export)

What it PRESERVES:
  • Pricing Rules tab (user inputs)            • learned_rules.json (A7 learning)
  • Pipeline Guide / How-To tabs               • learnings sheet in the xlsx

Run from project root:
  python scripts/reset_all_state.py            # clear everything
  python scripts/reset_all_state.py --trigger  # clear, then dispatch the pipeline
"""

import argparse
import json
import os
import sys

import httpx
import openpyxl
from dotenv import load_dotenv

load_dotenv()

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "code", "shared"))

from core.onedrive_excel_client import (  # noqa: E402
    ONEDRIVE_SHEET_NAME,
    ONEDRIVE_TABLE_NAME,
    ensure_action_dropdown,
    _session_headers,
    _wb_base,
)

_REPO_ROOT     = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
XLSX_STATE     = os.path.join(_REPO_ROOT, "data", "invoice_pipeline_state.xlsx")
JSON_STATE     = os.path.join(_REPO_ROOT, "data", "pipeline_state.json")

GITHUB_PAT  = os.getenv("GITHUB_PAT", "")
GITHUB_REPO = os.getenv("GITHUB_REPO", "Nexgen-AI-Agents/FTF--Survey-Invoicing-Billing-Delivery-E2E-")


# ── 1. OneDrive Approvals table ───────────────────────────────────────────────

def clear_onedrive_approvals() -> bool:
    """Delete every Approvals data row; restore the Action dropdown.

    Returns True ONLY if the clear is fully complete: all rows deleted (or none
    existed) AND the Action dropdown was successfully (re)applied. A partial clear
    (e.g. some rows survive a 423 lock) returns False so the caller does not falsely
    report a clean fresh-start.
    """
    print("[1/3] Clearing OneDrive Approvals table...")
    base = _wb_base()
    h    = _session_headers()
    r = httpx.get(
        f"{base}/worksheets/{ONEDRIVE_SHEET_NAME}/tables/{ONEDRIVE_TABLE_NAME}/rows",
        headers=h, timeout=30.0,
    )
    r.raise_for_status()
    rows = r.json().get("value", [])
    print(f"  {len(rows)} data rows found.")

    deleted = 0
    skipped_no_index = 0
    failed = 0
    # Delete from the highest index down so earlier indices stay valid.
    for row in sorted(rows, key=lambda x: x.get("index", -1), reverse=True):
        idx = row.get("index")
        if idx is None:
            skipped_no_index += 1
            continue
        resp = httpx.delete(
            f"{base}/worksheets/{ONEDRIVE_SHEET_NAME}/tables/{ONEDRIVE_TABLE_NAME}/rows/itemAt(index={idx})",
            headers=h, timeout=20.0,
        )
        if resp.is_success:
            deleted += 1
        else:
            failed += 1
            print(f"  [!] failed to delete row index={idx}: {resp.status_code} {resp.text[:120]}")

    rows_complete = (deleted == len(rows)) and skipped_no_index == 0 and failed == 0
    print(f"  Deleted {deleted}/{len(rows)} rows "
          f"(failed={failed}, no-index={skipped_no_index}). Header + schema preserved.")

    # Row deletion strips the Action-column dropdown — re-apply it (returns bool).
    dropdown_ok = ensure_action_dropdown()
    if dropdown_ok:
        print("  Action dropdown (Approve/Reject/On-hold) re-applied.")
    else:
        print("  [!] Action dropdown NOT re-applied — likely the file is open in Excel "
              "(423 lock). Close it and re-run, or it self-heals on the next pipeline run.")

    return rows_complete and dropdown_ok


# ── 2. Authoritative xlsx state store ─────────────────────────────────────────

def clear_xlsx_state() -> int:
    print("[2/3] Clearing pipeline_state sheet in invoice_pipeline_state.xlsx...")
    if not os.path.exists(XLSX_STATE):
        print(f"  [!] {XLSX_STATE} not found — skipping.")
        return 0

    wb = openpyxl.load_workbook(XLSX_STATE)
    if "pipeline_state" not in wb.sheetnames:
        print("  [!] 'pipeline_state' sheet missing — skipping.")
        return 0

    ws = wb["pipeline_state"]
    data_rows = ws.max_row - 1  # exclude header
    if data_rows > 0:
        ws.delete_rows(2, ws.max_row - 1)  # keep row 1 (header)
    wb.save(XLSX_STATE)
    print(f"  Removed {max(data_rows, 0)} data rows. Header kept. "
          f"Preserved sheets: {[s for s in wb.sheetnames if s != 'pipeline_state']}")
    return max(data_rows, 0)


# ── 3. JSON dashboard export ──────────────────────────────────────────────────

def clear_json_state() -> int:
    print("[3/3] Emptying pipeline_state.json (dashboard export)...")
    if not os.path.exists(JSON_STATE):
        print(f"  [!] {JSON_STATE} not found — writing fresh empty file.")
        state = {}
    else:
        with open(JSON_STATE) as f:
            state = json.load(f)
    before = len(state.get("orders", []))
    state["orders"] = []
    # Zero out known aggregate keys so the dashboard doesn't show non-zero totals
    # against zero orders until export_pipeline_json.py regenerates the file next run.
    for k in ("status_counts", "totals", "summary"):
        if isinstance(state.get(k), dict):
            state[k] = {}
    with open(JSON_STATE, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2, default=str)
    print(f"  orders[]: {before} -> 0 (aggregate keys cleared; full regen on next export run)")
    return before


def commit_and_push_state() -> bool:
    """Commit + push the emptied data/ files so the dispatched runner checks out the
    fresh state. Without this, --trigger fires a run that checks out origin/main (the OLD
    populated state) and resurrects the deleted orders. Returns True on a clean push."""
    import subprocess
    print("\n[commit] Committing + pushing emptied state before dispatch...")
    try:
        subprocess.run(["git", "add", "data/invoice_pipeline_state.xlsx", "data/pipeline_state.json"],
                       cwd=_REPO_ROOT, check=True)
        diff = subprocess.run(["git", "diff", "--cached", "--quiet"], cwd=_REPO_ROOT)
        if diff.returncode == 0:
            print("  no state changes to commit (already clean).")
            return True
        subprocess.run(["git", "commit", "-m", "chore: reset pipeline state to empty [skip ci]"],
                       cwd=_REPO_ROOT, check=True)
        subprocess.run(["git", "pull", "--rebase", "-X", "ours", "origin", "main"], cwd=_REPO_ROOT, check=True)
        subprocess.run(["git", "push", "origin", "main"], cwd=_REPO_ROOT, check=True)
        print("  pushed.")
        return True
    except Exception as exc:
        print(f"  [!] commit/push failed: {exc}")
        return False


def trigger_pipeline() -> None:
    # HIGH-3: never dispatch before the emptied state is on origin/main, or the runner
    # resurrects the deleted orders from the old committed state.
    if not commit_and_push_state():
        print("  [!] Aborting dispatch — emptied state was not pushed. Commit/push manually, then dispatch.")
        return
    print("\n[trigger] Dispatching invoice_pipeline workflow...")
    if not GITHUB_PAT:
        print("  [!] GITHUB_PAT not set — commit + dispatch manually.")
        return
    resp = httpx.post(
        f"https://api.github.com/repos/{GITHUB_REPO}/actions/workflows/invoice_pipeline.yml/dispatches",
        headers={
            "Authorization": f"Bearer {GITHUB_PAT}",
            "Accept": "application/vnd.github+json",
            "X-GitHub-Api-Version": "2022-11-28",
        },
        json={"ref": "main"},
        timeout=15.0,
    )
    print("  HTTP 204 — triggered." if resp.status_code == 204
          else f"  [!] {resp.status_code}: {resp.text[:200]}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Full pipeline state reset")
    parser.add_argument("--trigger", action="store_true", help="dispatch pipeline after reset")
    args = parser.parse_args()

    print("\n=== FULL PIPELINE STATE RESET ===\n")
    onedrive_ok = False
    try:
        onedrive_ok = clear_onedrive_approvals()
    except Exception as exc:
        print(f"  [!] OneDrive clear failed: {exc}")
    print()
    clear_xlsx_state()
    print()
    clear_json_state()

    # Only dispatch on a fully-clean reset — a partial OneDrive clear + dispatch would run
    # the pipeline against a split-brain (local empty, OneDrive populated).
    if args.trigger and onedrive_ok:
        trigger_pipeline()
    elif args.trigger and not onedrive_ok:
        print("\n[trigger] SKIPPED — reset was partial; not dispatching. Fix OneDrive, re-run, then dispatch.")

    if onedrive_ok:
        print("\nDone — all stores cleared cleanly. "
              + ("Pipeline dispatched against the fresh state." if args.trigger
                 else "Commit + push data/ (or re-run with --trigger) to start fresh from live FTF orders."))
    else:
        print("\n[!] PARTIAL RESET — OneDrive Approvals was not fully cleared (rows survived or the "
              "dropdown could not be re-applied; the file may be open in Excel). Local xlsx/JSON ARE "
              "empty. Close the file and re-run to avoid a split-brain (OneDrive populated, local empty).")
    sys.exit(0 if onedrive_ok else 1)
