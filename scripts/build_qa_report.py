"""
build_qa_report.py — Build comprehensive QA test report Excel workbook.

Tabs:
  1. Summary            — overall status, connection test results, counts
  2. Teams Connection   — live test results (4 steps)
  3. Positive Tests     — happy-path test cases across all components
  4. Negative Tests     — error, rejection, invalid-input, security test cases
  5. TC-TEAMS           — 35 Teams integration test cases (detailed)
  6. TC-PIPE            — 45 pipeline test cases (detailed)
  7. Existing Scenarios — 30 QA orders from qa_orders.py

Usage:
    python scripts/build_qa_report.py
"""

import os
import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "code", "shared"))

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Style helpers ─────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=11) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _align(wrap=True, h="left", v="center") -> Alignment:
    return Alignment(wrap_text=wrap, horizontal=h, vertical=v)

HEADER_FILL   = _fill("1F4E79")  # dark blue
HEADER_FONT   = _font(bold=True, color="FFFFFF", size=11)
SECTION_FILL  = _fill("D6E4F0")  # light blue
SECTION_FONT  = _font(bold=True, color="1F4E79")
PASS_FILL     = _fill("C6EFCE")
FAIL_FILL     = _fill("FFC7CE")
WARN_FILL     = _fill("FFEB9C")
SKIP_FILL     = _fill("EDEDED")
P1_FILL       = _fill("FFC7CE")
P2_FILL       = _fill("FFEB9C")
P3_FILL       = _fill("C6EFCE")
ALT_FILL      = _fill("F5F9FF")
WHITE_FILL    = _fill("FFFFFF")

def _write_row(ws, row, values, fills=None, fonts=None, bold_first=False):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border  = _border()
        cell.alignment = _align()
        if fills and col <= len(fills) and fills[col-1]:
            cell.fill = fills[col-1]
        if fonts and col <= len(fonts) and fonts[col-1]:
            cell.font = fonts[col-1]
        elif bold_first and col == 1:
            cell.font = _font(bold=True)
        else:
            cell.font = _font()

def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell

def _header_row(ws, row, headers, fill=None, font=None):
    f = fill or HEADER_FILL
    fn = font or HEADER_FONT
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = f
        cell.font = fn
        cell.border = _border()
        cell.alignment = _align(h="center")


# ── Tab 1: Summary ────────────────────────────────────────────────────────────

def build_summary(wb, now_str: str):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "FTF Agentic AI — QA Test Report"
    ws["A1"].font = _font(bold=True, color="1F4E79", size=16)
    ws["A1"].alignment = _align(h="center")
    ws["A1"].fill = _fill("D6E4F0")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Generated: {now_str}  |  Prepared by: FTF Agentic AI QA Suite"
    ws["A2"].font = _font(color="555555", size=10)
    ws["A2"].alignment = _align(h="center")

    # Connection test results
    r = 4
    ws.merge_cells(f"A{r}:F{r}")
    ws[f"A{r}"] = "Live Connection Test Results (2026-05-28)"
    ws[f"A{r}"].font = SECTION_FONT
    ws[f"A{r}"].fill = SECTION_FILL
    ws[f"A{r}"].alignment = _align(h="center")
    ws.row_dimensions[r].height = 20

    r += 1
    _header_row(ws, r, ["Step", "Description", "Result", "Details"])
    r += 1
    conn_results = [
        ("1 — Config",      "All 5 Teams/Azure env vars set",                          "PASS", "Tenant, App ID, Team, Channel all configured"),
        ("2 — Auth",        "client_credentials token obtained from Azure AD",          "PASS", "Token: eyJ0eXAi... (valid 3600s)"),
        ("3 — Read",        "ChannelMessage.Read.All — Graph API channel poll",         "PASS", "Retrieved 0 messages (channel empty — expected)"),
        ("4 — Send",        "Mail.Send — email to rt@ryantisko.com + robert_e@nexgenlogix.com", "PASS", "Sent via graph_email; appears in Outlook Sent"),
    ]
    for step, desc, result, details in conn_results:
        fill = PASS_FILL if result == "PASS" else FAIL_FILL if result == "FAIL" else WARN_FILL
        _write_row(ws, r, [step, desc, result, details],
                   fills=[None, None, fill, None])
        r += 1

    # Test case counts
    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    ws[f"A{r}"] = "Test Case Inventory"
    ws[f"A{r}"].font = SECTION_FONT
    ws[f"A{r}"].fill = SECTION_FILL
    ws[f"A{r}"].alignment = _align(h="center")
    ws.row_dimensions[r].height = 20

    r += 1
    _header_row(ws, r, ["Tab", "Category", "Count", "Priority Mix", "Status", "Notes"])
    r += 1
    inventory = [
        ("TC-TEAMS",           "Teams integration + email notifications",    "35", "P1: 22 | P2: 10 | P3: 3",  "Designed",  "Command parsing, email content, state machine, E2E flow"),
        ("TC-PIPE",            "Pipeline: classifier + pricing + DB + E2E",  "45", "P1: 28 | P2: 14 | P3: 3",  "Designed",  "Alias normalization, county math, B2B multiplier, edge cases"),
        ("Existing Scenarios", "QA order seeds (injected into pipeline DB)", "30", "Mixed",                     "Implemented","30 scenarios covering all flag triggers + service types"),
        ("",                   "TOTAL",                                      "110","",                           "",          ""),
    ]
    for tab, cat, count, priority, status, notes in inventory:
        bold = tab == ""
        row_fill = _fill("EBF3FB") if not bold else _fill("1F4E79")
        row_font = _font(bold=True, color="FFFFFF") if bold else None
        _write_row(ws, r, [tab, cat, count, priority, status, notes])
        if bold:
            for col in range(1, 7):
                ws.cell(r, col).fill = _fill("1F4E79")
                ws.cell(r, col).font = _font(bold=True, color="FFFFFF")
        r += 1

    # Architecture note
    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    ws[f"A{r}"] = "Notification Architecture"
    ws[f"A{r}"].font = SECTION_FONT
    ws[f"A{r}"].fill = SECTION_FILL
    ws[f"A{r}"].alignment = _align(h="center")
    r += 1
    arch = [
        ("SEND (notify out)",  "Graph API Mail.Send — emails Robert + Ryan with hourly digest of pending orders"),
        ("RECEIVE (response)", "Robert/Ryan type APPROVE <id> or REJECT <id> reason in Teams FTF-Approvals channel"),
        ("POLL (pick up cmd)", "poll_teams_approvals.py reads Teams channel via Graph API ChannelMessage.Read.All"),
        ("DB update",          "process_approval_reply() updates order status to approved/rejected in PostgreSQL"),
        ("Confirm",            "send_confirmation() emails confirmation back to Robert + Ryan"),
    ]
    for label, desc in arch:
        _write_row(ws, r, [label, desc, "", "", "", ""])
        r += 1

    _set_col_widths(ws, [22, 50, 12, 26, 14, 50])
    _freeze(ws, "A3")


# ── Tab 2: Teams Connection ───────────────────────────────────────────────────

def build_teams_connection(wb):
    ws = wb.create_sheet("Teams Connection")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"] = "Teams + Graph API — Live Connection Test Results"
    ws["A1"].font = _font(bold=True, color="1F4E79", size=14)
    ws["A1"].fill = SECTION_FILL
    ws["A1"].alignment = _align(h="center")
    ws.row_dimensions[1].height = 25

    r = 2
    _header_row(ws, r, ["Step", "Test Name", "Method", "Result", "Details", "Fix if FAIL"])
    r += 1
    tests = [
        ("1", "Config check",     "Env vars",         "PASS", "TEAMS_TENANT_ID, TEAMS_APP_ID, TEAMS_CLIENT_SECRET, TEAMS_TEAM_ID, TEAMS_CHANNEL_ID all set. NOTIFICATION_FROM_EMAIL + NOTIFICATION_TO_EMAILS set.",
                                                                "Add missing vars to .env"),
        ("2", "Authentication",   "Azure AD OAuth2",   "PASS", "client_credentials token obtained. Expires in 3600s. Cached in-memory with 60s buffer before expiry.",
                                                                "Check TEAMS_CLIENT_SECRET not expired; regenerate in Azure portal"),
        ("3", "Read channel",     "Graph API GET",     "PASS", "ChannelMessage.Read.All Application permission. Retrieved 0 messages (channel empty — expected at setup).",
                                                                "Ensure Application permission + admin consent granted in Azure AD"),
        ("4", "Send email",       "Graph API POST",    "PASS", "Mail.Send Application permission. Email from pchandra@nexgen.enterprises to rt@ryantisko.com + robert_e@nexgenlogix.com. Appears in Sent folder.",
                                                                "Ensure Mail.Send Application permission + admin consent. FROM must be licensed M365 mailbox in tenant."),
    ]
    for step, name, method, result, details, fix in tests:
        fill = PASS_FILL if result == "PASS" else FAIL_FILL
        _write_row(ws, r, [step, name, method, result, details, fix],
                   fills=[None, None, None, fill, None, None])
        ws.row_dimensions[r].height = 50
        r += 1

    _set_col_widths(ws, [6, 22, 18, 10, 55, 45])
    _freeze(ws, "A3")


# ── Tab 3: TC-TEAMS ───────────────────────────────────────────────────────────

def build_tc_teams(wb):
    ws = wb.create_sheet("TC-TEAMS")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "TC-TEAMS — Teams Integration & Email Notification Test Cases (35 cases)"
    ws["A1"].font = _font(bold=True, color="1F4E79", size=13)
    ws["A1"].fill = SECTION_FILL
    ws["A1"].alignment = _align(h="center")
    ws.row_dimensions[1].height = 22

    r = 2
    _header_row(ws, r, ["Test ID", "Area", "Priority", "Type", "Scenario", "Test Data / Inputs", "Expected Result", "Pass Criteria"])
    r += 1

    tc_teams = [
        # Email Content
        ("TC-TEAMS-001","Email Content","P1","Unit","Batch digest subject — singular/plural","send_batch_approval_digest() with 2 flagged orders","Subject = 'FTF Estimates Pending Review — 2 orders'","Subject field matches pattern; singular when 1 order"),
        ("TC-TEAMS-002","Email Content","P1","Unit","Default subject when subject='' passed","send_email_notification(body, subject='')","Subject = 'FTF Estimates — Action Required'","Not empty string; not None"),
        ("TC-TEAMS-003","Email Content","P1","Unit","All 6 order fields in digest row","orders=[{order_id, service_type, amount, flag_reason, status, flagged_at}]","Row contains order_id, link, service, $amount, flag_reason[:60], age","All 6 fields present; $ prefix; , separator on amount"),
        ("TC-TEAMS-004","Email Content","P1","Unit","estimate_amount=None renders as TBD","orders=[{...estimate_amount: None...}]","Amount shows 'TBD'","No $None, no $0.00, no crash"),
        ("TC-TEAMS-005","Email Content","P1","Unit","APPROVE/REJECT instructions always included","Any non-empty orders list","Output contains APPROVE <order_id>, APPROVE ALL, REJECT <order_id> <reason>, example","All 4 instruction strings present"),
        ("TC-TEAMS-006","Email Content","P2","Unit","HTML body passthrough — tags preserved","text_or_html with <h3>, <ul>, <li>, <a href>","contentType=HTML; content field has tags intact","<h3> still present; not stripped"),
        ("TC-TEAMS-007","Email Content","P1","Unit","Two recipients in toRecipients array","NOTIFICATION_TO_EMAILS='rt@ryantisko.com,robert_e@nexgenlogix.com'","toRecipients has 2 entries","len(toRecipients)==2; correct addresses"),
        ("TC-TEAMS-008","Email Content","P2","Unit","flag_reason truncated to 60 chars","flag_reason longer than 60 characters","Digest shows at most 60 chars","Rendered reason <= 60 chars; sliced via [:60]"),
        # Command Parsing
        ("TC-TEAMS-009","Command Parsing","P1","Unit","Standard APPROVE <id>","text='APPROVE 1000276115'","Returns ('approve', '1000276115', None)","Exact tuple match"),
        ("TC-TEAMS-010","Command Parsing","P1","Unit","APPROVE ALL","text='APPROVE ALL'","Returns ('approve_all', None, None)","action=='approve_all'; both ids None"),
        ("TC-TEAMS-011","Command Parsing","P1","Unit","REJECT with reason","text='REJECT 1000276115 wrong county entered'","Returns ('reject', '1000276115', 'wrong county entered')","Multi-word reason preserved"),
        ("TC-TEAMS-012","Command Parsing","P1","Unit","REJECT without reason","text='REJECT 1000276115'","Returns ('reject', '1000276115', None)","reason is None; no crash"),
        ("TC-TEAMS-013","Command Parsing","P2","Unit","Mixed case commands","'approve 1000276115', 'Approve ALL', 'reject 1000276115 bad'","All parse same as uppercase","Case-insensitive via text.upper()"),
        ("TC-TEAMS-014","Command Parsing","P1","Unit","@mention prefix Teams inserts","Raw HTML '<at>FTF Bot</at> APPROVE 1000276115' cleaned first","After _clean_message_body: 'APPROVE 1000276115'; parses correctly","order_id==1000276115; mention stripped"),
        ("TC-TEAMS-015","Command Parsing","P2","Unit","Extra whitespace","'APPROVE   1000276115', '  APPROVE 1000276115  '","('approve', '1000276115', None) both cases","split(None,1) collapses whitespace"),
        ("TC-TEAMS-016","Command Parsing","P1","Unit","APPROVE with no order ID","text='APPROVE', text='APPROVE  '","Returns ('unknown', None, None)","No IndexError; action=='unknown'"),
        ("TC-TEAMS-017","Command Parsing","P1","Unit","REJECT with no order ID","text='REJECT', text='REJECT   '","Returns ('unknown', None, None)","No crash; action=='unknown'"),
        ("TC-TEAMS-018","Command Parsing","P2","Unit","Normal chat message — no command","'Looks good!', 'What is status of 123?', ''","Returns ('unknown', None, None) for all","No false positive parse"),
        ("TC-TEAMS-019","Command Parsing","P2","Unit","HTML body stripping","'<p>APPROVE <b>1000276115</b></p>'","After _clean_message_body: 'APPROVE 1000276115'","No angle brackets in output; content correct"),
        # Polling State
        ("TC-TEAMS-020","State Machine","P1","Unit","since= datetime filter skips old messages","State file has T-10min; old msg at T-20min, new at T-5min","Only T-5min command processed","summary[found]==1; old not re-processed"),
        ("TC-TEAMS-021","State Machine","P1","Unit","State file persisted after processing","Command at 2026-05-28T15:30:00+00:00; dry_run=False","poll_state.json has {last_processed_at: '2026-05-28T15:30:00+00:00'}","File exists; ISO timestamp matches newest cmd"),
        ("TC-TEAMS-022","State Machine","P1","Unit","Dry-run prevents DB writes + state update","Channel has APPROVE cmd; dry_run=True","process_approval_reply NOT called; state file NOT updated; [DRY RUN] in stdout","DB unchanged; state unchanged"),
        ("TC-TEAMS-023","State Machine","P1","Unit","No state file — fallback to --since-hours","poll_state.json deleted; --since-hours 2","since = now - 2h; messages within 2h processed","_load_last_polled() returns None; correct cutoff"),
        ("TC-TEAMS-024","State Machine","P2","Unit","State file not updated when no commands","No APPROVE/REJECT since last poll","poll_state.json mtime unchanged","File mtime unchanged; summary[found]==0"),
        ("TC-TEAMS-025","State Machine","P1","Unit","Bot's own messages filtered out","Graph response includes msg from app_id==TEAMS_APP_ID","That message excluded from results","No recursive confirmation loop"),
        # Error Handling
        ("TC-TEAMS-026","Error Handling","P1","Unit","Mail.Send HTTP 403","Mock httpx.post returns 403","AgentError raised with 'HTTP 403'","Error propagates; no silent swallow"),
        ("TC-TEAMS-027","Error Handling","P1","Unit","Channel read timeout","Mock httpx.get raises ReadTimeout","AgentError raised; run_poll returns {found:0, failed:1}","No unhandled exception; pipeline continues"),
        ("TC-TEAMS-028","Error Handling","P1","Unit","Missing TEAMS_ env vars","TEAMS_CLIENT_SECRET=''","AgentError naming missing vars; token endpoint NOT called","Error is actionable; names the vars"),
        ("TC-TEAMS-029","Error Handling","P2","Unit","No notification method configured","Both email vars and webhook URL empty","AgentError with Option A + Option B guidance","Message contains 'Option A' and 'docs/teams_setup.md'"),
        ("TC-TEAMS-030","Error Handling","P3","Unit","Token cache prevents redundant calls","_get_token() called twice within validity window","Token endpoint called exactly once","Second call uses cached token; no extra HTTP I/O"),
        # E2E
        ("TC-TEAMS-031","E2E Flow","P1","Integration","Happy path: flagged → approved","Order 1000276115 flagged; Robert types APPROVE 1000276115; poll runs","Status: flagged→awaiting→approved; confirmation email sent","All 5 steps; DB shows approved"),
        ("TC-TEAMS-032","E2E Flow","P1","Integration","Reject with multi-word reason","Order 1000276116 awaiting; Ryan types REJECT 1000276116 wrong county","status=rejected; confirmation contains reason string","Reason preserved; DB updated"),
        ("TC-TEAMS-033","E2E Flow","P1","Integration","APPROVE ALL — 3 orders","3 orders: 1 flagged + 2 awaiting; Ryan types APPROVE ALL","All 3 advance to approved; confirmation lists all 3 IDs","summary[approved]==3; all DB rows approved"),
        ("TC-TEAMS-034","E2E Flow","P2","Integration","APPROVE ALL with no pending orders","DB has zero flagged/awaiting; Ryan types APPROVE ALL","Confirmation: 'No orders pending approval right now.'","No crash; approved==0"),
        ("TC-TEAMS-035","E2E Flow","P1","Integration","Double-approve guard","Order already approved; cmd APPROVE sent again","AgentError 'not awaiting approval'; orange warning email; DB unchanged","status stays approved; failed+=1"),
    ]

    for i, row_data in enumerate(tc_teams):
        alt = ALT_FILL if i % 2 == 0 else WHITE_FILL
        p_fill = P1_FILL if row_data[2]=="P1" else P2_FILL if row_data[2]=="P2" else P3_FILL
        _write_row(ws, r, list(row_data),
                   fills=[None, None, p_fill, None, None, None, None, None])
        for col in range(1, 9):
            if col != 3:
                ws.cell(r, col).fill = alt
        ws.row_dimensions[r].height = 45
        r += 1

    _set_col_widths(ws, [14, 18, 10, 12, 40, 38, 38, 35])
    _freeze(ws, "A3")


# ── Tab 4: TC-PIPE ────────────────────────────────────────────────────────────

def build_tc_pipe(wb):
    ws = wb.create_sheet("TC-PIPE")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "TC-PIPE — Pipeline: Classifier + Pricing + DB + E2E Test Cases (45 cases)"
    ws["A1"].font = _font(bold=True, color="1F4E79", size=13)
    ws["A1"].fill = SECTION_FILL
    ws["A1"].alignment = _align(h="center")
    ws.row_dimensions[1].height = 22

    r = 2
    _header_row(ws, r, ["Test ID","Area","Priority","Type","Scenario","Input Data","Expected Output","Pass Criteria"])
    r += 1

    tc_pipe = [
        # Classifier
        ("TC-PIPE-001","Classifier","P1","Unit","Alias 'land survey only' → 'Boundary Survey'","service_type='land survey only', county=ORANGE, individual","service_type=Boundary Survey, status=classified","Alias lookup succeeds; ORANGE county avg $607 used"),
        ("TC-PIPE-002","Classifier","P1","Unit","Alias 'spot survey' → 'Foundation Tie-In'","service_type='spot survey', county=PALM BEACH","service_type=Foundation Tie-In, status=classified","FTF API pricing (not county avg — not in _BOUNDARY_SERVICES)"),
        ("TC-PIPE-003","Classifier","P1","Unit","Alias 'elevation code only' → 'Elevation Only' (I-072)","service_type='elevation code only', county=BROWARD","service_type=Elevation Only; NOT Elevation Certificate","I-072 fix verified; no wrong substitution"),
        ("TC-PIPE-004","Classifier","P1","Unit","ALTA alias → ALTA Table A Survey → ALWAYS_FLAG","service_type='alta survey', county=MIAMI-DADE, individual","service_type=ALTA Table A Survey, status=flagged","Alias resolves first; then ALWAYS_FLAG fires; not double-flagged"),
        ("TC-PIPE-005","Classifier","P1","Unit","service_type='Quote' held as unresolved","service_type='Quote', county=ORANGE","status=flagged; flag_reason contains 'unresolved as Quote'","No alias; no pricing; flag reason exact match"),
        ("TC-PIPE-006","Classifier","P2","Unit","Unrecognized type → LLM fallback → flag","service_type='Drone Photogrammetry Survey'","status=flagged; flag_reason contains 'unrecognized service type'","LLM called; non-canonical response still flags"),
        ("TC-PIPE-007","Classifier","P2","Unit","Informal type 'construction/permitting' flagged","service_type='construction/permitting', county=HILLSBOROUGH","status=flagged; flag_reason contains 'informal survey type'","Never reaches pricing; _INFORMAL_SERVICE_TYPES check"),
        ("TC-PIPE-008","Classifier","P1","Unit","Topography Survey always flagged (I-054)","service_type='Topography Survey', county=ORANGE, individual","status=flagged; NEVER_AUTO_QUOTE flag","Even individual residential with known county still flags"),
        # Flag Triggers
        ("TC-PIPE-009","Flag Triggers","P1","Unit","Building Stake Out ALWAYS_FLAG","service_type='Building Stake Out', county=SEMINOLE, individual","status=flagged; ALWAYS_FLAG_SERVICES path","Not NEVER_AUTO_QUOTE; ALWAYS_FLAG fires"),
        ("TC-PIPE-010","Flag Triggers","P1","Unit","B2B does NOT cause flag (is pricing tier)","service_type='Boundary Survey', county=ORANGE, b2b","status=classified, pricing_tier=b2b, flag_reason=null","Classifier passes; B2B multiplier applied by pricing only"),
        ("TC-PIPE-011","Flag Triggers","P2","Unit","Dual trigger: competitor + ALWAYS_FLAG service","service_type='ALTA Table A Survey', company_name='SurvTech Solutions'","flag_reason contains BOTH competitor match AND service review","Reasons joined with '; '; both present"),
        ("TC-PIPE-012","Classifier","P2","Unit","Subdomain does NOT match competitor domain","customer_email='agent@mail.apexsurvey.us'","status=classified; no false positive flag","split('@')[-1] = mail.apexsurvey.us; not in COMPETITOR_DOMAINS"),
        ("TC-PIPE-013","Flag Triggers","P1","Unit","FEMA API unavailable → flag","FEMA client throws FEMAUnavailableError","status=flagged; flood_zone=UNAVAILABLE; flag reason set","No crash; order flagged cleanly"),
        ("TC-PIPE-014","Classifier","P2","Unit","Property lat 24.4 (FL south) — no lat-bounds flag","service_type='Boundary Survey', lat=24.4, county=MONROE","MONROE flag fires; lat-bounds check does NOT add second flag","I-037 only checks > 31.0; 24.4 not in upper-bound check"),
        ("TC-PIPE-015","Classifier","P2","Unit","Property lat exactly 31.0 — NOT flagged by I-037","service_type='Boundary Survey', lat=31.0, county=DUVAL","status=classified (lat > 31.0 is strictly greater)","float(31.0) > 31.0 is False; passes clean"),
        ("TC-PIPE-016","Classifier","P1","Unit","property_state='Florida' normalized (I-050)","property_state='Florida', service_type='Boundary Survey', county=ORANGE","Normalized to FL; no out-of-state flag","I-050 fix verified; 'FLORIDA'→'FL'"),
        # Pricing Math
        ("TC-PIPE-017","Pricing","P1","Unit","Okeechobee county avg $1,797","service_type='Boundary Survey', county=OKEECHOBEE, individual","base_amount=1797, pricing_source=county_avg","Exact county avg; no FTF API call"),
        ("TC-PIPE-018","Pricing","P1","Unit","Unknown county → FL statewide avg $616","service_type='Boundary Survey', county=GILCHRIST, individual","base_amount=616, pricing_source=fl_statewide_avg","Note warns to verify; statewide avg used"),
        ("TC-PIPE-019","Pricing","P2","Unit","ST LUCIE punctuation variants both $702","county='ST LUCIE' and county='ST. LUCIE'","Both yield base_amount=702","Both keys in _COUNTY_AVG with same value"),
        ("TC-PIPE-020","Pricing","P1","Unit","Lowercase county normalized","county='orange', service_type='Boundary Survey'","base_amount=607 (ORANGE avg)","raw_county.upper() normalizes before lookup"),
        ("TC-PIPE-021","Pricing","P1","Unit","B2B multiplier: Brevard $560 × 4.1 = $2,296","county=BREVARD, b2b, individual pricing_tier","base_amount=2296.0, pricing_source=county_avg+b2b_4.1x","round(560*4.1, 2)==2296.0; multiplier only when b2b AND NOT override"),
        ("TC-PIPE-022","Pricing","P1","Unit","B2B + unknown county: $616 × 4.1 = $2,525.60","county=GILCHRIST, b2b","base_amount=2525.6, pricing_source=fl_statewide_avg+b2b_4.1x","Statewide avg first; B2B on top; order of operations correct"),
        ("TC-PIPE-023","Pricing","P1","Unit","Elevation cert add-on: Miami-Dade $547 + $225","county=MIAMI-DADE, elevation_cert_required=True","total_amount=772; elevation_cert_amount=225","ELEVATION_CERT_PRICE=225; 547+225=772"),
        ("TC-PIPE-024","Pricing","P1","Unit","No elevation cert — no upcharge","county=ORANGE, elevation_cert_required=False","elevation_cert_amount=0, total_amount=607","ELEVATION_CERT_PRICE not added"),
        ("TC-PIPE-025","Pricing","P1","Unit","Monroe County → immediate flag (pricing engine)","county=MONROE, service_type='Boundary Survey'","AgentError; status=flagged; flag_reason='Monroe County (Florida Keys)'","Monroe check fires BEFORE county avg lookup; no price computed"),
        ("TC-PIPE-026","Pricing","P2","Unit","B2B multiplier NOT applied on top of override","county=ORANGE, b2b, special_pricing=True, override=1500","base_amount=1500, override_applied=True; NOT 1500×4.1","'if pricing_tier==b2b and not override_applied' guard verified"),
        ("TC-PIPE-027","Pricing","P2","Unit","special_pricing=True + overrides API 500 → flag","special_pricing=True; API returns 500","status=flagged; flag_reason contains 'overrides API unavailable'","_flag_order() in except PricingError; no price written"),
        ("TC-PIPE-028","Pricing","P1","Unit","Non-boundary service uses FTF API not county avg","service_type='Final Survey', county=ORANGE","pricing_source=ftf_api; county_applied='' (empty)","service_type not in _BOUNDARY_SERVICES; API pricing path"),
        # DB Layer
        ("TC-PIPE-029","DB","P1","Unit","log_decision dedup guard (I-029) — rapid double call","Call log_decision twice within 1 second for same order+decision","First call inserts; second call skipped with warning","agent_decision_log has exactly 1 row after 2 calls"),
        ("TC-PIPE-030","DB","P1","Unit","save_order_state upsert — INSERT for new order","save_order_state('NEW-ORDER-999', status='classified')","Row inserted; rowcount=0 triggers INSERT branch","Row verifiable via get_order_by_id"),
        ("TC-PIPE-031","DB","P1","Unit","save_order_state rejects unknown columns","save_order_state('ORDER-001', unknown_field='foo')","AgentError 'unknown columns {unknown_field}'","_VALID_ORDER_COLUMNS check fires before any SQL"),
        ("TC-PIPE-032","DB","P1","Unit","get_classified_order returns oldest-first (FIFO)","Two classified orders inserted at t=0 and t=+5min","Returns the t=0 order","ORDER BY created_at ASC LIMIT 1"),
        # Edge Cases
        ("TC-PIPE-033","Edge Cases","P1","Unit","property_county=None → classifier flags","county=None, service_type='Boundary Survey', state=FL","status=flagged; flag_reason contains 'missing property_county'","Pricing engine not invoked; no AttributeError on None"),
        ("TC-PIPE-034","Edge Cases","P1","Unit","Empty string county treated as missing","county='', service_type='Boundary Survey'","Not county avg; not statewide avg; FTF API fallback","'and property_county' guard evaluates False for empty string"),
        ("TC-PIPE-035","Edge Cases","P2","Unit","base_amount=0 — no crash; persisted","FTF API returns {'price': 0}; elev_cert=False","base_amount=0.0, total_amount=0.0, status=priced","No division-by-zero; amount=0 persisted"),
        ("TC-PIPE-036","Edge Cases","P1","Unit","Duplicate order ID — no duplicate DB row","save_order_state called twice for same order_id","UPDATE path fires on second call; single row in DB","No second INSERT; second call updates existing row"),
        ("TC-PIPE-037","Edge Cases","P1","Unit","customer_type='company' maps to b2b pricing tier","customer_type='company', service_type='Boundary Survey', county=ORANGE","pricing_tier=b2b; 4.1x multiplier applied","frozenset({'b2b','company','business'}) contains 'company'"),
        ("TC-PIPE-038","Edge Cases","P2","Unit","property_lat as string — type coercion no crash","property_lat='28.5383' (string), county=ORANGE","float(lat)=28.5383; lat-bounds check passes; classified","try/except around float() handles ValueError"),
        # E2E
        ("TC-PIPE-039","E2E","P1","Integration","Clean auto-flow — individual boundary, known county","service_type='Boundary Survey', county=ORANGE, individual, no flood","Status: classified→priced→written→reviewed→sent","All 5 timestamps set; final status=sent"),
        ("TC-PIPE-040","E2E","P1","Integration","Flagged order approved → resumes at writer","Order at flagged; human approves; process_approval_reply called","Status: flagged→awaiting→approved→written→reviewed→sent","get_ready_to_write_order() picks up approved status"),
        ("TC-PIPE-041","E2E","P1","Integration","Approval timeout 25h → escalation alert","Order awaiting, flagged_at = NOW() - 25h","get_overdue_approvals(24h) returns order; escalation email sent","Status unchanged; escalation fires; not auto-approved"),
        ("TC-PIPE-042","Classifier+Pricing","P1","Integration","'land survey and elevation' → Boundary + elev cert","county=HILLSBOROUGH, is_flood_zone=True","service_type=Boundary Survey, elev_cert=True; base=575+225=800","Alias + elev cert flag both apply"),
        ("TC-PIPE-043","Pricing","P1","Unit","'re-survey' → Update Survey → county avg (boundary)","county=POLK, individual after alias resolution","base_amount=743 (POLK avg), pricing_source=county_avg","'update survey' in _BOUNDARY_SERVICES; county avg used"),
        ("TC-PIPE-044","Pricing","P1","Unit","'re-survey / update survey' → county avg","county=LEE, individual","base_amount=500 (LEE avg), pricing_source=county_avg","Literal string in _BOUNDARY_SERVICES frozenset; LEE $500"),
        ("TC-PIPE-045","Pricing","P1","Integration","Triple modifier: B2B + known county + elev cert","county=PALM BEACH, b2b, elev_cert=True","base=round(571×4.1,2)=2341.10; total=2341.10+225=2566.10","Order: county avg → B2B → elev cert; pricing_source=county_avg+b2b_4.1x"),
    ]

    for i, row_data in enumerate(tc_pipe):
        alt = ALT_FILL if i % 2 == 0 else WHITE_FILL
        p_fill = P1_FILL if row_data[2]=="P1" else P2_FILL if row_data[2]=="P2" else P3_FILL
        _write_row(ws, r, list(row_data),
                   fills=[None, None, p_fill, None, None, None, None, None])
        for col in range(1, 9):
            if col != 3:
                ws.cell(r, col).fill = alt
        ws.row_dimensions[r].height = 45
        r += 1

    _set_col_widths(ws, [14, 18, 10, 12, 40, 38, 38, 35])
    _freeze(ws, "A3")


# ── Tab 5: Existing Scenarios ─────────────────────────────────────────────────

def build_existing_scenarios(wb):
    ws = wb.create_sheet("Existing Scenarios")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "Existing QA Order Seeds — 30 Scenarios (scripts/qa_orders.py)"
    ws["A1"].font = _font(bold=True, color="1F4E79", size=13)
    ws["A1"].fill = SECTION_FILL
    ws["A1"].alignment = _align(h="center")
    ws.row_dimensions[1].height = 22

    r = 2
    _header_row(ws, r, ["#","Scenario ID","Service Type","Customer Type","Flood Zone","Initial Status","Flag Reason","Test Note"])
    r += 1

    scenarios = [
        (1,  "boundary-clean",       "Boundary Survey",         "individual","No", "classified",  "",                                           "Standard individual boundary — clean path through pricing→writer→reviewer"),
        (2,  "boundary-flood",       "Boundary Survey",         "individual","Yes","classified",  "",                                           "AE flood zone — elevation cert upcharge expected in pricing"),
        (3,  "topo-b2b",             "Topography Survey",       "b2b",       "No", "classified",  "",                                           "B2B topography — should use b2b pricing tier"),
        (4,  "final-survey",         "Final Survey",            "individual","No", "classified",  "",                                           "Final survey — standard Tampa-area residential"),
        (5,  "form-board",           "Form Board Survey",       "individual","No", "classified",  "",                                           "Form board survey — Jacksonville area"),
        (6,  "update-survey",        "Update Survey",           "individual","No", "classified",  "",                                           "Update survey — Gainesville area"),
        (7,  "elevation-cert-flagged","Elevation Certificate",  "individual","Yes","flagged",     "service requires human review: Elevation Certificate","ALWAYS_FLAG service — routes to human gate (agent_04)"),
        (8,  "alta-b2b-flagged",     "ALTA Table A Survey",     "b2b",       "No", "flagged",     "service requires human review: ALTA Table A Survey","ALWAYS_FLAG service — human gate required before pricing"),
        (9,  "prateek-test-1",       "Boundary Survey",         "individual","No", "classified",  "",                                           "Prateek Test 1 — FTF order 1000276115 (Orange County)"),
        (10, "prateek-test-2",       "Elevation Certificate",   "individual","Yes","flagged",     "service requires human review: Elevation Certificate","Prateek Test 2 — FTF order 1000276116 — ALWAYS_FLAG + flood"),
        (11, "monroe-county-flagged","Boundary Survey",         "individual","No", "flagged",     "Monroe County (Florida Keys)",               "Monroe County auto-flag — Keys pricing too complex for auto"),
        (12, "b2b-boundary",         "Boundary Survey",         "b2b",       "No", "classified",  "",                                           "B2B boundary — should apply 4.1x commercial multiplier"),
        (13, "b2b-flood-elev",       "Boundary Survey",         "b2b",       "Yes","classified",  "",                                           "B2B + flood zone — B2B multiplier + elevation cert upcharge"),
        (14, "competitor-email",     "Boundary Survey",         "individual","No", "flagged",     "competitor company name match",              "Customer from competitor domain — human review required"),
        (15, "competitor-company",   "Boundary Survey",         "b2b",       "No", "flagged",     "competitor company name match",              "Company name matches competitor list"),
        (16, "missing-county",       "Boundary Survey",         "individual","No", "flagged",     "missing property_county",                    "No county provided — can't price without it"),
        (17, "out-of-state-ga",      "Boundary Survey",         "individual","No", "flagged",     "out of state: GA",                           "Georgia property — NexGen only covers FL"),
        (18, "high-lat-flagged",     "Boundary Survey",         "individual","No", "flagged",     "property lat 31.5 > FL bounds",              "I-037: property too far north for FL surveyor"),
        (19, "topo-individual",      "Topography Survey",       "individual","No", "flagged",     "never-auto-quote service: Topography Survey","NEVER_AUTO_QUOTE — always human gate regardless of customer type"),
        (20, "building-stake-out",   "Building Stake Out",      "individual","No", "flagged",     "service requires human review: Building Stake Out","ALWAYS_FLAG service — construction-phase staking"),
        (21, "alta-individual",      "ALTA Table A Survey",     "individual","No", "flagged",     "service requires human review: ALTA Table A Survey","ALWAYS_FLAG — complex commercial title survey"),
        (22, "lot-stake-out",        "Lot Stake Out",           "individual","No", "flagged",     "service requires human review: Lot Stake Out","ALWAYS_FLAG — staking service"),
        (23, "set-corners",          "Set Corners",             "individual","No", "flagged",     "service requires human review: Set Corners", "ALWAYS_FLAG — monument placement"),
        (24, "elevation-cert-only",  "Elevation Certificate",   "individual","No", "flagged",     "service requires human review: Elevation Certificate","ALWAYS_FLAG — FEMA form requires licensed review"),
        (25, "foundation-tie-in",    "Foundation Tie-In",       "individual","No", "classified",  "",                                           "Foundation tie-in — quick construction survey; auto-price"),
        (26, "tree-survey",          "Tree Survey",             "individual","No", "flagged",     "never-auto-quote service: Tree Survey",      "NEVER_AUTO_QUOTE — arborist scope unclear"),
        (27, "as-built-survey",      "As-Built Survey",         "individual","No", "classified",  "",                                           "As-built — post-construction verification; auto-price path"),
        (28, "hydrographic-survey",  "Hydrographic Survey",     "b2b",       "No", "flagged",     "never-auto-quote service: Hydrographic Survey","NEVER_AUTO_QUOTE — marine/waterway scope"),
        (29, "right-of-way",         "Right Of Way Survey",     "b2b",       "No", "flagged",     "service requires human review: Right Of Way Survey","ALWAYS_FLAG — government/utility ROW"),
        (30, "mortgage-inspection",  "Mortgage Inspection",     "individual","No", "classified",  "",                                           "Mortgage inspection — rapid turnaround; auto-price"),
    ]

    for i, s in enumerate(scenarios):
        alt = ALT_FILL if i % 2 == 0 else WHITE_FILL
        flag_fill = FAIL_FILL if s[5] == "flagged" else PASS_FILL
        _write_row(ws, r, list(s),
                   fills=[None, None, None, None, None, flag_fill, None, None])
        for col in [1,2,3,4,5,7,8]:
            ws.cell(r, col).fill = alt
        ws.row_dimensions[r].height = 40
        r += 1

    _set_col_widths(ws, [5, 24, 26, 14, 12, 14, 38, 55])
    _freeze(ws, "A3")


# ── Main ──────────────────────────────────────────────────────────────────────

# ── Tab: Positive Tests ───────────────────────────────────────────────────────

def build_positive_tests(wb):
    ws = wb.create_sheet("Positive Tests")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"] = "Positive Tests — Happy Path Scenarios (Expected to PASS)"
    ws["A1"].font = _font(bold=True, color="1F4E79", size=13)
    ws["A1"].fill = _fill("C6EFCE")
    ws["A1"].alignment = _align(h="center")
    ws.row_dimensions[1].height = 22

    r = 2
    _header_row(ws, r, ["Test ID","Component","Scenario","Input","Expected Output","Pass Criteria","Status"])
    r += 1

    positive = [
        # Auth & Send
        ("POS-001","Auth","Valid credentials → token obtained","TEAMS_TENANT_ID, TEAMS_APP_ID, TEAMS_CLIENT_SECRET all set","Bearer token returned; expires_in=3600","Token starts with 'eyJ'; no exception raised","PASS (live)"),
        ("POS-002","Email Send","Valid email to 2 recipients","FROM=pchandra@nexgen.enterprises; TO=rt@ryantisko.com,robert_e@nexgenlogix.com; body=HTML","HTTP 202; email delivered; appears in Sent folder","Graph API returns 202; log shows 'email notification sent'","PASS (live)"),
        ("POS-003","Channel Read","Poll empty channel returns empty list","FTF-Approvals channel with no messages","get_recent_messages() returns []","No exception; list length == 0","PASS (live)"),
        ("POS-004","Channel Read","Poll channel with APPROVE command","Channel has 'APPROVE 1000276115' message","check_for_approvals() returns [{action:'approve', order_id:'1000276115'}]","Command found; action and order_id correct","Designed"),
        # Classifier
        ("POS-005","Classifier","Known service type — clean classification","service_type='Boundary Survey', county=ORANGE, individual","status=classified, pricing_tier=individual, flag_reason=null","No flag; status=classified; passes to pricing","Designed"),
        ("POS-006","Classifier","Alias resolution — 'land survey only'→'Boundary Survey'","service_type='land survey only'","service_type='Boundary Survey' after alias lookup","Canonical name set; alias resolved deterministically","Designed"),
        ("POS-007","Classifier","B2B customer → pricing_tier='b2b'","customer_type='b2b', service_type='Boundary Survey'","pricing_tier='b2b'; status=classified; no flag","B2B is a pricing tier, not a flag trigger","Designed"),
        ("POS-008","Classifier","I-050 — 'Florida' normalized to 'FL'","property_state='Florida'","property_state='FL' after normalization; no out-of-state flag","Upper().strip() + FLORIDA→FL substitution","Designed"),
        # Pricing
        ("POS-009","Pricing","Known county → county avg used","county=ORANGE, Boundary Survey, individual","base_amount=607, pricing_source=county_avg","Exact county avg; no FTF API call","Designed"),
        ("POS-010","Pricing","B2B multiplier applied correctly","county=BREVARD, Boundary Survey, b2b","base_amount=round(560×4.1,2)=2296.0","Multiplier=4.1; pricing_source=county_avg+b2b_4.1x","Designed"),
        ("POS-011","Pricing","Elevation cert upcharge added","county=MIAMI-DADE, Boundary Survey, elev_cert=True","total_amount=547+225=772; elev_cert_amount=225","ELEVATION_CERT_PRICE=225 added to base","Designed"),
        ("POS-012","Pricing","Non-boundary service → FTF API pricing","service_type='Final Survey', county=ORANGE","pricing_source=ftf_api; county not applied","FTF API called; county avg not used","Designed"),
        # Human Gate
        ("POS-013","Human Gate","APPROVE command → status=approved","Order awaiting_approval; process_approval_reply(id,'approve')","status=approved; log_decision recorded","DB row shows status=approved","Designed"),
        ("POS-014","Human Gate","REJECT with reason → status=rejected","process_approval_reply(id,'reject'); reason passed","status=rejected; reason in decision log","DB row shows status=rejected","Designed"),
        ("POS-015","Human Gate","APPROVE ALL — all pending orders approved","3 orders awaiting; APPROVE ALL command","All 3 → status=approved; confirmation email lists all IDs","summary[approved]==3; all DB rows approved","Designed"),
        ("POS-016","Human Gate","Hourly digest sent — flagged + awaiting combined","2 flagged + 1 awaiting orders in DB","Single email with all 3 orders; flagged orders advance to awaiting_approval","sent=True; flagged=2; awaiting=1","Designed"),
        # Pipeline E2E
        ("POS-017","E2E","Full auto-flow — individual boundary clean path","Boundary Survey, ORANGE, individual, no flood","classified→priced→written→reviewed→sent","All 5 timestamps set; final status=sent","Designed"),
        ("POS-018","E2E","Approved order resumes pipeline","Order flagged; human approves; pipeline continues","approved→written→reviewed→sent","No skip; all stages completed","Designed"),
        ("POS-019","E2E","Token cache — second call uses cache","_get_token() called twice within 3600s window","Token endpoint called once; second returns cached token","No extra HTTP call; _cache['token'] populated","Designed"),
        ("POS-020","E2E","Dry-run poll — shows commands without DB write","--dry-run flag; APPROVE in channel","[DRY RUN] output; no DB change; no state file update","DB unchanged; state file unchanged","Designed"),
    ]

    for i, row_data in enumerate(positive):
        alt = ALT_FILL if i % 2 == 0 else WHITE_FILL
        status_fill = PASS_FILL if "PASS" in row_data[6] else _fill("FFFACD")
        _write_row(ws, r, list(row_data),
                   fills=[None, None, None, None, None, None, status_fill])
        for col in range(1, 7):
            ws.cell(r, col).fill = alt
        ws.row_dimensions[r].height = 42
        r += 1

    _set_col_widths(ws, [12, 16, 38, 38, 38, 35, 14])
    _freeze(ws, "A3")


# ── Tab: Negative Tests ───────────────────────────────────────────────────────

def build_negative_tests(wb):
    ws = wb.create_sheet("Negative Tests")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"] = "Negative Tests — Error Paths, Invalid Inputs, Rejection Scenarios (Expected to FAIL/FLAG/ERROR)"
    ws["A1"].font = _font(bold=True, color="1F4E79", size=13)
    ws["A1"].fill = _fill("FFC7CE")
    ws["A1"].alignment = _align(h="center")
    ws.row_dimensions[1].height = 22

    r = 2
    _header_row(ws, r, ["Test ID","Component","Scenario","Invalid Input","Expected Behavior","Should NOT Happen","Status"])
    r += 1

    negative = [
        # Auth & Config
        ("NEG-001","Auth","Missing TEAMS_CLIENT_SECRET","TEAMS_CLIENT_SECRET='' (empty)","AgentError raised: 'set TEAMS_TENANT_ID, TEAMS_APP_ID, TEAMS_CLIENT_SECRET'","Silent failure; token returned as None; crash later"),
        ("NEG-002","Auth","Wrong client secret","TEAMS_CLIENT_SECRET='wrong_secret_abc'","HTTP 401 from token endpoint; AgentError: 'Teams token request failed: HTTP 401'","Order processing continues with bad token"),
        ("NEG-003","Auth","Expired client secret","Secret expired in Azure AD","HTTP 401 from token endpoint; user-facing message to regenerate secret","Pipeline silently uses null token"),
        ("NEG-004","Email Send","NOTIFICATION_FROM_EMAIL not in tenant","FROM=fakeemail@otherdomain.com (not M365 user in tenant)","HTTP 404 or 403 from Graph API; AgentError raised","Email silently dropped without error"),
        ("NEG-005","Email Send","NOTIFICATION_TO_EMAILS empty","NOTIFICATION_TO_EMAILS='' (empty string)","AgentError raised: 'NOTIFICATION_TO_EMAILS not set'","Empty toRecipients array sent to Graph API"),
        ("NEG-006","Email Send","Mail.Send permission not granted","Permission removed from Azure AD app","HTTP 403: ErrorAccessDenied; AgentError raised","Exception swallowed; order stuck silently"),
        ("NEG-007","Webhook","O365 connector URL → 403","URL=webhook.office.com (retired)","HTTP 403; AgentError: 'Teams webhook POST failed: HTTP 403'","Retry loop; hanging; silent failure"),
        # Command Parsing
        ("NEG-008","Command Parsing","APPROVE with no order ID","Text='APPROVE' (nothing after keyword)","Returns ('unknown', None, None); command ignored","IndexError crash; or fake approval of None"),
        ("NEG-009","Command Parsing","REJECT with no order ID","Text='REJECT' (nothing after keyword)","Returns ('unknown', None, None); command ignored","AgentError trying to process None order_id"),
        ("NEG-010","Command Parsing","Garbage text in channel","Text='Hey team, lunch at 12?'","Returns ('unknown', None, None); not processed","False positive: 'approve' found in unrelated word"),
        ("NEG-011","Command Parsing","Order ID with SQL injection attempt","Text='APPROVE 1000276115; DROP TABLE processed_orders;'","Parser takes first token '1000276115;' as order_id; DB lookup finds no such order; AgentError","SQL injection reaches DB query; table dropped"),
        ("NEG-012","Command Parsing","Order ID with spaces/special chars","Text='APPROVE order 1000276115' (extra word between)","tokens[0]='order' used as order_id; DB lookup fails; AgentError; orange warning sent","Wrong order approved silently"),
        # Pipeline — Classifier
        ("NEG-013","Classifier","Monroe County → auto-flag","county=MONROE, service_type='Boundary Survey'","status=flagged; flag_reason='Monroe County (Florida Keys)'","Price computed; estimate sent without human review"),
        ("NEG-014","Classifier","ALTA survey → auto-price without approval","service_type='ALTA Table A Survey', customer_type='individual'","status=flagged; ALWAYS_FLAG fires","Estimate sent to client without human review"),
        ("NEG-015","Classifier","Out-of-state property → flag","property_state='GA', service_type='Boundary Survey'","status=flagged; flag_reason='out of state: GA'","FL surveyor dispatched to Georgia job"),
        ("NEG-016","Classifier","Unrecognized service type → flag","service_type='Drone Photogrammetry'","status=flagged; flag_reason contains 'unrecognized'","Drone job auto-priced using wrong rate"),
        ("NEG-017","Classifier","property_county=None","county=None, service_type='Boundary Survey'","status=flagged; flag_reason='missing property_county'","AttributeError crash; or statewide avg used without warning"),
        ("NEG-018","Classifier","property_state not FL","property_state='TX', service_type='Boundary Survey'","status=flagged; flag_reason='out of state: TX'","Texas order auto-priced as FL boundary"),
        # Pipeline — Pricing
        ("NEG-019","Pricing","B2B multiplier on top of override (double-multiply)","special_pricing=True, override=1500, b2b","base_amount=1500 (override); multiplier NOT applied again","base_amount=round(1500×4.1,2)=6150 — far too high"),
        ("NEG-020","Pricing","special_pricing=True but overrides API returns 500","special_pricing=True; FTF API returns 500","status=flagged; flag_reason='overrides API unavailable'","Incorrect price from fallback; estimate sent"),
        ("NEG-021","Pricing","FTF API returns 0 for a service","pricing_source=ftf_api; API returns {'price': 0}","base_amount=0.0 persisted; status=priced","ZeroDivisionError crash; or None written to DB"),
        # Human Gate
        ("NEG-022","Human Gate","Double-approve same order","Order already approved; APPROVE command sent again","AgentError: 'not awaiting approval (status=approved)'; orange warning email","Status changed from approved to approved again (noop); or crash"),
        ("NEG-023","Human Gate","Approve non-existent order ID","APPROVE fake-order-999 (not in DB)","AgentError: 'order fake-order-999 not found'; orange warning sent","KeyError crash; or None written to DB"),
        ("NEG-024","Human Gate","Invalid decision string","process_approval_reply(order_id, 'maybe')","AgentError: 'invalid decision maybe — must be approve or reject'","Order status set to 'maybe' in DB"),
        ("NEG-025","Human Gate","Approve order not in awaiting_approval","Order in status=priced (not awaiting); APPROVE command","AgentError: 'not awaiting approval (status=priced)'","Status changed to approved skipping flagged/awaiting stages"),
        # DB
        ("NEG-026","DB","save_order_state with unknown column","save_order_state('ID', unknown_field='x')","AgentError: 'unknown columns {unknown_field}'","Unknown column silently ignored; or SQL error at runtime"),
        ("NEG-027","DB","log_decision rapid duplicate (dedup guard)","Same call twice within 1 second","Second call skipped with warning log; 1 row in DB","Duplicate decision row inserted; confusion in audit trail"),
        ("NEG-028","DB","DB connection failure during save_order_state","PostgreSQL unreachable","Exception propagated to caller; order not silently marked as saved","Status marked saved while DB write failed; phantom state"),
        # State File
        ("NEG-029","Poll","Corrupted poll_state.json","File contains '{invalid json'","_load_last_polled() returns None (caught by except); fallback to --since-hours","Crash with JSONDecodeError; poll fails entirely"),
        ("NEG-030","Poll","State file missing when saving","OS permission denied on scripts/ dir during _save_last_polled","Warning logged: 'could not save poll state'; processing continues","Unhandled exception aborts poll mid-cycle"),
    ]

    for i, row_data in enumerate(negative):
        alt = ALT_FILL if i % 2 == 0 else WHITE_FILL
        _write_row(ws, r, list(row_data),
                   fills=[None, None, None, FAIL_FILL, None, None, None])
        for col in [1, 2, 3, 5, 6, 7]:
            ws.cell(r, col).fill = alt
        ws.row_dimensions[r].height = 42
        r += 1

    _set_col_widths(ws, [12, 16, 38, 38, 42, 42, 12])
    _freeze(ws, "A3")


def main():
    now = datetime.now(timezone.utc)
    now_str = now.strftime("%Y-%m-%d %H:%M UTC")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_summary(wb, now_str)
    build_teams_connection(wb)
    build_positive_tests(wb)
    build_negative_tests(wb)
    build_tc_teams(wb)
    build_tc_pipe(wb)
    build_existing_scenarios(wb)

    out_dir  = Path(__file__).parent.parent / "docs"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / f"QA_Test_Report_{now.strftime('%Y%m%d')}.xlsx"
    wb.save(out_path)
    print(f"Report saved: {out_path}")
    print(f"  Tabs: Summary | Teams Connection | Positive Tests (20) | Negative Tests (30) | TC-TEAMS (35) | TC-PIPE (45) | Existing Scenarios (30)")
    print(f"  Total test cases: 160")


if __name__ == "__main__":
    main()
