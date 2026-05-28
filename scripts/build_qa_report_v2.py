"""
build_qa_report_v2.py — Generate comprehensive QA Test Report v2 Excel workbook.

175 test cases across 10 tabs:
  TC-ORDER    (30)  — Order creation, pipeline routing, flag logic
  TC-APPROVAL (20)  — APPROVE command, DB updates, confirmations
  TC-REJECTION(15)  — REJECT command, reason storage, notifications
  TC-NEGATIVE (35)  — Edge cases, error handling, bad inputs
  TC-MONITOR  (15)  — Smart monitoring activation/deactivation
  TC-PRICING  (25)  — Pricing accuracy by lot/complexity factors
  TC-NOTIFY   (20)  — Notification delivery (Teams + email fallback)
  TC-E2E      (15)  — Full end-to-end scenarios

Output: docs/QA_Test_Report_v2_YYYYMMDD.xlsx

Usage:
    python scripts/build_qa_report_v2.py
"""

import os
import sys
from datetime import date

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

TODAY    = date.today().strftime("%Y%m%d")
OUT_FILE = os.path.join(os.path.dirname(__file__), "..", "docs", f"QA_Test_Report_v2_{TODAY}.xlsx")

# ── Styles ────────────────────────────────────────────────────────────────────

HDR_FILL  = PatternFill("solid", fgColor="1F3864")
SUBH_FILL = PatternFill("solid", fgColor="2E75B6")
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
SKIP_FILL = PatternFill("solid", fgColor="FFEB9C")
ALT_FILL  = PatternFill("solid", fgColor="EEF4FF")
SECT_FILL = PatternFill("solid", fgColor="BDD7EE")
NO_FILL   = PatternFill("none")

HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
BOLD_FONT = Font(bold=True, size=10)
NORM_FONT = Font(size=10)

CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(ws, headers, row=1, fill=HDR_FILL):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill  = fill
        cell.font  = HDR_FONT
        cell.alignment = CTR
        cell.border = _border()


def _row(ws, values, row, alt=True):
    fill = ALT_FILL if alt else NO_FILL
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill      = fill
        cell.font      = NORM_FONT
        cell.alignment = LFT
        cell.border    = _border()


def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _pf_drop(ws, col, first_row, last_row):
    dv = DataValidation(type="list", formula1='"PASS,FAIL,SKIP,PENDING"',
                        allow_blank=True, showDropDown=False)
    dv.sqref = f"{get_column_letter(col)}{first_row}:{get_column_letter(col)}{last_row}"
    ws.add_data_validation(dv)


def _freeze(ws, cell="B2"):
    ws.freeze_panes = cell


# ── Test Data ─────────────────────────────────────────────────────────────────

ORDER_HEADERS = [
    "TC_ID", "Scenario", "Customer Name", "Customer Email", "County",
    "Lot Size (ac)", "Service Type", "Pool", "Shed", "Patio",
    "Multi-Driveways", "Waterfront", "Corner", "Commercial", "Distance (mi)",
    "Exp Price Min ($)", "Exp Price Max ($)", "Should Flag",
    "Flag Reason", "Expected Pipeline Status",
    "Actual Status", "Pass/Fail", "Run Date", "Notes",
]

ORDER_CASES = [
    ("ORD-001","Standard residential 0.25ac","John Smith","john.smith@gmail.com","Broward",0.25,"Boundary Survey","N","N","N","N","N","N","N",5,300,400,"N","","priced","","PENDING","",""),
    ("ORD-002","Standard residential 0.5ac","Jane Doe","jane.doe@gmail.com","Broward",0.5,"Boundary Survey","N","N","N","N","N","N","N",5,400,500,"N","","priced","","PENDING","",""),
    ("ORD-003","Standard residential 1.0ac","Bob Johnson","bob.j@yahoo.com","Palm Beach",1.0,"Boundary Survey","N","N","N","N","N","N","N",5,550,700,"N","","priced","","PENDING","",""),
    ("ORD-004","Large lot 2.0ac","Alice Brown","alice.b@gmail.com","Orange",2.0,"Boundary Survey","N","N","N","N","N","N","N",5,700,950,"N","","priced","","PENDING","",""),
    ("ORD-005","Small lot 0.1ac","Mike Wilson","mike.w@hotmail.com","Broward",0.1,"Boundary Survey","N","N","N","N","N","N","N",5,250,350,"N","","priced","","PENDING","",""),
    ("ORD-006","0.5ac with pool","Sarah Davis","sarah.d@gmail.com","Miami-Dade",0.5,"Boundary Survey","Y","N","N","N","N","N","N",5,650,800,"N","","priced","","PENDING","",""),
    ("ORD-007","0.5ac with shed","Tom Miller","tom.m@yahoo.com","Broward",0.5,"Boundary Survey","N","Y","N","N","N","N","N",5,500,600,"N","","priced","","PENDING","",""),
    ("ORD-008","0.5ac with back patio","Linda Garcia","linda.g@gmail.com","Broward",0.5,"Boundary Survey","N","N","Y","N","N","N","N",5,500,600,"N","","priced","","PENDING","",""),
    ("ORD-009","0.5ac with multiple driveways","Carlos Lopez","carlos.l@gmail.com","Palm Beach",0.5,"Boundary Survey","N","N","N","Y","N","N","N",5,550,650,"N","","priced","","PENDING","",""),
    ("ORD-010","Full complexity (pool+shed+patio+driveways)","Nancy Lee","nancy.l@aol.com","Broward",0.5,"Boundary Survey","Y","Y","Y","Y","N","N","N",5,800,1100,"N","","priced","","PENDING","",""),
    ("ORD-011","Waterfront lot 0.5ac","Peter Kim","peter.k@gmail.com","Miami-Dade",0.5,"Boundary Survey","N","N","N","N","Y","N","N",5,500,650,"N","","priced","","PENDING","",""),
    ("ORD-012","Corner lot 0.5ac","Diana Chen","diana.c@yahoo.com","Broward",0.5,"Boundary Survey","N","N","N","N","N","Y","N",5,450,600,"N","","priced","","PENDING","",""),
    ("ORD-013","Elevation certificate (fixed $225)","Frank Reed","frank.r@gmail.com","Broward",0.25,"Elevation Certificate","N","N","N","N","N","N","N",5,225,225,"N","","priced","","PENDING","",""),
    ("ORD-014","Far location 30mi — upcharge","Helen Park","helen.p@gmail.com","Alachua",0.5,"Boundary Survey","N","N","N","N","N","N","N",30,550,700,"N","","priced","","PENDING","",""),
    ("ORD-015","Very far 75mi — remote flag","George Evans","george.e@aol.com","Marion",0.5,"Boundary Survey","N","N","N","N","N","N","N",75,700,950,"Y","remote_location","flagged","","PENDING","",""),
    ("ORD-016","Commercial lot — commercial flag","ABC Corp","info@abcbiz.com","Broward",0.5,"Boundary Survey","N","N","N","N","N","N","Y",5,1200,2000,"Y","commercial","flagged","","PENDING","",""),
    ("ORD-017","ALTA survey — alta_survey flag","XYZ Holdings","deals@xyzholdings.com","Miami-Dade",1.0,"ALTA/NSPS Survey","N","N","N","N","N","N","Y",5,2500,5000,"Y","alta_survey","flagged","","PENDING","",""),
    ("ORD-018","Topographic survey — high_value flag","James Turner","james.t@gmail.com","Orange",1.5,"Topographic Survey","N","N","N","N","N","N","N",10,1200,2000,"Y","high_value","flagged","","PENDING","",""),
    ("ORD-019","Competitor domain email","Rival Survey","admin@rival-survey.com","Broward",0.5,"Boundary Survey","N","N","N","N","N","N","N",5,400,500,"Y","competitor_domain","flagged","","PENDING","",""),
    ("ORD-020","Max complexity + max distance","Remote Rancher","ranch@cowboy.net","Jackson",2.0,"Boundary Survey","Y","Y","Y","Y","N","N","N",125,1500,2500,"Y","remote_high_complexity","flagged","","PENDING","",""),
    ("ORD-021","Repeat customer same service","John Smith","john.smith@gmail.com","Broward",0.5,"Boundary Survey","N","N","N","N","N","N","N",5,400,500,"N","","priced","","PENDING","",""),
    ("ORD-022","Multi-service (survey+elev cert)","Kate Walsh","kate.w@gmail.com","Broward",0.5,"Boundary Survey","N","N","N","N","N","N","N",5,625,750,"N","","priced","","PENDING","","Elev cert bundled"),
    ("ORD-023","Missing lot size — error expected","Error Test 1","test1@test.com","Broward",None,"Boundary Survey","N","N","N","N","N","N","N",5,None,None,"Y","missing_required_field","error","","PENDING","","lot_size=null"),
    ("ORD-024","Missing county — error expected","Error Test 2","test2@test.com",None,0.5,"Boundary Survey","N","N","N","N","N","N","N",5,None,None,"Y","missing_county","error","","PENDING","","county=null"),
    ("ORD-025","Invalid service type — error","Error Test 3","test3@test.com","Broward",0.5,"Unknown XYZ Service","N","N","N","N","N","N","N",5,None,None,"Y","invalid_service_type","error","","PENDING","",""),
    ("ORD-026","Out-of-state (Georgia) — flagged","Georgia Guy","ga.guy@gmail.com","Fulton (GA)",0.5,"Boundary Survey","N","N","N","N","N","N","N",5,None,None,"Y","out_of_state","flagged","","PENDING","","Non-FL address"),
    ("ORD-027","Subdivision plat — large job flag","Developer LLC","dev@developer.com","Broward",5.0,"Subdivision Plat","N","N","N","N","N","N","Y",10,8000,15000,"Y","subdivision_plat","flagged","","PENDING","",""),
    ("ORD-028","Refund request — blocked to Jessica","Unhappy Client","unhappy@gmail.com","Broward",None,"N/A - Refund Request","N","N","N","N","N","N","N",None,None,None,"Y","refund_request_jessica","blocked","","PENDING","","NEVER AI action"),
    ("ORD-029","Miami-Dade county modifier","Maria Santos","maria.s@gmail.com","Miami-Dade",0.25,"Boundary Survey","N","N","N","N","N","N","N",12,350,450,"N","","priced","","PENDING","",""),
    ("ORD-030","Waterfront+pool+patio combo","Luxury Owner","luxury@bighouse.com","Palm Beach",1.0,"Boundary Survey","Y","N","Y","N","Y","N","N",5,900,1300,"Y","high_value","flagged","","PENDING","","Multi-factor premium"),
]

APPROVAL_HEADERS = [
    "TC_ID", "Scenario", "Linked Order ID", "DB Prerequisite (status)",
    "Teams Command", "Sent By", "Expected DB Status After",
    "Expected Teams Confirmation", "Expected Next Action", "Auto/Manual",
    "Actual DB Status", "Actual Confirmation Received", "Pass/Fail", "Run Date", "Notes",
]

APPROVAL_CASES = [
    ("APP-001","Approve valid order - Robert","QA-ORD-001","awaiting_approval","APPROVE QA-ORD-001","Robert Engle","approved","[APPROVED] QA-ORD-001 — Estimate will be sent","Order picked up by Writer next cycle","Manual","","","PENDING","",""),
    ("APP-002","Approve valid order - Ryan","QA-ORD-002","awaiting_approval","APPROVE QA-ORD-002","Ryan Tisko","approved","[APPROVED] confirmation","Order advanced to Writer","Manual","","","PENDING","",""),
    ("APP-003","APPROVE ALL — 3 pending orders","QA-ORD-003,QA-ORD-004,QA-ORD-005","3x awaiting_approval","APPROVE ALL","Robert Engle","all 3 → approved","Bulk confirmation: 3 approved","All 3 advanced to Writer","Manual","","","PENDING","",""),
    ("APP-004","APPROVE ALL — 0 pending (graceful)","None","0 awaiting_approval","APPROVE ALL","Robert Engle","no change","No orders pending approval right now","None","Auto","","","PENDING","",""),
    ("APP-005","Approve — verify audit log created","QA-ORD-006","awaiting_approval","APPROVE QA-ORD-006","Ryan Tisko","approved","Confirmation sent","agent_decision_log row exists","Auto","","","PENDING","","Check DB directly"),
    ("APP-006","Approve already-approved (idempotent)","QA-ORD-007","approved","APPROVE QA-ORD-007","Robert Engle","approved (unchanged)","AgentError handled gracefully","No re-processing","Auto","","","PENDING","",""),
    ("APP-007","Approve non-existent order ID","QA-FAKE-999","not in DB","APPROVE QA-FAKE-999","Robert Engle","no change","AgentError: order not found","No action","Auto","","","PENDING","",""),
    ("APP-008","Approve with @mention prefix","QA-ORD-008","awaiting_approval","@FTFBot APPROVE QA-ORD-008","Ryan Tisko","approved","Confirmation sent","Order advanced","Manual","","","PENDING","","Teams auto-inserts @mention"),
    ("APP-009","Approve last pending → monitor goes idle","QA-ORD-009","awaiting_approval (last one)","APPROVE QA-ORD-009","Robert Engle","approved","Confirmation sent","Monitor switches to IDLE","Manual","","","PENDING","",""),
    ("APP-010","Approve one from batch of 3 (partial)","QA-ORD-010,QA-ORD-011,QA-ORD-012","3x awaiting_approval","APPROVE QA-ORD-010","Ryan Tisko","QA-ORD-010 approved; others unchanged","Single confirmation","1 advanced; 2 still pending","Manual","","","PENDING","",""),
    ("APP-011","Approve — case-insensitive command","QA-ORD-013","awaiting_approval","approve qa-ord-013","Robert Engle","approved","Confirmation sent","Order advanced","Auto","","","PENDING","","Lowercase should work"),
    ("APP-012","Approve order in wrong status (flagged)","QA-ORD-014","flagged","APPROVE QA-ORD-014","Robert Engle","flagged (unchanged)","AgentError: not awaiting_approval","No action","Auto","","","PENDING","",""),
    ("APP-013","Approve with extra trailing text","QA-ORD-015","awaiting_approval","APPROVE QA-ORD-015 looks correct to me","Ryan Tisko","approved","Confirmation sent","Order advanced","Auto","","","PENDING","","Extra text after ID ignored"),
    ("APP-014","Detect APPROVE within 60s latency","QA-ORD-016","awaiting_approval","APPROVE QA-ORD-016","Robert Engle","approved","Confirmation within 1 poll cycle","Latency ≤ 60s","Manual","","","PENDING","","Time the detection"),
    ("APP-015","APPROVE ALL — only awaiting not flagged","QA-ORD-017 (awaiting), QA-ORD-018 (flagged)","1x awaiting + 1x flagged","APPROVE ALL","Robert Engle","QA-ORD-017 approved; QA-ORD-018 still flagged","Confirms 1 approved","1 advanced; 1 unchanged","Auto","","","PENDING","",""),
    ("APP-016","Approve within 24h window","QA-ORD-019","awaiting_approval 22h old","APPROVE QA-ORD-019","Ryan Tisko","approved","Confirmation sent","Order advanced","Manual","","","PENDING","",""),
    ("APP-017","Approve after 25h (past timeout — escalated)","QA-ORD-020","awaiting_approval 25h old","APPROVE QA-ORD-020","Robert Engle","approved (late)","Confirmation; escalation already sent","Order advanced","Auto","","","PENDING","","Verify escalation was sent"),
    ("APP-018","DB updated immediately (not next cycle)","QA-ORD-021","awaiting_approval","APPROVE QA-ORD-021","Ryan Tisko","approved (immediate write)","Confirmation sent","DB status visible before next pipeline run","Manual","","","PENDING","",""),
    ("APP-019","Confirmation visible in FTF-Approvals channel","QA-ORD-022","awaiting_approval","APPROVE QA-ORD-022","Robert Engle","approved","[APPROVED] visible in Teams channel","Order advanced","Manual","","","PENDING","","Check Teams visually"),
    ("APP-020","Approved order picked up by Writer next cycle","QA-ORD-023","awaiting_approval","APPROVE QA-ORD-023","Ryan Tisko","approved → written (next cycle)","Confirmation sent","get_ready_to_write_order() picks up approved","Auto","","","PENDING","","Run pipeline cycle after"),
]

REJECTION_HEADERS = [
    "TC_ID", "Scenario", "Linked Order ID", "DB Prerequisite",
    "Teams Command", "Sent By", "Rejection Reason",
    "Expected DB Status", "Expected Notification", "Auto/Manual",
    "Actual DB Status", "Notification Received", "Pass/Fail", "Run Date", "Notes",
]

REJECTION_CASES = [
    ("REJ-001","Reject with reason - Robert","QA-REJ-001","awaiting_approval","REJECT QA-REJ-001 pricing seems off","Robert Engle","pricing seems off","rejected","[REJECTED] + reason in Teams","Manual","","","PENDING","",""),
    ("REJ-002","Reject with reason - Ryan","QA-REJ-002","awaiting_approval","REJECT QA-REJ-002 wrong service type","Ryan Tisko","wrong service type","rejected","[REJECTED] + reason in Teams","Manual","","","PENDING","",""),
    ("REJ-003","Reject without reason (empty)","QA-REJ-003","awaiting_approval","REJECT QA-REJ-003","Robert Engle","(empty)","rejected","[REJECTED] confirmation, reason blank","Auto","","","PENDING","",""),
    ("REJ-004","Reject with long multi-word reason","QA-REJ-004","awaiting_approval","REJECT QA-REJ-004 customer wants to switch from boundary to topo","Ryan Tisko","customer wants to switch from boundary to topo","rejected","Full reason stored in audit log","Auto","","","PENDING","",""),
    ("REJ-005","Reject non-existent order","QA-FAKE-888","not in DB","REJECT QA-FAKE-888 bad data","Robert Engle","bad data","no change","AgentError handled gracefully","Auto","","","PENDING","",""),
    ("REJ-006","Reject already-rejected (idempotent)","QA-REJ-006","rejected","REJECT QA-REJ-006 dup","Robert Engle","dup","rejected (unchanged)","AgentError or silent","Auto","","","PENDING","",""),
    ("REJ-007","Reject — verify audit log has reason","QA-REJ-007","awaiting_approval","REJECT QA-REJ-007 price too high for lot size","Ryan Tisko","price too high for lot size","rejected","agent_decision_log.reason matches input","Auto","","","PENDING","","Query DB to verify"),
    ("REJ-008","Reject last pending → monitor goes idle","QA-REJ-008","awaiting_approval (last)","REJECT QA-REJ-008 on hold pending crew schedule","Robert Engle","on hold pending crew schedule","rejected","Confirmation; monitor switches to IDLE","Manual","","","PENDING","",""),
    ("REJ-009","Reject confirmation visible in Teams","QA-REJ-009","awaiting_approval","REJECT QA-REJ-009 missing easement info","Ryan Tisko","missing easement info","rejected","[REJECTED] visible in FTF-Approvals channel","Manual","","","PENDING","","Check Teams visually"),
    ("REJ-010","Reject in wrong status (flagged, not awaiting)","QA-REJ-010","flagged","REJECT QA-REJ-010 skip it","Robert Engle","skip it","flagged (unchanged)","AgentError: not awaiting_approval","Auto","","","PENDING","",""),
    ("REJ-011","Reject — case-insensitive command","QA-REJ-011","awaiting_approval","reject qa-rej-011 bad address","Ryan Tisko","bad address","rejected","Confirmation sent","Auto","","","PENDING","",""),
    ("REJ-012","Reject with @mention prefix","QA-REJ-012","awaiting_approval","@FTFBot REJECT QA-REJ-012 no go","Robert Engle","no go","rejected","Confirmation sent","Manual","","","PENDING","",""),
    ("REJ-013","Detect REJECT within 60s latency","QA-REJ-013","awaiting_approval","REJECT QA-REJ-013 cost issue","Ryan Tisko","cost issue","rejected","Detected within 60s of typing","Manual","","","PENDING","","Time the detection"),
    ("REJ-014","REJECT ALL (unsupported — no crash)","Multiple awaiting","3x awaiting_approval","REJECT ALL","Robert Engle","N/A","no change","Not supported — no action, no crash","Auto","","","PENDING","",""),
    ("REJ-015","Reason preserved exactly in DB","QA-REJ-015","awaiting_approval","REJECT QA-REJ-015 surveyor notes conflict with county records","Robert Engle","surveyor notes conflict with county records","rejected","DB reason column matches input verbatim","Auto","","","PENDING","",""),
]

NEGATIVE_HEADERS = [
    "TC_ID", "Category", "Scenario", "Input / Action",
    "Expected System Behavior", "Severity", "Auto/Manual",
    "Actual Behavior", "Pass/Fail", "Run Date", "Notes",
]

NEGATIVE_CASES = [
    ("NEG-001","Command Parsing","Typo in APPROVE keyword","APPROV QA-001","Ignored — no action taken","Low","Auto","","PENDING","",""),
    ("NEG-002","Command Parsing","Typo in REJECT keyword","REJET QA-001 reason","Ignored — no action taken","Low","Auto","","PENDING","",""),
    ("NEG-003","Command Parsing","APPROVE with no order ID","APPROVE","Ignored — no order_id extracted","Low","Auto","","PENDING","",""),
    ("NEG-004","Command Parsing","REJECT with no order ID","REJECT","Ignored","Low","Auto","","PENDING","",""),
    ("NEG-005","Command Parsing","General chat message (no command)","hey when is the crew going out?","Ignored","Low","Manual","","PENDING","","Post in FTF-Approvals"),
    ("NEG-006","Command Parsing","Code block format in Teams","```APPROVE QA-001```","Parsed correctly (HTML stripped before _parse_command)","Medium","Auto","","PENDING","","Test _clean_message_body()"),
    ("NEG-007","Command Parsing","Emoji before command","[emoji] APPROVE QA-001","Parsed correctly","Low","Auto","","PENDING","",""),
    ("NEG-008","Command Parsing","Two IDs on one line","APPROVE QA-001 QA-002","Only first ID (QA-001) parsed; QA-002 ignored","Medium","Auto","","PENDING","","Current parser behaviour"),
    ("NEG-009","Command Parsing","All lowercase","approve qa-001","Handled (case-insensitive re.search)","Low","Auto","","PENDING","",""),
    ("NEG-010","Business Rules","Refund request order in pipeline","Order with 'refund' in description/service","Blocked — routed to Jessica only; AI takes no action","Critical","Auto","","PENDING","","Hard rule"),
    ("NEG-011","Business Rules","Competitor domain email","customer@rival-survey.com","Flagged as competitor_domain","High","Auto","","PENDING","",""),
    ("NEG-012","Business Rules","Out-of-state address","Address in Atlanta GA","Flagged or blocked — non-FL order","High","Auto","","PENDING","",""),
    ("NEG-013","Business Rules","APPROVE ALL with 0 pending","APPROVE ALL, no awaiting_approval orders","Graceful: '0 orders pending' message, no error","Medium","Auto","","PENDING","",""),
    ("NEG-014","Business Rules","REJECT ALL (unsupported)","REJECT ALL","Not supported — no action, no crash","Medium","Auto","","PENDING","","_parse_command returns unknown"),
    ("NEG-015","Business Rules","Programmatic auto-approve (no human)","process_approval_reply() called from prod code directly","Only valid in QA test mode; never from live pipeline","Critical","Auto","","PENDING","","Design guard needed"),
    ("NEG-016","Order Validation","Missing lot size","lot_size = null","Error status — pricing skipped","High","Auto","","PENDING","",""),
    ("NEG-017","Order Validation","Missing county","county = null","Error status — classification skipped","High","Auto","","PENDING","",""),
    ("NEG-018","Order Validation","Invalid service type","service_type = 'Unknown XYZ'","Error status — not classified","High","Auto","","PENDING","",""),
    ("NEG-019","Order Validation","Duplicate order (same addr+type+customer)","Submit identical order twice","Second flagged as duplicate","Medium","Auto","","PENDING","",""),
    ("NEG-020","Order Validation","Zero-dollar estimate","estimate_amount = 0","Flagged as suspicious_pricing","High","Auto","","PENDING","",""),
    ("NEG-021","Order Validation","Negative estimate amount","estimate_amount = -500","Error or flagged — never sent to customer","High","Auto","","PENDING","",""),
    ("NEG-022","Order Validation","HTML injection in customer name","<script>alert(1)</script> in name field","Sanitized before display in Teams/email","Critical","Auto","","PENDING","","Check send_channel_message output"),
    ("NEG-023","Order Validation","10k-char field value","customer_name = 10000 chars","Truncated or error — no crash, no Teams API failure","Medium","Auto","","PENDING","",""),
    ("NEG-024","System Resilience","Logic App webhook unreachable","TEAMS_INCOMING_WEBHOOK_URL = invalid URL","Falls back to email automatically; error logged","High","Auto","","PENDING","",""),
    ("NEG-025","System Resilience","Both webhook and email fail","Webhook down + Mail.Send returns 403","AgentError logged; order stays in awaiting; no data lost","Critical","Auto","","PENDING","",""),
    ("NEG-026","System Resilience","DB offline during approval processing","DB drops during process_approval_reply()","AgentError raised; no partial DB update; order stays awaiting","Critical","Auto","","PENDING","",""),
    ("NEG-027","System Resilience","Graph API rate limit (429) during poll","Rapid polls exceed rate limit","Backs off gracefully; logs 429; retries next cycle","High","Auto","","PENDING","",""),
    ("NEG-028","System Resilience","Graph API token expired mid-session","Token TTL elapses during active poll","_get_token() auto-refreshes on next call","Medium","Auto","","PENDING","","Token cache + 60s buffer"),
    ("NEG-029","Authorization","APPROVE from bot message","Application sender posts APPROVE in channel","Ignored — sender_is_app=True filtered in get_recent_messages","High","Auto","","PENDING","",""),
    ("NEG-030","Authorization","APPROVE from unknown user","APPROVE by user not in [Robert, Ryan]","Currently processed (no whitelist) — flag as gap","Medium","Manual","","PENDING","","Whitelist recommended"),
    ("NEG-031","Monitoring","Poll when 0 awaiting orders","No awaiting_approval rows in DB","Monitor does NOT call Teams API (idle path taken)","High","Auto","","PENDING","",""),
    ("NEG-032","Monitoring","Duplicate command re-processed","Same APPROVE message re-read on next poll cycle","Skipped — poll_state timestamp guard prevents re-processing","High","Auto","","PENDING","",""),
    ("NEG-033","Monitoring","Stale poll_state.json","State file timestamp from 7 days ago","Loads state; polls last 7 days; large fetch but no crash","Medium","Auto","","PENDING","",""),
    ("NEG-034","Monitoring","poll_state.json deleted mid-run","File deleted while monitor running","Falls back to default 2h lookback gracefully","Low","Auto","","PENDING","",""),
    ("NEG-035","Monitoring","DB error during awaiting count check","DB offline during _get_awaiting_count()","Returns 0; logs error; monitor takes idle path","High","Auto","","PENDING","",""),
]

MONITOR_HEADERS = [
    "TC_ID", "Scenario", "Awaiting Orders (before)", "Trigger Action",
    "Expected Monitor State", "Expected Poll Interval", "Expected Idle Interval",
    "Actual Monitor State", "Actual Interval", "Pass/Fail", "Run Date", "Notes",
]

MONITOR_CASES = [
    ("MON-001","No pending orders — no Teams API calls",0,"Monitor startup","IDLE","N/A","300s","","","PENDING","",""),
    ("MON-002","1 order enters awaiting — polling activates","0 → 1","Order status → awaiting_approval","ACTIVE","60s","N/A","","","PENDING","",""),
    ("MON-003","APPROVE detected within 60s of typing",1,"Human types APPROVE in Teams","ACTIVE — command detected","≤60s latency","N/A","","","PENDING","","Time detection"),
    ("MON-004","Last order approved — polling stops","1 → 0","APPROVE QA-XXX processed","IDLE","N/A","300s","","","PENDING","",""),
    ("MON-005","Multiple orders — stays active after 1 resolved",3,"APPROVE QA-001 (2 remain)","ACTIVE (2 still pending)","60s","N/A","","","PENDING","",""),
    ("MON-006","APPROVE ALL — all resolved — idle","3 → 0","APPROVE ALL processed","IDLE","N/A","300s","","","PENDING","",""),
    ("MON-007","Monitor restart — recovers from DB","N/A","Process restarted (kill & restart)","ACTIVE (if pending) or IDLE (per DB count)","Per DB count","Per DB count","","","PENDING","","State is in DB not file"),
    ("MON-008","No duplicate processing of same APPROVE","1","APPROVE typed; same message re-read next poll","Command processed once only","60s","N/A","","","PENDING","","poll_state timestamp guards this"),
    ("MON-009","Timeout — 24h order detected","1","Order 25h old in awaiting_approval","Monitor active; escalation triggered separately","60s","N/A","","","PENDING","","run_escalation_check()"),
    ("MON-010","Active poll interval timing",1,"Timed measurement over 3 cycles","60s ± 2s between polls","60 ± 2s","N/A","","","PENDING","",""),
    ("MON-011","Idle check interval timing",0,"Timed measurement over 2 idle cycles","300s ± 5s between DB checks","N/A","300 ± 5s","","","PENDING","",""),
    ("MON-012","loop_state DB row updated on each cycle","Any","Each poll/idle cycle","loop_state.last_run_at updated every cycle","N/A","N/A","","","PENDING","","Query loop_state table"),
    ("MON-013","Graph API error during poll — retry next cycle",1,"Teams API returns 503","Error logged; sleeps ACTIVE_POLL_SECS; retries","60s","N/A","","","PENDING","",""),
    ("MON-014","IDLE→ACTIVE transition logged","0 → 1","New order enters awaiting","Log: 'ACTIVE: N order(s) awaiting approval'","N/A","N/A","","","PENDING","","Check log output"),
    ("MON-015","ACTIVE→IDLE transition logged","1 → 0","Last order approved/rejected","Log: 'IDLE: no orders awaiting approval'","N/A","N/A","","","PENDING","","Check log output"),
]

PRICING_HEADERS = [
    "TC_ID", "Scenario", "Service Type", "Lot Size (ac)", "Pool", "Shed",
    "Patio", "Multi-Drive", "Waterfront", "Corner", "Commercial",
    "County", "Distance (mi)", "Exp Price Min ($)", "Exp Price Max ($)",
    "Tolerance (%)", "Actual Price ($)", "Within Tolerance", "Pass/Fail", "Notes",
]

PRICING_CASES = [
    ("PRC-001","Smallest standard lot","Boundary Survey",0.1,"N","N","N","N","N","N","N","Broward",5,250,350,15,"","","PENDING",""),
    ("PRC-002","Quarter acre basic","Boundary Survey",0.25,"N","N","N","N","N","N","N","Broward",5,300,400,15,"","","PENDING",""),
    ("PRC-003","Half acre basic","Boundary Survey",0.5,"N","N","N","N","N","N","N","Broward",5,400,500,15,"","","PENDING",""),
    ("PRC-004","1 acre basic","Boundary Survey",1.0,"N","N","N","N","N","N","N","Broward",5,550,700,15,"","","PENDING",""),
    ("PRC-005","2 acres basic","Boundary Survey",2.0,"N","N","N","N","N","N","N","Broward",5,700,950,15,"","","PENDING",""),
    ("PRC-006","Pool upcharge","Boundary Survey",0.5,"Y","N","N","N","N","N","N","Broward",5,650,800,15,"","","PENDING","vs PRC-003 baseline"),
    ("PRC-007","Shed upcharge","Boundary Survey",0.5,"N","Y","N","N","N","N","N","Broward",5,500,600,15,"","","PENDING",""),
    ("PRC-008","Patio upcharge","Boundary Survey",0.5,"N","N","Y","N","N","N","N","Broward",5,500,600,15,"","","PENDING",""),
    ("PRC-009","Multi-driveway upcharge","Boundary Survey",0.5,"N","N","N","Y","N","N","N","Broward",5,550,650,15,"","","PENDING",""),
    ("PRC-010","Full complexity (pool+shed+patio+drives)","Boundary Survey",0.5,"Y","Y","Y","Y","N","N","N","Broward",5,800,1100,15,"","","PENDING",""),
    ("PRC-011","Waterfront upcharge","Boundary Survey",0.5,"N","N","N","N","Y","N","N","Broward",5,500,650,15,"","","PENDING",""),
    ("PRC-012","Corner lot upcharge","Boundary Survey",0.5,"N","N","N","N","N","Y","N","Broward",5,450,600,15,"","","PENDING",""),
    ("PRC-013","Commercial multiplier (2-3x)","Boundary Survey",0.5,"N","N","N","N","N","N","Y","Broward",5,1200,2000,20,"","","PENDING",""),
    ("PRC-014","Elevation cert fixed price","Elevation Certificate",0.25,"N","N","N","N","N","N","N","Broward",5,225,225,0,"","","PENDING","Must be exactly $225"),
    ("PRC-015","ALTA survey premium","ALTA/NSPS Survey",1.0,"N","N","N","N","N","N","Y","Broward",5,2500,5000,20,"","","PENDING",""),
    ("PRC-016","Topographic survey","Topographic Survey",1.0,"N","N","N","N","N","N","N","Broward",5,1000,2000,20,"","","PENDING",""),
    ("PRC-017","As-built survey","As-Built Survey",0.5,"N","N","N","N","N","N","N","Broward",5,600,900,20,"","","PENDING",""),
    ("PRC-018","Far distance 30mi","Boundary Survey",0.5,"N","N","N","N","N","N","N","Alachua",30,550,700,15,"","","PENDING",""),
    ("PRC-019","Very far 75mi","Boundary Survey",0.5,"N","N","N","N","N","N","N","Marion",75,700,950,20,"","","PENDING",""),
    ("PRC-020","Remote 125mi + full complexity","Boundary Survey",2.0,"Y","Y","Y","Y","N","N","N","Jackson",125,1500,2500,20,"","","PENDING","Max scenario"),
    ("PRC-021","Miami-Dade county modifier","Boundary Survey",0.5,"N","N","N","N","N","N","N","Miami-Dade",12,450,600,15,"","","PENDING",""),
    ("PRC-022","Palm Beach county modifier","Boundary Survey",0.5,"N","N","N","N","N","N","N","Palm Beach",8,420,560,15,"","","PENDING",""),
    ("PRC-023","Orange county modifier","Boundary Survey",0.5,"N","N","N","N","N","N","N","Orange",20,430,580,15,"","","PENDING",""),
    ("PRC-024","Waterfront+pool+corner premium","Boundary Survey",1.0,"Y","N","N","N","Y","Y","N","Palm Beach",5,900,1300,20,"","","PENDING",""),
    ("PRC-025","Zero distance baseline","Boundary Survey",0.5,"N","N","N","N","N","N","N","Broward",0,350,450,15,"","","PENDING","Edge: dist=0"),
]

NOTIFY_HEADERS = [
    "TC_ID", "Scenario", "Trigger Event", "Expected Channel",
    "Primary Method", "Fallback Method", "Expected Content Keywords",
    "Latency Req", "Actual Method", "Content Match (Y/N)",
    "Latency (s)", "Pass/Fail", "Run Date", "Notes",
]

NOTIFY_CASES = [
    ("NOT-001","Order flagged → Teams alert","Order status → awaiting_approval","Teams FTF-Approvals","Logic App webhook","Graph API email","order_id, estimate_amount, service_type, FTF link, APPROVE/REJECT instructions","≤30s","","","","PENDING","",""),
    ("NOT-002","APPROVE → confirmation in Teams","APPROVE command detected","Teams FTF-Approvals","Logic App webhook","Email","[APPROVED], order_id, sender name","≤30s","","","","PENDING","",""),
    ("NOT-003","REJECT → confirmation in Teams","REJECT command detected","Teams FTF-Approvals","Logic App webhook","Email","[REJECTED], order_id, reason, sender name","≤30s","","","","PENDING","",""),
    ("NOT-004","24h timeout escalation","Order 25h in awaiting_approval","Teams + Email","Logic App + Mail.Send","None","ESCALATION, order_id, hours_overdue","≤30s","","","","PENDING","",""),
    ("NOT-005","Hourly batch digest","Hourly pipeline cycle","Email (Robert + Ryan)","Graph API Mail.Send","None","pending orders list, amounts, APPROVE/REJECT instructions","Per cycle","","","","PENDING","",""),
    ("NOT-006","Webhook fail → email fallback","Logic App returns non-2xx","Email (Robert + Ryan)","Graph API Mail.Send (fallback)","None","Same content as Teams notification","≤60s","","","","PENDING","",""),
    ("NOT-007","Email appears in Outlook sent folder","Email notification sent","Outlook Sent Items","Graph API Mail.Send","N/A","saveToSentItems=True in payload","≤30s","","","","PENDING","",""),
    ("NOT-008","Email sent to BOTH Robert AND Ryan","Any notification","Both inboxes","Graph API Mail.Send","N/A","Both addresses in toRecipients list","N/A","","","","PENDING","",""),
    ("NOT-009","Teams message contains FTF portal link","Order flagged notification","Teams FTF-Approvals","Logic App webhook","Email","fieldtofinish.jobs/admin/orders/{id}","N/A","","","","PENDING","",""),
    ("NOT-010","Teams message contains APPROVE/REJECT instructions","Order flagged notification","Teams FTF-Approvals","Logic App webhook","Email","APPROVE <order_id> OR REJECT <order_id> <reason>","N/A","","","","PENDING","",""),
    ("NOT-011","Email sender is pchandra@nexgen.enterprises","Any email notification","Email inbox","Graph API Mail.Send","N/A","From: pchandra@nexgen.enterprises","N/A","","","","PENDING","",""),
    ("NOT-012","No raw HTML tags visible in Teams message","Any Teams notification","Teams FTF-Approvals","Logic App webhook","N/A","Plain text only (HTML stripped before POST)","N/A","","","","PENDING","",""),
    ("NOT-013","Notification includes estimate amount","Order flagged notification","Teams + Email","Logic App + Mail.Send","N/A","$X,XXX.XX format","N/A","","","","PENDING","",""),
    ("NOT-014","No duplicate notifications same order","Order re-run without status change","Teams FTF-Approvals","Logic App webhook","N/A","Single notification per state change","N/A","","","","PENDING","",""),
    ("NOT-015","Confirmation includes approver name","APPROVE/REJECT processed","Teams FTF-Approvals","Logic App webhook","N/A","Robert Engle OR Ryan Tisko in message","N/A","","","","PENDING","",""),
    ("NOT-016","Special chars in name escaped","Customer name has <, >, & chars","Teams + Email","Both","N/A","No raw HTML special chars visible to recipient","N/A","","","","PENDING","",""),
    ("NOT-017","QA test orders labeled [QA TEST]","QA test order flagged","Teams FTF-Approvals","Logic App webhook","N/A","[QA TEST] prefix or marker in notification","N/A","","","","PENDING","","Design guidance"),
    ("NOT-018","Notification latency < 30s","Order → awaiting_approval","Teams FTF-Approvals","Logic App webhook","N/A","Message in Teams within 30s of DB update","≤30s","","","","PENDING","",""),
    ("NOT-019","Teams message within length limits","Any notification (long order list)","Teams FTF-Approvals","Logic App webhook","N/A","No truncation; Teams max ~28KB per message","N/A","","","","PENDING","",""),
    ("NOT-020","Email subject includes order ID","Any email notification","Email inbox","Graph API Mail.Send","N/A","Order ID visible in email subject line","N/A","","","","PENDING","",""),
]

E2E_HEADERS = [
    "TC_ID", "Scenario", "Preconditions",
    "Step 1", "Step 2", "Step 3", "Step 4", "Step 5",
    "Expected Final Status", "Est. Duration",
    "Actual Final Status", "Actual Duration (s)", "Pass/Fail", "Run Date", "Notes",
]

E2E_CASES = [
    ("E2E-001","Simple approve full flow","QA order in DB, Teams + Logic App connected","Set order to awaiting_approval","notify_human() sends Teams alert","Robert sees message, types APPROVE <id>","Monitor detects command (≤60s)","process_approval_reply() updates DB","approved","~5 min","","","PENDING","",""),
    ("E2E-002","Simple reject full flow","QA order in DB, Teams connected","Set order to awaiting_approval","notify_human() sends Teams alert","Ryan types REJECT <id> reason here","Monitor detects (≤60s)","DB status → rejected, reason stored","rejected","~5 min","","","PENDING","",""),
    ("E2E-003","Batch APPROVE ALL","3 QA orders in awaiting_approval","3 orders set to awaiting","Batch digest sent","Robert types APPROVE ALL","Monitor processes all 3","All 3 confirmed in Teams","all approved","~10 min","","","PENDING","",""),
    ("E2E-004","Mixed batch (1 approve, 2 reject)","3 QA orders awaiting","Robert types APPROVE QA-A","Ryan types REJECT QA-B reason","Robert types REJECT QA-C reason","Monitor processes all 3 commands","Verify all statuses in DB","1 approved, 2 rejected","~15 min","","","PENDING","",""),
    ("E2E-005","Competitor flag end-to-end","Order with @rival-survey.com email","Classifier flags competitor_domain","Order → flagged status","Teams alert sent","Robert reviews manually","No auto-advance — human decides","flagged (human decides)","~5 min","","","PENDING","",""),
    ("E2E-006","24h timeout escalation","Order in awaiting 25h (simulate via DB)","Order at 24h mark","run_escalation_check() runs","Escalation notification sent to Teams+email","Robert/Ryan notified urgently","Order still awaiting pending human action","awaiting_approval (escalated)","~25h (sim)","","","PENDING","",""),
    ("E2E-007","Smart monitor activation/deactivation","No pending orders initially","Monitor starts in idle","QA order added → awaiting_approval","Monitor DB check detects count > 0","Monitor starts polling every 60s","Order approved → monitor back to idle","approved → idle","~10 min","","","PENDING","",""),
    ("E2E-008","Full auto pipeline (no flag)","FTF order with no flag triggers","Order → pending → classified → priced","Writer generates estimate","Reviewer validates (4 checks)","Sender dispatches 8AM-6PM ET","No human gate needed","sent","~60 min","","","PENDING","","1 pipeline cycle"),
    ("E2E-009","Full pipeline with human gate","Order triggers commercial flag","Order → classified → flagged → awaiting_approval","Teams notification sent","Robert approves","Order → approved → written → reviewed → sent","Customer receives estimate","sent","~2h","","","PENDING","",""),
    ("E2E-010","Webhook fail → email fallback E2E","Logic App URL set to invalid","Order flagged → awaiting_approval","notify_human() tries Logic App → 4xx error","Auto-falls back to Graph API email","Robert/Ryan receive email notification","Manually type APPROVE/REJECT in Teams","awaiting_approval (notified via email)","~5 min","","","PENDING","",""),
    ("E2E-011","Refund request blocked","Order with refund keyword in description","Order enters pipeline","Refund detected by classifier","Routed to Jessica ONLY","AI takes no action","Jessica handles manually","blocked / jessica_notified","~5 min","","","PENDING","","Hard rule"),
    ("E2E-012","Command detection latency measurement","1 QA order awaiting, monitor active","Monitor running (60s cycle)","APPROVE typed at T=0s","Monitor polls at T≤60s","Command detected and processed","Confirmation sent to Teams","approved (within 60s)","~2 min","","","PENDING","","Stopwatch the detection"),
    ("E2E-013","Overnight batch — morning bulk review","10 orders in awaiting overnight","Morning: 10 orders awaiting","Robert types APPROVE ALL","Monitor processes all 10 in one loop","10 confirmations sent to Teams","All audit log entries created","all 10 approved","~5 min","","","PENDING","",""),
    ("E2E-014","Reject → order re-enters pipeline","Order rejected; crew updates FTF","Order rejected in DB","Crew updates order info in FTF portal","Order re-fetched as pending by Monitor agent","Re-classified → re-flagged → awaiting_approval","Second approval recorded in audit log","approved (second attempt)","~2h","","","PENDING","",""),
    ("E2E-015","Stress test — 20 simultaneous approvals","20 QA orders in awaiting_approval","APPROVE ALL typed once","Monitor loops through all 20","All 20 status → approved","20 confirmations sent to Teams","20 audit log entries in DB","all 20 approved","~10 min","","","PENDING","","Test loop stability"),
]


# ── Summary Data ──────────────────────────────────────────────────────────────

CATEGORY_SUMMARY = [
    ("TC-ORDER",    "Order creation, flag logic, pipeline routing", 30),
    ("TC-APPROVAL", "APPROVE command, DB updates, confirmations",   20),
    ("TC-REJECTION","REJECT command, reason storage, notifications", 15),
    ("TC-NEGATIVE", "Edge cases, error handling, bad inputs",        35),
    ("TC-MONITOR",  "Smart monitoring activation/deactivation",      15),
    ("TC-PRICING",  "Pricing accuracy by lot/complexity factors",    25),
    ("TC-NOTIFY",   "Notification delivery (Teams + email fallback)",20),
    ("TC-E2E",      "Full end-to-end scenarios",                     15),
]

TOTAL_TC = sum(r[2] for r in CATEGORY_SUMMARY)

GAPS_AND_RISKS = [
    ("CRITICAL", "No sender whitelist — any Teams user can APPROVE/REJECT", "Add APPROVED_SENDERS = ['Robert Engle', 'Ryan Tisko'] to poll logic"),
    ("CRITICAL", "APPROVE ALL has no safety confirmation", "Show count before executing: 'Approving N orders — proceeding'"),
    ("HIGH",     "run_escalation_check() not wired into orchestrator", "Call from agent_01_orchestrator.py after human gate step"),
    ("HIGH",     "poll_state.json lost on server restart", "Migrate last_processed_at to loop_state DB table"),
    ("MEDIUM",   "No QA test order isolation prefix", "Use 'QA-' order ID prefix; add cleanup function after tests"),
    ("MEDIUM",   "Audit log lacks Teams message ID", "Store cmd['message_id'] in agent_decision_log for full traceability"),
    ("LOW",      "Notification has no duplicate guard", "Track sent notification IDs to prevent double-alerts on re-runs"),
]


# ── Builder Functions ─────────────────────────────────────────────────────────

def build_summary(wb):
    ws = wb.create_sheet("Summary", 0)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 55
    ws.row_dimensions[1].height = 28

    # Title
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = f"FTF Agentic AI — QA Test Report v2  |  {date.today().isoformat()}  |  Total: {TOTAL_TC} test cases"
    t.font  = Font(bold=True, color="FFFFFF", size=13)
    t.fill  = HDR_FILL
    t.alignment = CTR

    ws.cell(row=3, column=1, value="Category").font = BOLD_FONT
    ws.cell(row=3, column=2, value="Description").font = BOLD_FONT
    ws.cell(row=3, column=3, value="Test Cases").font = BOLD_FONT
    for col in [1, 2, 3]:
        ws.cell(row=3, column=col).fill   = SUBH_FILL
        ws.cell(row=3, column=col).font   = HDR_FONT
        ws.cell(row=3, column=col).alignment = CTR

    for i, (cat, desc, cnt) in enumerate(CATEGORY_SUMMARY, 4):
        ws.cell(row=i, column=1, value=cat).font  = BOLD_FONT
        ws.cell(row=i, column=2, value=desc).font = NORM_FONT
        ws.cell(row=i, column=3, value=cnt).font  = NORM_FONT
        ws.cell(row=i, column=3).alignment = CTR
        fill = ALT_FILL if i % 2 == 0 else NO_FILL
        for col in [1, 2, 3]:
            ws.cell(row=i, column=col).fill   = fill
            ws.cell(row=i, column=col).border = _border()

    total_row = 4 + len(CATEGORY_SUMMARY)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, size=11)
    ws.cell(row=total_row, column=3, value=TOTAL_TC).font = Font(bold=True, size=11, color="1F3864")
    ws.cell(row=total_row, column=3).alignment = CTR

    # Gaps & Risks table
    gap_start = total_row + 3
    ws.merge_cells(f"A{gap_start}:D{gap_start}")
    g = ws.cell(row=gap_start, column=1, value="Identified Gaps & Risks")
    g.font = Font(bold=True, color="FFFFFF", size=11)
    g.fill = PatternFill("solid", fgColor="C00000")
    g.alignment = CTR

    ws.cell(row=gap_start+1, column=1, value="Severity").fill = HDR_FILL
    ws.cell(row=gap_start+1, column=1).font = HDR_FONT
    ws.cell(row=gap_start+1, column=2, value="Gap / Risk").fill = HDR_FILL
    ws.cell(row=gap_start+1, column=2).font = HDR_FONT
    ws.cell(row=gap_start+1, column=4, value="Recommended Fix").fill = HDR_FILL
    ws.cell(row=gap_start+1, column=4).font = HDR_FONT

    for j, (sev, gap, fix) in enumerate(GAPS_AND_RISKS, gap_start+2):
        sev_fill = FAIL_FILL if sev == "CRITICAL" else (
                   PatternFill("solid", fgColor="FFCC99") if sev == "HIGH" else
                   SKIP_FILL if sev == "MEDIUM" else ALT_FILL)
        ws.cell(row=j, column=1, value=sev).fill = sev_fill
        ws.cell(row=j, column=1).font = Font(bold=True, size=10)
        ws.cell(row=j, column=1).alignment = CTR
        ws.cell(row=j, column=2, value=gap).font = NORM_FONT
        ws.cell(row=j, column=2).alignment = LFT
        ws.merge_cells(f"C{j}:D{j}")
        c3 = ws.cell(row=j, column=3, value=fix)
        c3.font      = NORM_FONT
        c3.alignment = LFT

    # Key flow note
    note_row = gap_start + 2 + len(GAPS_AND_RISKS) + 2
    ws.merge_cells(f"A{note_row}:D{note_row}")
    note = ws.cell(row=note_row, column=1,
        value="POST-APPROVAL ROUTING: get_ready_to_write_order() picks up BOTH 'priced' AND 'approved' orders — "
              "approved orders automatically advance to Writer on next pipeline cycle. No gap.")
    note.font = Font(bold=True, color="375623", size=10)
    note.fill = PASS_FILL
    note.alignment = LFT


def _build_generic_sheet(wb, sheet_name, headers, cases, pf_col_index, widths):
    ws = wb.create_sheet(sheet_name)
    _hdr(ws, headers)
    _freeze(ws, f"B2")
    for i, row in enumerate(cases, 2):
        _row(ws, row, i, alt=(i % 2 == 0))
    _pf_drop(ws, pf_col_index, 2, 1 + len(cases))
    _widths(ws, widths)
    ws.row_dimensions[1].height = 30
    # Colour Pass/Fail cells that are already filled
    pf_col = get_column_letter(pf_col_index)
    for r in range(2, 2 + len(cases)):
        cell = ws[f"{pf_col}{r}"]
        if cell.value == "PASS":
            cell.fill = PASS_FILL
        elif cell.value == "FAIL":
            cell.fill = FAIL_FILL
        elif cell.value == "SKIP":
            cell.fill = SKIP_FILL
    return ws


def build_exec_log(wb):
    ws = wb.create_sheet("Execution Log")
    headers = [
        "Run ID", "Date / Time", "Runner", "TC Sheet", "TC_ID",
        "Result (PASS/FAIL/SKIP)", "Actual Output", "Executor", "Duration (s)", "Notes",
    ]
    _hdr(ws, headers)
    _freeze(ws)
    _pf_drop(ws, 6, 2, 200)
    _widths(ws, [10, 18, 20, 14, 14, 18, 45, 16, 12, 35])
    ws.row_dimensions[1].height = 28


def main():
    wb = Workbook()
    wb.remove(wb.active)

    build_summary(wb)

    _build_generic_sheet(wb, "TC-ORDER", ORDER_HEADERS, ORDER_CASES,
        pf_col_index=22,
        widths=[10,35,18,28,14,10,20,6,6,6,12,10,7,11,12,14,14,11,20,22,14,10,12,30])

    _build_generic_sheet(wb, "TC-APPROVAL", APPROVAL_HEADERS, APPROVAL_CASES,
        pf_col_index=13,
        widths=[10,38,20,22,38,14,22,35,28,12,18,28,10,12,30])

    _build_generic_sheet(wb, "TC-REJECTION", REJECTION_HEADERS, REJECTION_CASES,
        pf_col_index=13,
        widths=[10,38,14,22,45,14,32,16,30,12,16,25,10,12,30])

    _build_generic_sheet(wb, "TC-NEGATIVE", NEGATIVE_HEADERS, NEGATIVE_CASES,
        pf_col_index=9,
        widths=[10,20,38,38,42,10,12,30,10,12,30])

    _build_generic_sheet(wb, "TC-MONITOR", MONITOR_HEADERS, MONITOR_CASES,
        pf_col_index=10,
        widths=[10,40,16,30,22,18,18,22,16,10,12,30])

    _build_generic_sheet(wb, "TC-PRICING", PRICING_HEADERS, PRICING_CASES,
        pf_col_index=18,
        widths=[10,28,20,10,6,6,6,10,10,7,11,14,12,14,14,10,14,14,10,20])

    _build_generic_sheet(wb, "TC-NOTIFY", NOTIFY_HEADERS, NOTIFY_CASES,
        pf_col_index=12,
        widths=[10,35,30,22,20,20,42,12,16,14,12,10,12,30])

    _build_generic_sheet(wb, "TC-E2E", E2E_HEADERS, E2E_CASES,
        pf_col_index=13,
        widths=[10,32,28,30,30,30,30,30,22,12,22,14,10,12,30])

    build_exec_log(wb)

    os.makedirs(os.path.dirname(OUT_FILE), exist_ok=True)
    wb.save(OUT_FILE)
    print(f"Saved: {OUT_FILE}")
    print(f"Total test cases: {TOTAL_TC}")
    for cat, _, cnt in CATEGORY_SUMMARY:
        print(f"  {cat}: {cnt}")


if __name__ == "__main__":
    main()
